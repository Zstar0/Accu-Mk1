"""
FastAPI backend for Accu-Mk1.
Provides REST API for scientific calculations, database access, and audit logging.
"""

import asyncio
import json
import os
import re
import secrets
import subprocess
import sys
from contextlib import asynccontextmanager
from datetime import datetime, date, time, timezone
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Literal, Optional, Union
from uuid import UUID

# App version: prefer APP_VERSION env var (set by Docker build-arg),
# fall back to reading package.json (works in local dev).
def _read_app_version() -> str:
    if v := os.environ.get("APP_VERSION"):
        return v
    try:
        pkg = Path(__file__).resolve().parent.parent / "package.json"
        return json.loads(pkg.read_text())["version"]
    except Exception:
        return "0.0.0"

APP_VERSION = _read_app_version()

from fastapi import FastAPI, Body, Depends, Form, HTTPException, Header, Query, Request, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field, validator
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import select, desc, delete, update, func, extract
from sqlalchemy.exc import IntegrityError

from database import get_db, init_db
from sla_engine import BusinessSchedule, compute_business_minutes, sla_status_dict
from models import AuditLog, Settings, Job, Sample, Result, Instrument, AnalysisService, HplcMethod, Peptide, PeptideAnalyte, CalibrationCurve, HPLCAnalysis, User, SharePointFileCache, WizardSession, WizardMeasurement, peptide_methods, blend_components, ServiceGroup, service_group_members, SamplePriority, Worksheet, WorksheetItem, instrument_methods, SampleAnalyteAlias, SlaTier, SlaPriorityTier, BusinessHoursConfig, LabHoliday, LimsSample, LimsSubSample, LimsBox, FlagType
from auth import (
    get_current_user, require_admin, create_access_token,
    verify_password, get_password_hash, seed_admin_user,
    require_internal_service_token,
    UserCreate, UserRead, UserUpdate, MeUpdate, PasswordChange, TokenResponse,
    SenaiteCredentials,
)
from models_peptide_request import (
    PeptideRequestCreate, PeptideRequest, PeptideRequestList,
    PeptideRequestUpdate, PeptideRequestRetract,
    PeptideRequestSyncApplyRequest, StatusLogEntry,
)
from peptide_request_repo import PeptideRequestRepository
from status_log_repo import StatusLogRepository
from clickup_user_mapping_repo import ClickUpUserMappingRepository
from peptide_request_config import get_config as get_peptide_request_config
from clickup_client import ClickUpClient
from clickup_webhook import verify_signature, dispatch_event
from peptide_request_sync import (
    apply_actions as peptide_request_apply_sync_actions,
    compute_diff as peptide_request_compute_sync_diff,
)
from parsers import parse_txt_file
from parsers.peakdata_csv_parser import parse_hplc_files, calculate_purity
from calculations import CalculationEngine
from calculations.calibration import calculate_calibration_curve
from calculations.hplc_processor import (
    process_hplc_analysis, AnalysisInput, WeightInputs, CalibrationParams, PeptideParams
)
from file_watcher import FileWatcher
from sub_samples.routes import router as sub_samples_router
import sub_samples.service as sub_service
from sub_samples.service import derive_base_demand
from sub_samples import senaite
from sub_samples.registry_debug import diff_registry_vs_senaite
from lims_analyses.routes import router as lims_analyses_router
from families.routes import router as families_router  # Phase 5b
from boxes.routes import router as boxes_router
from boxes.service import box_label_code
from packaging_photos.routes import router as packaging_photos_router
from flags.routes import router as flags_router
from slack_notify.routes import router as slack_prefs_router
from slack_notify.interactions import router as slack_interactions_router
from workflow.routes import router as workflow_router

import logging

logger = logging.getLogger(__name__)


# --- API Key Configuration ---

# API key can be set via environment variable, or uses a default for development
# In production, set ACCU_MK1_API_KEY to a secure random value
API_KEY = os.environ.get("ACCU_MK1_API_KEY", "ak_dev_accumark_2024")


async def verify_api_key(x_api_key: Optional[str] = Header(None, alias="X-API-Key")):
    """
    Validate API key from X-API-Key header.
    Returns None if valid, raises HTTPException if invalid.
    """
    if not x_api_key:
        raise HTTPException(
            status_code=401,
            detail="API key required. Add your API key in Settings.",
            headers={"WWW-Authenticate": "API-Key"}
        )
    
    # Constant-time comparison to prevent timing attacks
    if not secrets.compare_digest(x_api_key, API_KEY):
        raise HTTPException(
            status_code=401,
            detail="Invalid API key. Check your API key in Settings.",
            headers={"WWW-Authenticate": "API-Key"}
        )
    
    return x_api_key



# --- Pydantic schemas ---

class HealthResponse(BaseModel):
    """Health check response."""
    status: str
    version: str


class AuditLogCreate(BaseModel):
    """Schema for creating an audit log entry."""
    operation: str
    entity_type: str
    entity_id: Optional[str] = None
    details: Optional[dict] = None


class AuditLogResponse(BaseModel):
    """Schema for audit log response."""
    id: int
    timestamp: datetime
    operation: str
    entity_type: str
    entity_id: Optional[str]
    details: Optional[dict]
    created_at: datetime

    class Config:
        from_attributes = True


class SettingUpdate(BaseModel):
    """Schema for updating a setting."""
    value: str


class SettingResponse(BaseModel):
    """Schema for setting response."""
    id: int
    key: str
    value: str
    updated_at: datetime

    class Config:
        from_attributes = True


# --- Import schemas ---

class ParsePreviewResponse(BaseModel):
    """Schema for file parse preview response."""
    filename: str
    headers: list[str]
    rows: list[dict[str, Union[str, int, float, None]]]
    row_count: int
    errors: list[str]


class BatchImportRequest(BaseModel):
    """Schema for batch import request."""
    file_paths: list[str]


class FileData(BaseModel):
    """Schema for file data from browser."""
    filename: str
    headers: list[str]
    rows: list[dict[str, Union[str, int, float, None]]]
    row_count: int


class BatchImportDataRequest(BaseModel):
    """Schema for batch import with pre-parsed data from browser."""
    files: list[FileData]


class SampleSummary(BaseModel):
    """Summary of a created sample."""
    id: int
    filename: str
    row_count: int


class ImportResultResponse(BaseModel):
    """Schema for batch import result response."""
    job_id: int
    samples_created: int
    samples: list[SampleSummary]
    errors: list[str]


class JobResponse(BaseModel):
    """Schema for job response."""
    id: int
    status: str
    source_directory: Optional[str]
    created_at: datetime
    completed_at: Optional[datetime]

    class Config:
        from_attributes = True


class SampleResponse(BaseModel):
    """Schema for sample response."""
    id: int
    job_id: int
    filename: str
    status: str
    input_data: Optional[dict]
    rejection_reason: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True


class RejectRequest(BaseModel):
    """Schema for sample rejection request."""
    reason: str


# --- Calculation schemas ---

class CalculationResultResponse(BaseModel):
    """Schema for a single calculation result."""
    calculation_type: str
    input_summary: dict
    output_values: dict
    warnings: list[str]
    success: bool
    error: Optional[str] = None


class CalculationSummaryResponse(BaseModel):
    """Schema for calculation summary response."""
    sample_id: int
    results: list[CalculationResultResponse]
    total_calculations: int
    successful: int
    failed: int


class CalculationPreviewRequest(BaseModel):
    """Schema for calculation preview request."""
    data: dict
    calculation_type: str


class ResultResponse(BaseModel):
    """Schema for stored result response."""
    id: int
    sample_id: int
    calculation_type: str
    input_data: Optional[dict]
    output_data: Optional[dict]
    created_at: datetime

    class Config:
        from_attributes = True


class SampleWithResultsResponse(BaseModel):
    """Schema for sample with calculated results flattened for UI."""
    id: int
    job_id: int
    filename: str
    status: str
    rejection_reason: Optional[str] = None
    created_at: datetime
    # Flattened calculation results for easy UI consumption
    purity: Optional[float] = None
    retention_time: Optional[float] = None
    compound_id: Optional[str] = None
    has_results: bool = False

    class Config:
        from_attributes = True


# --- Default settings ---

DEFAULT_SETTINGS = {
    "report_directory": "",
    "column_mappings": '{"peak_area": "Area", "retention_time": "RT", "compound_name": "Name"}',
    "compound_ranges": '{}',
    "calibration_slope": "1.0",
    "calibration_intercept": "0.0",
    "scale_host": "",      # Empty = scale disabled; set to IP address to enable
    "scale_port": "8001",  # Default MT-SICS TCP port for Excellence/XSR series
}


# --- App lifecycle ---

def seed_default_settings(db: Session):
    """Seed default settings if they don't exist."""
    for key, value in DEFAULT_SETTINGS.items():
        existing = db.execute(select(Settings).where(Settings.key == key)).scalar_one_or_none()
        if not existing:
            setting = Settings(key=key, value=value)
            db.add(setting)
    db.commit()


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Initialize database on startup and seed defaults."""
    init_db()
    from flags import seams as _flag_seams
    _flag_seams.register_mk1_entities()
    # Flag attachments reuse the S3 blob store used by vial photos when
    # configured (module purity: the adapter lives here, not in flags/).
    if os.environ.get("MK1_PHOTO_S3_BUCKET"):
        from sub_samples.photo_storage import S3PhotoStorage, PhotoNotFoundError

        class _S3FlagAttachmentStorage:
            def __init__(self):
                self._s3 = S3PhotoStorage(
                    prefix=os.environ.get("MK1_FLAG_ATTACH_S3_PREFIX", "flag-attachments/"))

            def save(self, flag_id, data, filename):
                return self._s3.save_photo(flag_id, data, filename)

            def fetch(self, key):
                try:
                    return self._s3.fetch_photo(key)
                except PhotoNotFoundError as e:
                    raise _flag_seams.AttachmentNotFound(str(e))

            def delete(self, key):
                self._s3.delete_photo(key)

        _flag_seams.set_attachment_storage(_S3FlagAttachmentStorage())
    import asyncio as _asyncio
    from flags import bus as _flag_bus
    _flag_bus.BUS.set_loop(_asyncio.get_running_loop())
    _flag_seams.set_event_sink(_flag_bus.SSEEventSink(_flag_bus.BUS))
    # Slack DM notifications (spec 2026-07-02) — dormant without the token.
    from slack_notify.notifier import maybe_start as _slack_maybe_start
    _slack_notifier_task = _slack_maybe_start(_flag_bus.BUS)
    # IS event-stream incremental sync (workflow state system, slice 3 Task 5)
    # — the only IS→Mk1 puller (spec §7); dormant without IS DB config or
    # when disabled via MK1_IS_EVENT_SYNC_ENABLED=0. Guarded: a startup issue
    # here must never take down the rest of the app.
    try:
        from workflow.is_event_stream import maybe_start as _is_sync_maybe_start
        _is_sync_task = _is_sync_maybe_start(app)
    except Exception:
        logger.warning("workflow.is_sync_start_failed", exc_info=True)
        _is_sync_task = None
    # Flag scheduler (Slice 5) — in-process ticker; jobs registered below.
    from datetime import timedelta as _timedelta
    from flags.scheduler import Scheduler as _Scheduler
    from database import SessionLocal as _SessionLocal
    _flag_scheduler = _Scheduler(_SessionLocal)
    # Job registration is appended by later Slice-5 tasks (recurring, digest, GC)
    # immediately BELOW, before start(). Registering zero jobs is harmless.
    from flags import recurring as _recurring

    def _recurring_job(now):
        db = _SessionLocal()
        try:
            _recurring.run_due(db, now=now)
        finally:
            db.close()
    _flag_scheduler.register("recurring_mint", interval=_timedelta(minutes=5),
                             fn=_recurring_job)
    # Morning digest — token-gated (needs the Slack client). Ticks every ~15 min;
    # digest.run dedupes to one DM per user per lab-local day.
    if os.getenv("MK1_SLACK_BOT_TOKEN"):
        from slack_notify.client import SlackClient as _SlackClient
        from slack_notify import digest as _digest
        _digest_base = os.getenv("MK1_PUBLIC_URL",
                                 "https://accumk1.valenceanalytical.com")

        async def _digest_job(now):
            await _digest.run(_SessionLocal,
                              _SlackClient(os.environ["MK1_SLACK_BOT_TOKEN"]),
                              _digest_base, now=now)
        _flag_scheduler.register("slack_digest", interval=_timedelta(minutes=15),
                                 fn=_digest_job)
    # Orphaned-attachment GC — always registered (no Slack env needed); hourly is
    # plenty for a 24h TTL. Lives in flags/ (zero Slack coupling).
    from flags import attachments_gc as _attachments_gc

    def _gc_job(now):
        db = _SessionLocal()
        try:
            _attachments_gc.gc_orphaned_attachments(db, now=now)
        finally:
            db.close()
    _flag_scheduler.register("attachment_gc", interval=_timedelta(hours=1),
                             fn=_gc_job)
    # State-change watches poller (Plan 6) — polls the host `state` seam every
    # ~2 min and fires armed watches once. Job fn takes `now` (the ticker calls
    # fn(now=now)); run_watch_poll owns its own Session via _watch_poll_job.
    from flags import watches as _flag_watches
    _flag_scheduler.register("flag_watch_poller", interval=_timedelta(minutes=2),
                             fn=_flag_watches._watch_poll_job)
    _flag_scheduler.start()
    # Seed default settings and admin user
    from database import SessionLocal
    db = SessionLocal()
    try:
        seed_default_settings(db)
        seed_admin_user(db)
    finally:
        db.close()

    # --- Scale Bridge (Phase 2) ---
    from scale_bridge import ScaleBridge, SCALE_PORT_DEFAULT
    import logging as _logging
    _logger = _logging.getLogger(__name__)

    scale_host = os.environ.get("SCALE_HOST")
    scale_port = int(os.environ.get("SCALE_PORT", str(SCALE_PORT_DEFAULT)))
    if scale_host:
        app.state.scale_bridge = ScaleBridge(host=scale_host, port=scale_port)
        await app.state.scale_bridge.start()
        _logger.info(f"ScaleBridge started: {scale_host}:{scale_port}")
    else:
        app.state.scale_bridge = None
        _logger.info("SCALE_HOST not set — scale bridge disabled (manual-entry mode)")

    yield

    # --- Scale Bridge shutdown ---
    if getattr(app.state, 'scale_bridge', None) is not None:
        await app.state.scale_bridge.stop()


# --- FastAPI app ---

app = FastAPI(
    title="Accu-Mk1 Backend",
    description="Backend API for lab purity calculations and SENAITE integration",
    version=APP_VERSION,
    lifespan=lifespan,
)

# CORS configuration for browser and Tauri frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:1420",      # Tauri dev server
        "http://127.0.0.1:1420",
        "http://localhost:5173",      # Vite default
        "http://127.0.0.1:5173",
        "http://localhost:3100",      # Docker local test (nginx)
        "http://127.0.0.1:3100",
        "http://localhost:3101",      # Docker local test (Vite dev fallback when 3100 is leased)
        "http://127.0.0.1:3101",
        "https://accumk1.valenceanalytical.com",  # Production
        "tauri://localhost",          # Tauri production (v1)
        "https://tauri.localhost",    # Tauri production (v2)
        "http://tauri.localhost",     # Tauri production fallback
    ],
    # accumark-stack platform mounts the frontend on a per-stack host port
    # (e.g. 5532 for subvial). Accept any localhost/127.0.0.1 dev port.
    allow_origin_regex=r"^http://(localhost|127\.0\.0\.1):\d+$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global file watcher instance
file_watcher = FileWatcher()

# Register sub-samples router
app.include_router(sub_samples_router)
app.include_router(lims_analyses_router)
app.include_router(families_router)
app.include_router(boxes_router)
app.include_router(packaging_photos_router)
app.include_router(flags_router)
app.include_router(slack_prefs_router)
app.include_router(slack_interactions_router)
app.include_router(workflow_router)

# --- Endpoints ---

@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Health check endpoint to verify backend is running."""
    return HealthResponse(status="ok", version=APP_VERSION)


# --- Auth Endpoints ---

@app.post("/auth/login", response_model=TokenResponse)
async def login(
    form_data: UserCreate,
    db: Session = Depends(get_db),
):
    """Authenticate user and return JWT access token."""
    user = db.query(User).filter(User.email == form_data.email).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Account is deactivated",
        )

    access_token = create_access_token(data={"sub": str(user.id)})
    return TokenResponse(
        access_token=access_token,
        user=_user_to_read(user),
    )


@app.get("/auth/me", response_model=UserRead)
async def get_me(current_user=Depends(get_current_user)):
    """Get current authenticated user info."""
    return _user_to_read(current_user)


@app.patch("/auth/me", response_model=UserRead)
async def update_me(
    data: MeUpdate,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Self-serve update of the caller's own name fields. Empty string clears
    to NULL. Cannot change role / active / email (not in MeUpdate)."""
    fields = data.model_dump(exclude_unset=True)
    if "first_name" in fields:
        v = (fields["first_name"] or "").strip()
        current_user.first_name = v or None
    if "last_name" in fields:
        v = (fields["last_name"] or "").strip()
        current_user.last_name = v or None
    db.commit()
    db.refresh(current_user)
    return _user_to_read(current_user)


@app.put("/auth/change-password")
async def change_password(
    data: PasswordChange,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Change current user's password (requires current password)."""
    if not verify_password(data.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    if len(data.new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    current_user.hashed_password = get_password_hash(data.new_password)
    db.commit()
    return {"message": "Password changed successfully"}


# --- Senaite Credentials ---

@app.put("/auth/senaite-credentials")
async def set_senaite_credentials(
    data: SenaiteCredentials,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Store the user's Senaite password (encrypted). Validates against Senaite first."""
    if not SENAITE_URL:
        raise HTTPException(400, "Senaite integration is not configured")
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=SENAITE_TIMEOUT) as client:
            auth_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/auth"
            print(f"[INFO] Validating Senaite credentials for {current_user.email}")
            resp = await client.get(
                auth_url,
                auth=httpx.BasicAuth(current_user.email, data.password),
            )
            print(f"[INFO] Senaite auth response: {resp.status_code}")
            if resp.status_code == 401:
                raise HTTPException(400, "Senaite authentication failed — check your password")
            resp.raise_for_status()
    except httpx.RequestError as e:
        raise HTTPException(502, f"Cannot reach Senaite: {e}")
    except httpx.HTTPStatusError as e:
        raise HTTPException(502, f"Senaite returned error: {e.response.status_code}")

    current_user.senaite_password_encrypted = _encrypt_senaite_password(data.password)
    db.commit()
    return {"message": "Senaite credentials saved", "senaite_configured": True}


@app.delete("/auth/senaite-credentials")
async def clear_senaite_credentials(
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Clear the user's stored Senaite password."""
    current_user.senaite_password_encrypted = None
    db.commit()
    return {"message": "Senaite credentials cleared", "senaite_configured": False}


# --- Admin User Management ---

@app.get("/auth/users", response_model=list[UserRead])
async def list_users(
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """List all users (admin only)."""
    users = db.query(User).order_by(User.created_at.desc()).all()
    return [_user_to_read(u) for u in users]


@app.get("/auth/directory")
async def user_directory(
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Lightweight id/email/name list for ALL users (active + inactive) so the
    FE can resolve historical analyst emails to names. Auth-only, not admin."""
    rows = db.execute(
        select(User.id, User.email, User.first_name, User.last_name)
        .order_by(User.email)
    ).all()
    return [
        {"id": r.id, "email": r.email, "first_name": r.first_name, "last_name": r.last_name}
        for r in rows
    ]


@app.post("/auth/users", response_model=UserRead)
async def create_user(
    data: UserCreate,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Create a new user (admin only)."""
    existing = db.query(User).filter(User.email == data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    if data.role not in ("standard", "admin"):
        raise HTTPException(status_code=400, detail="Role must be 'standard' or 'admin'")

    if len(data.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    user = User(
        email=data.email,
        hashed_password=get_password_hash(data.password),
        role=data.role,
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return _user_to_read(user)


@app.put("/auth/users/{user_id}", response_model=UserRead)
async def update_user(
    user_id: int,
    data: UserUpdate,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Update a user (admin only). Can change role and active status."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if data.role is not None:
        if data.role not in ("standard", "admin"):
            raise HTTPException(status_code=400, detail="Role must be 'standard' or 'admin'")
        user.role = data.role

    if data.is_active is not None:
        user.is_active = data.is_active

    if data.email is not None:
        existing = db.query(User).filter(User.email == data.email, User.id != user_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
        user.email = data.email

    if data.first_name is not None:
        user.first_name = data.first_name.strip() or None
    if data.last_name is not None:
        user.last_name = data.last_name.strip() or None

    db.commit()
    db.refresh(user)
    return _user_to_read(user)


@app.post("/auth/users/{user_id}/reset-password")
async def admin_reset_password(
    user_id: int,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Reset a user's password (admin only). Returns temporary password."""
    import secrets as _secrets

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    temp_password = _secrets.token_urlsafe(12)
    user.hashed_password = get_password_hash(temp_password)
    db.commit()

    print(f"\n[ADMIN RESET] Password reset for {user.email}: {temp_password}\n")

    return {
        "message": f"Password reset for {user.email}",
        "temporary_password": temp_password,
    }


@app.post("/audit", response_model=AuditLogResponse)
async def create_audit_log(
    audit_data: AuditLogCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Create a new audit log entry."""
    audit_log = AuditLog(
        operation=audit_data.operation,
        entity_type=audit_data.entity_type,
        entity_id=audit_data.entity_id,
        details=audit_data.details,
    )
    db.add(audit_log)
    db.commit()
    db.refresh(audit_log)
    return audit_log


@app.get("/audit", response_model=list[AuditLogResponse])
async def get_audit_logs(
    limit: int = 50,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get recent audit log entries."""
    stmt = select(AuditLog).order_by(desc(AuditLog.created_at)).limit(limit)
    result = db.execute(stmt)
    return result.scalars().all()


@app.get("/samples/{sample_id}/retest-info")
async def get_sample_retest_info(
    sample_id: str,
    _current_user=Depends(get_current_user),
):
    """
    Retest relationship metadata for a sample.

    Returns:
      - is_retest: this sample was created as a retest of another
      - source_sample_id, source_order_id, this_order_id, retest_created_at:
          populated when is_retest=True
      - retested_as: list of samples that retest THIS one (chain-forward)

    Reads from the integration-service Postgres directly. Cheap — bounded
    queries against indexed columns + JSONB lateral expansions.
    """
    from psycopg2.extras import RealDictCursor

    result = {
        "sample_id": sample_id,
        "is_retest": False,
        "source_sample_id": None,
        "source_order_id": None,
        "this_order_id": None,
        "retest_created_at": None,
        "retested_as": [],
    }

    try:
        with get_integration_db() as int_conn:
            with int_conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Is this sample itself a retest?
                cur.execute(
                    """
                    SELECT
                      os.order_id::text AS order_id,
                      os.created_at,
                      os.retest_of_order_id,
                      s.value->>'retest_of_senaite_id' AS source_sample_id
                    FROM order_submissions os,
                         jsonb_array_elements(os.payload->'samples') s
                    WHERE os.is_retest = TRUE
                      AND s.value->>'retest_of_senaite_id' IS NOT NULL
                      AND os.sample_results->(s.value->>'number')->>'senaite_id' = %s
                    LIMIT 1
                    """,
                    [sample_id],
                )
                row = cur.fetchone()
                if row:
                    result["is_retest"] = True
                    result["source_sample_id"] = row["source_sample_id"]
                    result["source_order_id"] = row["retest_of_order_id"]
                    result["this_order_id"] = int(row["order_id"]) if row["order_id"] else None
                    result["retest_created_at"] = (
                        row["created_at"].isoformat() if row["created_at"] else None
                    )

                # Forward-chain: samples that retest THIS one
                cur.execute(
                    """
                    SELECT
                      os.order_id::text AS order_id,
                      os.created_at,
                      os.sample_results->(s.value->>'number')->>'senaite_id' AS new_sample_id
                    FROM order_submissions os,
                         jsonb_array_elements(os.payload->'samples') s
                    WHERE os.is_retest = TRUE
                      AND s.value->>'retest_of_senaite_id' = %s
                      AND os.sample_results->(s.value->>'number')->>'senaite_id' IS NOT NULL
                    ORDER BY os.created_at
                    """,
                    [sample_id],
                )
                for r in cur.fetchall():
                    result["retested_as"].append({
                        "sample_id": r["new_sample_id"],
                        "order_id": int(r["order_id"]) if r["order_id"] else None,
                        "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                    })
    except Exception:
        # Integration DB unavailable — return the empty shell so callers can render gracefully
        pass

    return result


def _activity_bucket_label(kind, role):
    """Helper to map analysis bucket kind/role to display label."""
    if kind == "variance":
        return "Variance"
    if role in (None, "xtra"):
        return "Extra"
    return role


@app.get("/samples/{sample_id}/activity")
async def get_sample_activity(
    sample_id: str,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Federated activity timeline for a sample.

    Pulls events from multiple Mk1 tables + Integration DB and returns
    a unified, reverse-chronological activity stream.
    """
    events: list[dict] = []

    # --- Mk1 DB: wizard_sessions (prep started / completed) ---
    sessions = db.execute(
        select(WizardSession).where(WizardSession.sample_id_label == sample_id)
    ).scalars().all()
    for s in sessions:
        events.append({
            "timestamp": s.created_at.isoformat() if s.created_at else None,
            "event": "prep_started",
            "label": "Sample prep started",
            "details": {"session_id": s.id, "status": s.status},
            "source": "wizard_sessions",
        })
        if s.completed_at:
            events.append({
                "timestamp": s.completed_at.isoformat(),
                "event": "prep_completed",
                "label": "Sample prep completed",
                "details": {"session_id": s.id},
                "source": "wizard_sessions",
            })

    # --- Mk1 DB: sample_preps (prep records with user attribution) ---
    try:
        from mk1_db import ensure_sample_preps_table, get_mk1_db
        from psycopg2.extras import RealDictCursor
        ensure_sample_preps_table()
        with get_mk1_db() as mk1_conn:
            with mk1_conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, sample_id, senaite_sample_id, status, created_at, created_by_email "
                    "FROM sample_preps WHERE senaite_sample_id = %s ORDER BY created_at",
                    [sample_id],
                )
                for row in cur.fetchall():
                    events.append({
                        "timestamp": row["created_at"].isoformat() if row["created_at"] else None,
                        "event": "prep_record_created",
                        "label": "Prep record created",
                        "details": {
                            "prep_id": row["sample_id"],
                            "status": row["status"],
                            "by": row["created_by_email"],
                        },
                        "source": "sample_preps",
                    })
    except Exception:
        logger.warning("sample_preps unavailable for activity timeline", exc_info=True)

    # --- Mk1 DB: worksheet_items (added to worksheet) ---
    items = db.execute(
        select(WorksheetItem).where(WorksheetItem.sample_id == sample_id)
    ).scalars().all()
    for item in items:
        ws = db.execute(
            select(Worksheet).where(Worksheet.id == item.worksheet_id)
        ).scalar_one_or_none()
        analyst_email = None
        if item.assigned_analyst_id:
            analyst_user = db.execute(
                select(User).where(User.id == item.assigned_analyst_id)
            ).scalar_one_or_none()
            analyst_email = analyst_user.email if analyst_user else None
        created_by_email = None
        if ws and ws.created_by:
            ws_creator = db.execute(
                select(User).where(User.id == ws.created_by)
            ).scalar_one_or_none()
            created_by_email = ws_creator.email if ws_creator else None
        events.append({
            "timestamp": item.added_at.isoformat() if item.added_at else None,
            "event": "added_to_worksheet",
            "label": f"Added to worksheet {ws.title if ws else item.worksheet_id}",
            "details": {
                "worksheet_id": item.worksheet_id,
                "worksheet_title": ws.title if ws else None,
                "analyst": analyst_email,
                "created_by": created_by_email,
            },
            "source": "worksheet_items",
        })

    # --- Mk1 DB: hplc_analyses (results processed) ---
    analyses = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.sample_id_label == sample_id)
    ).scalars().all()
    for a in analyses:
        peptide = db.execute(
            select(Peptide).where(Peptide.id == a.peptide_id)
        ).scalar_one_or_none()
        events.append({
            "timestamp": a.created_at.isoformat() if a.created_at else None,
            "event": "hplc_analysis",
            "label": f"HPLC analysis — {peptide.abbreviation if peptide else 'unknown'}",
            "details": {
                "analysis_id": a.id,
                "peptide": peptide.abbreviation if peptide else None,
                "purity": a.purity_percent,
                "identity_conforms": a.identity_conforms,
                "processed_by": a.processed_by_email,
            },
            "source": "hplc_analyses",
        })

    # --- Mk1 DB: lims_analysis_promotions (analysis promoted from vials) ---
    from lims_analyses.service import list_promotions_for_parent
    for p in list_promotions_for_parent(db, sample_id):
        events.append({
            "timestamp": p.promoted_at.isoformat(),
            "event": "analysis_promoted",
            "label": (
                f"{p.keyword} promoted from "
                f"{', '.join(s.sample_id or '?' for s in p.sources)}"
            ),
            "details": {
                "keyword": p.keyword,
                "result_value": p.result_value,
                "by": p.promoted_by_email,
                "sources": [s.model_dump() for s in p.sources],
            },
            "source": "lims_analysis_promotions",
        })

    # --- Mk1 DB: variance replicate verifications (one event per vial) ---
    # Variance vials terminate in `variance_verified` (they feed the series, they
    # are never promoted), so without this they were invisible in the timeline.
    from lims_analyses.service import list_variance_verifications_for_parent
    for v in list_variance_verifications_for_parent(db, sample_id):
        n = v["count"]
        events.append({
            "timestamp": v["occurred_at"].isoformat() if v["occurred_at"] else None,
            "event": "variance_verified",
            "label": (
                f"Variance replicates verified — {v['vial_sample_id']} "
                f"({n} analys{'is' if n == 1 else 'es'})"
            ),
            "details": {
                "by": v["by_email"],
                "vial": v["vial_sample_id"],
                "count": n,
            },
            "source": "lims_analysis_transitions",
        })

    # --- Mk1 DB: variance-set lock/unlock (audit_logs, append-only) ---
    var_audits = db.execute(
        select(AuditLog).where(
            AuditLog.entity_type == "variance_set",
            AuditLog.entity_id == sample_id,
        ).order_by(AuditLog.timestamp)
    ).scalars().all()
    for a in var_audits:
        a_details = a.details or {}
        uid = a_details.get("user_id")
        by_email = None
        if uid:
            u = db.execute(select(User).where(User.id == uid)).scalar_one_or_none()
            by_email = u.email if u else None
        locked = a.operation == "variance_set_locked"
        n = a_details.get("selected_vials")
        events.append({
            "timestamp": a.timestamp.isoformat() if a.timestamp else None,
            "event": a.operation,
            "label": ("Variance set locked" if locked else "Variance set unlocked")
                     + (f" — {n} vials" if locked and n else ""),
            "details": {
                "by": by_email,
                "user_id": uid,
                **({"selected_vials": n} if locked and n is not None else {}),
            },
            "source": "audit_logs",
        })

    # --- Integration DB: sample_status_events ---
    try:
        with get_integration_db() as int_conn:
            with int_conn.cursor(cursor_factory=RealDictCursor) as cur:
                # --- Retest creation (this sample IS a retest) ---
                # Find the retest order whose sample_results includes this sample_id
                # AND whose payload identifies the source sample via retest_of_senaite_id.
                cur.execute(
                    """
                    SELECT
                      os.order_id::text AS order_id,
                      os.created_at,
                      s.value->>'retest_of_senaite_id' AS source_sample_id,
                      os.retest_of_order_id
                    FROM order_submissions os,
                         jsonb_array_elements(os.payload->'samples') s
                    WHERE os.is_retest = TRUE
                      AND s.value->>'retest_of_senaite_id' IS NOT NULL
                      AND os.sample_results->(s.value->>'number')->>'senaite_id' = %s
                    LIMIT 1
                    """,
                    [sample_id],
                )
                row = cur.fetchone()
                if row:
                    events.append({
                        "timestamp": row["created_at"].isoformat() if row["created_at"] else None,
                        "event": "retest_created",
                        "label": f"Retest of {row['source_sample_id']} created — order #{row['order_id']}",
                        "details": {
                            "source_sample_id": row["source_sample_id"],
                            "this_order_id": int(row["order_id"]) if row["order_id"] else None,
                            "source_order_id": row["retest_of_order_id"],
                        },
                        "source": "order_submissions",
                    })

                # --- Retested as (other samples that retest THIS one) ---
                cur.execute(
                    """
                    SELECT
                      os.order_id::text AS order_id,
                      os.created_at,
                      os.sample_results->(s.value->>'number')->>'senaite_id' AS new_sample_id
                    FROM order_submissions os,
                         jsonb_array_elements(os.payload->'samples') s
                    WHERE os.is_retest = TRUE
                      AND s.value->>'retest_of_senaite_id' = %s
                      AND os.sample_results->(s.value->>'number')->>'senaite_id' IS NOT NULL
                    ORDER BY os.created_at
                    """,
                    [sample_id],
                )
                for r in cur.fetchall():
                    events.append({
                        "timestamp": r["created_at"].isoformat() if r["created_at"] else None,
                        "event": "retested_as",
                        "label": f"Retested as {r['new_sample_id']} — order #{r['order_id']}",
                        "details": {
                            "new_sample_id": r["new_sample_id"],
                            "retest_order_id": int(r["order_id"]) if r["order_id"] else None,
                        },
                        "source": "order_submissions",
                    })

                cur.execute(
                    "SELECT transition, new_status, event_timestamp, wp_notified, created_at "
                    "FROM sample_status_events WHERE sample_id = %s ORDER BY created_at",
                    [sample_id],
                )
                for row in cur.fetchall():
                    ts = row["created_at"]
                    events.append({
                        "timestamp": ts.isoformat() if ts else None,
                        "event": "status_change",
                        "label": f"Status → {row['new_status']} ({row['transition']})",
                        "details": {
                            "transition": row["transition"],
                            "new_status": row["new_status"],
                            "wp_notified": row["wp_notified"],
                        },
                        "source": "sample_status_events",
                    })

                # --- Integration DB: coa_generations ---
                cur.execute(
                    "SELECT generation_number, verification_code, status, published_at, superseded_at, created_at "
                    "FROM coa_generations WHERE sample_id = %s ORDER BY created_at",
                    [sample_id],
                )
                for row in cur.fetchall():
                    vcode = row["verification_code"]
                    events.append({
                        "timestamp": row["created_at"].isoformat() if row["created_at"] else None,
                        "event": "coa_generated",
                        "label": f"COA v{row['generation_number']} generated",
                        "details": {
                            "generation_number": row["generation_number"],
                            "status": row["status"],
                            "verification_code": vcode,
                        },
                        "source": "coa_generations",
                    })
                    if row["published_at"]:
                        events.append({
                            "timestamp": row["published_at"].isoformat(),
                            "event": "coa_published",
                            "label": f"COA v{row['generation_number']} published",
                            "details": {
                                "generation_number": row["generation_number"],
                                "verification_code": vcode,
                            },
                            "source": "coa_generations",
                        })
                    if row["superseded_at"]:
                        events.append({
                            "timestamp": row["superseded_at"].isoformat(),
                            "event": "coa_superseded",
                            "label": f"COA v{row['generation_number']} superseded",
                            "details": {
                                "generation_number": row["generation_number"],
                                "verification_code": vcode,
                            },
                            "source": "coa_generations",
                        })
    except Exception:
        pass  # Integration DB unavailable — return Mk1 events only

    # --- Mk1 DB: sub-sample activity (Section A + B from spec) ---
    # Fan-out: when sample_id is a vial id, runs for that vial only (unchanged).
    # When sample_id is a parent id, runs Sections A+B for every family vial.
    from models import (
        LimsAnalysis,
        LimsAnalysisTransition,
        LimsAnalysisPromotion,
        LimsSubSampleEvent,
    )
    direct_sub = db.execute(
        select(LimsSubSample).where(LimsSubSample.sample_id == sample_id)
    ).scalar_one_or_none()
    if direct_sub is not None:
        family_subs = [direct_sub]
    else:
        parent = db.execute(
            select(LimsSample).where(LimsSample.sample_id == sample_id)
        ).scalar_one_or_none()
        family_subs = list(parent.sub_samples)[:64] if parent is not None else []
    for sub_row in family_subs:
        # Section A1: lims_analysis_transitions for analyses on this sub-sample
        # Join: lims_analyses → lims_analysis_transitions; email join for user.
        analyses_on_sub = db.execute(
            select(LimsAnalysis).where(
                LimsAnalysis.lims_sub_sample_pk == sub_row.id
            )
        ).scalars().all()
        analysis_ids = [a.id for a in analyses_on_sub]
        keyword_by_id = {a.id: a.keyword for a in analyses_on_sub}

        if analysis_ids:
            transitions = db.execute(
                select(LimsAnalysisTransition).where(
                    LimsAnalysisTransition.analysis_id.in_(analysis_ids)
                )
            ).scalars().all()
            for t in transitions:
                kw = keyword_by_id.get(t.analysis_id, "?")
                actor_email = None
                if t.user_id:
                    actor = db.execute(
                        select(User).where(User.id == t.user_id)
                    ).scalar_one_or_none()
                    actor_email = actor.email if actor else None

                if t.from_state is None and t.reason == "initial insert":
                    # Seeded / manually added
                    label = f"Analysis added: {kw}"
                    event_name = "analysis_added"
                    details: dict = {"keyword": kw, "by": actor_email, "vial": sub_row.sample_id}
                else:
                    label = f"{kw}: {t.from_state}→{t.to_state}"
                    event_name = "analysis_transition"
                    details = {
                        "keyword": kw,
                        "from": t.from_state,
                        "to": t.to_state,
                        "kind": t.transition_kind,
                        "reason": t.reason,
                        "by": actor_email,
                        "vial": sub_row.sample_id,
                    }
                events.append({
                    "timestamp": t.occurred_at.isoformat() if t.occurred_at else None,
                    "event": event_name,
                    "label": label,
                    "details": details,
                    "source": "lims_analysis_transitions",
                })

            # Section A2: lims_analysis_promotions (vial side — source_analysis_id)
            promotions = db.execute(
                select(LimsAnalysisPromotion).where(
                    LimsAnalysisPromotion.source_analysis_id.in_(analysis_ids)
                )
            ).scalars().all()
            for p in promotions:
                src_kw = keyword_by_id.get(p.source_analysis_id, "?")
                promoter_email = None
                if p.promoted_by_user_id:
                    promoter = db.execute(
                        select(User).where(User.id == p.promoted_by_user_id)
                    ).scalar_one_or_none()
                    promoter_email = promoter.email if promoter else None
                events.append({
                    "timestamp": p.promoted_at.isoformat() if p.promoted_at else None,
                    "event": "analysis_promoted_to_parent",
                    "label": f"Promoted {src_kw} to parent",
                    "details": {
                        "keyword": src_kw,
                        "parent_analysis_id": p.parent_analysis_id,
                        "contribution_kind": p.contribution_kind,
                        "by": promoter_email,
                        "vial": sub_row.sample_id,
                    },
                    "source": "lims_analysis_promotions",
                })

        # Section B: lims_sub_sample_events (role changes, remarks, analysis_removed)
        sub_events = db.execute(
            select(LimsSubSampleEvent).where(
                LimsSubSampleEvent.sub_sample_pk == sub_row.id
            )
        ).scalars().all()
        for se in sub_events:
            actor_email = None
            if se.user_id:
                actor = db.execute(
                    select(User).where(User.id == se.user_id)
                ).scalar_one_or_none()
                actor_email = actor.email if actor else None

            if se.event == "role_assigned":
                d = se.details or {}
                if d.get("kind_from") != d.get("kind_to"):
                    label = f"Bucket: {_activity_bucket_label(d.get('kind_from'), d.get('from'))} → {_activity_bucket_label(d.get('kind_to'), d.get('to'))}"
                else:
                    label = f"Role: {d.get('from')} → {d.get('to')}"
            elif se.event == "remarks_updated":
                label = "Remarks updated"
            elif se.event == "analysis_removed":
                d = se.details or {}
                label = f"Analysis removed: {d.get('keyword', '?')}"
            elif se.event == "worksheet_assigned":
                d = se.details or {}
                ws_label = d.get("worksheet_title") or f"#{d.get('worksheet_id')}"
                analyst = d.get("analyst_email") or "unassigned"
                label = f"Added to worksheet {ws_label} — analyst {analyst}"
            elif se.event == "worksheet_removed":
                d = se.details or {}
                ws_label = d.get("worksheet_title") or f"#{d.get('worksheet_id')}"
                label = f"Removed from worksheet {ws_label}"
            elif se.event == "worksheet_analyst_changed":
                d = se.details or {}
                label = (
                    f"Worksheet analyst: {d.get('from_email') or '—'} → "
                    f"{d.get('to_email') or '—'}"
                )
            elif se.event == "box_assigned":
                d = se.details or {}
                label = f"Boxed: {d.get('box_label') or '?'}"
            elif se.event == "box_moved":
                d = se.details or {}
                label = f"Box: {d.get('from_box_label') or '?'} → {d.get('to_box_label') or '?'}"
            elif se.event == "box_removed":
                d = se.details or {}
                reason = d.get("reason")
                bl = d.get("box_label") or "?"
                if reason == "stored":
                    label = f"Box stored: {bl}"
                elif reason == "box_deleted":
                    label = f"Removed from box {bl} (box deleted)"
                else:
                    label = f"Unboxed from {bl}"
            else:
                label = se.event

            event_details = dict(se.details or {})
            event_details["by"] = actor_email
            event_details["vial"] = sub_row.sample_id
            events.append({
                "timestamp": se.created_at.isoformat() if se.created_at else None,
                "event": se.event,
                "label": label,
                "details": event_details,
                "source": "lims_sub_sample_events",
            })

    # Sort all events reverse-chronological, nulls last
    events.sort(key=lambda e: e["timestamp"] or "", reverse=True)

    return {"sample_id": sample_id, "events": events, "count": len(events)}


# --- Settings Endpoints ---

# Keys only admins may write/delete. The UI hides these toggles from
# non-admins, but the gate must live server-side too.
ADMIN_ONLY_SETTING_KEYS = {"checkin_multi_order_enabled"}


@app.get("/settings", response_model=list[SettingResponse])
async def get_settings(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get all settings."""
    stmt = select(Settings).order_by(Settings.key)
    result = db.execute(stmt)
    return result.scalars().all()


@app.get("/settings/{key}", response_model=SettingResponse)
async def get_setting(key: str, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get a single setting by key."""
    stmt = select(Settings).where(Settings.key == key)
    setting = db.execute(stmt).scalar_one_or_none()
    if not setting:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")
    return setting


@app.put("/settings/{key}", response_model=SettingResponse)
async def update_setting(
    key: str,
    data: SettingUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Create or update a setting by key. Admin-only keys require an admin caller."""
    if key in ADMIN_ONLY_SETTING_KEYS and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="admin only")
    stmt = select(Settings).where(Settings.key == key)
    setting = db.execute(stmt).scalar_one_or_none()

    if setting:
        # Update existing
        setting.value = data.value
    else:
        # Create new
        setting = Settings(key=key, value=data.value)
        db.add(setting)

    db.commit()
    db.refresh(setting)
    return setting


@app.delete("/settings/{key}")
async def delete_setting(key: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Delete a setting by key. Admin-only keys require an admin caller."""
    if key in ADMIN_ONLY_SETTING_KEYS and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="admin only")
    stmt = select(Settings).where(Settings.key == key)
    setting = db.execute(stmt).scalar_one_or_none()
    if not setting:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")

    db.delete(setting)
    db.commit()
    return {"message": f"Setting '{key}' deleted"}


# --- File Watcher Endpoints ---

@app.get("/watcher/status")
async def get_watcher_status(_current_user=Depends(get_current_user)):
    """Get file watcher status."""
    return file_watcher.status()


@app.post("/watcher/start")
async def start_watcher(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Start file watcher using report_directory from settings."""
    # Get report_directory from settings
    setting = db.execute(
        select(Settings).where(Settings.key == "report_directory")
    ).scalar_one_or_none()
    if not setting or not setting.value:
        raise HTTPException(400, "report_directory not configured")

    if not os.path.isdir(setting.value):
        raise HTTPException(400, f"Directory does not exist: {setting.value}")

    file_watcher.start(setting.value)
    return {"status": "started", "watching": setting.value}


@app.post("/watcher/stop")
async def stop_watcher(_current_user=Depends(get_current_user)):
    """Stop file watcher."""
    file_watcher.stop()
    return {"status": "stopped"}


@app.get("/watcher/files")
async def get_detected_files(_current_user=Depends(get_current_user)):
    """Get and clear list of detected files."""
    files = file_watcher.get_detected_files()
    return {"files": files, "count": len(files)}


# --- Import Endpoints ---

def _get_column_mappings(db: Session) -> dict:
    """Get column mappings from settings."""
    stmt = select(Settings).where(Settings.key == "column_mappings")
    setting = db.execute(stmt).scalar_one_or_none()
    if setting and setting.value:
        try:
            return json.loads(setting.value)
        except json.JSONDecodeError:
            return {}
    return {}


@app.post("/import/file", response_model=ParsePreviewResponse)
async def import_file_preview(
    file_path: str,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Preview a single file parse without saving.

    Returns the first rows of parsed data for user review.
    """
    column_mappings = _get_column_mappings(db)
    result = parse_txt_file(file_path, column_mappings)

    # Limit preview to first 10 rows
    preview_rows = result.rows[:10]

    return ParsePreviewResponse(
        filename=result.filename,
        headers=result.raw_headers,
        rows=preview_rows,
        row_count=result.row_count,
        errors=result.errors,
    )


@app.post("/import/batch", response_model=ImportResultResponse)
async def import_batch(
    request: BatchImportRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Import multiple files and create a job with samples.

    Creates:
    - One Job record for the batch
    - One Sample record per file with parsed data stored in input_data
    """
    errors: list[str] = []
    samples: list[SampleSummary] = []
    column_mappings = _get_column_mappings(db)

    # Determine source directory from first file
    source_dir = None
    if request.file_paths:
        first_path = Path(request.file_paths[0])
        source_dir = str(first_path.parent) if first_path.parent else None

    # Create Job
    job = Job(
        status="pending",
        source_directory=source_dir,
    )
    db.add(job)
    db.flush()  # Get job.id without committing

    # Create audit log for job creation
    audit_log = AuditLog(
        operation="create",
        entity_type="job",
        entity_id=str(job.id),
        details={"file_count": len(request.file_paths)},
    )
    db.add(audit_log)

    # Process each file
    for file_path in request.file_paths:
        result = parse_txt_file(file_path, column_mappings)

        if result.errors:
            # Include file-specific errors in response
            for error in result.errors:
                errors.append(f"{result.filename}: {error}")

        # Create Sample with parsed data
        sample = Sample(
            job_id=job.id,
            filename=result.filename,
            status="pending" if not result.errors else "error",
            input_data={
                "rows": result.rows,
                "headers": result.raw_headers,
                "row_count": result.row_count,
            },
        )
        db.add(sample)
        db.flush()

        samples.append(SampleSummary(
            id=sample.id,
            filename=result.filename,
            row_count=result.row_count,
        ))

        # Create audit log for sample creation
        sample_audit = AuditLog(
            operation="create",
            entity_type="sample",
            entity_id=str(sample.id),
            details={
                "job_id": job.id,
                "filename": result.filename,
                "row_count": result.row_count,
            },
        )
        db.add(sample_audit)

    # Update job status based on results
    if errors:
        job.status = "completed_with_errors"
    else:
        job.status = "imported"

    db.commit()

    return ImportResultResponse(
        job_id=job.id,
        samples_created=len(samples),
        samples=samples,
        errors=errors,
    )


@app.post("/import/batch-data", response_model=ImportResultResponse)
async def import_batch_data(
    request: BatchImportDataRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Import pre-parsed file data from browser.

    This endpoint accepts already-parsed data from the frontend,
    useful when files are selected via browser file input (no file path access).
    """
    errors: list[str] = []
    samples: list[SampleSummary] = []

    # Create Job
    job = Job(
        status="pending",
        source_directory="browser-upload",
    )
    db.add(job)
    db.flush()

    # Create audit log for job creation
    audit_log = AuditLog(
        operation="create",
        entity_type="job",
        entity_id=str(job.id),
        details={"file_count": len(request.files), "source": "browser-upload"},
    )
    db.add(audit_log)

    # Process each file's data
    for file_data in request.files:
        # Create Sample with parsed data
        sample = Sample(
            job_id=job.id,
            filename=file_data.filename,
            status="pending",
            input_data={
                "rows": file_data.rows,
                "headers": file_data.headers,
                "row_count": file_data.row_count,
            },
        )
        db.add(sample)
        db.flush()

        samples.append(SampleSummary(
            id=sample.id,
            filename=file_data.filename,
            row_count=file_data.row_count,
        ))

        # Create audit log for sample creation
        sample_audit = AuditLog(
            operation="create",
            entity_type="sample",
            entity_id=str(sample.id),
            details={
                "job_id": job.id,
                "filename": file_data.filename,
                "row_count": file_data.row_count,
            },
        )
        db.add(sample_audit)

    # Update job status
    job.status = "imported"
    db.commit()

    return ImportResultResponse(
        job_id=job.id,
        samples_created=len(samples),
        samples=samples,
        errors=errors,
    )


# --- Job and Sample Endpoints ---

@app.get("/jobs", response_model=list[JobResponse])
async def get_jobs(
    limit: int = 50,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get recent jobs."""
    stmt = select(Job).order_by(desc(Job.created_at)).limit(limit)
    result = db.execute(stmt)
    return result.scalars().all()


@app.get("/jobs/{job_id}", response_model=JobResponse)
async def get_job(job_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get a single job by ID."""
    stmt = select(Job).where(Job.id == job_id)
    job = db.execute(stmt).scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
    return job


@app.get("/jobs/{job_id}/samples", response_model=list[SampleResponse])
async def get_job_samples(job_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get all samples for a job."""
    stmt = select(Sample).where(Sample.job_id == job_id).order_by(Sample.id)
    result = db.execute(stmt)
    return result.scalars().all()


@app.get("/jobs/{job_id}/samples-with-results", response_model=list[SampleWithResultsResponse])
async def get_job_samples_with_results(job_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """
    Get all samples for a job with their calculation results flattened.

    Returns samples with key calculation values (purity, retention_time, compound_id)
    extracted from Result records for easy UI consumption in batch review tables.
    """
    # Verify job exists
    job_stmt = select(Job).where(Job.id == job_id)
    job = db.execute(job_stmt).scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    # Get all samples for the job
    stmt = select(Sample).where(Sample.job_id == job_id).order_by(Sample.id)
    samples = db.execute(stmt).scalars().all()

    # Build response with flattened results
    response: list[SampleWithResultsResponse] = []
    for sample in samples:
        # Get the most recent purity result for this sample
        result_stmt = (
            select(Result)
            .where(Result.sample_id == sample.id)
            .where(Result.calculation_type == "purity")
            .order_by(desc(Result.created_at))
            .limit(1)
        )
        purity_result = db.execute(result_stmt).scalar_one_or_none()

        # Extract values from result output_data
        purity: Optional[float] = None
        retention_time: Optional[float] = None
        compound_id: Optional[str] = None
        has_results = False

        if purity_result and purity_result.output_data:
            has_results = True
            values = purity_result.output_data.get("values", {})
            # Extract purity percentage
            if "purity_percent" in values:
                purity = values["purity_percent"]
            # Extract matched compound info
            if "matched_compound" in values:
                compound_id = values["matched_compound"]
            if "retention_time" in values:
                retention_time = values["retention_time"]

        response.append(SampleWithResultsResponse(
            id=sample.id,
            job_id=sample.job_id,
            filename=sample.filename,
            status=sample.status,
            rejection_reason=sample.rejection_reason,
            created_at=sample.created_at,
            purity=purity,
            retention_time=retention_time,
            compound_id=compound_id,
            has_results=has_results,
        ))

    return response


@app.get("/samples", response_model=list[SampleResponse])
async def get_samples(
    limit: int = 50,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get recent samples."""
    stmt = select(Sample).order_by(desc(Sample.created_at)).limit(limit)
    result = db.execute(stmt)
    return result.scalars().all()


@app.get("/samples/{sample_id}", response_model=SampleResponse)
async def get_sample(sample_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get a single sample by ID."""
    stmt = select(Sample).where(Sample.id == sample_id)
    sample = db.execute(stmt).scalar_one_or_none()
    if not sample:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")
    return sample


@app.put("/samples/{sample_id}/approve", response_model=SampleResponse)
async def approve_sample(sample_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """
    Approve a sample.

    Sets status to 'approved' and clears any rejection reason.
    Creates an audit log entry.
    """
    stmt = select(Sample).where(Sample.id == sample_id)
    sample = db.execute(stmt).scalar_one_or_none()
    if not sample:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")

    old_status = sample.status
    sample.status = "approved"
    sample.rejection_reason = None

    # Create audit log
    audit = AuditLog(
        operation="approve",
        entity_type="sample",
        entity_id=str(sample_id),
        details={"old_status": old_status, "new_status": "approved"},
    )
    db.add(audit)
    db.commit()
    db.refresh(sample)

    return sample


@app.put("/samples/{sample_id}/reject", response_model=SampleResponse)
async def reject_sample(
    sample_id: int,
    request: RejectRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Reject a sample with a reason.

    Sets status to 'rejected' and stores the rejection reason.
    Creates an audit log entry.
    """
    stmt = select(Sample).where(Sample.id == sample_id)
    sample = db.execute(stmt).scalar_one_or_none()
    if not sample:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")

    old_status = sample.status
    sample.status = "rejected"
    sample.rejection_reason = request.reason

    # Create audit log
    audit = AuditLog(
        operation="reject",
        entity_type="sample",
        entity_id=str(sample_id),
        details={
            "old_status": old_status,
            "new_status": "rejected",
            "reason": request.reason,
        },
    )
    db.add(audit)
    db.commit()
    db.refresh(sample)

    return sample


# --- Calculation Endpoints ---

def _get_calculation_settings(db: Session) -> dict:
    """Load all settings relevant to calculations as a dict."""
    stmt = select(Settings)
    settings = db.execute(stmt).scalars().all()
    return {s.key: s.value for s in settings}


@app.get("/calculations/types", response_model=list[str])
async def get_calculation_types(_current_user=Depends(get_current_user)):
    """Get list of available calculation types."""
    return CalculationEngine.get_available_types()


@app.post("/calculate/{sample_id}", response_model=CalculationSummaryResponse)
async def calculate_sample(
    sample_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Run all applicable calculations for a sample.

    Loads sample data, runs calculations based on settings,
    stores results in Result table, and returns summary.
    """
    # Load sample
    stmt = select(Sample).where(Sample.id == sample_id)
    sample = db.execute(stmt).scalar_one_or_none()
    if not sample:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")

    if not sample.input_data:
        raise HTTPException(
            status_code=400,
            detail=f"Sample {sample_id} has no input data"
        )

    # Load settings
    settings = _get_calculation_settings(db)

    # Create engine and run calculations
    engine = CalculationEngine(settings)
    calc_results = engine.calculate_all(sample.input_data)

    # Store results and create audit logs
    stored_results: list[CalculationResultResponse] = []
    for calc_result in calc_results:
        # Create Result record
        result = Result(
            sample_id=sample_id,
            calculation_type=calc_result.calculation_type,
            input_data=calc_result.input_summary,
            output_data={
                "values": calc_result.output_values,
                "warnings": calc_result.warnings,
                "success": calc_result.success,
                "error": calc_result.error,
            },
        )
        db.add(result)
        db.flush()

        # Create audit log
        audit = AuditLog(
            operation="calculate",
            entity_type="result",
            entity_id=str(result.id),
            details={
                "sample_id": sample_id,
                "calculation_type": calc_result.calculation_type,
                "success": calc_result.success,
            },
        )
        db.add(audit)

        stored_results.append(CalculationResultResponse(
            calculation_type=calc_result.calculation_type,
            input_summary=calc_result.input_summary,
            output_values=calc_result.output_values,
            warnings=calc_result.warnings,
            success=calc_result.success,
            error=calc_result.error,
        ))

    # Update sample status
    sample.status = "calculated"
    db.commit()

    successful = sum(1 for r in stored_results if r.success)
    failed = len(stored_results) - successful

    return CalculationSummaryResponse(
        sample_id=sample_id,
        results=stored_results,
        total_calculations=len(stored_results),
        successful=successful,
        failed=failed,
    )


@app.post("/calculate/preview", response_model=CalculationResultResponse)
async def preview_calculation(
    request: CalculationPreviewRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Run a calculation without saving (for testing/preview).

    Useful for testing formulas with custom data before applying to samples.
    """
    settings = _get_calculation_settings(db)
    engine = CalculationEngine(settings)

    try:
        result = engine.calculate(request.data, request.calculation_type)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return CalculationResultResponse(
        calculation_type=result.calculation_type,
        input_summary=result.input_summary,
        output_values=result.output_values,
        warnings=result.warnings,
        success=result.success,
        error=result.error,
    )


@app.get("/samples/{sample_id}/results", response_model=list[ResultResponse])
async def get_sample_results(
    sample_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get all calculation results for a sample."""
    # Verify sample exists
    stmt = select(Sample).where(Sample.id == sample_id)
    sample = db.execute(stmt).scalar_one_or_none()
    if not sample:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")

    stmt = select(Result).where(Result.sample_id == sample_id).order_by(Result.created_at)
    results = db.execute(stmt).scalars().all()
    return results


# --- HPLC Analysis Endpoints ---

class HPLCFileInput(BaseModel):
    """A single file's content for HPLC parsing."""
    filename: str
    content: str


class HPLCParseBrowserRequest(BaseModel):
    """Request to parse HPLC PeakData files from browser upload."""
    files: list[HPLCFileInput]


class PeakResponse(BaseModel):
    """Response for a single chromatographic peak."""
    height: float
    area: float
    area_percent: float
    begin_time: float
    end_time: float
    retention_time: float
    is_solvent_front: bool
    is_main_peak: bool


class InjectionResponse(BaseModel):
    """Response for one injection's parsed data."""
    injection_name: str
    peptide_label: str = ""
    peaks: list[PeakResponse]
    total_area: float
    main_peak_index: int


class PurityResponse(BaseModel):
    """Purity calculation result."""
    purity_percent: Optional[float]
    individual_values: list[float]
    injection_names: list[str]
    rsd_percent: Optional[float]
    error: Optional[str] = None


class StandardInjectionResponse(BaseModel):
    """Standard injection reference data parsed from _std_ files."""
    analyte_label: str
    main_peak_rt: float
    main_peak_area_pct: float
    source_sample_id: str
    filename: str


class HPLCParseResponse(BaseModel):
    """Response from parsing HPLC files."""
    injections: list[InjectionResponse]
    purity: PurityResponse
    errors: list[str]
    warnings: list[str] = []
    detected_peptides: list[str] = []
    standard_injections: list[StandardInjectionResponse] = []


@app.post("/hplc/parse-files", response_model=HPLCParseResponse)
async def parse_hplc_peakdata(request: HPLCParseBrowserRequest, _current_user=Depends(get_current_user)):
    """
    Parse HPLC PeakData CSV files and calculate purity.

    Accepts file contents from browser upload, parses peak tables,
    identifies main peaks (excluding solvent front), and averages
    Area% across injections for purity calculation.
    """
    files_data = [{"filename": f.filename, "content": f.content} for f in request.files]
    result = parse_hplc_files(files_data)

    # Calculate purity from parsed injections
    purity = calculate_purity(result.injections)

    # Build response
    injections_resp = []
    for inj in result.injections:
        peaks_resp = [
            PeakResponse(
                height=p.height,
                area=p.area,
                area_percent=p.area_percent,
                begin_time=p.begin_time,
                end_time=p.end_time,
                retention_time=p.retention_time,
                is_solvent_front=p.is_solvent_front,
                is_main_peak=p.is_main_peak,
            )
            for p in inj.peaks
        ]
        injections_resp.append(InjectionResponse(
            injection_name=inj.injection_name,
            peptide_label=inj.peptide_label,
            peaks=peaks_resp,
            total_area=inj.total_area,
            main_peak_index=inj.main_peak_index,
        ))

    # Collect unique peptide labels (non-empty) for blend detection
    detected_peptides = sorted(set(
        inj.peptide_label for inj in result.injections if inj.peptide_label
    ))

    # Map standard injections to response model
    std_inj_resp = [
        StandardInjectionResponse(
            analyte_label=si.analyte_label,
            main_peak_rt=si.main_peak_rt,
            main_peak_area_pct=si.main_peak_area_pct,
            source_sample_id=si.source_sample_id,
            filename=si.filename,
        )
        for si in result.standard_injections
    ]

    return HPLCParseResponse(
        injections=injections_resp,
        purity=PurityResponse(**purity),
        errors=result.errors,
        warnings=result.warnings,
        detected_peptides=detected_peptides,
        standard_injections=std_inj_resp,
    )


# --- Peptide & Calibration Endpoints ---

# ─── HPLC Method schemas ───

# ─── Instrument schemas ───

class InstrumentBrief(BaseModel):
    """Minimal instrument info embedded in method responses."""
    id: int
    name: str
    model: Optional[str] = None

    class Config:
        from_attributes = True


class InstrumentResponse(BaseModel):
    """Full instrument response."""
    id: int
    name: str
    senaite_id: Optional[str] = None
    senaite_uid: Optional[str] = None
    instrument_type: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ─── Analysis Service schemas ───

class AnalysisServiceResponse(BaseModel):
    """Full analysis service response."""
    id: int
    title: str
    keyword: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    methods: Optional[list] = None
    peptide_name: Optional[str] = None
    peptide_id: Optional[int] = None
    senaite_id: Optional[str] = None
    senaite_uid: Optional[str] = None
    active: bool
    created_at: datetime
    updated_at: datetime
    result_type: Optional[str] = None
    result_options: Optional[list] = None
    variance_capable: bool

    class Config:
        from_attributes = True


# ─── Service Group schemas ───

class ServiceGroupCreate(BaseModel):
    """Schema for creating a service group."""
    name: str
    description: Optional[str] = None
    color: str = "blue"
    sort_order: int = 0
    is_default: bool = False
    sla_tier_id: Optional[int] = None


class ServiceGroupUpdate(BaseModel):
    """Schema for updating a service group."""
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None
    sort_order: Optional[int] = None
    is_default: Optional[bool] = None
    sla_tier_id: Optional[int] = None


class ServiceGroupResponse(BaseModel):
    """Schema for service group response."""
    id: int
    name: str
    description: Optional[str]
    color: str
    sort_order: int
    is_default: bool = False
    sla_tier_id: Optional[int] = None
    member_count: int = 0
    member_ids: list[int] = []
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class ServiceGroupMembersRequest(BaseModel):
    """Schema for setting service group membership."""
    analysis_service_ids: list[int]


# ─── SLA tier schemas (sub-project A, revised to tiers) ───

# Priority tiers mirror SamplePriority/WorksheetItem.priority. Validated here at
# the API edge — the DB columns are unconstrained VARCHAR.
SlaPriority = Literal["normal", "high", "expedited"]


class SlaTierCreate(BaseModel):
    name: str
    target_minutes: int
    business_hours_only: bool = False
    is_default: bool = False
    amber_threshold_percent: int = Field(20, ge=1, le=100)


class SlaTierUpdate(BaseModel):
    name: Optional[str] = None
    target_minutes: Optional[int] = None
    business_hours_only: Optional[bool] = None
    is_default: Optional[bool] = None
    amber_threshold_percent: Optional[int] = Field(None, ge=1, le=100)


class SlaTierResponse(BaseModel):
    id: int
    name: str
    target_minutes: int
    business_hours_only: bool
    is_default: bool
    amber_threshold_percent: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# Consumed by the /sla-priority-tiers endpoints.
# Multi-tier follow-on: `service_group_id` scopes the override to a single
# service group. NULL means the row applies globally (precedence falls through
# (priority, group_id) → (priority, NULL) → group's own tier → default).
class SlaPriorityTierResponse(BaseModel):
    id: int
    priority: str
    sla_tier_id: int
    service_group_id: int | None = None

    class Config:
        from_attributes = True


class SlaPriorityTierSet(BaseModel):
    sla_tier_id: int
    # Omit / null => the global override row for this priority. Specify a group
    # id to scope the override to that group (e.g. expedited + HPLC group).
    service_group_id: int | None = None


# ── D2: bulk per-sample priority lookup ────────────────────────────────────

class SamplePriorityLookupRequest(BaseModel):
    sample_uids: list[str] = Field(..., min_length=1, max_length=500)


class SamplePriorityResponseItem(BaseModel):
    sample_uid: str
    priority: SlaPriority

    class Config:
        from_attributes = True


class SamplePriorityLookupResponse(BaseModel):
    items: list[SamplePriorityResponseItem]


class BusinessHoursConfigResponse(BaseModel):
    open_time: time
    close_time: time
    timezone: str
    working_days: list[int]

    class Config:
        from_attributes = True


class BusinessHoursConfigUpdate(BaseModel):
    open_time: time
    close_time: time
    timezone: str
    working_days: list[int]


class LabHolidayResponse(BaseModel):
    id: int
    holiday_date: date
    name: str
    source: str

    class Config:
        from_attributes = True


class LabHolidayCreate(BaseModel):
    holiday_date: date
    name: str


class SlaStatusRequestItem(BaseModel):
    key: str
    received_at: Optional[datetime] = None
    target_minutes: int
    business_hours_only: bool = False
    # Historical mode for published samples: when set, the server uses this as
    # the "now" instead of `datetime.utcnow()`. Lets the UI render frozen-in-time
    # SLA results ("took 28h, Met/Missed") on the Sample Details header.
    now_override: Optional[datetime] = None


class SlaStatusRequest(BaseModel):
    items: list[SlaStatusRequestItem]


class SlaStatusResultItem(BaseModel):
    key: str
    status: Optional[dict] = None


class SlaStatusResponse(BaseModel):
    items: list[SlaStatusResultItem]


# ─── HPLC Method schemas ───

class MethodCreate(BaseModel):
    """Schema for creating an HPLC method."""
    name: str
    senaite_id: Optional[str] = None
    instrument_ids: list[int] = []
    size_peptide: Optional[str] = None
    starting_organic_pct: Optional[float] = None
    temperature_mct_c: Optional[float] = None
    dissolution: Optional[str] = None
    notes: Optional[str] = None


class MethodUpdate(BaseModel):
    """Schema for updating an HPLC method."""
    name: Optional[str] = None
    senaite_id: Optional[str] = None
    instrument_ids: Optional[list[int]] = None
    size_peptide: Optional[str] = None
    starting_organic_pct: Optional[float] = None
    temperature_mct_c: Optional[float] = None
    dissolution: Optional[str] = None
    notes: Optional[str] = None
    active: Optional[bool] = None


class PeptideBrief(BaseModel):
    """Minimal peptide info for method responses."""
    id: int
    name: str
    abbreviation: str

    class Config:
        from_attributes = True


class MethodBrief(BaseModel):
    """Minimal method info for peptide responses."""
    id: int
    name: str
    senaite_id: Optional[str] = None
    instrument_ids: list[int] = []
    instruments: list[InstrumentBrief] = []

    class Config:
        from_attributes = True


class MethodResponse(BaseModel):
    """Full HPLC method response with common peptides."""
    id: int
    name: str
    senaite_id: Optional[str] = None
    instrument_ids: list[int] = []
    instruments: list[InstrumentBrief] = []
    size_peptide: Optional[str] = None
    starting_organic_pct: Optional[float] = None
    temperature_mct_c: Optional[float] = None
    dissolution: Optional[str] = None
    notes: Optional[str] = None
    active: bool
    created_at: datetime
    updated_at: datetime
    common_peptides: list[PeptideBrief] = []

    class Config:
        from_attributes = True


# ─── Peptide schemas ───

class AnalyteInput(BaseModel):
    """One analyte slot for peptide create/update."""
    slot: int
    analysis_service_id: int
    sample_id: Optional[str] = None
    component_peptide_id: Optional[int] = None


class AnalyteResponse(BaseModel):
    """Analyte slot in peptide response — includes denormalized service fields."""
    id: int
    slot: int
    analysis_service_id: int
    sample_id: Optional[str] = None
    peptide_name: Optional[str] = None
    service_title: Optional[str] = None
    component_peptide_id: Optional[int] = None
    component_abbreviation: Optional[str] = None

    class Config:
        from_attributes = True


class ComponentBrief(BaseModel):
    """Brief component peptide info for blend responses."""
    id: int
    name: str
    abbreviation: str
    vial_number: int = 1
    hplc_aliases: Optional[list[str]] = None

    class Config:
        from_attributes = True


class PeptideCreate(BaseModel):
    """Schema for creating a peptide."""
    name: str
    abbreviation: str
    analytes: list[AnalyteInput] = []
    is_blend: bool = False
    component_ids: list[int] = []
    analyte_class: str = "peptide"  # 'peptide' | 'additive'


class PeptideUpdate(BaseModel):
    """Schema for updating a peptide."""
    name: Optional[str] = None
    abbreviation: Optional[str] = None
    active: Optional[bool] = None
    prep_vial_count: Optional[int] = None
    hplc_aliases: Optional[list[str]] = None  # Alternate names used in HPLC filenames
    display_aliases: Optional[list[str]] = None  # Approved customer-facing COA display aliases
    method_ids: Optional[list[int]] = None  # Set all method assignments (one per instrument)
    analytes: Optional[list[AnalyteInput]] = None
    component_ids: Optional[list[int]] = None
    component_vial_assignments: Optional[dict[str, int]] = None  # {"component_id": vial_number}
    analyte_class: Optional[str] = None


class CalibrationCurveResponse(BaseModel):
    """Schema for calibration curve response."""
    id: int
    peptide_id: int
    peptide_analyte_id: Optional[int] = None
    reference_rt: Optional[float] = None
    rt_tolerance: float = 0.5
    diluent_density: float = 997.1
    slope: float
    intercept: float
    r_squared: float
    standard_data: Optional[dict]
    source_filename: Optional[str]
    source_path: Optional[str] = None
    source_date: Optional[datetime] = None
    sharepoint_url: Optional[str] = None
    is_active: bool
    created_at: datetime
    # Standard identification metadata
    source_sample_id: Optional[str] = None
    instrument: Optional[str] = None
    instrument_id: Optional[int] = None
    vendor: Optional[str] = None
    lot_number: Optional[str] = None
    batch_number: Optional[str] = None
    cap_color: Optional[str] = None
    run_date: Optional[datetime] = None
    # Wizard fields (populated when creating standards in AccuMk1)
    standard_weight_mg: Optional[float] = None
    stock_concentration_ug_ml: Optional[float] = None
    diluent: Optional[str] = None
    column_type: Optional[str] = None
    wavelength_nm: Optional[float] = None
    flow_rate_ml_min: Optional[float] = None
    injection_volume_ul: Optional[float] = None
    operator: Optional[str] = None
    notes: Optional[str] = None
    # Phase 09: Chromatogram storage
    chromatogram_data: Optional[dict] = None
    source_sharepoint_folder: Optional[str] = None
    # User tracking
    created_by_user_id: Optional[int] = None
    created_by_email: Optional[str] = None
    updated_by_user_id: Optional[int] = None
    updated_by_email: Optional[str] = None

    class Config:
        from_attributes = True


class InstrumentSummary(BaseModel):
    """Per-instrument calibration curve count for a peptide."""
    instrument: str  # "1260", "1290", or "unknown"
    instrument_id: Optional[int] = None
    curve_count: int


class PeptideResponse(BaseModel):
    """Schema for peptide response."""
    id: int
    name: str
    abbreviation: str
    active: bool
    is_blend: bool = False
    analyte_class: str = "peptide"
    prep_vial_count: int = 1
    hplc_aliases: Optional[list[str]] = None
    display_aliases: Optional[list[str]] = None
    created_at: datetime
    updated_at: datetime
    methods: list[MethodBrief] = []
    active_calibration: Optional[CalibrationCurveResponse] = None
    calibration_summary: list[InstrumentSummary] = []
    analytes: list[AnalyteResponse] = []
    components: list[ComponentBrief] = []

    class Config:
        from_attributes = True


class CalibrationDataInput(BaseModel):
    """Schema for manual calibration data entry."""
    concentrations: list[float]
    areas: list[float]
    rts: Optional[list[float]] = None
    source_filename: Optional[str] = None
    analyte_id: Optional[int] = None
    instrument: Optional[str] = None
    instrument_id: Optional[int] = None
    notes: Optional[str] = None


class StandardCalibrationInput(BaseModel):
    """Input for auto-creating calibration curve from standard HPLC processing."""
    sample_prep_id: str                        # e.g. "P-0136" — the standard prep
    concentrations: list[float]                # ug/mL per dilution level
    areas: list[float]                         # Peak areas per dilution level
    rts: Optional[list[float]] = None          # Retention times per dilution level
    chromatogram_data: Optional[dict] = None   # {times: [], signals: []} from DAD1A CSV
    source_sharepoint_folder: Optional[str] = None  # SharePoint folder path
    vendor: Optional[str] = None               # Manufacturer from standard prep metadata
    notes: Optional[str] = None                # Notes from standard prep metadata
    instrument: Optional[str] = None           # HPLC instrument identifier


def _resolve_instrument(db: Session, name: Optional[str] = None, inst_id: Optional[int] = None) -> tuple[Optional[str], Optional[int]]:
    """Resolve instrument name ↔ ID. Returns (name, id) tuple."""
    if inst_id and not name:
        inst = db.execute(select(Instrument).where(Instrument.id == inst_id)).scalar_one_or_none()
        if inst:
            return inst.name, inst.id
    elif name and not inst_id:
        inst = db.execute(select(Instrument).where(Instrument.name == name)).scalar_one_or_none()
        if inst:
            return inst.name, inst.id
    return name, inst_id


def _cal_to_response(cal: CalibrationCurve, include_blobs: bool = True) -> CalibrationCurveResponse:
    """Convert CalibrationCurve model to response with SharePoint URL.

    Set include_blobs=False to skip loading large JSON fields (chromatogram_data, standard_data)
    for list views where they're not needed.
    """
    if not include_blobs:
        # Build response without loading deferred chromatogram_data (3+ MB per curve)
        # standard_data is included (small — just arrays of numbers for the chart/table)
        resp = CalibrationCurveResponse(
            id=cal.id, peptide_id=cal.peptide_id, peptide_analyte_id=cal.peptide_analyte_id,
            reference_rt=cal.reference_rt, rt_tolerance=cal.rt_tolerance, diluent_density=cal.diluent_density,
            slope=cal.slope, intercept=cal.intercept, r_squared=cal.r_squared,
            standard_data=cal.standard_data, chromatogram_data=None,
            source_filename=cal.source_filename, source_path=cal.source_path,
            source_date=cal.source_date, sharepoint_url=cal.sharepoint_url,
            is_active=cal.is_active, created_at=cal.created_at,
            source_sample_id=cal.source_sample_id, vendor=cal.vendor,
            instrument=cal.instrument, instrument_id=cal.instrument_id,
            notes=cal.notes,
            created_by_user_id=cal.created_by_user_id, created_by_email=cal.created_by_email,
            updated_by_user_id=cal.updated_by_user_id, updated_by_email=cal.updated_by_email,
        )
    else:
        resp = CalibrationCurveResponse.model_validate(cal)
    # Prefer stored webUrl from Graph API; fall back to computed URL for legacy records
    if cal.sharepoint_url:
        resp.sharepoint_url = cal.sharepoint_url
    elif cal.source_path:
        from sharepoint import get_sharepoint_file_url
        resp.sharepoint_url = get_sharepoint_file_url(cal.source_path)
    return resp


def _get_active_calibration(db: Session, peptide_id: int) -> Optional[CalibrationCurveResponse]:
    """Get the active calibration curve for a peptide."""
    stmt = (
        select(CalibrationCurve)
        .where(CalibrationCurve.peptide_id == peptide_id)
        .where(CalibrationCurve.is_active == True)
        .order_by(desc(CalibrationCurve.created_at))
        .limit(1)
    )
    cal = db.execute(stmt).scalar_one_or_none()
    if cal:
        return _cal_to_response(cal)
    return None


def _instrument_to_brief(instrument) -> Optional[InstrumentBrief]:
    """Convert Instrument model to brief response."""
    if instrument is None:
        return None
    return InstrumentBrief.model_validate(instrument)


def _method_to_brief(method: HplcMethod) -> MethodBrief:
    """Convert HplcMethod model to brief response with instruments."""
    brief = MethodBrief(
        id=method.id,
        name=method.name,
        senaite_id=method.senaite_id,
        instrument_ids=[i.id for i in method.instruments],
        instruments=[_instrument_to_brief(i) for i in method.instruments],
    )
    return brief


def _build_analyte_responses(analytes) -> list[AnalyteResponse]:
    """Convert PeptideAnalyte ORM objects to AnalyteResponse dicts."""
    results = []
    for a in analytes:
        svc = a.analysis_service
        comp = a.component_peptide
        results.append(AnalyteResponse(
            id=a.id,
            slot=a.slot,
            analysis_service_id=a.analysis_service_id,
            sample_id=a.sample_id,
            peptide_name=svc.peptide_name if svc else None,
            service_title=svc.title if svc else None,
            component_peptide_id=a.component_peptide_id,
            component_abbreviation=comp.abbreviation if comp else None,
        ))
    return results


def _peptide_to_response(db: Session, peptide: Peptide) -> PeptideResponse:
    """Convert Peptide model to response with active calibration, methods, analytes, and components."""
    resp = PeptideResponse.model_validate(peptide)
    resp.active_calibration = _get_active_calibration(db, peptide.id)
    resp.methods = [_method_to_brief(m) for m in peptide.methods]
    resp.analytes = _build_analyte_responses(peptide.analytes)
    if peptide.is_blend:
        resp.components = _build_component_briefs(db, peptide.id)
    return resp


def _build_component_briefs(db: Session, blend_id: int) -> list[ComponentBrief]:
    """Build ComponentBrief list with vial_number from blend_components junction table."""
    rows = db.execute(
        select(Peptide, blend_components.c.vial_number)
        .join(blend_components, blend_components.c.component_id == Peptide.id)
        .where(blend_components.c.blend_id == blend_id)
        .order_by(blend_components.c.display_order)
    ).all()
    return [
        ComponentBrief(id=p.id, name=p.name, abbreviation=p.abbreviation, vial_number=vn or 1, hplc_aliases=p.hplc_aliases)
        for p, vn in rows
    ]


def _method_to_response(method: HplcMethod) -> MethodResponse:
    """Convert HplcMethod model to response with common peptides and instruments."""
    resp = MethodResponse.model_validate(method)
    resp.instrument_ids = [i.id for i in method.instruments]
    resp.instruments = [_instrument_to_brief(i) for i in method.instruments]
    resp.common_peptides = [PeptideBrief.model_validate(p) for p in method.peptides]
    return resp


# ─── Instrument Endpoints ───

@app.get("/instruments", response_model=list[InstrumentResponse])
async def get_instruments(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get all instruments."""
    instruments = db.execute(select(Instrument).order_by(Instrument.name)).scalars().all()
    return [InstrumentResponse.model_validate(i) for i in instruments]


@app.post("/instruments/sync")
async def sync_instruments(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Sync instruments from Senaite. Adds new instruments, does not overwrite existing."""
    import httpx as _httpx

    if not SENAITE_URL:
        raise HTTPException(400, "SENAITE_URL not configured")

    try:
        resp = _httpx.get(
            f"{SENAITE_URL}/senaite/@@API/senaite/v1/instrument",
            auth=_httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            timeout=SENAITE_TIMEOUT,
            params={"limit": 100},
        )
        resp.raise_for_status()
    except Exception as e:
        raise HTTPException(502, f"Failed to reach Senaite: {e}")

    items = resp.json().get("items", [])
    created = 0
    updated = 0
    for item in items:
        senaite_id = item.get("id")
        if not senaite_id:
            continue

        # Extract instrument type and manufacturer from nested objects or title
        title = item.get("title", senaite_id)
        inst_type = None
        brand = None
        model = item.get("Model")
        if isinstance(item.get("InstrumentType"), dict):
            inst_type = item["InstrumentType"].get("title")
        if isinstance(item.get("Manufacturer"), dict):
            brand = item["Manufacturer"].get("title")

        # Auto-parse from title when SENAITE fields are empty
        # e.g. "HPLC 1290b" → type=HPLC, model=1290, brand=Agilent
        if not inst_type and title.upper().startswith("HPLC"):
            inst_type = "HPLC"
            brand = brand or "Agilent"
        if not model:
            import re as _re
            m = _re.search(r'(\d{4})', title)
            if m:
                model = m.group(1)

        existing = db.execute(select(Instrument).where(Instrument.senaite_id == senaite_id)).scalar_one_or_none()
        if existing:
            # Backfill missing fields on existing instruments
            changed = False
            if not existing.instrument_type and inst_type:
                existing.instrument_type = inst_type
                changed = True
            if not existing.brand and brand:
                existing.brand = brand
                changed = True
            if not existing.model and model:
                existing.model = model
                changed = True
            if changed:
                updated += 1
            continue

        instrument = Instrument(
            name=title,
            senaite_id=senaite_id,
            senaite_uid=item.get("uid"),
            instrument_type=inst_type,
            brand=brand,
            model=model,
        )
        db.add(instrument)
        created += 1

    db.commit()
    total = db.execute(select(func.count()).select_from(Instrument)).scalar()
    return {"created": created, "updated": updated, "total": total}


# ─── Analysis Service Endpoints ───


def _extract_peptide_name(title: str) -> Optional[str]:
    """Derive peptide name from analysis service title.
    'AICAR – Identity (HPLC)' → 'AICAR'
    'BPC157 – Purity (HPLC)' → 'BPC157'
    """
    import re
    match = re.match(r'^(.+?)\s*[–\-]\s*(?:Purity|Identity|Quantity)\b', title)
    return match.group(1).strip() if match else None


@app.get("/analysis-services", response_model=list[AnalysisServiceResponse])
async def get_analysis_services(
    search: Optional[str] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """List all analysis services. Optional search by title, keyword, or category. Optional exact category filter."""
    query = select(AnalysisService).order_by(AnalysisService.title)
    if category:
        query = query.where(AnalysisService.category == category)
    if search:
        q = f"%{search}%"
        query = query.where(
            AnalysisService.title.ilike(q)
            | AnalysisService.keyword.ilike(q)
            | AnalysisService.category.ilike(q)
            | AnalysisService.peptide_name.ilike(q)
        )
    services = db.execute(query).scalars().all()
    return [AnalysisServiceResponse.model_validate(s) for s in services]


class AnalysisServicePeptideUpdate(BaseModel):
    peptide_id: Optional[int] = None  # null to clear the link


@app.put("/analysis-services/{service_id}/peptide", response_model=AnalysisServiceResponse)
async def update_analysis_service_peptide(
    service_id: int,
    data: AnalysisServicePeptideUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Link or unlink a peptide to an analysis service."""
    service = db.execute(
        select(AnalysisService).where(AnalysisService.id == service_id)
    ).scalar_one_or_none()
    if not service:
        raise HTTPException(404, f"Analysis service {service_id} not found")

    if data.peptide_id is not None:
        peptide = db.execute(
            select(Peptide).where(Peptide.id == data.peptide_id)
        ).scalar_one_or_none()
        if not peptide:
            raise HTTPException(404, f"Peptide {data.peptide_id} not found")
        service.peptide_id = peptide.id
        service.peptide_name = peptide.name
    else:
        service.peptide_id = None
        service.peptide_name = None

    db.commit()
    db.refresh(service)
    return AnalysisServiceResponse.model_validate(service)


class AnalysisServiceResultTypeUpdate(BaseModel):
    result_type: Optional[str] = None
    result_options: Optional[list] = None


@app.patch("/analysis-services/{service_id}/result-type", response_model=AnalysisServiceResponse)
async def update_analysis_service_result_type(
    service_id: int,
    data: AnalysisServiceResultTypeUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Set a service's result type + options (local-authoritative once set)."""
    service = db.execute(
        select(AnalysisService).where(AnalysisService.id == service_id)
    ).scalar_one_or_none()
    if not service:
        raise HTTPException(404, f"Analysis service {service_id} not found")
    if "result_type" in data.model_fields_set:
        service.result_type = data.result_type
    if "result_options" in data.model_fields_set:
        service.result_options = data.result_options
    db.commit()
    db.refresh(service)
    return AnalysisServiceResponse.model_validate(service)


class AnalysisServiceVarianceCapableUpdate(BaseModel):
    variance_capable: bool


@app.patch("/analysis-services/{service_id}/variance-capable", response_model=AnalysisServiceResponse)
async def update_analysis_service_variance_capable(
    service_id: int,
    data: AnalysisServiceVarianceCapableUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Lab-managed toggle: mark an analyte as a variance figure. Mk1-owned —
    never touched by the SENAITE sync."""
    service = db.execute(
        select(AnalysisService).where(AnalysisService.id == service_id)
    ).scalar_one_or_none()
    if not service:
        raise HTTPException(404, f"Analysis service {service_id} not found")
    service.variance_capable = data.variance_capable
    db.commit()
    db.refresh(service)
    return AnalysisServiceResponse.model_validate(service)


def _parse_service_result_options(raw) -> list[dict]:
    """SENAITE ResultOptions [{ResultValue, ResultText}] -> [{value, label}]."""
    out: list[dict] = []
    if raw and isinstance(raw, list):
        for opt in raw:
            if isinstance(opt, dict) and opt.get("ResultValue") is not None:
                out.append({
                    "value": str(opt["ResultValue"]),
                    "label": str(opt.get("ResultText") or opt["ResultValue"]),
                })
    return out


def _apply_service_result_type(svc, item: dict) -> None:
    """Seed svc.result_type / result_options from a SENAITE service item, but
    ONLY when svc.result_type is NULL (local-wins). No-op otherwise."""
    if svc.result_type is not None:
        return
    rtype = item.get("ResultType") or item.get("getResultType")
    if not rtype:
        return
    svc.result_type = str(rtype)
    svc.result_options = _parse_service_result_options(
        item.get("ResultOptions") or item.get("getResultOptions") or []
    ) or None


@app.post("/analysis-services/sync")
async def sync_analysis_services(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Sync analysis services from Senaite. Adds new services and reconciles a
    service SENAITE recreated under a new id/UID by adopting the existing row;
    never overwrites local result-type edits."""
    import httpx as _httpx

    if not SENAITE_URL:
        raise HTTPException(400, "SENAITE_URL not configured")

    try:
        resp = _httpx.get(
            f"{SENAITE_URL}/senaite/@@API/senaite/v1/search",
            auth=_httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            timeout=SENAITE_TIMEOUT,
            params={"portal_type": "AnalysisService", "limit": 500, "complete": "true"},
        )
        resp.raise_for_status()
    except Exception as e:
        raise HTTPException(502, f"Failed to reach Senaite: {e}")

    items = resp.json().get("items", [])

    # Pre-fetch category UID → title map (Category field is a reference, not a string)
    category_map: dict[str, str] = {}
    try:
        cat_resp = _httpx.get(
            f"{SENAITE_URL}/senaite/@@API/senaite/v1/search",
            auth=_httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            timeout=SENAITE_TIMEOUT,
            params={"portal_type": "AnalysisCategory", "limit": 200},
        )
        cat_resp.raise_for_status()
        for cat_item in cat_resp.json().get("items", []):
            if cat_item.get("uid") and cat_item.get("title"):
                category_map[cat_item["uid"]] = cat_item["title"]
    except Exception:
        pass  # Non-fatal — categories will just be None

    created = 0
    updated = 0
    # senaite_ids present in THIS pull — used to spot orphaned Mk1 rows whose
    # SENAITE object was deleted/recreated, so we adopt instead of cloning.
    current_ids = {it.get("id") for it in items if it.get("id")}
    for item in items:
        senaite_id = item.get("id")
        if not senaite_id:
            continue

        title = item.get("title", senaite_id)

        # Resolve category title via UID lookup
        category = None
        cat = item.get("Category")
        if isinstance(cat, dict):
            category = category_map.get(cat.get("uid", ""))
        elif isinstance(cat, str):
            category = cat
        if not category:
            cat_title = item.get("getCategoryTitle")
            if isinstance(cat_title, str):
                category = cat_title

        # Extract methods list
        raw_methods = item.get("Methods") or item.get("getMethods") or []
        methods_list = []
        if isinstance(raw_methods, list):
            for m in raw_methods:
                if isinstance(m, dict):
                    methods_list.append({"uid": m.get("uid", ""), "title": m.get("title", "")})

        existing = db.execute(
            select(AnalysisService).where(AnalysisService.senaite_id == senaite_id)
        ).scalar_one_or_none()

        if existing:
            # Back-fill category if it was missing
            if not existing.category and category:
                existing.category = category
                updated += 1
            _apply_service_result_type(existing, item)  # local-wins seed
            continue

        # SENAITE can delete+recreate a service under a new id/UID (same keyword).
        # Matching only by senaite_id would clone the keyword and orphan the old
        # row (the TB500 promote-502 incident). Adopt an orphaned row — same
        # keyword, stale senaite_id absent from this pull — preserving its id and
        # all lims_analyses / peptide_analytes references. .first() tolerates any
        # pre-existing duplicates.
        kw = item.get("getKeyword") or item.get("Keyword")
        orphan = None
        if kw and current_ids:
            orphan = db.execute(
                select(AnalysisService)
                .where(
                    AnalysisService.keyword == kw,
                    AnalysisService.senaite_id.isnot(None),
                    AnalysisService.senaite_id.not_in(current_ids),
                )
                .order_by(AnalysisService.id)
            ).scalars().first()
        if orphan is not None:
            orphan.senaite_id = senaite_id
            orphan.senaite_uid = item.get("uid")
            orphan.title = title
            if category:
                orphan.category = category
            if methods_list:
                orphan.methods = methods_list
            _apply_service_result_type(orphan, item)
            updated += 1
            continue

        svc = AnalysisService(
            title=title,
            keyword=item.get("getKeyword") or item.get("Keyword"),
            category=category,
            unit=item.get("getUnit") or item.get("Unit"),
            methods=methods_list if methods_list else None,
            peptide_name=_extract_peptide_name(title),
            senaite_id=senaite_id,
            senaite_uid=item.get("uid"),
        )
        db.add(svc)
        _apply_service_result_type(svc, item)
        created += 1

    db.commit()
    total = db.execute(select(func.count()).select_from(AnalysisService)).scalar()
    return {"created": created, "total": total}


# ─── HPLC Method Endpoints ───

@app.get("/hplc/methods", response_model=list[MethodResponse])
async def get_methods(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get all HPLC methods with their common peptides and instruments."""
    methods = db.execute(
        select(HplcMethod)
        .options(joinedload(HplcMethod.instruments), joinedload(HplcMethod.peptides))
        .order_by(HplcMethod.name)
    ).scalars().unique().all()
    return [_method_to_response(m) for m in methods]


@app.post("/hplc/methods", response_model=MethodResponse, status_code=201)
async def create_method(data: MethodCreate, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Create a new HPLC method."""
    existing = db.execute(select(HplcMethod).where(HplcMethod.name == data.name)).scalar_one_or_none()
    if existing:
        raise HTTPException(400, f"Method with name '{data.name}' already exists")

    if data.senaite_id:
        dup = db.execute(select(HplcMethod).where(HplcMethod.senaite_id == data.senaite_id)).scalar_one_or_none()
        if dup:
            raise HTTPException(400, f"Method with Senaite ID '{data.senaite_id}' already exists")

    method = HplcMethod(**data.model_dump(exclude={"instrument_ids"}))
    if data.instrument_ids:
        instruments = db.execute(select(Instrument).where(Instrument.id.in_(data.instrument_ids))).scalars().all()
        method.instruments = list(instruments)
    db.add(method)
    db.commit()
    db.refresh(method)
    return _method_to_response(method)


@app.put("/hplc/methods/{method_id}", response_model=MethodResponse)
async def update_method(method_id: int, data: MethodUpdate, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Update an HPLC method."""
    method = db.execute(
        select(HplcMethod).options(joinedload(HplcMethod.instruments))
        .where(HplcMethod.id == method_id)
    ).scalars().unique().first()
    if not method:
        raise HTTPException(404, f"Method {method_id} not found")

    update_data = data.model_dump(exclude_unset=True)
    instrument_ids = update_data.pop("instrument_ids", None)
    for field, value in update_data.items():
        setattr(method, field, value)
    if instrument_ids is not None:
        instruments = db.execute(select(Instrument).where(Instrument.id.in_(instrument_ids))).scalars().all() if instrument_ids else []
        method.instruments = list(instruments)

    db.commit()
    db.refresh(method)
    return _method_to_response(method)


@app.delete("/hplc/methods/{method_id}")
async def delete_method(method_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Delete an HPLC method. Junction rows cascade-delete automatically."""
    method = db.execute(select(HplcMethod).where(HplcMethod.id == method_id)).scalar_one_or_none()
    if not method:
        raise HTTPException(404, f"Method {method_id} not found")

    db.delete(method)
    db.commit()
    return {"message": f"Method '{method.name}' deleted"}


# ─── Peptide Endpoints ───

@app.get("/peptides", response_model=list[PeptideResponse])
async def get_peptides(
    analyte_class: Optional[str] = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get all peptides with their active calibration curves and per-instrument summary.

    Optional ?analyte_class=peptide|additive filters by class. Default returns all classes.
    """
    query = (
        select(Peptide)
        .options(
            joinedload(Peptide.methods).joinedload(HplcMethod.instruments),
            joinedload(Peptide.analytes).joinedload(PeptideAnalyte.analysis_service),
            joinedload(Peptide.analytes).joinedload(PeptideAnalyte.component_peptide),
            joinedload(Peptide.components),
        )
        .order_by(Peptide.abbreviation)
    )
    if analyte_class:
        query = query.where(Peptide.analyte_class == analyte_class)
    peptides = db.execute(query).scalars().unique().all()

    # Batch 1: per-instrument curve counts for all peptides in one query
    summary_rows = db.execute(
        select(
            CalibrationCurve.peptide_id,
            CalibrationCurve.instrument_id,
            func.coalesce(Instrument.model, CalibrationCurve.instrument, "unknown").label("instrument"),
            func.count().label("curve_count"),
        )
        .outerjoin(Instrument, CalibrationCurve.instrument_id == Instrument.id)
        .group_by(
            CalibrationCurve.peptide_id,
            CalibrationCurve.instrument_id,
            Instrument.model,
            CalibrationCurve.instrument,
        )
    ).all()
    summary_map: dict[int, list[InstrumentSummary]] = {}
    for row in summary_rows:
        summary_map.setdefault(row.peptide_id, []).append(
            InstrumentSummary(instrument=row.instrument, instrument_id=row.instrument_id, curve_count=row.curve_count)
        )

    # Batch 2: all active calibration curves in one query, keep first per peptide
    # Defer chromatogram_data (3+ MB each) — not needed for peptide list view
    from sqlalchemy.orm import defer
    active_cals = db.execute(
        select(CalibrationCurve)
        .options(defer(CalibrationCurve.chromatogram_data))
        .where(CalibrationCurve.is_active == True)
        .order_by(CalibrationCurve.peptide_id, desc(CalibrationCurve.created_at))
    ).scalars().all()
    active_cal_map: dict[int, CalibrationCurveResponse] = {}
    for cal in active_cals:
        if cal.peptide_id not in active_cal_map:
            active_cal_map[cal.peptide_id] = _cal_to_response(cal, include_blobs=False)

    results = []
    for p in peptides:
        resp = PeptideResponse.model_validate(p)
        resp.active_calibration = active_cal_map.get(p.id)
        resp.calibration_summary = sorted(summary_map.get(p.id, []), key=lambda x: x.instrument)
        resp.methods = [_method_to_brief(m) for m in p.methods]
        resp.analytes = _build_analyte_responses(p.analytes)
        if p.is_blend:
            resp.components = _build_component_briefs(db, p.id)
            # Aggregate calibration summaries from component peptides
            blend_summary: dict[tuple, int] = {}
            for comp in p.components:
                for s in summary_map.get(comp.id, []):
                    key = (s.instrument_id, s.instrument)
                    blend_summary[key] = blend_summary.get(key, 0) + s.curve_count
            resp.calibration_summary = sorted(
                [InstrumentSummary(instrument=inst, instrument_id=iid, curve_count=cnt) for (iid, inst), cnt in blend_summary.items()],
                key=lambda x: x.instrument,
            )
        results.append(resp)
    return results


@app.post("/peptides", response_model=PeptideResponse, status_code=201)
async def create_peptide(data: PeptideCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Create a new peptide."""
    # Check uniqueness
    existing = db.execute(
        select(Peptide).where(Peptide.abbreviation == data.abbreviation)
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(400, f"Peptide with abbreviation '{data.abbreviation}' already exists")

    peptide = Peptide(
        name=data.name,
        abbreviation=data.abbreviation,
        is_blend=data.is_blend,
        analyte_class=data.analyte_class,
        created_by_user_id=current_user.id,
        created_by_email=current_user.email,
        updated_by_user_id=current_user.id,
        updated_by_email=current_user.email,
    )
    db.add(peptide)
    db.flush()  # Get peptide.id without committing

    # Link blend components and auto-create analyte slots from components
    if data.is_blend and data.component_ids:
        components = db.execute(
            select(Peptide).where(Peptide.id.in_(data.component_ids))
        ).scalars().all()
        if any(c.is_blend for c in components):
            raise HTTPException(400, "Blends cannot contain other blends")
        # Preserve requested order
        comp_map = {c.id: c for c in components}
        slot_num = 1
        for order, comp_id in enumerate(data.component_ids):
            db.execute(blend_components.insert().values(
                blend_id=peptide.id, component_id=comp_id, display_order=order,
            ))
            # Auto-create analyte slot from component's slot-1 analyte
            comp = comp_map.get(comp_id)
            if comp:
                comp_analyte = db.execute(
                    select(PeptideAnalyte).where(
                        PeptideAnalyte.peptide_id == comp_id,
                        PeptideAnalyte.slot == 1,
                    )
                ).scalar_one_or_none()
                if comp_analyte:
                    db.add(PeptideAnalyte(
                        peptide_id=peptide.id,
                        analysis_service_id=comp_analyte.analysis_service_id,
                        sample_id=comp_analyte.sample_id,
                        slot=slot_num,
                        component_peptide_id=comp_id,
                    ))
                    slot_num += 1
    else:
        # Create analyte slot rows (non-blend)
        for a in data.analytes:
            db.add(PeptideAnalyte(
                peptide_id=peptide.id,
                analysis_service_id=a.analysis_service_id,
                sample_id=a.sample_id,
                slot=a.slot,
                component_peptide_id=a.component_peptide_id,
            ))

    db.commit()
    db.refresh(peptide)
    return _peptide_to_response(db, peptide)


@app.delete("/peptides/wipe-all")
async def wipe_all_peptides(db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Delete ALL peptide standards, calibration curves, and SharePoint file cache."""
    cache_deleted = db.execute(delete(SharePointFileCache)).rowcount
    curves_deleted = db.execute(delete(CalibrationCurve)).rowcount
    peptides_deleted = db.execute(delete(Peptide)).rowcount
    db.commit()
    return {
        "message": f"Wiped {peptides_deleted} peptides, {curves_deleted} curves, and {cache_deleted} cached file records",
        "peptides_deleted": peptides_deleted,
        "curves_deleted": curves_deleted,
        "cache_deleted": cache_deleted,
    }


@app.post("/peptides/seed-from-services")
async def seed_peptides_from_services(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Seed the Peptides list from Analysis Services.
    Creates one peptide per unique peptide_name found in services with
    category 'Peptide Identity' and title containing 'Identity (HPLC)'.
    Each matching service is linked as an analyte (slots 1-4).
    Existing peptides (by abbreviation) are skipped.
    """
    # 1. Query matching analysis services
    services = db.execute(
        select(AnalysisService)
        .where(AnalysisService.category == "Peptide Identity")
        .where(AnalysisService.title.contains("Identity (HPLC)"))
        .where(AnalysisService.active == True)
        .order_by(AnalysisService.title)
    ).scalars().all()

    if not services:
        return {"created": 0, "skipped": 0, "message": "No matching analysis services found"}

    # 2. Group services by peptide_name
    from collections import defaultdict
    grouped: dict[str, list] = defaultdict(list)
    for svc in services:
        pname = (svc.peptide_name or "").strip()
        if pname:
            grouped[pname].append(svc)

    # 3. Get existing peptide abbreviations
    existing_abbrs = set(
        row[0] for row in db.execute(select(Peptide.abbreviation)).all()
    )

    created = 0
    skipped = 0
    for pname, svcs in sorted(grouped.items()):
        abbr = pname.upper()
        if abbr in existing_abbrs:
            skipped += 1
            continue

        # Create the peptide
        peptide = Peptide(name=pname, abbreviation=abbr)
        db.add(peptide)
        db.flush()

        # Link services as analytes (max 4 slots)
        for slot_idx, svc in enumerate(svcs[:4], start=1):
            db.add(PeptideAnalyte(
                peptide_id=peptide.id,
                analysis_service_id=svc.id,
                slot=slot_idx,
            ))

        created += 1
        existing_abbrs.add(abbr)

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "total_services": len(services),
        "message": f"Created {created} peptide(s), skipped {skipped} existing",
    }


@app.put("/peptides/{peptide_id}", response_model=PeptideResponse)
async def update_peptide(peptide_id: int, data: PeptideUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Update a peptide. method_ids sets all method assignments (one per instrument)."""
    peptide = db.execute(
        select(Peptide).options(
            joinedload(Peptide.methods).joinedload(HplcMethod.instruments),
            joinedload(Peptide.analytes).joinedload(PeptideAnalyte.analysis_service),
            joinedload(Peptide.analytes).joinedload(PeptideAnalyte.component_peptide),
            joinedload(Peptide.components),
        )
        .where(Peptide.id == peptide_id)
    ).scalars().unique().one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    update_data = data.model_dump(exclude_unset=True)
    method_ids = update_data.pop("method_ids", None)
    analytes_data = update_data.pop("analytes", None)
    component_ids = update_data.pop("component_ids", None)
    vial_assignments = update_data.pop("component_vial_assignments", None)

    # Update scalar fields (includes prep_vial_count if provided)
    for field, value in update_data.items():
        setattr(peptide, field, value)

    # User tracking — stamp updated_by, backfill created_by if missing
    peptide.updated_by_user_id = current_user.id
    peptide.updated_by_email = current_user.email
    if not peptide.created_by_user_id:
        peptide.created_by_user_id = current_user.id
        peptide.created_by_email = current_user.email

    # Update vial assignments for blend components (without replacing components)
    if vial_assignments is not None and component_ids is None:
        for comp_id_str, vial_num in vial_assignments.items():
            db.execute(
                blend_components.update()
                .where(blend_components.c.blend_id == peptide.id)
                .where(blend_components.c.component_id == int(comp_id_str))
                .values(vial_number=vial_num)
            )
        # If vial count was reset to 1, normalize all to vial 1
        if update_data.get("prep_vial_count") == 1:
            db.execute(
                blend_components.update()
                .where(blend_components.c.blend_id == peptide.id)
                .values(vial_number=1)
            )

    # Update blend components if provided (delete-and-replace); auto-rebuild analytes
    if component_ids is not None:
        db.execute(blend_components.delete().where(blend_components.c.blend_id == peptide.id))
        # Also rebuild analyte slots from components
        db.execute(delete(PeptideAnalyte).where(PeptideAnalyte.peptide_id == peptide.id))
        if component_ids:
            components = db.execute(
                select(Peptide).where(Peptide.id.in_(component_ids))
            ).scalars().all()
            if any(c.is_blend for c in components):
                raise HTTPException(400, "Blends cannot contain other blends")
            comp_map = {c.id: c for c in components}
            slot_num = 1
            for order, comp_id in enumerate(component_ids):
                db.execute(blend_components.insert().values(
                    blend_id=peptide.id, component_id=comp_id, display_order=order,
                ))
                comp = comp_map.get(comp_id)
                if comp:
                    comp_analyte = db.execute(
                        select(PeptideAnalyte).where(
                            PeptideAnalyte.peptide_id == comp_id,
                            PeptideAnalyte.slot == 1,
                        )
                    ).scalar_one_or_none()
                    if comp_analyte:
                        db.add(PeptideAnalyte(
                            peptide_id=peptide.id,
                            analysis_service_id=comp_analyte.analysis_service_id,
                            sample_id=comp_analyte.sample_id,
                            slot=slot_num,
                            component_peptide_id=comp_id,
                        ))
                        slot_num += 1
    elif analytes_data is not None:
        # Update analyte slots if provided (delete-and-replace) — non-blend only
        db.execute(delete(PeptideAnalyte).where(PeptideAnalyte.peptide_id == peptide.id))
        for a in analytes_data:
            db.add(PeptideAnalyte(
                peptide_id=peptide.id,
                analysis_service_id=a["analysis_service_id"],
                sample_id=a.get("sample_id"),
                slot=a["slot"],
                component_peptide_id=a.get("component_peptide_id"),
            ))

    # Update method assignments if provided
    if method_ids is not None:
        if method_ids:
            methods = db.execute(
                select(HplcMethod).options(joinedload(HplcMethod.instruments))
                .where(HplcMethod.id.in_(method_ids))
            ).scalars().unique().all()
            if len(methods) != len(method_ids):
                raise HTTPException(400, "One or more method IDs not found")
            # Enforce one method per instrument: two methods can't share an instrument
            seen_instrument_ids: set[int] = set()
            for m in methods:
                for inst in m.instruments:
                    if inst.id in seen_instrument_ids:
                        raise HTTPException(400, "Cannot assign multiple methods for the same instrument")
                    seen_instrument_ids.add(inst.id)
            peptide.methods = list(methods)
        else:
            peptide.methods = []

    db.commit()
    db.refresh(peptide)
    return _peptide_to_response(db, peptide)


@app.delete("/peptides/{peptide_id}")
async def delete_peptide(peptide_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Delete a peptide and all its calibration curves."""
    peptide = db.execute(select(Peptide).where(Peptide.id == peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    db.delete(peptide)
    db.commit()
    return {"message": f"Peptide '{peptide.abbreviation}' deleted"}


@app.get("/peptides/{peptide_id}/calibrations", response_model=list[CalibrationCurveResponse])
async def get_calibrations(peptide_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get all calibration curves for a peptide (newest first)."""
    peptide = db.execute(select(Peptide).where(Peptide.id == peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    from sqlalchemy.orm import defer
    stmt = (
        select(CalibrationCurve)
        .options(defer(CalibrationCurve.chromatogram_data))
        .where(CalibrationCurve.peptide_id == peptide_id)
        .order_by(desc(CalibrationCurve.created_at))
    )
    cals = db.execute(stmt).scalars().all()
    return [_cal_to_response(c, include_blobs=False) for c in cals]


@app.get("/peptides/{peptide_id}/blend-calibrations")
async def get_blend_calibrations(peptide_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get calibration curves for all component peptides of a blend, grouped by component."""
    peptide = db.execute(
        select(Peptide).options(joinedload(Peptide.components))
        .where(Peptide.id == peptide_id)
    ).scalars().unique().one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")
    if not peptide.is_blend:
        raise HTTPException(400, "Not a blend peptide")

    result = {}
    comp_ids = [c.id for c in peptide.components]
    if comp_ids:
        cals = db.execute(
            select(CalibrationCurve)
            .where(CalibrationCurve.peptide_id.in_(comp_ids))
            .order_by(desc(CalibrationCurve.created_at))
        ).scalars().all()
        # Group by component
        cal_map: dict[int, list] = {cid: [] for cid in comp_ids}
        for cal in cals:
            cal_map[cal.peptide_id].append(_cal_to_response(cal))

        for comp in peptide.components:
            result[comp.abbreviation] = {
                "peptide_id": comp.id,
                "name": comp.name,
                "calibrations": cal_map.get(comp.id, []),
            }

    return result


@app.post("/peptides/{peptide_id}/calibrations", response_model=CalibrationCurveResponse, status_code=201)
async def create_calibration(
    peptide_id: int,
    data: CalibrationDataInput,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Create a calibration curve from concentration/area pairs.

    Calculates linear regression and stores the curve.
    Automatically sets this as the active calibration for the peptide.
    """
    peptide = db.execute(select(Peptide).where(Peptide.id == peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    try:
        regression = calculate_calibration_curve(data.concentrations, data.areas)
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Resolve analyte if provided
    resolved_analyte_id: Optional[int] = None
    if data.analyte_id:
        analyte = db.execute(
            select(PeptideAnalyte).where(
                PeptideAnalyte.id == data.analyte_id,
                PeptideAnalyte.peptide_id == peptide_id,
            )
        ).scalar_one_or_none()
        if analyte:
            resolved_analyte_id = analyte.id

    # Build standard_data with optional RTs
    std_data: dict = {
        "concentrations": data.concentrations,
        "areas": data.areas,
    }
    if data.rts:
        std_data["rts"] = data.rts

    # Compute reference RT from provided RTs
    avg_rt = None
    if data.rts and len(data.rts) > 0:
        avg_rt = round(sum(data.rts) / len(data.rts), 4)

    # Resolve instrument_id → name (or vice versa) for the curve
    resolved_instrument_name = data.instrument
    resolved_instrument_id = data.instrument_id
    if resolved_instrument_id and not resolved_instrument_name:
        inst = db.execute(select(Instrument).where(Instrument.id == resolved_instrument_id)).scalar_one_or_none()
        if inst:
            resolved_instrument_name = inst.name
    elif resolved_instrument_name and not resolved_instrument_id:
        inst = db.execute(select(Instrument).where(Instrument.name == resolved_instrument_name)).scalar_one_or_none()
        if inst:
            resolved_instrument_id = inst.id

    # Deactivate existing active curves for this peptide on the same instrument
    deactivate_query = (
        select(CalibrationCurve)
        .where(CalibrationCurve.peptide_id == peptide_id)
        .where(CalibrationCurve.is_active == True)
    )
    if resolved_instrument_id is not None:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id == resolved_instrument_id)
    else:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id.is_(None))
    for cal in db.execute(deactivate_query).scalars().all():
        cal.is_active = False

    # Create new active curve
    from datetime import datetime, timezone
    curve = CalibrationCurve(
        peptide_id=peptide_id,
        peptide_analyte_id=resolved_analyte_id,
        slope=regression["slope"],
        intercept=regression["intercept"],
        r_squared=regression["r_squared"],
        standard_data=std_data,
        source_filename=data.source_filename or "Manual entry",
        source_date=datetime.now(timezone.utc).isoformat(),
        instrument=resolved_instrument_name,
        instrument_id=resolved_instrument_id,
        reference_rt=avg_rt,
        notes=data.notes,
        is_active=True,
        created_by_user_id=current_user.id,
        created_by_email=current_user.email,
        updated_by_user_id=current_user.id,
        updated_by_email=current_user.email,
    )
    db.add(curve)
    db.commit()
    db.refresh(curve)
    return curve


@app.post("/peptides/{peptide_id}/calibrations/from-standard", response_model=CalibrationCurveResponse, status_code=201)
async def create_calibration_from_standard(
    peptide_id: int,
    data: StandardCalibrationInput,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Auto-create a calibration curve from standard HPLC processing results.

    Receives computed concentration/area/RT data from standard processing,
    validates the source sample prep is a standard, computes regression,
    and creates a fully-linked CalibrationCurve with provenance fields.
    """
    # 1. Validate peptide exists
    peptide = db.execute(select(Peptide).where(Peptide.id == peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    # 2. Validate sample_prep_id exists and is a standard
    from mk1_db import ensure_sample_preps_table, get_mk1_db
    from psycopg2.extras import RealDictCursor
    ensure_sample_preps_table()
    with get_mk1_db() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT id, sample_id, is_standard, instrument_name FROM sample_preps WHERE sample_id = %s OR senaite_sample_id = %s",
                [data.sample_prep_id, data.sample_prep_id],
            )
            prep_row = cur.fetchone()
    if not prep_row:
        raise HTTPException(400, f"Sample prep '{data.sample_prep_id}' not found")
    if not prep_row.get("is_standard"):
        raise HTTPException(400, f"Sample prep '{data.sample_prep_id}' is not a standard")
    # Use instrument from request, falling back to what's stored on the sample prep
    resolved_instrument_name = data.instrument or prep_row.get("instrument_name")
    # Resolve to instrument_id FK
    resolved_instrument_id = None
    if resolved_instrument_name:
        inst_row = db.execute(
            select(Instrument).where(Instrument.name == resolved_instrument_name)
        ).scalar_one_or_none()
        if inst_row:
            resolved_instrument_id = inst_row.id
            resolved_instrument_name = inst_row.name  # normalize

    # 3. Resolve analyte — use first PeptideAnalyte for this peptide
    first_analyte = db.execute(
        select(PeptideAnalyte)
        .where(PeptideAnalyte.peptide_id == peptide_id)
        .order_by(PeptideAnalyte.slot)
        .limit(1)
    ).scalar_one_or_none()
    resolved_analyte_id = first_analyte.id if first_analyte else None

    # 4. Calculate regression
    try:
        regression = calculate_calibration_curve(data.concentrations, data.areas)
    except ValueError as e:
        raise HTTPException(400, str(e))

    # 5. Deactivate existing active curves for this peptide on the same instrument
    deactivate_query = (
        select(CalibrationCurve)
        .where(CalibrationCurve.peptide_id == peptide_id)
        .where(CalibrationCurve.is_active == True)
    )
    if resolved_instrument_id is not None:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id == resolved_instrument_id)
    else:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id.is_(None))
    for cal in db.execute(deactivate_query).scalars().all():
        cal.is_active = False

    # 6. Compute reference RT from provided RTs
    avg_rt = None
    if data.rts and len(data.rts) > 0:
        avg_rt = round(sum(data.rts) / len(data.rts), 4)

    # 7. Build standard_data dict
    std_data: dict = {
        "concentrations": data.concentrations,
        "areas": data.areas,
    }
    if data.rts:
        std_data["rts"] = data.rts

    # 8. Create CalibrationCurve with full provenance
    from datetime import datetime, timezone
    curve = CalibrationCurve(
        peptide_id=peptide_id,
        peptide_analyte_id=resolved_analyte_id,
        slope=regression["slope"],
        intercept=regression["intercept"],
        r_squared=regression["r_squared"],
        standard_data=std_data,
        reference_rt=avg_rt,
        source_sample_id=data.sample_prep_id,
        chromatogram_data=data.chromatogram_data,
        source_sharepoint_folder=data.source_sharepoint_folder,
        vendor=data.vendor,
        notes=data.notes,
        instrument=resolved_instrument_name,
        instrument_id=resolved_instrument_id,
        source_filename=f"Standard: {data.sample_prep_id}",
        source_date=datetime.now(timezone.utc).isoformat(),
        is_active=True,
        created_by_user_id=current_user.id,
        created_by_email=current_user.email,
        updated_by_user_id=current_user.id,
        updated_by_email=current_user.email,
    )
    db.add(curve)
    db.commit()
    db.refresh(curve)
    return _cal_to_response(curve)


@app.get("/peptides/{peptide_id}/calibrations/{calibration_id}", response_model=CalibrationCurveResponse)
async def get_calibration(
    peptide_id: int,
    calibration_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Get a single calibration curve with full data (including chromatogram_data)."""
    cal = db.execute(
        select(CalibrationCurve)
        .where(CalibrationCurve.id == calibration_id, CalibrationCurve.peptide_id == peptide_id)
    ).scalar_one_or_none()
    if not cal:
        raise HTTPException(status_code=404, detail="Calibration not found")
    return _cal_to_response(cal)  # include_blobs=True (default) — full data


@app.post("/peptides/{peptide_id}/calibrations/{calibration_id}/activate", response_model=CalibrationCurveResponse)
async def activate_calibration(
    peptide_id: int,
    calibration_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Set a specific calibration curve as the active one for a peptide.
    Deactivates all other curves for that peptide.
    """
    peptide = db.execute(select(Peptide).where(Peptide.id == peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {peptide_id} not found")

    target = db.execute(
        select(CalibrationCurve)
        .where(CalibrationCurve.id == calibration_id)
        .where(CalibrationCurve.peptide_id == peptide_id)
    ).scalar_one_or_none()
    if not target:
        raise HTTPException(404, f"Calibration {calibration_id} not found for peptide {peptide_id}")

    # Deactivate curves for this peptide on the same instrument
    deactivate_query = select(CalibrationCurve).where(CalibrationCurve.peptide_id == peptide_id)
    if target.instrument_id is not None:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id == target.instrument_id)
    else:
        deactivate_query = deactivate_query.where(CalibrationCurve.instrument_id.is_(None))
    all_cals = db.execute(deactivate_query).scalars().all()
    for cal in all_cals:
        cal.is_active = False

    # Activate the target
    target.is_active = True
    target.updated_by_user_id = current_user.id
    target.updated_by_email = current_user.email
    if not target.created_by_user_id:
        target.created_by_user_id = current_user.id
        target.created_by_email = current_user.email

    # Update curve's reference RT from its own RT data if not already set
    if target.reference_rt is None and target.standard_data and target.standard_data.get("rts"):
        rts = target.standard_data["rts"]
        if rts:
            target.reference_rt = round(sum(rts) / len(rts), 4)

    db.commit()
    db.refresh(target)
    return _cal_to_response(target)


class CalibrationCurveUpdate(BaseModel):
    """Partial update schema for a calibration curve."""
    reference_rt: Optional[float] = None
    rt_tolerance: Optional[float] = None
    diluent_density: Optional[float] = None
    instrument: Optional[str] = None
    instrument_id: Optional[int] = None
    peptide_analyte_id: Optional[int] = None
    notes: Optional[str] = None
    source_sample_id: Optional[str] = None
    vendor: Optional[str] = None
    standard_data: Optional[dict] = None  # {concentrations, areas, rts?, excluded_indices?}

    class Config:
        from_attributes = True


@app.patch("/peptides/{peptide_id}/calibrations/{calibration_id}", response_model=CalibrationCurveResponse)
async def update_calibration(
    peptide_id: int,
    calibration_id: int,
    body: CalibrationCurveUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Partial update of a calibration curve's editable fields."""
    target = db.execute(
        select(CalibrationCurve)
        .where(CalibrationCurve.id == calibration_id)
        .where(CalibrationCurve.peptide_id == peptide_id)
    ).scalar_one_or_none()
    if not target:
        raise HTTPException(404, f"Calibration {calibration_id} not found for peptide {peptide_id}")

    updates = body.model_dump(exclude_unset=True)

    # When instrument_id is set, resolve the instrument name for the legacy string field
    if "instrument_id" in updates and updates["instrument_id"] is not None:
        inst = db.execute(
            select(Instrument).where(Instrument.id == updates["instrument_id"])
        ).scalar_one_or_none()
        if inst:
            updates["instrument"] = inst.name

    # When source_sample_id is set/changed, auto-fetch chromatogram from SharePoint
    new_sample_id = updates.get("source_sample_id")
    if new_sample_id and new_sample_id != target.source_sample_id:
        try:
            import sharepoint as sp
            sample_files = await sp.get_sample_files(new_sample_id)
            if sample_files and sample_files.get("chromatogram_files"):
                import re as _re
                chrom_by_conc: dict[str, dict] = {}
                for chrom_file in sample_files["chromatogram_files"]:
                    fname = chrom_file["name"]
                    # Skip blanks
                    if "blank" in fname.lower():
                        continue
                    # Extract concentration from filename like P-0309_Std_250.dx_DAD1A.CSV
                    conc_match = _re.search(r'_Std_(\d+)\.', fname)
                    conc_label = conc_match.group(1) if conc_match else fname.split(".")[0]
                    try:
                        content_bytes, _ = await sp.download_file(chrom_file["id"])
                        csv_text = content_bytes.decode("utf-8", errors="replace")
                        times = []
                        signals = []
                        for line in csv_text.splitlines():
                            line = line.strip()
                            if not line:
                                continue
                            parts = line.split(",", 1)
                            if len(parts) != 2:
                                continue
                            try:
                                times.append(float(parts[0]))
                                signals.append(float(parts[1]))
                            except ValueError:
                                continue
                        if times:
                            chrom_by_conc[conc_label] = {"times": times, "signals": signals}
                    except Exception:
                        continue  # skip individual file failures
                if chrom_by_conc:
                    updates["chromatogram_data"] = chrom_by_conc
                    updates["source_sharepoint_folder"] = sample_files["sample"]["path"]
        except Exception as e:
            # Chromatogram fetch is best-effort — log and continue
            import logging
            logging.getLogger(__name__).warning(
                f"Chromatogram auto-fetch failed for sample '{new_sample_id}': {e}"
            )

    # When standard_data is updated, recalculate regression from non-excluded points
    if "standard_data" in updates and updates["standard_data"]:
        from calculations.calibration import calculate_calibration_curve as _calc_curve
        sd = updates["standard_data"]
        concs = sd.get("concentrations", [])
        areas = sd.get("areas", [])
        excluded = set(sd.get("excluded_indices", []))
        # Filter to only included points
        inc_concs = [c for i, c in enumerate(concs) if i not in excluded]
        inc_areas = [a for i, a in enumerate(areas) if i not in excluded]
        if len(inc_concs) >= 2:
            try:
                reg = _calc_curve(inc_concs, inc_areas)
                updates["slope"] = reg["slope"]
                updates["intercept"] = reg["intercept"]
                updates["r_squared"] = reg["r_squared"]
            except Exception:
                pass  # Keep existing regression if recalc fails
        # Recalculate reference_rt from non-excluded RTs
        rts = sd.get("rts", [])
        inc_rts = [r for i, r in enumerate(rts) if i not in excluded and i < len(rts)]
        if inc_rts:
            updates["reference_rt"] = round(sum(inc_rts) / len(inc_rts), 4)

    for field, value in updates.items():
        setattr(target, field, value)

    # User tracking
    target.updated_by_user_id = current_user.id
    target.updated_by_email = current_user.email
    if not target.created_by_user_id:
        target.created_by_user_id = current_user.id
        target.created_by_email = current_user.email

    db.commit()
    db.refresh(target)
    return _cal_to_response(target)


@app.delete("/peptides/{peptide_id}/calibrations/{calibration_id}", status_code=204)
async def delete_calibration(
    peptide_id: int,
    calibration_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Delete a single calibration curve."""
    target = db.execute(
        select(CalibrationCurve)
        .where(CalibrationCurve.id == calibration_id)
        .where(CalibrationCurve.peptide_id == peptide_id)
    ).scalar_one_or_none()
    if not target:
        raise HTTPException(404, f"Calibration {calibration_id} not found for peptide {peptide_id}")

    # Nullify FK references in wizard_sessions before deleting
    db.execute(
        update(WizardSession)
        .where(WizardSession.calibration_curve_id == calibration_id)
        .values(calibration_curve_id=None)
    )
    db.delete(target)
    db.commit()
    return Response(status_code=204)


# --- Full HPLC Analysis Endpoint ---

class HPLCWeightsInput(BaseModel):
    """Five balance weights from the tech."""
    stock_vial_empty: float
    stock_vial_with_diluent: float
    dil_vial_empty: float
    dil_vial_with_diluent: float
    dil_vial_with_diluent_and_sample: float


class HPLCAnalyzeRequest(BaseModel):
    """Request to run a full HPLC analysis."""
    sample_id_label: str
    peptide_id: int
    calibration_curve_id: Optional[int] = None  # If provided, use this specific curve
    weights: HPLCWeightsInput
    injections: list[dict]  # Parsed injection data from /hplc/parse-files
    # Provenance fields (Phase 10.5)
    sample_prep_id: Optional[int] = None
    instrument_id: Optional[int] = None
    source_sharepoint_folder: Optional[str] = None
    chromatogram_data: Optional[dict] = None
    run_group_id: Optional[str] = None
    # Phase 13: standard injection reference RTs keyed by analyte label
    # Format: {"BPC157": {"rt": 10.165, "source_sample_id": "P-0111"}, ...}
    standard_injection_rts: Optional[dict[str, dict]] = None
    # Phase 13.5: Audit trail — debug log and source file archive
    debug_log: Optional[list[dict]] = None  # [{level: str, msg: str}] from frontend buildDebugLines()
    source_files: Optional[list[dict]] = None  # [{filename: str, content: str, sha256: str}] archived source files


class HPLCAnalysisResponse(BaseModel):
    """Full analysis result."""
    id: int
    sample_id_label: str
    peptide_id: int
    peptide_abbreviation: str
    purity_percent: Optional[float]
    quantity_mg: Optional[float]
    identity_conforms: Optional[bool]
    identity_rt_delta: Optional[float]
    dilution_factor: Optional[float]
    stock_volume_ml: Optional[float]
    avg_main_peak_area: Optional[float]
    concentration_ug_ml: Optional[float]
    calculation_trace: Optional[dict]
    raw_data: Optional[dict] = None
    created_at: datetime
    # Provenance fields (Phase 10.5)
    calibration_curve_id: Optional[int] = None
    sample_prep_id: Optional[int] = None
    instrument_id: Optional[int] = None
    source_sharepoint_folder: Optional[str] = None
    chromatogram_data: Optional[dict] = None
    run_group_id: Optional[str] = None
    # Phase 13: identity reference source
    identity_reference_source: Optional[str] = None      # "standard_injection" or "calibration_curve"
    identity_reference_source_id: Optional[str] = None  # e.g. "P-0111" for standard injection
    # Phase 13.5: Audit trail
    debug_log: Optional[list[dict]] = None


def _analysis_to_response(analysis: "HPLCAnalysis", peptide_abbreviation: str) -> "HPLCAnalysisResponse":
    """Convert an HPLCAnalysis ORM object to an HPLCAnalysisResponse."""
    identity_trace = (analysis.calculation_trace or {}).get("identity", {})
    return HPLCAnalysisResponse(
        id=analysis.id,
        sample_id_label=analysis.sample_id_label,
        peptide_id=analysis.peptide_id,
        peptide_abbreviation=peptide_abbreviation,
        purity_percent=analysis.purity_percent,
        quantity_mg=analysis.quantity_mg,
        identity_conforms=analysis.identity_conforms,
        identity_rt_delta=analysis.identity_rt_delta,
        dilution_factor=analysis.dilution_factor,
        stock_volume_ml=analysis.stock_volume_ml,
        avg_main_peak_area=analysis.avg_main_peak_area,
        concentration_ug_ml=analysis.concentration_ug_ml,
        calculation_trace=analysis.calculation_trace,
        raw_data=analysis.raw_data,
        created_at=analysis.created_at,
        calibration_curve_id=analysis.calibration_curve_id,
        sample_prep_id=analysis.sample_prep_id,
        instrument_id=analysis.instrument_id,
        source_sharepoint_folder=analysis.source_sharepoint_folder,
        chromatogram_data=analysis.chromatogram_data,
        run_group_id=analysis.run_group_id,
        # Phase 13: identity reference source (extracted from calculation_trace)
        identity_reference_source=identity_trace.get("reference_source"),
        identity_reference_source_id=identity_trace.get("reference_source_id"),
        # Phase 13.5: Audit trail
        debug_log=analysis.debug_log,
    )


@app.post("/hplc/analyze", response_model=HPLCAnalysisResponse, status_code=201)
async def run_hplc_analysis(
    request: HPLCAnalyzeRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Run a complete HPLC analysis: purity + quantity + identity.

    Requires parsed injection data (from /hplc/parse-files), peptide selection,
    and 5 balance weights for dilution factor calculation.
    """
    # Load peptide
    peptide = db.execute(select(Peptide).where(Peptide.id == request.peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(404, f"Peptide {request.peptide_id} not found")

    # Load calibration curve: use specific one if requested, else fall back to active
    if request.calibration_curve_id:
        cal = db.execute(
            select(CalibrationCurve)
            .where(CalibrationCurve.id == request.calibration_curve_id)
            .where(CalibrationCurve.peptide_id == peptide.id)
        ).scalar_one_or_none()
        if not cal:
            raise HTTPException(404, f"Calibration curve {request.calibration_curve_id} not found for peptide '{peptide.abbreviation}'")
    else:
        # Look up active curve matching the request's instrument
        cal_query = (
            select(CalibrationCurve)
            .where(CalibrationCurve.peptide_id == peptide.id)
            .where(CalibrationCurve.is_active == True)
        )
        if request.instrument_id:
            cal_query = cal_query.where(CalibrationCurve.instrument_id == request.instrument_id)
        cal = db.execute(
            cal_query.order_by(desc(CalibrationCurve.created_at)).limit(1)
        ).scalar_one_or_none()
        if not cal:
            if request.instrument_id:
                inst = db.execute(select(Instrument).where(Instrument.id == request.instrument_id)).scalar_one_or_none()
                inst_name = inst.name if inst else f"ID {request.instrument_id}"
                raise HTTPException(400, f"No active calibration curve for peptide '{peptide.abbreviation}' on instrument '{inst_name}'. Star a curve for this instrument first.")
            raise HTTPException(400, f"No active calibration curve for peptide '{peptide.abbreviation}'")

    # Resolve reference RT: prefer curve setting, fall back to calibration standard RTs
    ref_rt = cal.reference_rt
    if ref_rt is None and cal.standard_data:
        cal_rts = cal.standard_data.get("rts", [])
        if cal_rts:
            ref_rt = round(sum(cal_rts) / len(cal_rts), 4)
            # Persist on the curve so future analyses don't need this fallback
            cal.reference_rt = ref_rt
            db.flush()

    # Phase 13: Resolve standard injection RT for this analyte using alias-aware matching
    # Standard files use labels like "BPC157" but peptide abbreviation is "BPC-157".
    # Normalize both by stripping non-alphanumeric chars, then also check hplc_aliases.
    std_injection_rt: Optional[float] = None
    std_injection_source: Optional[str] = None
    if request.standard_injection_rts:
        def _normalize_label(s: str) -> str:
            return re.sub(r"[^a-zA-Z0-9]", "", s).upper()

        peptide_abbr_norm = _normalize_label(peptide.abbreviation)
        peptide_aliases = getattr(peptide, "hplc_aliases", None) or []
        alias_norms = {_normalize_label(a) for a in peptide_aliases}

        for label, std_info in request.standard_injection_rts.items():
            label_norm = _normalize_label(label)
            if label_norm == peptide_abbr_norm or label_norm in alias_norms:
                std_injection_rt = std_info.get("rt")
                std_injection_source = std_info.get("source_sample_id")
                break

        # Fallback: single-peptide standard injection with no analyte label
        # e.g. P-0416_Inj_1_Std_PeakData.csv — no label between _Std_ and _PeakData
        if std_injection_rt is None:
            unlabeled = [
                (lbl, info) for lbl, info in request.standard_injection_rts.items()
                if not lbl  # empty string = no analyte label
            ]
            if len(unlabeled) == 1:
                std_injection_rt = unlabeled[0][1].get("rt")
                std_injection_source = unlabeled[0][1].get("source_sample_id")

    # Build analysis input
    analysis_input = AnalysisInput(
        injections=request.injections,
        weights=WeightInputs(
            stock_vial_empty=request.weights.stock_vial_empty,
            stock_vial_with_diluent=request.weights.stock_vial_with_diluent,
            dil_vial_empty=request.weights.dil_vial_empty,
            dil_vial_with_diluent=request.weights.dil_vial_with_diluent,
            dil_vial_with_diluent_and_sample=request.weights.dil_vial_with_diluent_and_sample,
        ),
        calibration=CalibrationParams(slope=cal.slope, intercept=cal.intercept),
        peptide=PeptideParams(
            reference_rt=ref_rt,
            rt_tolerance=cal.rt_tolerance,
            diluent_density=cal.diluent_density,
            standard_injection_rt=std_injection_rt,
            standard_injection_source=std_injection_source,
        ),
    )

    # Run analysis
    result = process_hplc_analysis(analysis_input)

    # Store in database
    analysis = HPLCAnalysis(
        sample_id_label=request.sample_id_label,
        peptide_id=peptide.id,
        stock_vial_empty=request.weights.stock_vial_empty,
        stock_vial_with_diluent=request.weights.stock_vial_with_diluent,
        dil_vial_empty=request.weights.dil_vial_empty,
        dil_vial_with_diluent=request.weights.dil_vial_with_diluent,
        dil_vial_with_diluent_and_sample=request.weights.dil_vial_with_diluent_and_sample,
        dilution_factor=result.get("dilution_factor"),
        stock_volume_ml=result.get("stock_volume_ml"),
        avg_main_peak_area=result.get("avg_main_peak_area"),
        concentration_ug_ml=result.get("concentration_ug_ml"),
        purity_percent=result.get("purity_percent"),
        quantity_mg=result.get("quantity_mg"),
        identity_conforms=result.get("identity_conforms"),
        identity_rt_delta=result.get("identity_rt_delta"),
        calculation_trace=result.get("calculation_trace"),
        raw_data={
            "injections": request.injections,
            **({"source_files": request.source_files} if request.source_files else {}),
            **({"file_manifest": [
                {"filename": f["filename"], "sha256": f["sha256"], "size": len(f.get("content", ""))}
                for f in request.source_files
            ]} if request.source_files else {}),
        },
        # Phase 13.5: Audit trail
        debug_log=request.debug_log,
        # Phase 10.5: Provenance fields
        calibration_curve_id=cal.id,
        sample_prep_id=request.sample_prep_id,
        instrument_id=request.instrument_id,
        source_sharepoint_folder=request.source_sharepoint_folder,
        chromatogram_data=request.chromatogram_data,
        run_group_id=request.run_group_id,
        # User tracking
        processed_by_user_id=current_user.id,
        processed_by_email=current_user.email,
    )
    db.add(analysis)

    # Audit log
    audit = AuditLog(
        operation="hplc_analysis",
        entity_type="hplc_analysis",
        entity_id=request.sample_id_label,
        details={
            "peptide": peptide.abbreviation,
            "purity": result.get("purity_percent"),
            "quantity_mg": result.get("quantity_mg"),
            "identity_conforms": result.get("identity_conforms"),
        },
    )
    db.add(audit)
    db.commit()
    db.refresh(analysis)

    # Bridge: a vial-scoped sample prep pushes its HPLC result onto the vial's
    # lims_analyses row(s) and submits. The analysis above is already committed
    # (db.commit at the line prior), so a bridge failure can never lose it — we
    # roll back any partial bridge mutation and continue.
    if request.sample_prep_id is not None:
        try:
            import mk1_db
            _prep = mk1_db.get_sample_prep(request.sample_prep_id)
            _sub_pk = _prep.get("lims_sub_sample_pk") if _prep else None
            if _sub_pk is not None:
                from lims_analyses.prep_bridge import bridge_prep_result_to_vial, bridge_blend_aggregates
                bridge_prep_result_to_vial(
                    db,
                    lims_sub_sample_pk=_sub_pk,
                    analysis=analysis,
                    peptide=peptide,
                    user_id=current_user.id,
                )
                # Fill blend aggregates (BLEND-PUR / PEPT-Total) once this save
                # completes the per-component set; no-op on partial blends/singles.
                bridge_blend_aggregates(db, lims_sub_sample_pk=_sub_pk, user_id=current_user.id)
        except Exception:
            db.rollback()
            logger.exception("prep_bridge: failed for sample_prep_id=%s", request.sample_prep_id)

    return _analysis_to_response(analysis, peptide.abbreviation)


@app.post("/hplc/sample-preps/{prep_id}/bridge")
def rebridge_sample_prep(
    prep_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Re-run the vial-prep result bridge for every HPLC analysis on this prep.

    Idempotent (only 'unassigned' lims_analyses rows are touched). Used by the
    flyout's vial-results view Auto-fill. 404 unknown prep; 409 when the prep
    is parent-scoped or has no HPLC analyses recorded yet."""
    from lims_analyses.prep_bridge import rebridge_prep
    try:
        submitted = rebridge_prep(db, prep_id=prep_id, user_id=current_user.id)
    except LookupError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    return {"submitted": submitted, "count": len(submitted)}


# --- HPLC Analysis History Endpoints ---

class HPLCAnalysisListItem(BaseModel):
    """Summary item for analysis history list."""
    id: int
    sample_id_label: str
    peptide_abbreviation: str
    purity_percent: Optional[float]
    quantity_mg: Optional[float]
    identity_conforms: Optional[bool]
    created_at: datetime

    class Config:
        from_attributes = True


class HPLCAnalysisListResponse(BaseModel):
    """Paginated list of analyses."""
    items: list[HPLCAnalysisListItem]
    total: int


@app.get("/hplc/analyses", response_model=HPLCAnalysisListResponse)
async def get_hplc_analyses(
    search: Optional[str] = None,
    peptide_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    List HPLC analyses with optional search and filtering.

    Query params:
    - search: Filter by sample_id_label (partial match)
    - peptide_id: Filter by peptide
    - limit/offset: Pagination
    """
    from sqlalchemy import func

    base = select(HPLCAnalysis)
    count_base = select(func.count(HPLCAnalysis.id))

    if search:
        base = base.where(HPLCAnalysis.sample_id_label.ilike(f"%{search}%"))
        count_base = count_base.where(HPLCAnalysis.sample_id_label.ilike(f"%{search}%"))
    if peptide_id is not None:
        base = base.where(HPLCAnalysis.peptide_id == peptide_id)
        count_base = count_base.where(HPLCAnalysis.peptide_id == peptide_id)

    total = db.execute(count_base).scalar() or 0

    stmt = base.order_by(desc(HPLCAnalysis.created_at)).offset(offset).limit(limit)
    analyses = db.execute(stmt).scalars().all()

    # Build list items with peptide abbreviation
    items = []
    for a in analyses:
        peptide = db.execute(select(Peptide).where(Peptide.id == a.peptide_id)).scalar_one_or_none()
        items.append(HPLCAnalysisListItem(
            id=a.id,
            sample_id_label=a.sample_id_label,
            peptide_abbreviation=peptide.abbreviation if peptide else "?",
            purity_percent=a.purity_percent,
            quantity_mg=a.quantity_mg,
            identity_conforms=a.identity_conforms,
            created_at=a.created_at,
        ))

    return HPLCAnalysisListResponse(items=items, total=total)


@app.get("/hplc/analyses/by-sample-prep/{sample_prep_id}", response_model=list[HPLCAnalysisResponse])
async def get_analyses_by_sample_prep(
    sample_prep_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Return all HPLC analyses recorded for a given sample_prep_id.

    Ordered newest-first. Returns empty list if none found.
    NOTE: Registered before /{analysis_id} routes to prevent FastAPI treating
    the literal segment "by-sample-prep" as an integer path parameter.
    """
    analyses = db.execute(
        select(HPLCAnalysis)
        .where(HPLCAnalysis.sample_prep_id == sample_prep_id)
        .order_by(desc(HPLCAnalysis.created_at))
    ).scalars().all()

    result = []
    for a in analyses:
        peptide = db.execute(select(Peptide).where(Peptide.id == a.peptide_id)).scalar_one_or_none()
        result.append(_analysis_to_response(a, peptide.abbreviation if peptide else "?"))
    return result


@app.post("/hplc/analyses/{analysis_id}/chromatogram-image")
async def render_chromatogram_image(
    analysis_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Render chromatogram PNG via Integration Service and return it directly.

    Used by the frontend to display a preview before uploading to SENAITE.
    """
    analysis = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.id == analysis_id)
    ).scalar_one_or_none()
    if not analysis:
        raise HTTPException(404, f"HPLC Analysis {analysis_id} not found")

    chrom = analysis.chromatogram_data
    if not chrom or not chrom.get("times") or not chrom.get("signals"):
        raise HTTPException(400, "No chromatogram data stored on this analysis")

    render_url = f"{INTEGRATION_SERVICE_URL}/v1/chromatogram/render"
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=30.0) as client:
            render_resp = await client.post(
                render_url,
                json={
                    "times": chrom["times"],
                    "signals": chrom["signals"],
                    "sample_id": analysis.sample_id_label,
                },
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            render_resp.raise_for_status()
    except httpx.HTTPStatusError as e:
        raise HTTPException(502, f"Integration Service render failed: {e.response.status_code}")
    except httpx.RequestError as e:
        raise HTTPException(502, f"Integration Service unreachable: {e}")

    return Response(
        content=render_resp.content,
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="chromatogram_{analysis.sample_id_label}.png"'},
    )


@app.get("/hplc/chromatogram-status")
async def get_chromatogram_status(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Return set of sample_prep_ids that have chromatogram_data stored."""
    from sqlalchemy import text as sa_text
    rows = db.execute(
        sa_text("SELECT DISTINCT sample_prep_id FROM hplc_analyses WHERE sample_prep_id IS NOT NULL AND chromatogram_data IS NOT NULL")
    ).fetchall()
    return {"prep_ids_with_chromatogram": [r[0] for r in rows]}


@app.post("/hplc/analyses/{analysis_id}/refetch-chromatogram")
async def refetch_chromatogram_data(
    analysis_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Re-fetch chromatogram data from SharePoint for an existing analysis.

    Reads the stored injections to find the one with the largest main peak,
    downloads the matching DAD1A CSV from SharePoint, downsamples to 800 points,
    and updates chromatogram_data on the analysis record.
    """
    analysis = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.id == analysis_id)
    ).scalar_one_or_none()
    if not analysis:
        raise HTTPException(404, f"HPLC Analysis {analysis_id} not found")

    # Find the injection with the largest main peak from stored raw_data
    raw_data = analysis.raw_data or {}
    injections = raw_data.get("injections", [])
    if not injections:
        raise HTTPException(400, "No injection data stored on this analysis")

    best_area = -1.0
    best_inj_name = ""
    for inj in injections:
        inj_name = inj.get("injection_name", "")
        peaks = inj.get("peaks", [])
        main_idx = inj.get("main_peak_index", 0)
        if 0 <= main_idx < len(peaks):
            area = peaks[main_idx].get("area", 0)
            if area > best_area:
                best_area = area
                best_inj_name = inj_name

    if not best_inj_name:
        raise HTTPException(400, "Could not determine best injection from stored data")

    # Build match tokens from the injection name
    # "TB17-23_Inj_1" → prefix "TB17-23_Inj_1", analyte label "TB17-23"
    # "P-0390_Inj1_PeakData" → prefix "P-0390_Inj1", sample "P-0390"
    best_prefix = re.sub(r'_PeakData$', '', best_inj_name, flags=re.IGNORECASE)
    # Extract analyte label: everything before _Inj (for blend injections like "TB17-23_Inj_1")
    analyte_match = re.match(r'^(.+?)_Inj', best_inj_name, re.IGNORECASE)
    analyte_label = analyte_match.group(1) if analyte_match else ""

    # Find chromatogram files on SharePoint
    import sharepoint as sp
    sample_id = analysis.sample_id_label
    sample_files = await sp.get_sample_files(sample_id)
    if not sample_files or not sample_files.get("chromatogram_files"):
        raise HTTPException(404, f"No chromatogram files found on SharePoint for {sample_id}")

    # Filter out blanks
    chrom_files = [cf for cf in sample_files["chromatogram_files"]
                   if "blank" not in cf["name"].lower()]

    if not chrom_files:
        raise HTTPException(404, f"No non-blank chromatogram files found for {sample_id}")

    # Match strategy (in priority order):
    # 1. Exact prefix match: "P-0390_Inj1" in filename
    # 2. Analyte label match: "TB17-23" in filename (for blends)
    # 3. Sample ID match: "PB-0078" in filename (fallback to first non-blank)
    matched_file = None

    # Strategy 1: exact prefix
    for cf in chrom_files:
        cf_stem = re.sub(r'\.dx_DAD1A\.CSV$', '', cf["name"], flags=re.IGNORECASE)
        if cf_stem.lower() == best_prefix.lower():
            matched_file = cf
            break

    # Strategy 2: analyte label in filename
    if not matched_file and analyte_label:
        for cf in chrom_files:
            if analyte_label.lower() in cf["name"].lower():
                matched_file = cf
                break

    # Strategy 3: sample ID in filename (first non-blank)
    if not matched_file:
        for cf in chrom_files:
            if sample_id.lower() in cf["name"].lower():
                matched_file = cf
                break

    # Strategy 4: just take first non-blank
    if not matched_file:
        matched_file = chrom_files[0]

    # Download and parse
    file_bytes, _ = await sp.download_file(matched_file["id"])
    csv_text = file_bytes.decode("utf-8", errors="replace")

    times = []
    signals = []
    for line in csv_text.strip().splitlines():
        parts = line.split(",")
        if len(parts) >= 2:
            try:
                times.append(float(parts[0]))
                signals.append(float(parts[1]))
            except ValueError:
                continue

    if len(times) < 10:
        raise HTTPException(400, f"Chromatogram file too small: {len(times)} points")

    # Downsample to 800 points using LTTB (pure Python — no numpy needed)
    target = 800
    if len(times) > target:
        data = list(zip(times, signals))
        sampled = [data[0]]
        bucket_size = (len(data) - 2) / (target - 2)
        prev_idx = 0
        for i in range(1, target - 1):
            avg_start = int((i + 0) * bucket_size) + 1
            avg_end = min(int((i + 1) * bucket_size) + 1, len(data))
            avg_x = sum(d[0] for d in data[avg_start:avg_end]) / max(1, avg_end - avg_start)
            avg_y = sum(d[1] for d in data[avg_start:avg_end]) / max(1, avg_end - avg_start)
            rng_start = int(i * bucket_size) + 1
            rng_end = min(int((i + 1) * bucket_size) + 1, len(data))
            best_idx = rng_start
            max_area = -1.0
            for j in range(rng_start, rng_end):
                area = abs(
                    (data[prev_idx][0] - avg_x) * (data[j][1] - data[prev_idx][1])
                    - (data[prev_idx][0] - data[j][0]) * (avg_y - data[prev_idx][1])
                )
                if area > max_area:
                    max_area = area
                    best_idx = j
            sampled.append(data[best_idx])
            prev_idx = best_idx
        sampled.append(data[-1])
        times = [p[0] for p in sampled]
        signals = [p[1] for p in sampled]

    # Update the analysis record
    analysis.chromatogram_data = {"times": times, "signals": signals}
    db.commit()

    return {
        "success": True,
        "message": f"Chromatogram updated for {sample_id} from {matched_file['name']}",
        "points": len(times),
        "matched_file": matched_file["name"],
        "matched_injection": best_prefix,
    }


@app.post("/hplc/analyses/{analysis_id}/chromatogram-to-senaite")
async def upload_chromatogram_to_senaite(
    analysis_id: int,
    sample_uid: str = Query(..., description="SENAITE sample UID to attach image to"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Generate chromatogram CSV and upload to SENAITE as HPLC Graph attachment.

    1. Loads chromatogram_data from the HPLC analysis record
    2. Builds a CSV from the times/signals arrays
    3. Uploads CSV to SENAITE as "HPLC Graph" attachment (not rendered in COA)
    """
    analysis = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.id == analysis_id)
    ).scalar_one_or_none()
    if not analysis:
        raise HTTPException(404, f"HPLC Analysis {analysis_id} not found")

    chrom = analysis.chromatogram_data
    if not chrom or not chrom.get("times") or not chrom.get("signals"):
        raise HTTPException(400, "No chromatogram data stored on this analysis")

    # Step 1: Build CSV from chromatogram data
    import io, csv as csv_mod
    times = chrom["times"]
    signals = chrom["signals"]
    buf = io.StringIO()
    writer = csv_mod.writer(buf)
    for t, s in zip(times, signals):
        writer.writerow([t, s])
    csv_bytes = buf.getvalue().encode("utf-8")

    # Step 2: Upload CSV to SENAITE as HPLC Graph attachment
    if SENAITE_URL is None:
        raise HTTPException(503, "SENAITE not configured")

    filename = f"chromatogram_{analysis.sample_id_label}.csv"
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(60.0, connect=10.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            # Resolve sample Plone URL
            api_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/analysisrequest/{sample_uid}"
            meta_resp = await client.get(api_url)
            meta_resp.raise_for_status()
            items = meta_resp.json().get("items", [])
            if not items:
                raise HTTPException(404, "Sample not found in SENAITE")
            sample_url = items[0].get("url") or items[0].get("absolute_url")
            if not sample_url:
                raise HTTPException(502, "Could not resolve sample URL")

            # GET page for CSRF token + attachment type UID
            page_resp = await client.get(sample_url)
            page_html = page_resp.text

            auth_match = re.search(r'name="_authenticator"\s+value="([^"]+)"', page_html)
            authenticator = auth_match.group(1) if auth_match else ""

            type_pattern = re.compile(
                r'<option\s+value="([^"]+)"[^>]*>\s*' + re.escape("HPLC Graph") + r'\s*</option>',
                re.IGNORECASE,
            )
            type_match = type_pattern.search(page_html)
            attachment_type_uid = type_match.group(1) if type_match else ""

            # POST attachment
            form_url = f"{sample_url}/@@attachments_view/add"
            form_data = {
                "submitted": "1",
                "_authenticator": authenticator,
                "AttachmentType": attachment_type_uid,
                "Analysis": "",
                "AttachmentKeys": "",
                "RenderInReport:boolean": "False",
                "RenderInReport:boolean:default": "False",
                "addARAttachment": "Add Attachment",
            }
            files = {
                "AttachmentFile_file": (filename, csv_bytes, "text/csv"),
            }
            headers = {
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Referer": sample_url,
            }
            att_resp = await client.post(form_url, data=form_data, files=files, headers=headers)
            if att_resp.status_code not in (200, 301, 302):
                raise HTTPException(502, f"SENAITE attachment upload returned {att_resp.status_code}")

    except HTTPException:
        raise
    except httpx.TimeoutException:
        raise HTTPException(504, "SENAITE request timed out")
    except Exception as e:
        raise HTTPException(502, f"SENAITE upload failed: {e}")

    return {
        "success": True,
        "message": f"Chromatogram CSV uploaded to SENAITE for {analysis.sample_id_label}",
        "filename": filename,
        "size_bytes": len(csv_bytes),
    }


@app.delete("/hplc/analyses/{analysis_id}")
async def delete_hplc_analysis(analysis_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Delete an HPLC analysis and its related audit log entries."""
    analysis = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.id == analysis_id)
    ).scalar_one_or_none()
    if not analysis:
        raise HTTPException(404, f"HPLC Analysis {analysis_id} not found")

    sample_label = analysis.sample_id_label

    # Delete related audit logs
    audit_logs = db.execute(
        select(AuditLog).where(
            AuditLog.entity_type == "hplc_analysis",
            AuditLog.entity_id == sample_label,
        )
    ).scalars().all()
    for log in audit_logs:
        db.delete(log)

    db.delete(analysis)
    db.commit()
    return {"message": f"Analysis {analysis_id} ({sample_label}) deleted"}


@app.get("/hplc/analyses/{analysis_id}", response_model=HPLCAnalysisResponse)
async def get_hplc_analysis(analysis_id: int, db: Session = Depends(get_db), _current_user=Depends(get_current_user)):
    """Get full detail of a single HPLC analysis including calculation trace."""
    analysis = db.execute(
        select(HPLCAnalysis).where(HPLCAnalysis.id == analysis_id)
    ).scalar_one_or_none()
    if not analysis:
        raise HTTPException(404, f"HPLC Analysis {analysis_id} not found")

    peptide = db.execute(
        select(Peptide).where(Peptide.id == analysis.peptide_id)
    ).scalar_one_or_none()

    return _analysis_to_response(analysis, peptide.abbreviation if peptide else "?")


# --- Peptide Seed from Lab Folder ---


class SeedPeptidesResponse(BaseModel):
    """Response from running the peptide seed/scan."""
    success: bool
    output: str
    errors: str


# ── Filename metadata parser ──

def _parse_filename_metadata(filename: str, path: str = "") -> dict:
    """
    Extract structured metadata from a standard curve filename and path.

    Returns a dict with keys:
      instrument, vendor, lot_number, batch_number, cap_color, run_date
    All values are None if not detected.
    """
    import re
    from datetime import datetime as dt

    stem = filename.removesuffix(".xlsx").removesuffix(".xls")
    upper = stem.upper()

    # --- Instrument: from directory path ---
    instrument = None
    if "\\1260" in path or "/1260" in path:
        instrument = "1260"
    elif "\\1290" in path or "/1290" in path:
        instrument = "1290"

    # --- Vendor ---
    vendor = None
    vendor_map = [
        (r"cayman", "Cayman"),
        (r"targetmol", "Targetmol"),
        (r"shanghai.?sigma.?audley|ssa(?=_)", "Shanghai Sigma Audley"),
        (r"hyb(?=_|\b)", "HYB"),
        (r"polaris", "Polaris"),
        (r"achemblock", "AChemBlocks"),
        (r"astatech", "AstaTech"),
        (r"levi(?=_|\b)", "Levi"),
        (r"valor(?=_|\b)", "Valor"),
        (r"drpeptide|dr.peptide", "Dr. Peptide"),
    ]
    for pattern, name in vendor_map:
        if re.search(pattern, stem, re.IGNORECASE):
            vendor = name
            break

    # --- Cap color: word immediately before "Cap" ---
    cap_color = None
    cap_match = re.search(r"([A-Za-z]+)cap", stem, re.IGNORECASE)
    if cap_match:
        cap_color = cap_match.group(1).capitalize() + "Cap"

    # --- Date: look for YYYYMMDD or MMDDYYYY or YYYY-MM-DD ---
    run_date = None
    # ISO with hyphens: 2026-02-03
    iso_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", stem)
    if iso_match:
        try:
            run_date = dt(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
        except ValueError:
            pass
    if run_date is None:
        # 8-digit block: try YYYYMMDD first, then MMDDYYYY
        for m in re.finditer(r"\b(\d{8})\b", stem):
            s = m.group(1)
            y, mo, d = int(s[:4]), int(s[4:6]), int(s[6:])
            if 2020 <= y <= 2035 and 1 <= mo <= 12 and 1 <= d <= 31:
                try:
                    run_date = dt(y, mo, d)
                    break
                except ValueError:
                    pass
            # Try MMDDYYYY
            mo2, d2, y2 = int(s[:2]), int(s[2:4]), int(s[4:])
            if 2020 <= y2 <= 2035 and 1 <= mo2 <= 12 and 1 <= d2 <= 31:
                try:
                    run_date = dt(y2, mo2, d2)
                    break
                except ValueError:
                    pass

    # --- Lot number ---
    lot_number = None
    # Hash-prefixed: #63162
    hash_match = re.search(r"#(\d{4,})", stem)
    if hash_match:
        lot_number = f"#{hash_match.group(1)}"
    else:
        # Standalone 5-6+ digit number not part of a date
        for m in re.finditer(r"\b(\d{5,7})\b", stem):
            val = m.group(1)
            # Skip if it looks like part of a date we already parsed
            if run_date and str(run_date.year) in val:
                continue
            lot_number = val
            break

    # --- Batch number: Targetmol-style codes (e.g. T20561L, TP2328L) ---
    batch_number = None
    batch_match = re.search(r"\b(T[P]?\d{4,}[A-Z]?)\b", stem, re.IGNORECASE)
    if batch_match:
        batch_number = batch_match.group(1).upper()

    return {
        "instrument": instrument,
        "vendor": vendor,
        "lot_number": lot_number,
        "batch_number": batch_number,
        "cap_color": cap_color,
        "run_date": run_date,
    }


# ── Excel calibration parsing helpers (shared with seed script) ──

def _is_number(v) -> bool:
    """Check if a value is a numeric value."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _to_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v))


def _try_extract_calibration(ws, filename: str, max_rows: int = 25) -> dict | None:
    """
    Try to extract calibration data from a worksheet by scanning for headers
    or falling back to known fixed layouts.
    """
    # 1. Dynamic Header Scan
    conc_col = None
    area_col = None
    rt_col = None
    header_row = None

    # Scan top 20 rows for headers using substring matching
    # This catches variants like "Actual Concentration", "Target Conc. (µg/mL)", etc.
    found_headers = False

    # Patterns: if ANY keyword appears in the cell value, it's a match.
    # Listed most-specific first to avoid false positives.
    _conc_keywords = ("actual concentration", "actual (ug", "actual (\u00b5g",
                      "concentration", "target conc", "target (ug", "target (\u00b5g",
                      "std. conc", "std conc", "amount", "cal level", "level")
    _area_keywords = ("peak area", "area")
    _rt_keywords   = ("ret. time", "ret time", "ret_time", "rt")
    # Words that disqualify a cell even if it contains a keyword
    _area_exclude  = ("area %", "area%", "area purity", "purity")

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue
            val = cell.value.lower().strip()

            # --- Concentration column (prefer "actual" over "target") ---
            if conc_col is None or ("actual" in val and "target" not in val):
                for kw in _conc_keywords:
                    if kw in val:
                        conc_col = cell.column
                        break

            # --- Area column (exclude "Area %" / "Area Purity") ---
            if area_col is None:
                if not any(ex in val for ex in _area_exclude):
                    for kw in _area_keywords:
                        if kw in val:
                            area_col = cell.column
                            break

            # --- RT column ---
            if rt_col is None:
                for kw in _rt_keywords:
                    if kw in val:
                        rt_col = cell.column
                        break

        if conc_col and area_col:
            header_row = r_idx
            found_headers = True
            break
            
    # DEBUG: Log if headers not found for KPV
    if not found_headers and "KPV" in filename and ws.title not in ("Sequence", "Instrument Method"):
         # Grab first row as sample
         row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
         print(f"[DEBUG-KPV] {filename}/{ws.title}: No headers found. Row 1: {row1}")
    
    # If headers found, extract data below
    if found_headers:
        concentrations = []
        areas = []
        rts = []
        
        # Scan data rows below header (stop after 2 consecutive empty rows)
        consecutive_empty = 0
        for r in range(header_row + 1, header_row + 1 + max_rows):
            conc_val = ws.cell(row=r, column=conc_col).value
            area_val = ws.cell(row=r, column=area_col).value
            
            if not _is_number(conc_val) or not _is_number(area_val):
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
                continue
            
            c = _to_float(conc_val)
            a = _to_float(area_val)
            
            if c <= 0 or a <= 0:
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
                continue
            
            consecutive_empty = 0  # Reset on valid row
            concentrations.append(c)
            areas.append(a)
            
            if rt_col:
                rt_val = ws.cell(row=r, column=rt_col).value
                if _is_number(rt_val) and _to_float(rt_val) > 0:
                    rts.append(_to_float(rt_val))

        if len(concentrations) >= 3:
            # Check for linearity/validity (simple check)
            max_c = max(concentrations)
            min_c = min(concentrations)
            if max_c > min_c: # Just ensure some spread
                 return {
                    "concentrations": concentrations,
                    "areas": areas,
                    "rts": rts,
                    "format": "dynamic_header",
                    "n_points": len(concentrations),
                }

    # 2. Fallback to fixed layouts (if dynamic failed)
    layouts = [
        ("new_S_U", 2, 19, 21),     # B, S, U
        ("old_J_L", 2, 10, 12),     # B, J, L
        ("older_B_G", 2, 7, None),  # B, G, no RT
    ]

    for fmt_name, c_col, a_col, r_col in layouts:
        concentrations = []
        areas = []
        rts = []

        for row in range(2, 2 + max_rows):
            conc_val = ws.cell(row=row, column=c_col).value
            area_val = ws.cell(row=row, column=a_col).value

            if not _is_number(conc_val) or not _is_number(area_val):
                continue

            conc = _to_float(conc_val)
            area = _to_float(area_val)

            if conc <= 0 or area <= 0:
                continue

            concentrations.append(conc)
            areas.append(area)

            if r_col is not None:
                rt_val = ws.cell(row=row, column=r_col).value
                if _is_number(rt_val) and _to_float(rt_val) > 0:
                    rts.append(_to_float(rt_val))

        if len(concentrations) >= 3:
            # Simple validity check: range of areas should be somewhat large spread
            # if max_area <= min_area * 1.5, likely just noise or same sample repeated
            max_conc_area = areas[concentrations.index(max(concentrations))]
            min_conc_area = areas[concentrations.index(min(concentrations))]
            if max_conc_area <= min_conc_area * 1.05: # lenient
                continue

            return {
                "concentrations": concentrations,
                "areas": areas,
                "rts": rts,
                "format": fmt_name,
                "n_points": len(concentrations),
            }

    return None


def _parse_peakdata_csv(data: bytes, filename: str) -> dict | None:
    """
    Parse a PeakData CSV file exported from the HPLC instrument.

    Expected format:
        Height,Area,Area%,Peak Begin Time,Peak End Time,RT [min]
        505.4477,7268.6459,98.0781,3.314,4.205,3.477
        9.8865,142.4372,1.9219,4.205,4.967,4.292
        Sum,7411.0831,,,,

    Returns the dominant peak (highest Area%) as:
        {"area": float, "rt": float, "height": float}
    or None if no valid peaks found.
    """
    import csv
    from io import StringIO

    try:
        text = data.decode("utf-8", errors="replace")
        reader = csv.DictReader(StringIO(text))

        best_peak = None
        best_area_pct = -1.0

        for row in reader:
            # Skip the "Sum" row
            height_str = (row.get("Height") or "").strip()
            if not height_str or height_str.lower() == "sum":
                continue

            try:
                area = float(row.get("Area", 0))
                area_pct = float(row.get("Area%", 0))
                rt = float(row.get("RT [min]", 0))
                height = float(height_str)
            except (ValueError, TypeError):
                continue

            if area > 0 and area_pct > best_area_pct:
                best_area_pct = area_pct
                best_peak = {"area": area, "rt": rt, "height": height}

        return best_peak
    except Exception as e:
        print(f"[WARN] Failed to parse PeakData CSV {filename}: {e}")
        return None


def _build_curve_from_peakdata_csvs(files: list[tuple[str, bytes, str]]) -> dict | None:
    """
    Build a calibration curve from multiple PeakData CSV files.

    Args:
        files: List of (concentration_str, file_bytes, filename) tuples

    Returns:
        dict with keys: concentrations, areas, rts
        or None if fewer than 3 valid data points
    """
    concentrations = []
    areas = []
    rts = []

    for conc_str, data, filename in files:
        try:
            conc = float(conc_str)
        except ValueError:
            continue

        peak = _parse_peakdata_csv(data, filename)
        if not peak or peak["area"] <= 0:
            continue

        concentrations.append(conc)
        areas.append(peak["area"])
        if peak["rt"] > 0:
            rts.append(peak["rt"])

    if len(concentrations) < 3:
        return None

    return {
        "concentrations": concentrations,
        "areas": areas,
        "rts": rts,
    }


def _parse_calibration_excel_bytes(data: bytes, filename: str) -> dict | None:
    """Parse calibration data from Excel file bytes (downloaded from SharePoint)."""
    import openpyxl
    from io import BytesIO

    skip_sheets = {"Dissolution method", "Dissolution Method"}

    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return None

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        try:
            result = _try_extract_calibration(ws, filename)
        except Exception as e:
            # Catch worksheet access errors
            print(f"[ERROR] failed parsing sheet {sheet_name} in {filename}: {e}")
            continue

        if result:
            result["sheet"] = sheet_name
            result["filename"] = filename
            wb.close()
            return result

    wb.close()
    return None


@app.post("/hplc/seed-peptides", response_model=SeedPeptidesResponse)
async def seed_peptides_from_sharepoint(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Scan the SharePoint Peptides folder and import peptides + ALL calibration curves.

    - Discovers peptide folders dynamically (no hardcoded map).
    - Scans entire peptide folder for .xlsx files.
    - Parses each for concentration/area/RT data.
    - Creates new peptides and imports ALL valid calibration curves.
    - For existing peptides, imports any new calibrations not yet in DB.
    - Most recent calibration (by filename sort) is set as active.
    """
    import sharepoint as sp
    from calculations.calibration import calculate_calibration_curve

    log_lines = []
    error_lines = []

    def log(msg: str):
        print(msg)  # Ensure it shows in docker logs
        log_lines.append(msg)

    try:
        # 1. List all folders in the Peptides root
        peptide_folders = await sp.list_folder("")
        peptide_dirs = [f for f in peptide_folders if f["type"] == "folder"]
        log(f"Found {len(peptide_dirs)} folders in Peptides root\n")

        # 2. Get existing peptides from DB
        existing_peptides = {
            p.abbreviation: p
            for p in db.execute(select(Peptide).order_by(Peptide.abbreviation)).scalars().all()
        }
        log(f"Existing peptides in DB: {len(existing_peptides)} ({', '.join(existing_peptides.keys()) or 'none'})\n")

        created = 0
        calibrations_added = 0
        skipped = 0
        no_cal_data = []

        for folder in sorted(peptide_dirs, key=lambda f: f["name"]):
            folder_name = folder["name"]
            abbreviation = folder_name.strip()

            # Skip known non-peptide folders
            skip_names = {"Templates", "Blends", "_Templates", "Archive"}
            if folder_name in skip_names:
                log(f"[SKIP] {folder_name}/ — non-peptide folder")
                continue

            log(f"\n--- {folder_name} ---")

            # 3. Find ALL Excel files in the peptide folder
            cal_files = []
            try:
                all_xlsx = await sp.list_files_recursive(
                    folder_name,
                    extensions=[".xlsx"],
                    root="peptides",
                )
                for item in all_xlsx:
                    fn = item["name"]
                    if fn.startswith("~$"):
                        continue
                    # Skip Agilent data exports
                    if ".dx_" in fn or "_PeakData" in fn:
                        continue
                    cal_files.append(item)
            except Exception as e:
                log(f"  [ERROR] Failed to list files: {e}")

            # 4. Download and parse calibration files (skip already-imported paths)
            cal_files.sort(key=lambda f: f["name"])
            parsed_cals = []

            # Pre-load known source_paths to skip downloads (from file cache + calibration curves)
            known_paths = set(
                r[0] for r in db.execute(
                    select(SharePointFileCache.source_path)
                    .where(SharePointFileCache.peptide_abbreviation == abbreviation)
                ).all()
            )
            peptide = existing_peptides.get(abbreviation)
            if peptide:
                # Also include paths from calibration curves (covers pre-cache data)
                known_paths.update(
                    r[0] for r in db.execute(
                        select(CalibrationCurve.source_path)
                        .where(CalibrationCurve.peptide_id == peptide.id)
                        .where(CalibrationCurve.source_path.isnot(None))
                    ).all()
                )

            for cal_file in cal_files:
                file_path = cal_file.get("path", cal_file["name"])
                if file_path in known_paths:
                    log(f"  [SKIP] Already imported: {cal_file['name']}")
                    continue
                try:
                    file_bytes, filename = await sp.download_file(cal_file["id"])
                    result = _parse_calibration_excel_bytes(file_bytes, filename)
                    produced_cal = result is not None
                    if result:
                        result["_sharepoint_path"] = file_path
                        result["_last_modified"] = cal_file.get("last_modified")
                        result["_web_url"] = cal_file.get("web_url")
                        parsed_cals.append(result)
                    # Cache this path so we never re-download it
                    db.add(SharePointFileCache(
                        source_path=file_path,
                        peptide_abbreviation=abbreviation,
                        produced_calibration=produced_cal,
                    ))
                except Exception as e:
                    log(f"  [WARN] Failed to download/parse {cal_file['name']}: {e}")

            # 5. Get or create peptide
            peptide = existing_peptides.get(abbreviation)
            is_new = peptide is None

            if is_new:
                # Extract reference RT from most recent calibration
                ref_rt = None
                if parsed_cals:
                    last_cal = parsed_cals[-1]
                    if last_cal.get("rts"):
                        ref_rt = round(sum(last_cal["rts"]) / len(last_cal["rts"]), 4)

                peptide = Peptide(
                    name=folder_name,
                    abbreviation=abbreviation,
                )
                db.add(peptide)
                db.flush()
                created += 1
                existing_peptides[abbreviation] = peptide
                log(f"  Created peptide (id={peptide.id})")
            else:
                log(f"  [EXISTS] {abbreviation} (id={peptide.id})")

            # 6. Get existing calibration source filenames for this peptide
            existing_cal_filenames = set()
            existing_cals = db.execute(
                select(CalibrationCurve)
                .where(CalibrationCurve.peptide_id == peptide.id)
            ).scalars().all()
            for ec in existing_cals:
                if ec.source_filename:
                    existing_cal_filenames.add(ec.source_filename)

            # 7. Import all parsed calibrations that aren't already in DB
            new_cals_for_peptide = []
            for cal_data in parsed_cals:
                source = f"{cal_data['filename']}[{cal_data['sheet']}]"
                if source in existing_cal_filenames:
                    log(f"  [EXISTS] Calibration: {source}")
                    continue

                try:
                    cal_result = calculate_calibration_curve(
                        cal_data["concentrations"],
                        cal_data["areas"],
                    )
                    cal_curve = CalibrationCurve(
                        peptide_id=peptide.id,
                        slope=cal_result["slope"],
                        intercept=cal_result["intercept"],
                        r_squared=cal_result["r_squared"],
                        standard_data={
                            "concentrations": cal_data["concentrations"],
                            "areas": cal_data["areas"],
                            "rts": cal_data.get("rts", []),
                        },
                        source_filename=source,
                        source_path=cal_data.get("_sharepoint_path"),
                        source_date=(
                            datetime.fromisoformat(cal_data["_last_modified"].replace("Z", "+00:00"))
                            if cal_data.get("_last_modified")
                            else None
                        ),
                        sharepoint_url=cal_data.get("_web_url"),
                        is_active=False,  # Will activate the most recent below
                    )
                    db.add(cal_curve)
                    new_cals_for_peptide.append(cal_curve)
                    calibrations_added += 1
                    log(f"  + Calibration: {source} "
                        f"(slope={cal_result['slope']:.4f}, R²={cal_result['r_squared']:.6f})")
                except Exception as e:
                    error_lines.append(f"Calibration error for {abbreviation} ({source}): {e}")
                    log(f"  [ERROR] Calibration {source}: {e}")

            # 8. Set the most recent calibration as active
            if new_cals_for_peptide:
                # Deactivate all existing
                for ec in existing_cals:
                    ec.is_active = False
                # Activate the last (most recent by filename sort)
                db.flush()
                new_cals_for_peptide[-1].is_active = True
                log(f"  ✓ Active: {new_cals_for_peptide[-1].source_filename}")
                # Set reference RT on the active curve from its own RT data
                active_data = new_cals_for_peptide[-1].standard_data
                if active_data and active_data.get("rts"):
                    new_cals_for_peptide[-1].reference_rt = round(sum(active_data["rts"]) / len(active_data["rts"]), 4)
                    log(f"  Updated reference RT: {new_cals_for_peptide[-1].reference_rt}")

            if not parsed_cals and is_new:
                no_cal_data.append(folder_name)
                log(f"  No calibration data found")

            if not is_new and not new_cals_for_peptide and not parsed_cals:
                skipped += 1

        db.commit()

        # Summary
        log(f"\n{'=' * 60}")
        log(f"SUMMARY")
        log(f"  Peptides created:      {created}")
        log(f"  Calibrations added:    {calibrations_added}")
        log(f"  Skipped (no changes):  {skipped}")
        log(f"  No calibration data:   {len(no_cal_data)}")
        if no_cal_data:
            log(f"    {', '.join(no_cal_data)}")
        log(f"  Total in DB:           {len(existing_peptides)}")

        return SeedPeptidesResponse(
            success=True,
            output="\n".join(log_lines),
            errors="\n".join(error_lines),
        )
    except Exception as e:
        db.rollback()
        return SeedPeptidesResponse(
            success=False,
            output="\n".join(log_lines),
            errors=f"SharePoint scan failed: {e}\n" + "\n".join(error_lines),
        )


@app.get("/hplc/seed-peptides/stream")
async def seed_peptides_stream(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    SSE streaming version of seed-peptides.
    Imports ALL calibration curves per peptide.
    Sends each log line as a server-sent event for real-time progress display.
    """
    from starlette.responses import StreamingResponse
    import sharepoint as sp
    import asyncio
    from calculations.calibration import calculate_calibration_curve

    async def event_generator():
        def send_event(event_type: str, data: dict) -> str:
            """Format an SSE event and log to console."""
            if event_type == "log":
                import sys
                msg = data.get("message", "")
                level = data.get("level", "info")
                # Print to stderr to ensure it shows up in docker logs immediately
                print(f"[{level.upper()}] {msg}", file=sys.stderr)
            
            payload = json.dumps(data)
            return f"event: {event_type}\ndata: {payload}\n\n"

        created = 0
        calibrations_added = 0
        skipped = 0
        no_cal_data = []
        error_lines = []

        try:
            # 1. List all folders in the Peptides root
            yield send_event("log", {"message": "Connecting to SharePoint...", "level": "info"})
            await asyncio.sleep(0)  # flush

            peptide_folders = await sp.list_folder("")
            peptide_dirs = [f for f in peptide_folders if f["type"] == "folder"]
            total = len(peptide_dirs)
            yield send_event("log", {"message": f"Found {total} folders in Peptides root", "level": "info"})
            yield send_event("progress", {"current": 0, "total": total, "phase": "scanning"})

            # 2. Get existing peptides from DB
            existing_peptides = {
                p.abbreviation: p
                for p in db.execute(select(Peptide).order_by(Peptide.abbreviation)).scalars().all()
            }
            yield send_event("log", {
                "message": f"Existing peptides in DB: {len(existing_peptides)} ({', '.join(existing_peptides.keys()) or 'none'})",
                "level": "info",
            })

            processed = 0
            for folder in sorted(peptide_dirs, key=lambda f: f["name"]):
                processed += 1
                folder_name = folder["name"]
                abbreviation = folder_name.strip()

                # Skip known non-peptide folders
                skip_names = {"Templates", "Blends", "_Templates", "Archive"}
                if folder_name in skip_names:
                    yield send_event("log", {"message": f"[SKIP] {folder_name}/ — non-peptide folder", "level": "dim"})
                    yield send_event("progress", {"current": processed, "total": total, "phase": "scanning"})
                    continue

                yield send_event("log", {"message": f"--- {folder_name} ---", "level": "heading"})
                yield send_event("progress", {"current": processed, "total": total, "phase": f"Processing {folder_name}"})

                # 3. Find ALL Excel files in the peptide folder
                cal_files = []
                try:
                    yield send_event("log", {"message": f"  Scanning for Excel files...", "level": "dim"})
                    all_xlsx = await sp.list_files_recursive(
                        folder_name,
                        extensions=[".xlsx"],
                        root="peptides",
                    )
                    for item in all_xlsx:
                        fn = item["name"]
                        if fn.startswith("~$"):
                            continue
                        if ".dx_" in fn or "_PeakData" in fn:
                            continue
                        cal_files.append(item)
                    if cal_files:
                        yield send_event("log", {"message": f"  Found {len(cal_files)} Excel file(s)", "level": "info"})
                except Exception as e:
                    yield send_event("log", {"message": f"  [ERROR] File listing failed: {e}", "level": "error"})
                    import sys
                    print(f"[ERROR] list_files_recursive failed for {folder_name}: {e}", file=sys.stderr)

                # 4. Download and parse calibration files (skip already-imported paths)
                cal_files.sort(key=lambda f: f["name"])
                parsed_cals = []

                # Pre-load known source_paths to skip downloads (from file cache + calibration curves)
                known_paths = set(
                    r[0] for r in db.execute(
                        select(SharePointFileCache.source_path)
                        .where(SharePointFileCache.peptide_abbreviation == abbreviation)
                    ).all()
                )
                peptide = existing_peptides.get(abbreviation)
                if peptide:
                    # Also include paths from calibration curves (covers pre-cache data)
                    known_paths.update(
                        r[0] for r in db.execute(
                            select(CalibrationCurve.source_path)
                            .where(CalibrationCurve.peptide_id == peptide.id)
                            .where(CalibrationCurve.source_path.isnot(None))
                        ).all()
                    )

                skipped_count = 0
                for cal_file in cal_files:
                    file_path = cal_file.get("path", cal_file["name"])
                    if file_path in known_paths:
                        skipped_count += 1
                        continue
                    try:
                        yield send_event("log", {"message": f"  Downloading {cal_file['name']}...", "level": "dim"})
                        file_bytes, filename = await sp.download_file(cal_file["id"])
                        result = _parse_calibration_excel_bytes(file_bytes, filename)
                        produced_cal = result is not None
                        if result:
                            result["_sharepoint_path"] = file_path
                            result["_last_modified"] = cal_file.get("last_modified")
                            result["_web_url"] = cal_file.get("web_url")
                            parsed_cals.append(result)
                        # Cache this path so we never re-download it
                        db.add(SharePointFileCache(
                            source_path=file_path,
                            peptide_abbreviation=abbreviation,
                            produced_calibration=produced_cal,
                        ))
                    except Exception as e:
                        yield send_event("log", {"message": f"  [WARN] Failed: {cal_file['name']}: {e}", "level": "warn"})

                if skipped_count:
                    yield send_event("log", {"message": f"  [SKIP] {skipped_count} file(s) already imported", "level": "dim"})

                # 5. Get or create peptide
                peptide = existing_peptides.get(abbreviation)
                is_new = peptide is None

                if is_new:
                    peptide = Peptide(
                        name=folder_name,
                        abbreviation=abbreviation,
                    )
                    db.add(peptide)
                    db.flush()
                    created += 1
                    existing_peptides[abbreviation] = peptide
                    yield send_event("log", {"message": f"  ✓ Created peptide (id={peptide.id})", "level": "success"})
                else:
                    yield send_event("log", {
                        "message": f"  [EXISTS] {abbreviation} (id={peptide.id})",
                        "level": "dim",
                    })

                # 6. Check existing calibration source filenames
                existing_cal_filenames = set()
                existing_cals = db.execute(
                    select(CalibrationCurve)
                    .where(CalibrationCurve.peptide_id == peptide.id)
                ).scalars().all()
                for ec in existing_cals:
                    if ec.source_filename:
                        existing_cal_filenames.add(ec.source_filename)

                # 7. Import all new calibrations
                new_cals_for_peptide = []
                for cal_data in parsed_cals:
                    source = f"{cal_data['filename']}[{cal_data['sheet']}]"
                    if source in existing_cal_filenames:
                        yield send_event("log", {"message": f"  [EXISTS] Cal: {source}", "level": "dim"})
                        continue

                    try:
                        cal_result = calculate_calibration_curve(
                            cal_data["concentrations"],
                            cal_data["areas"],
                        )
                        cal_curve = CalibrationCurve(
                            peptide_id=peptide.id,
                            slope=cal_result["slope"],
                            intercept=cal_result["intercept"],
                            r_squared=cal_result["r_squared"],
                            standard_data={
                                "concentrations": cal_data["concentrations"],
                                "areas": cal_data["areas"],
                                "rts": cal_data.get("rts", []),
                            },
                            source_filename=source,
                            source_path=cal_data.get("_sharepoint_path"),
                            source_date=(
                                datetime.fromisoformat(cal_data["_last_modified"].replace("Z", "+00:00"))
                                if cal_data.get("_last_modified")
                                else None
                            ),
                            sharepoint_url=cal_data.get("_web_url"),
                            is_active=False,
                        )
                        db.add(cal_curve)
                        new_cals_for_peptide.append(cal_curve)
                        calibrations_added += 1
                        yield send_event("log", {
                            "message": f"  + Cal: {source} (slope={cal_result['slope']:.4f}, R²={cal_result['r_squared']:.6f})",
                            "level": "success",
                        })
                    except Exception as e:
                        error_lines.append(f"Calibration error for {abbreviation} ({source}): {e}")
                        yield send_event("log", {"message": f"  ✗ Cal error {source}: {e}", "level": "error"})

                # 8. Set the most recent as active
                if new_cals_for_peptide:
                    for ec in existing_cals:
                        ec.is_active = False
                    db.flush()
                    new_cals_for_peptide[-1].is_active = True
                    yield send_event("log", {
                        "message": f"  ✓ Active: {new_cals_for_peptide[-1].source_filename}",
                        "level": "success",
                    })
                    # Set reference RT on the active curve from its own RT data
                    active_data = new_cals_for_peptide[-1].standard_data
                    if active_data and active_data.get("rts"):
                        new_cals_for_peptide[-1].reference_rt = round(sum(active_data["rts"]) / len(active_data["rts"]), 4)
                        yield send_event("log", {
                            "message": f"  Updated reference RT: {new_cals_for_peptide[-1].reference_rt}",
                            "level": "success",
                        })

                if not parsed_cals and is_new:
                    no_cal_data.append(folder_name)
                    yield send_event("log", {
                        "message": f"  No calibration data found",
                        "level": "warn",
                    })

                if not is_new and not new_cals_for_peptide and not parsed_cals:
                    skipped += 1

                # Commit after each peptide so progress is saved incrementally
                db.commit()
                yield send_event("refresh", {})

            # Summary
            yield send_event("log", {"message": "=" * 50, "level": "info"})
            yield send_event("log", {"message": "SUMMARY", "level": "heading"})
            yield send_event("log", {"message": f"  Peptides created:    {created}", "level": "success" if created else "info"})
            yield send_event("log", {"message": f"  Calibrations added:  {calibrations_added}", "level": "success" if calibrations_added else "info"})
            yield send_event("log", {"message": f"  Skipped (no changes):{skipped}", "level": "dim"})
            yield send_event("log", {"message": f"  No calibration data: {len(no_cal_data)}", "level": "warn" if no_cal_data else "info"})
            yield send_event("log", {"message": f"  Total in DB:         {len(existing_peptides)}", "level": "info"})

            yield send_event("done", {
                "success": True,
                "created": created,
                "calibrations": calibrations_added,
                "skipped": skipped,
                "total": len(existing_peptides),
            })

        except Exception as e:
            db.rollback()
            yield send_event("log", {"message": f"✗ FAILED: {e}", "level": "error"})
            yield send_event("done", {"success": False, "error": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/hplc/rebuild-standards/stream")
async def rebuild_standards_stream(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Hard-wipe all peptide standard curves and rebuild from SharePoint.

    - Deletes ALL existing Peptide and CalibrationCurve records.
    - Scans SharePoint Peptides folder for _Std_ / STD_ files only.
    - Skips prep sheets (blank Peak Areas), COA summary sheets, template files.
    - Extracts structured metadata from filename + path (instrument, vendor, lot, etc.).
    - Deduplicates within each peptide+instrument by peak area fingerprint.
    - Sets the most recent curve (by run_date) as active per peptide+instrument.
    """
    from starlette.responses import StreamingResponse
    import sharepoint as sp
    import asyncio
    import re
    from calculations.calibration import calculate_calibration_curve

    async def event_generator():
        def send_event(event_type: str, data: dict) -> str:
            if event_type == "log":
                import sys
                print(f"[{data.get('level','info').upper()}] {data.get('message','')}", file=sys.stderr)
            return f"event: {event_type}\ndata: {json.dumps(data)}\n\n"

        skip_folder_names = {"Templates", "Blends", "_Templates", "Archive",
                              "1_SampleName_Std_GreenCap_Cayman_YearMonthDay_template",
                              "Organic_synthesis_Checks"}
        skip_file_fragments = ["DAD1A", "YearMonthDay", "template", "Template",
                                "Master sheet", "Calibration_Curve_Template",
                                "P-###"]

        created_peptides = 0
        created_curves = 0
        skipped_no_data = 0
        skipped_dup = 0
        errors = []

        try:
            # ── 1. Hard wipe ──────────────────────────────────────────────────────
            yield send_event("log", {"message": "Wiping existing peptide standards...", "level": "warn"})
            await asyncio.sleep(0)
            db.execute(delete(CalibrationCurve))
            db.execute(delete(Peptide))
            db.commit()
            yield send_event("log", {"message": "✓ Wipe complete", "level": "info"})

            # ── 2. List peptide folders ───────────────────────────────────────────
            yield send_event("log", {"message": "Scanning SharePoint Peptides folder...", "level": "info"})
            peptide_folders = await sp.list_folder("")
            peptide_dirs = [f for f in peptide_folders
                            if f["type"] == "folder" and f["name"] not in skip_folder_names]
            total = len(peptide_dirs)
            yield send_event("log", {"message": f"Found {total} peptide folders", "level": "info"})
            yield send_event("progress", {"current": 0, "total": total, "phase": "scanning"})

            # ── 3. Process each peptide folder ────────────────────────────────────
            peptide_map: dict[str, Peptide] = {}  # abbreviation → Peptide

            for idx, folder in enumerate(sorted(peptide_dirs, key=lambda f: f["name"]), 1):
                folder_name = folder["name"]
                yield send_event("log", {"message": f"── {folder_name}", "level": "info"})
                yield send_event("progress", {"current": idx, "total": total, "phase": folder_name})
                await asyncio.sleep(0)

                # List all xlsx files in this peptide folder
                try:
                    all_xlsx = await sp.list_files_recursive(
                        folder_name,
                        extensions=[".xlsx"],
                        root="peptides",
                    )
                except Exception as e:
                    yield send_event("log", {"message": f"  [ERROR] listing {folder_name}: {e}", "level": "error"})
                    errors.append(f"{folder_name}: {e}")
                    continue

                # Filter out temp files and known non-data fragments
                std_files = []
                for item in all_xlsx:
                    fn = item["name"]
                    if fn.startswith("~$"):
                        continue
                    if any(frag in fn for frag in skip_file_fragments):
                        continue
                    std_files.append(item)

                # Always ensure a Peptide stub exists for every folder
                if folder_name not in peptide_map:
                    peptide = Peptide(name=folder_name, abbreviation=folder_name)
                    db.add(peptide)
                    db.flush()
                    peptide_map[folder_name] = peptide
                    created_peptides += 1

                if not std_files:
                    yield send_event("log", {"message": f"  No standard files found", "level": "info"})
                    continue

                yield send_event("log", {"message": f"  {len(std_files)} standard file(s) found", "level": "info"})

                # Track fingerprints seen this peptide to deduplicate
                # Key: (instrument, frozenset of rounded areas)
                seen_fingerprints: set = set()

                folder_curves = 0

                for item in std_files:
                    fn = item["name"]
                    item_path = item.get("path", "")
                    item_id = item["id"]

                    # Download file
                    try:
                        file_bytes, _ = await sp.download_file(item_id)
                    except Exception as e:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: download failed ({e})", "level": "warn"})
                        skipped_no_data += 1
                        continue

                    # Parse calibration data
                    cal = _parse_calibration_excel_bytes(file_bytes, fn)
                    if not cal:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: no parseable curve data", "level": "info"})
                        skipped_no_data += 1
                        continue

                    # Parse filename metadata (instrument from path)
                    meta = _parse_filename_metadata(fn, item_path)
                    instrument = meta["instrument"]

                    # Deduplication fingerprint: instrument + sorted rounded areas
                    fp = (instrument, tuple(sorted(round(a, 1) for a in cal["areas"])))
                    if fp in seen_fingerprints:
                        yield send_event("log", {"message": f"  [DUP]  {fn}", "level": "info"})
                        skipped_dup += 1
                        continue
                    seen_fingerprints.add(fp)

                    # Regression
                    try:
                        regression = calculate_calibration_curve(cal["concentrations"], cal["areas"])
                    except Exception as e:
                        yield send_event("log", {"message": f"  [ERROR] {fn}: regression failed ({e})", "level": "error"})
                        errors.append(f"{fn}: {e}")
                        continue

                    peptide = peptide_map[folder_name]

                    # Create CalibrationCurve with full metadata
                    curve = CalibrationCurve(
                        peptide_id=peptide.id,
                        slope=regression["slope"],
                        intercept=regression["intercept"],
                        r_squared=regression["r_squared"],
                        standard_data={
                            "concentrations": cal["concentrations"],
                            "areas": cal["areas"],
                            "rts": cal.get("rts", []),
                        },
                        source_filename=fn,
                        source_path=item_path,
                        sharepoint_url=item.get("webUrl"),
                        source_date=item.get("lastModified"),
                        is_active=False,  # set active after all curves loaded
                        # Metadata from filename — resolve instrument_id
                        instrument=meta["instrument"],
                        instrument_id=_resolve_instrument(db, name=meta["instrument"])[1] if meta["instrument"] else None,
                        vendor=meta["vendor"],
                        lot_number=meta["lot_number"],
                        batch_number=meta["batch_number"],
                        cap_color=meta["cap_color"],
                        run_date=meta["run_date"],
                    )
                    db.add(curve)
                    folder_curves += 1
                    created_curves += 1
                    yield send_event("log", {
                        "message": f"  [OK]   {fn} ({instrument or '?'}, {meta['vendor'] or '?'}, R²={regression['r_squared']:.4f})",
                        "level": "info",
                    })

                db.flush()

                # Set active: most recent curve by run_date (fallback: source_date) per instrument
                curves_for_peptide = (
                    db.execute(
                        select(CalibrationCurve)
                        .where(CalibrationCurve.peptide_id == peptide_map.get(folder_name, Peptide()).id
                               if folder_name in peptide_map else CalibrationCurve.id == -1)
                    ).scalars().all()
                )
                # Group by instrument, activate newest in each group
                by_instrument: dict[str, list] = {}
                for c in curves_for_peptide:
                    key = c.instrument or "unknown"
                    by_instrument.setdefault(key, []).append(c)
                for inst_curves in by_instrument.values():
                    newest = max(
                        inst_curves,
                        key=lambda c: (c.run_date or c.source_date or c.created_at)
                    )
                    newest.is_active = True

            db.commit()

            yield send_event("log", {"message": f"\n✓ Rebuild complete", "level": "info"})
            yield send_event("log", {"message": f"  {created_peptides} peptides created", "level": "info"})
            yield send_event("log", {"message": f"  {created_curves} curves imported", "level": "info"})
            yield send_event("log", {"message": f"  {skipped_no_data} files skipped (no data)", "level": "info"})
            yield send_event("log", {"message": f"  {skipped_dup} duplicates skipped", "level": "info"})
            if errors:
                yield send_event("log", {"message": f"  {len(errors)} errors", "level": "warn"})
            yield send_event("done", {
                "success": True,
                "peptides": created_peptides,
                "curves": created_curves,
                "skipped_no_data": skipped_no_data,
                "skipped_dup": skipped_dup,
            })

        except Exception as e:
            db.rollback()
            yield send_event("log", {"message": f"✗ FAILED: {e}", "level": "error"})
            yield send_event("done", {"success": False, "error": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/hplc/import-standards/stream")
async def import_standards_stream(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Incremental import of peptide standard curves from LIMS CSV PeakData files.

    Scans the LIMS CSVs SharePoint folder for *_Std_*_PeakData.csv files,
    groups them by sample ID, builds calibration curves from each group.

    - Loads already-processed folder keys from SharePointFileCache.
    - Skips folders already seen in previous imports.
    - For new folders: finds or creates Peptide records, adds CalibrationCurve records.
    - Re-evaluates is_active for any peptide that received new curves.
    """
    from starlette.responses import StreamingResponse
    import sharepoint as sp
    import asyncio
    import re
    from calculations.calibration import calculate_calibration_curve

    async def event_generator():
        def send_event(event_type: str, data: dict) -> str:
            if event_type == "log":
                import sys
                print(f"[{data.get('level','info').upper()}] {data.get('message','')}", file=sys.stderr)
            return f"event: {event_type}\ndata: {json.dumps(data)}\n\n"

        new_curves = 0
        skipped_cached = 0
        skipped_no_data = 0
        skipped_dup = 0
        errors = []

        try:
            # ── 1. Load already-processed cache keys ───────────────────────────────
            yield send_event("log", {"message": "Loading known cache keys...", "level": "info"})
            known_paths: set[str] = set()

            existing_paths = db.execute(
                select(CalibrationCurve.source_path).where(CalibrationCurve.source_path.isnot(None))
            ).scalars().all()
            known_paths.update(p for p in existing_paths if p)

            cache_paths = db.execute(select(SharePointFileCache.source_path)).scalars().all()
            known_paths.update(p for p in cache_paths if p)

            yield send_event("log", {"message": f"Skipping {len(known_paths)} already-processed entries", "level": "info"})

            # ── 2. Scan LIMS CSV folders for Std_ PeakData files ───────────────────
            yield send_event("log", {"message": "Scanning LIMS CSVs for Std_ PeakData...", "level": "heading"})
            await asyncio.sleep(0)

            try:
                lims_top = await sp.list_lims_folder("")
                lims_dirs = [f for f in lims_top if f["type"] == "folder"]
            except Exception as e:
                yield send_event("log", {"message": f"  [ERROR] listing LIMS root: {e}", "level": "error"})
                errors.append(f"LIMS root: {e}")
                lims_dirs = []

            lims_total = len(lims_dirs)
            yield send_event("log", {"message": f"Found {lims_total} LIMS sample folders", "level": "info"})

            # Matches: P-0111_Std_100_PeakData.csv, P-0111_Std_1_PeakData.csv, etc.
            # Pattern: {SampleID}_Std_{Concentration}[_optional]_PeakData.csv
            # Note: .dx_DAD1A.CSV files are raw chromatogram data and NOT parseable
            std_peakdata_re = re.compile(
                r"^(.+?)_Std_(\d+(?:\.\d+)?)_.*PeakData\.csv$", re.IGNORECASE
            )

            for lims_idx, lims_dir in enumerate(sorted(lims_dirs, key=lambda f: f["name"]), 1):
                dir_name = lims_dir["name"]
                cache_key = f"lims-std:{dir_name}"

                yield send_event("progress", {"current": lims_idx, "total": lims_total, "phase": f"LIMS: {dir_name}"})
                await asyncio.sleep(0)

                # Already processed this folder?
                if cache_key in known_paths:
                    skipped_cached += 1
                    continue

                # Parse folder: "P-0111 BPC-157" → peptide_abbr = "BPC-157"
                parts = dir_name.split(" ", 1)
                if len(parts) < 2:
                    known_paths.add(cache_key)
                    db.merge(SharePointFileCache(
                        source_path=cache_key, peptide_abbreviation=dir_name, produced_calibration=False,
                    ))
                    continue
                peptide_abbr = parts[1].strip()

                # Recursively list all CSVs in this sample folder
                try:
                    all_csvs = await sp.list_files_recursive(
                        dir_name, extensions=[".csv"], root="lims",
                    )
                except Exception as e:
                    yield send_event("log", {"message": f"  [ERROR] listing {dir_name}: {e}", "level": "error"})
                    errors.append(f"LIMS {dir_name}: {e}")
                    continue

                # Group Std_ PeakData files by sample ID
                by_sample: dict[str, list[tuple[str, dict]]] = {}
                for csv_item in all_csvs:
                    m_csv = std_peakdata_re.match(csv_item["name"])
                    if m_csv:
                        sid_key = m_csv.group(1)
                        conc_val = m_csv.group(2)
                        by_sample.setdefault(sid_key, []).append((conc_val, csv_item))

                known_paths.add(cache_key)

                if not by_sample:
                    db.merge(SharePointFileCache(
                        source_path=cache_key, peptide_abbreviation=peptide_abbr, produced_calibration=False,
                    ))
                    continue

                total_std_files = sum(len(v) for v in by_sample.values())
                yield send_event("log", {
                    "message": f"── {dir_name}  ({total_std_files} Std_ file(s) across {len(by_sample)} sample(s))",
                    "level": "info",
                })

                # Match to existing peptide — skip if not found
                lims_peptide = db.execute(
                    select(Peptide).where(func.lower(Peptide.abbreviation) == peptide_abbr.lower())
                ).scalar_one_or_none()
                if not lims_peptide:
                    yield send_event("log", {"message": f"  [SKIP] No peptide '{peptide_abbr}' in database", "level": "dim"})
                    db.merge(SharePointFileCache(
                        source_path=cache_key, peptide_abbreviation=peptide_abbr, produced_calibration=False,
                    ))
                    continue

                # Auto-link to slot 1 analyte if available
                lims_first_analyte = db.execute(
                    select(PeptideAnalyte).where(
                        PeptideAnalyte.peptide_id == lims_peptide.id,
                        PeptideAnalyte.slot == 1,
                    )
                ).scalar_one_or_none()

                # Seed fingerprints from existing curves for dedup
                lims_fps: set = set()
                for fp_row in db.execute(
                    select(CalibrationCurve.standard_data, CalibrationCurve.instrument)
                    .where(CalibrationCurve.peptide_id == lims_peptide.id)
                ).all():
                    if fp_row.standard_data and "areas" in fp_row.standard_data:
                        fp_inst = fp_row.instrument or "unknown"
                        lims_fps.add(
                            (fp_inst, tuple(sorted(round(a, 1) for a in fp_row.standard_data["areas"])))
                        )

                folder_lims_new = 0

                for sid_key, conc_items in by_sample.items():
                    # Download all PeakData CSVs for this sample
                    csv_tuples: list[tuple[str, bytes, str]] = []
                    file_dates: list[str] = []
                    for conc_str, csv_item in sorted(conc_items, key=lambda x: float(x[0])):
                        fn = csv_item["name"]
                        try:
                            file_bytes, _ = await sp.download_file(csv_item["id"])
                            csv_tuples.append((conc_str, file_bytes, fn))
                            if csv_item.get("last_modified"):
                                file_dates.append(csv_item["last_modified"])
                            yield send_event("log", {"message": f"  Downloaded {fn} (Std {conc_str} µg/mL)", "level": "info"})
                        except Exception as e:
                            yield send_event("log", {"message": f"  [SKIP] {fn}: download failed ({e})", "level": "warn"})

                    cal = _build_curve_from_peakdata_csvs(csv_tuples)
                    if not cal:
                        yield send_event("log", {"message": f"  [SKIP] {sid_key}: not enough data points", "level": "dim"})
                        skipped_no_data += 1
                        continue

                    # Detect instrument: try path first, then Senaite lookup
                    inst_detected = None
                    sample_path = conc_items[0][1].get("path", "") if conc_items else ""
                    if "1260" in sample_path:
                        inst_detected = "1260"
                    elif "1290" in sample_path:
                        inst_detected = "1290"
                    else:
                        yield send_event("log", {"message": f"  Looking up instrument for {sid_key} via SENAITE...", "level": "dim"})
                        inst_detected = await _resolve_instrument_from_senaite(sid_key)
                        if inst_detected:
                            yield send_event("log", {"message": f"  Instrument from SENAITE: {inst_detected}", "level": "info"})
                        else:
                            yield send_event("log", {"message": f"  [WARN] No instrument found for {sid_key}", "level": "warn"})

                    fp_key = (inst_detected or "unknown", tuple(sorted(round(a, 1) for a in cal["areas"])))
                    if fp_key in lims_fps:
                        yield send_event("log", {"message": f"  [DUP] {sid_key}", "level": "dim"})
                        skipped_dup += 1
                        continue
                    lims_fps.add(fp_key)

                    try:
                        regression = calculate_calibration_curve(cal["concentrations"], cal["areas"])
                    except Exception as e:
                        yield send_event("log", {"message": f"  [ERROR] {sid_key}: regression failed ({e})", "level": "error"})
                        errors.append(f"LIMS {sid_key}: {e}")
                        continue

                    avg_rt = round(sum(cal["rts"]) / len(cal["rts"]), 4) if cal["rts"] else None
                    earliest_date = min(file_dates) if file_dates else None

                    lims_curve = CalibrationCurve(
                        peptide_id=lims_peptide.id,
                        peptide_analyte_id=lims_first_analyte.id if lims_first_analyte else None,
                        slope=regression["slope"],
                        intercept=regression["intercept"],
                        r_squared=regression["r_squared"],
                        standard_data={
                            "concentrations": cal["concentrations"],
                            "areas": cal["areas"],
                            "rts": cal["rts"],
                        },
                        source_filename=f"{sid_key}_PeakData (LIMS CSV)",
                        source_path=f"{dir_name}/{sid_key}",
                        source_date=earliest_date,
                        source_sample_id=sid_key,
                        is_active=False,
                        reference_rt=avg_rt,
                        instrument=inst_detected,
                        instrument_id=_resolve_instrument(db, name=inst_detected)[1] if inst_detected else None,
                    )
                    db.add(lims_curve)
                    folder_lims_new += 1
                    new_curves += 1

                    yield send_event("log", {
                        "message": f"  [OK] {sid_key}: R²={regression['r_squared']:.4f} ({len(cal['concentrations'])} pts, {inst_detected or '?'})",
                        "level": "info",
                    })

                db.merge(SharePointFileCache(
                    source_path=cache_key, peptide_abbreviation=peptide_abbr,
                    produced_calibration=(folder_lims_new > 0),
                ))
                db.flush()

                # Re-evaluate is_active for this peptide if new curves were added
                if folder_lims_new > 0:
                    lims_all = db.execute(
                        select(CalibrationCurve).where(CalibrationCurve.peptide_id == lims_peptide.id)
                    ).scalars().all()
                    for c in lims_all:
                        c.is_active = False
                    by_inst: dict[str, list] = {}
                    for c in lims_all:
                        by_inst.setdefault(c.instrument or "unknown", []).append(c)
                    for inst_curves in by_inst.values():
                        newest = max(inst_curves, key=lambda c: (c.run_date or c.source_date or c.created_at))
                        newest.is_active = True

            db.commit()

            yield send_event("log", {"message": "\n✓ Import complete", "level": "info"})
            yield send_event("log", {"message": f"  {new_curves} new curves imported", "level": "info"})
            yield send_event("log", {"message": f"  {skipped_cached} already cached (skipped)", "level": "info"})
            yield send_event("log", {"message": f"  {skipped_no_data} skipped (no data)", "level": "info"})
            yield send_event("log", {"message": f"  {skipped_dup} duplicates skipped", "level": "info"})
            if errors:
                yield send_event("log", {"message": f"  {len(errors)} errors", "level": "warn"})
            yield send_event("done", {
                "success": True,
                "new_curves": new_curves,
                "skipped_cached": skipped_cached,
                "skipped_no_data": skipped_no_data,
                "skipped_dup": skipped_dup,
            })

        except Exception as e:
            db.rollback()
            yield send_event("log", {"message": f"✗ FAILED: {e}", "level": "error"})
            yield send_event("done", {"success": False, "error": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.get("/hplc/peptides/{peptide_id}/resync/stream")
async def resync_peptide_stream(
    peptide_id: int,
    analyte_id: Optional[int] = None,
    sample_id: Optional[str] = None,
    folder_path: Optional[str] = None,
    folder_root: Optional[str] = None,
    instrument: Optional[str] = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    SSE streaming re-sync of a single peptide from SharePoint.
    Clears the file cache for this peptide so all files are re-downloaded and re-parsed.

    Optional query params:
      analyte_id — PeptideAnalyte ID to link new curves to
      sample_id — Senaite sample ID to save on the analyte record
    """
    from starlette.responses import StreamingResponse
    import sharepoint as sp
    import asyncio
    from calculations.calibration import calculate_calibration_curve

    # Look up peptide first (outside generator so we can 404 early)
    peptide = db.get(Peptide, peptide_id)
    if not peptide:
        raise HTTPException(status_code=404, detail="Peptide not found")

    # Resolve analyte and persist sample_id if provided
    resolved_analyte_id: Optional[int] = None
    if analyte_id:
        analyte = db.execute(
            select(PeptideAnalyte).where(
                PeptideAnalyte.id == analyte_id,
                PeptideAnalyte.peptide_id == peptide_id,
            )
        ).scalar_one_or_none()
        if analyte:
            resolved_analyte_id = analyte.id
            if sample_id and sample_id.strip():
                analyte.sample_id = sample_id.strip()
                db.commit()

    abbreviation = peptide.abbreviation
    requested_instrument = (instrument or "").strip() or None
    requested_instrument_id: Optional[int] = None
    if requested_instrument:
        _, requested_instrument_id = _resolve_instrument(db, name=requested_instrument)

    async def event_generator():
        import re

        def send_event(event_type: str, data: dict) -> str:
            if event_type == "log":
                import sys
                msg = data.get("message", "")
                level = data.get("level", "info")
                print(f"[{level.upper()}] {msg}", file=sys.stderr)
            payload = json.dumps(data)
            return f"event: {event_type}\ndata: {payload}\n\n"

        calibrations_added = 0

        try:
            yield send_event("log", {"message": f"Re-syncing {abbreviation}...", "level": "heading"})
            await asyncio.sleep(0)

            # 1. Clear existing curves for this peptide (full re-import)
            deleted_curves = db.execute(
                delete(CalibrationCurve).where(CalibrationCurve.peptide_id == peptide.id)
            )
            deleted_cache = db.execute(
                delete(SharePointFileCache).where(SharePointFileCache.peptide_abbreviation == abbreviation)
            )
            db.flush()
            yield send_event("log", {
                "message": f"Cleared {deleted_curves.rowcount} existing curve(s) and {deleted_cache.rowcount} cache entry(ies)",
                "level": "info",
            })

            new_curves: list[CalibrationCurve] = []

            # ── Branch: LIMS CSV PeakData pipeline (sample_id provided) ──
            if sample_id and sample_id.strip():
                sid = sample_id.strip()
                yield send_event("log", {"message": f"Searching LIMS CSV folder for sample {sid}...", "level": "info"})

                folder = await sp.find_lims_sample_folder(sid)
                if not folder:
                    yield send_event("log", {"message": f"[ERROR] No LIMS folder found for sample ID '{sid}'", "level": "error"})
                    yield send_event("done", {"success": False, "error": f"LIMS folder not found for {sid}"})
                    return

                yield send_event("log", {"message": f"Found folder: {folder['name']}", "level": "info"})

                # Recursively list all CSVs (Std_ files may be in subfolders)
                all_files = await sp.list_files_recursive(
                    folder["path"], extensions=[".csv"], root="lims",
                )
                peakdata_pattern = re.compile(
                    rf"^{re.escape(sid)}_Std_(\d+(?:\.\d+)?)_.*PeakData\.csv$", re.IGNORECASE
                )

                peakdata_files = []
                for item in all_files:
                    m = peakdata_pattern.match(item["name"])
                    if m:
                        peakdata_files.append((m.group(1), item))

                if not peakdata_files:
                    yield send_event("log", {"message": f"[ERROR] No PeakData CSV files found matching {sid}_Std_*_PeakData.csv", "level": "error"})
                    yield send_event("done", {"success": False, "error": "No PeakData CSVs found"})
                    return

                yield send_event("log", {"message": f"Found {len(peakdata_files)} PeakData CSV file(s)", "level": "info"})
                yield send_event("progress", {"current": 0, "total": len(peakdata_files), "phase": "downloading"})

                # Download all PeakData CSVs
                csv_tuples: list[tuple[str, bytes, str]] = []
                file_dates: list[str] = []
                for idx, (conc_str, item) in enumerate(sorted(peakdata_files, key=lambda x: float(x[0]))):
                    fn = item["name"]
                    yield send_event("progress", {"current": idx + 1, "total": len(peakdata_files), "phase": f"Downloading {fn}"})
                    try:
                        file_bytes, _ = await sp.download_file(item["id"])
                        csv_tuples.append((conc_str, file_bytes, fn))
                        if item.get("last_modified"):
                            file_dates.append(item["last_modified"])
                        yield send_event("log", {"message": f"  Downloaded {fn} (Std {conc_str} µg/mL)", "level": "info"})
                    except Exception as e:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: download failed ({e})", "level": "warn"})

                # Build curve from PeakData CSVs
                cal = _build_curve_from_peakdata_csvs(csv_tuples)
                if not cal:
                    yield send_event("log", {"message": "[ERROR] Not enough valid data points to build curve (need >= 3)", "level": "error"})
                    yield send_event("done", {"success": False, "error": "Fewer than 3 valid data points"})
                    return

                yield send_event("log", {
                    "message": f"Built curve from {len(cal['concentrations'])} data points",
                    "level": "info",
                })

                # Run regression
                try:
                    regression = calculate_calibration_curve(cal["concentrations"], cal["areas"])
                except Exception as e:
                    yield send_event("log", {"message": f"[ERROR] Regression failed: {e}", "level": "error"})
                    yield send_event("done", {"success": False, "error": f"Regression failed: {e}"})
                    return

                # Compute average RT
                avg_rt = round(sum(cal["rts"]) / len(cal["rts"]), 4) if cal["rts"] else None

                # Use earliest file modification date as source_date
                earliest_date = min(file_dates) if file_dates else None

                curve = CalibrationCurve(
                    peptide_id=peptide.id,
                    peptide_analyte_id=resolved_analyte_id,
                    slope=regression["slope"],
                    intercept=regression["intercept"],
                    r_squared=regression["r_squared"],
                    standard_data={
                        "concentrations": cal["concentrations"],
                        "areas": cal["areas"],
                        "rts": cal["rts"],
                    },
                    source_filename=f"{sid}_PeakData (LIMS CSV)",
                    source_path=folder["path"],
                    source_date=earliest_date,
                    source_sample_id=sid,
                    is_active=True,
                    reference_rt=avg_rt,
                    instrument=requested_instrument,
                    instrument_id=requested_instrument_id,
                )

                # Fetch chromatogram DAD1A files for the curve (best-effort)
                chrom_pattern_sid = re.compile(
                    rf"^{re.escape(sid)}_Std_(\d+(?:\.\d+)?)[\._].*DAD1A\.CSV$", re.IGNORECASE
                )
                chrom_by_conc: dict[str, dict] = {}
                for item in all_files:
                    cm = chrom_pattern_sid.match(item["name"])
                    if cm:
                        conc_label = cm.group(1)
                        try:
                            content_bytes, _ = await sp.download_file(item["id"])
                            csv_text = content_bytes.decode("utf-8", errors="replace")
                            times, signals = [], []
                            for cline in csv_text.splitlines():
                                cline = cline.strip()
                                if not cline:
                                    continue
                                parts = cline.split(",", 1)
                                if len(parts) != 2:
                                    continue
                                try:
                                    times.append(float(parts[0]))
                                    signals.append(float(parts[1]))
                                except ValueError:
                                    continue
                            if times:
                                chrom_by_conc[conc_label] = {"times": times, "signals": signals}
                        except Exception:
                            pass  # best-effort
                if chrom_by_conc:
                    curve.chromatogram_data = chrom_by_conc
                    curve.source_sharepoint_folder = folder["path"]

                db.add(curve)
                new_curves.append(curve)
                calibrations_added = 1

                chrom_msg = f", {len(chrom_by_conc)} chromatogram(s)" if chrom_by_conc else ""
                yield send_event("log", {
                    "message": f"  [OK] Curve: R²={regression['r_squared']:.6f}, slope={regression['slope']:.4f}, intercept={regression['intercept']:.4f}{chrom_msg}",
                    "level": "success",
                })
                if avg_rt:
                    yield send_event("log", {"message": f"  Reference RT: {avg_rt} min", "level": "success"})

                # Log each data point
                for i, (c, a) in enumerate(zip(cal["concentrations"], cal["areas"])):
                    rt_str = f", RT={cal['rts'][i]:.3f}" if i < len(cal["rts"]) else ""
                    yield send_event("log", {"message": f"    Std {c} µg/mL → Area {a:.2f}{rt_str}", "level": "dim"})

            # ── Branch: Folder path scan (folder_path provided, no sample_id) ──
            elif folder_path and folder_path.strip():
                fpath = folder_path.strip()
                folder_name = fpath.rsplit("/", 1)[-1] if "/" in fpath else fpath
                yield send_event("log", {"message": f"Scanning folder: {fpath}...", "level": "info"})

                scan_root = folder_root if folder_root in ("lims", "peptides") else "lims"
                all_files = await sp.list_files_recursive(fpath, extensions=[".csv"], root=scan_root)

                # Match any *_Std_{concentration}_*_PeakData.csv (any sample ID prefix)
                std_pattern = re.compile(
                    r"^(.+?)_Std_(\d+(?:\.\d+)?)_.*PeakData\.csv$", re.IGNORECASE
                )

                peakdata_files = []
                for item in all_files:
                    m = std_pattern.match(item["name"])
                    if m:
                        peakdata_files.append((m.group(2), item))  # (concentration_str, item)

                if not peakdata_files:
                    yield send_event("log", {"message": "[ERROR] No *_Std_*_PeakData.csv files found in folder", "level": "error"})
                    yield send_event("done", {"success": False, "error": "No Std PeakData CSVs found in folder"})
                    return

                yield send_event("log", {"message": f"Found {len(peakdata_files)} Std PeakData file(s)", "level": "info"})
                yield send_event("progress", {"current": 0, "total": len(peakdata_files), "phase": "downloading"})

                csv_tuples: list[tuple[str, bytes, str]] = []
                file_dates: list[str] = []
                for idx, (conc_str, item) in enumerate(sorted(peakdata_files, key=lambda x: float(x[0]))):
                    fn = item["name"]
                    yield send_event("progress", {"current": idx + 1, "total": len(peakdata_files), "phase": f"Downloading {fn}"})
                    try:
                        file_bytes, _ = await sp.download_file(item["id"])
                        csv_tuples.append((conc_str, file_bytes, fn))
                        if item.get("last_modified"):
                            file_dates.append(item["last_modified"])
                        yield send_event("log", {"message": f"  Downloaded {fn} (Std {conc_str} µg/mL)", "level": "info"})
                    except Exception as e:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: download failed ({e})", "level": "warn"})

                # Build curve from PeakData CSVs
                cal = _build_curve_from_peakdata_csvs(csv_tuples)
                if not cal:
                    yield send_event("log", {"message": "[ERROR] Not enough valid data points to build curve (need >= 3)", "level": "error"})
                    yield send_event("done", {"success": False, "error": "Fewer than 3 valid data points"})
                    return

                yield send_event("log", {
                    "message": f"Built curve from {len(cal['concentrations'])} data points",
                    "level": "info",
                })

                # Run regression
                try:
                    regression = calculate_calibration_curve(cal["concentrations"], cal["areas"])
                except Exception as e:
                    yield send_event("log", {"message": f"[ERROR] Regression failed: {e}", "level": "error"})
                    yield send_event("done", {"success": False, "error": f"Regression failed: {e}"})
                    return

                # Compute average RT
                avg_rt = round(sum(cal["rts"]) / len(cal["rts"]), 4) if cal["rts"] else None

                # Use earliest file modification date as source_date
                earliest_date = min(file_dates) if file_dates else None

                curve = CalibrationCurve(
                    peptide_id=peptide.id,
                    peptide_analyte_id=resolved_analyte_id,
                    slope=regression["slope"],
                    intercept=regression["intercept"],
                    r_squared=regression["r_squared"],
                    standard_data={
                        "concentrations": cal["concentrations"],
                        "areas": cal["areas"],
                        "rts": cal["rts"],
                    },
                    source_filename=f"{folder_name}_PeakData (folder scan)",
                    source_path=fpath,
                    source_date=earliest_date,
                    source_sample_id=None,
                    is_active=True,
                    reference_rt=avg_rt,
                    instrument=requested_instrument,
                    instrument_id=requested_instrument_id,
                )

                # Fetch chromatogram DAD1A files from the same folder (best-effort)
                chrom_pattern_folder = re.compile(
                    r".*_Std_(\d+(?:\.\d+)?)[\._].*DAD1A\.CSV$", re.IGNORECASE
                )
                folder_chrom: dict[str, dict] = {}
                for item in all_files:
                    cm = chrom_pattern_folder.match(item["name"])
                    if cm:
                        conc_label = cm.group(1)
                        try:
                            content_bytes, _ = await sp.download_file(item["id"])
                            csv_text = content_bytes.decode("utf-8", errors="replace")
                            times, signals = [], []
                            for cline in csv_text.splitlines():
                                cline = cline.strip()
                                if not cline:
                                    continue
                                parts = cline.split(",", 1)
                                if len(parts) != 2:
                                    continue
                                try:
                                    times.append(float(parts[0]))
                                    signals.append(float(parts[1]))
                                except ValueError:
                                    continue
                            if times:
                                folder_chrom[conc_label] = {"times": times, "signals": signals}
                        except Exception:
                            pass  # best-effort
                if folder_chrom:
                    curve.chromatogram_data = folder_chrom
                    curve.source_sharepoint_folder = fpath

                db.add(curve)
                new_curves.append(curve)
                calibrations_added = 1

                chrom_msg = f", {len(folder_chrom)} chromatogram(s)" if folder_chrom else ""
                yield send_event("log", {
                    "message": f"  [OK] Curve: R²={regression['r_squared']:.6f}, slope={regression['slope']:.4f}, intercept={regression['intercept']:.4f}{chrom_msg}",
                    "level": "success",
                })
                if avg_rt:
                    yield send_event("log", {"message": f"  Reference RT: {avg_rt} min", "level": "success"})

                for i, (c, a) in enumerate(zip(cal["concentrations"], cal["areas"])):
                    rt_str = f", RT={cal['rts'][i]:.3f}" if i < len(cal["rts"]) else ""
                    yield send_event("log", {"message": f"    Std {c} µg/mL → Area {a:.2f}{rt_str}", "level": "dim"})

            # ── Fallback: Legacy xlsx pipeline (no sample_id) ──
            else:
                skip_file_fragments = ["DAD1A", "YearMonthDay", "template", "Template",
                                        "Master sheet", "Calibration_Curve_Template", "P-###"]

                yield send_event("log", {"message": "Scanning SharePoint for calibration files...", "level": "info"})
                std_files = []
                try:
                    all_xlsx = await sp.list_files_recursive(
                        abbreviation, extensions=[".xlsx"], root="peptides",
                    )
                    for item in all_xlsx:
                        fn = item["name"]
                        if fn.startswith("~$"):
                            continue
                        if any(frag in fn for frag in skip_file_fragments):
                            continue
                        std_files.append(item)
                    yield send_event("log", {"message": f"Found {len(std_files)} standard file(s)", "level": "info"})
                    yield send_event("progress", {"current": 0, "total": len(std_files), "phase": "downloading"})
                except Exception as e:
                    yield send_event("log", {"message": f"[ERROR] File listing failed: {e}", "level": "error"})
                    yield send_event("done", {"success": False, "error": str(e)})
                    return

                seen_fingerprints: set = set()

                for idx, item in enumerate(sorted(std_files, key=lambda f: f["name"])):
                    fn = item["name"]
                    item_path = item.get("path", "")
                    item_id = item["id"]

                    yield send_event("progress", {"current": idx + 1, "total": len(std_files), "phase": f"Downloading {fn}"})

                    try:
                        file_bytes, _ = await sp.download_file(item_id)
                    except Exception as e:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: download failed ({e})", "level": "warn"})
                        db.merge(SharePointFileCache(
                            source_path=item_path, peptide_abbreviation=abbreviation, produced_calibration=False
                        ))
                        continue

                    cal = _parse_calibration_excel_bytes(file_bytes, fn)
                    db.merge(SharePointFileCache(
                        source_path=item_path, peptide_abbreviation=abbreviation, produced_calibration=bool(cal)
                    ))

                    if not cal:
                        yield send_event("log", {"message": f"  [SKIP] {fn}: no parseable curve data", "level": "dim"})
                        continue

                    meta = _parse_filename_metadata(fn, item_path)
                    instrument = meta["instrument"]
                    fp = (instrument, tuple(sorted(round(a, 1) for a in cal["areas"])))
                    if fp in seen_fingerprints:
                        yield send_event("log", {"message": f"  [DUP]  {fn}", "level": "dim"})
                        continue
                    seen_fingerprints.add(fp)

                    try:
                        regression = calculate_calibration_curve(cal["concentrations"], cal["areas"])
                    except Exception as e:
                        yield send_event("log", {"message": f"  [ERROR] {fn}: regression failed ({e})", "level": "error"})
                        continue

                    curve = CalibrationCurve(
                        peptide_id=peptide.id,
                        peptide_analyte_id=resolved_analyte_id,
                        slope=regression["slope"],
                        intercept=regression["intercept"],
                        r_squared=regression["r_squared"],
                        standard_data={"concentrations": cal["concentrations"], "areas": cal["areas"], "rts": cal.get("rts", [])},
                        source_filename=fn,
                        source_path=item_path,
                        sharepoint_url=item.get("webUrl"),
                        source_date=item.get("lastModified"),
                        is_active=False,
                        instrument=meta["instrument"],
                        instrument_id=_resolve_instrument(db, name=meta["instrument"])[1] if meta["instrument"] else None,
                        vendor=meta["vendor"],
                        lot_number=meta["lot_number"],
                        batch_number=meta["batch_number"],
                        cap_color=meta["cap_color"],
                        run_date=meta["run_date"],
                    )
                    db.add(curve)
                    new_curves.append(curve)
                    calibrations_added += 1
                    yield send_event("log", {
                        "message": f"  [OK]   {fn} ({instrument or '?'}, {meta['vendor'] or '?'}, R²={regression['r_squared']:.4f})",
                        "level": "success",
                    })

                db.flush()

                # Set active: newest per instrument by run_date
                if new_curves:
                    by_instrument: dict[str, list] = {}
                    for c in new_curves:
                        by_instrument.setdefault(c.instrument or "unknown", []).append(c)
                    for inst_curves in by_instrument.values():
                        newest = max(inst_curves, key=lambda c: (c.run_date or c.source_date or c.created_at))
                        newest.is_active = True
                        yield send_event("log", {
                            "message": f"  ✓ Active ({newest.instrument or '?'}): {newest.source_filename}",
                            "level": "success",
                        })

                    # Update reference RT from the active curve (prefer 1290 > 1260 > unknown)
                    priority_order = ["1290", "1260", "unknown"]
                    active_for_rt = None
                    for inst_key in priority_order:
                        group = by_instrument.get(inst_key)
                        if group:
                            active_for_rt = max(group, key=lambda c: (c.run_date or c.source_date or c.created_at))
                            break
                    if active_for_rt and active_for_rt.standard_data and active_for_rt.standard_data.get("rts"):
                        rts = active_for_rt.standard_data["rts"]
                        active_for_rt.reference_rt = round(sum(rts) / len(rts), 4)
                        yield send_event("log", {"message": f"  Updated reference RT: {active_for_rt.reference_rt}", "level": "success"})

            db.commit()
            yield send_event("refresh", {})

            yield send_event("log", {
                "message": f"Done — {calibrations_added} curve(s) imported",
                "level": "success" if calibrations_added else "info",
            })
            yield send_event("done", {"success": True, "calibrations": calibrations_added})

        except Exception as e:
            db.rollback()
            yield send_event("log", {"message": f"✗ FAILED: {e}", "level": "error"})
            yield send_event("done", {"success": False, "error": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


class DilutionRow(BaseModel):
    """One dilution level's weights from the Excel file."""
    label: str
    concentration: Optional[str] = None
    dil_vial_empty: float
    dil_vial_with_diluent: float
    dil_vial_with_diluent_and_sample: float


class TechCalibrationData(BaseModel):
    """Calibration curve data extracted from the tech's working Excel file."""
    concentrations: list[float]
    areas: list[float]
    slope: float
    intercept: float
    r_squared: float
    n_points: int
    matching_curve_ids: list[int] = []  # IDs of stored curves that match


class AnalyteWeights(BaseModel):
    """Weight data for a single analyte in a blend."""
    sheet_name: str
    stock_vial_empty: Optional[float] = None
    stock_vial_with_diluent: Optional[float] = None
    dilution_rows: list[DilutionRow] = []


class WeightExtractionResponse(BaseModel):
    """Extracted weight data from a lab Excel file."""
    found: bool
    folder_name: Optional[str] = None
    peptide_folder: Optional[str] = None
    excel_filename: Optional[str] = None
    stock_vial_empty: Optional[float] = None
    stock_vial_with_diluent: Optional[float] = None
    dilution_rows: list[DilutionRow] = []
    tech_calibration: Optional[TechCalibrationData] = None
    analytes: list[AnalyteWeights] = []
    error: Optional[str] = None


def _extract_weights_from_excel_bytes(data: bytes) -> dict:
    """
    Parse a lab HPLC Excel file (bytes) for stock + dilution weights.

    For blend Excel files with per-analyte tabs, collects data from ALL sheets
    into an 'analytes' list. The top-level fields are populated from the first
    sheet that has data.

    Tries multiple layout strategies in order per sheet:
    1. F/G/H columns with "Stock" label in col E
    2. "Peptide Sample Stock Preparation" section header
    3. Header-label scan: rows with "vial and cap" / "vial cap and diluent" headers
    4. Alternate layout: A/B labels for stock, C/E/H for dilution data
    """
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    result = {
        "stock_vial_empty": None,
        "stock_vial_with_diluent": None,
        "dilution_rows": [],
        "analytes": [],
    }

    max_scan_row = 70  # Some files have weight data past row 40

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # --- Strategy 1: "Sample" sheet layout (F/G/H columns, Stock in row with "Stock" in col E) ---
        stock_empty = None
        stock_diluent = None
        dilutions = []

        for row in range(1, max_scan_row):
            e_val = ws.cell(row=row, column=5).value  # col E
            f_val = ws.cell(row=row, column=6).value  # col F
            g_val = ws.cell(row=row, column=7).value  # col G
            h_val = ws.cell(row=row, column=8).value  # col H

            # Check for stock row
            if e_val and isinstance(e_val, str) and "stock" in e_val.lower():
                if isinstance(f_val, (int, float)) and isinstance(g_val, (int, float)):
                    stock_empty = float(f_val)
                    stock_diluent = float(g_val)
                continue

            # Check for dilution data rows (need F, G, H all numeric)
            if (isinstance(f_val, (int, float)) and f_val > 2000
                    and isinstance(g_val, (int, float)) and g_val > f_val
                    and isinstance(h_val, (int, float)) and h_val >= g_val):
                conc_label = str(e_val) if e_val else f"Row {row}"
                dilutions.append({
                    "label": conc_label,
                    "concentration": conc_label,
                    "dil_vial_empty": float(f_val),
                    "dil_vial_with_diluent": float(g_val),
                    "dil_vial_with_diluent_and_sample": float(h_val),
                })

        if dilutions:
            # Record this sheet's data as an analyte
            result["analytes"].append({
                "sheet_name": sheet_name,
                "stock_vial_empty": stock_empty,
                "stock_vial_with_diluent": stock_diluent,
                "dilution_rows": dilutions,
            })
            # Populate top-level from the first sheet that has data
            if result["stock_vial_empty"] is None:
                result["stock_vial_empty"] = stock_empty
                result["stock_vial_with_diluent"] = stock_diluent
                result["dilution_rows"] = dilutions
            continue  # Try next sheet (blend support)

        # --- Strategy 2: "Peptide Sample Stock Preparation" section ---
        # Files have a standard-curve section first (rows 1-36) then a sample-prep
        # section starting with a "Peptide Sample Stock Preparation" header.
        # Layout:
        #   Row 43: "Weight Sample Vial and cap (mg)" | 5501.68    (A/B label-value pairs)
        #   Row 45: "Weight of Sample Vial cap and Diluent (mg)" | 8505.75
        #   Row 53: header row with "Weight Vial and cap (mg)" in B, "...Diluent (mg)" in D,
        #           "...Diluent and sample (mg)" in G
        #   Row 54: data row with values in B, D, G
        sample_section_start = None
        for row in range(1, max_scan_row):
            a_val = ws.cell(row=row, column=1).value
            if (a_val and isinstance(a_val, str)
                    and "peptide sample stock preparation" in a_val.lower()):
                sample_section_start = row
                break

        if sample_section_start:
            # Extract stock vial weights from A/B label-value pairs
            sample_stock_empty = None
            sample_stock_diluent = None
            for row in range(sample_section_start, min(sample_section_start + 15, max_scan_row)):
                a_val = ws.cell(row=row, column=1).value
                b_val = ws.cell(row=row, column=2).value
                if not a_val or not isinstance(a_val, str):
                    continue
                lower = a_val.lower()
                if "weight" in lower and "vial" in lower and "cap" in lower:
                    if "diluent" in lower:
                        # "Weight of Sample Vial cap and Diluent (mg)" → stock with diluent
                        if isinstance(b_val, (int, float)):
                            sample_stock_diluent = float(b_val)
                    elif "sample" in lower:
                        # "Weight Sample Vial and cap (mg)" → stock empty
                        if isinstance(b_val, (int, float)):
                            sample_stock_empty = float(b_val)

            # Find the dilution header row: scan for a row where multiple columns
            # have weight-related header strings (e.g. "Weight Vial and cap")
            dil_header_row = None
            for row in range(sample_section_start + 5, min(sample_section_start + 20, max_scan_row)):
                weight_headers = 0
                for col in range(1, 10):
                    hdr = ws.cell(row=row, column=col).value
                    if hdr and isinstance(hdr, str) and "weight" in hdr.lower() and "vial" in hdr.lower():
                        weight_headers += 1
                if weight_headers >= 2:
                    dil_header_row = row
                    break

            sample_dilutions = []
            if dil_header_row:
                # Map columns by header content
                empty_col = None
                diluent_col = None
                sample_col = None
                for col in range(1, 12):
                    hdr = ws.cell(row=dil_header_row, column=col).value
                    if not hdr or not isinstance(hdr, str):
                        continue
                    h_lower = hdr.lower()
                    if "sample" in h_lower and "diluent" in h_lower:
                        sample_col = col
                    elif "diluent" in h_lower and "weight" in h_lower:
                        diluent_col = col
                    elif "vial" in h_lower and "cap" in h_lower and "diluent" not in h_lower:
                        empty_col = col

                if empty_col and diluent_col and sample_col:
                    for data_row in range(dil_header_row + 1, dil_header_row + 5):
                        ev = ws.cell(row=data_row, column=empty_col).value
                        dv = ws.cell(row=data_row, column=diluent_col).value
                        sv = ws.cell(row=data_row, column=sample_col).value
                        if (isinstance(ev, (int, float)) and ev > 1000
                                and isinstance(dv, (int, float)) and dv > ev
                                and isinstance(sv, (int, float)) and sv >= dv):
                            # Try to get the target concentration from the section above
                            conc_label = None
                            for scan_row in range(sample_section_start, dil_header_row):
                                scan_a = ws.cell(row=scan_row, column=1).value
                                if scan_a and isinstance(scan_a, str) and "target conc" in scan_a.lower():
                                    scan_b = ws.cell(row=scan_row, column=2).value
                                    if scan_b is not None:
                                        conc_label = str(int(scan_b) if isinstance(scan_b, float) and scan_b == int(scan_b) else scan_b)
                            sample_dilutions.append({
                                "label": conc_label or f"Row {data_row}",
                                "concentration": conc_label,
                                "dil_vial_empty": float(ev),
                                "dil_vial_with_diluent": float(dv),
                                "dil_vial_with_diluent_and_sample": float(sv),
                            })
                        else:
                            break

            if sample_dilutions:
                result["analytes"].append({
                    "sheet_name": sheet_name,
                    "stock_vial_empty": sample_stock_empty,
                    "stock_vial_with_diluent": sample_stock_diluent,
                    "dilution_rows": sample_dilutions,
                })
                if result["stock_vial_empty"] is None:
                    result["stock_vial_empty"] = sample_stock_empty
                    result["stock_vial_with_diluent"] = sample_stock_diluent
                    result["dilution_rows"] = sample_dilutions
                continue

        # --- Strategy 3: Header-label scan ---
        # Look for header rows containing weight-related labels, then read data from the row below.
        # Handles layouts like:
        #   Row 55: "Weight Vial and cap (mg)" | "Weight of Vial cap and Diluent (mg)" | "Weight of ... and sample (mg)"
        #   Row 56: 2758.79                    | 4139.42                                | 4262.24
        #   Row 59: "Weight Sample Vial and cap (mg)" | "Weight of Vial cap and Diluent (mg)"
        #   Row 60: 5462                              | 6450.26
        dil_header_row = None
        stock_header_row = None

        for row in range(1, max_scan_row):
            a_val = ws.cell(row=row, column=1).value
            if not a_val or not isinstance(a_val, str):
                continue
            lower = a_val.lower()

            # Dilution weights header: contains "vial" + "cap" but NOT "sample vial"
            if "weight" in lower and "vial" in lower and "cap" in lower and "sample" not in lower:
                # Check if col B or C also has a weight-related header (confirming this is a header row)
                b_val = ws.cell(row=row, column=2).value
                c_val = ws.cell(row=row, column=3).value
                has_dil_header = False
                for check in (b_val, c_val):
                    if check and isinstance(check, str) and "diluent" in check.lower():
                        has_dil_header = True
                        break
                if has_dil_header:
                    dil_header_row = row

            # Stock weights header: contains "sample vial" + "cap"
            if "weight" in lower and "sample vial" in lower and "cap" in lower:
                b_val = ws.cell(row=row, column=2).value
                if b_val and isinstance(b_val, str) and "diluent" in b_val.lower():
                    stock_header_row = row

        label_dilutions = []
        if dil_header_row:
            # Determine which columns have data by checking the header row
            # Find columns containing "vial and cap", "diluent", "sample" in the header
            empty_col = None
            diluent_col = None
            sample_col = None
            for col in range(1, 10):
                hdr = ws.cell(row=dil_header_row, column=col).value
                if not hdr or not isinstance(hdr, str):
                    continue
                h_lower = hdr.lower()
                if "sample" in h_lower and "diluent" in h_lower:
                    sample_col = col
                elif "diluent" in h_lower:
                    diluent_col = col
                elif "vial" in h_lower and "cap" in h_lower:
                    empty_col = col

            if empty_col and diluent_col and sample_col:
                # Read data row(s) below header
                for data_row in range(dil_header_row + 1, dil_header_row + 5):
                    ev = ws.cell(row=data_row, column=empty_col).value
                    dv = ws.cell(row=data_row, column=diluent_col).value
                    sv = ws.cell(row=data_row, column=sample_col).value
                    if (isinstance(ev, (int, float)) and ev > 1000
                            and isinstance(dv, (int, float)) and dv > ev
                            and isinstance(sv, (int, float)) and sv >= dv):
                        label_dilutions.append({
                            "label": f"Row {data_row}",
                            "concentration": f"Row {data_row}",
                            "dil_vial_empty": float(ev),
                            "dil_vial_with_diluent": float(dv),
                            "dil_vial_with_diluent_and_sample": float(sv),
                        })
                    else:
                        break  # Stop on first non-data row

        if stock_header_row:
            # Read stock data from the row below the stock header
            for data_row in range(stock_header_row + 1, stock_header_row + 3):
                sv_empty = ws.cell(row=data_row, column=1).value
                sv_dil = ws.cell(row=data_row, column=2).value
                if isinstance(sv_empty, (int, float)) and isinstance(sv_dil, (int, float)):
                    stock_empty = float(sv_empty)
                    stock_diluent = float(sv_dil)
                    break

        if label_dilutions:
            result["analytes"].append({
                "sheet_name": sheet_name,
                "stock_vial_empty": stock_empty,
                "stock_vial_with_diluent": stock_diluent,
                "dilution_rows": label_dilutions,
            })
            if result["stock_vial_empty"] is None:
                result["stock_vial_empty"] = stock_empty
                result["stock_vial_with_diluent"] = stock_diluent
                result["dilution_rows"] = label_dilutions
            continue

        # --- Strategy 4: Alternate layout (A/B labels for stock, C/E/H for dilution) ---
        stock_empty = None
        stock_diluent = None
        for row in range(1, max_scan_row):
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value

            if a_val and isinstance(a_val, str):
                lower = a_val.lower()
                if "stock vial+cap" in lower or "stock vial + cap" in lower:
                    if isinstance(b_val, (int, float)):
                        stock_empty = float(b_val)
                elif "stock peptide+vial" in lower or "stock vial+cap+diluent" in lower:
                    if isinstance(b_val, (int, float)):
                        stock_diluent = float(b_val)

        alt_dilutions = []
        for row in range(1, max_scan_row):
            a_val = ws.cell(row=row, column=1).value
            c_val = ws.cell(row=row, column=3).value
            e_val = ws.cell(row=row, column=5).value
            h_val = ws.cell(row=row, column=8).value

            if (isinstance(c_val, (int, float)) and c_val > 2000
                    and isinstance(e_val, (int, float)) and e_val > c_val
                    and isinstance(h_val, (int, float)) and h_val >= e_val):
                label = str(a_val) if a_val else f"Row {row}"
                if "stock" in label.lower():
                    if stock_empty is None:
                        stock_empty = float(c_val)
                        stock_diluent = float(e_val)
                else:
                    alt_dilutions.append({
                        "label": label,
                        "concentration": label,
                        "dil_vial_empty": float(c_val),
                        "dil_vial_with_diluent": float(e_val),
                        "dil_vial_with_diluent_and_sample": float(h_val),
                    })

        if alt_dilutions:
            result["analytes"].append({
                "sheet_name": sheet_name,
                "stock_vial_empty": stock_empty,
                "stock_vial_with_diluent": stock_diluent,
                "dilution_rows": alt_dilutions,
            })
            if result["stock_vial_empty"] is None:
                result["stock_vial_empty"] = stock_empty
                result["stock_vial_with_diluent"] = stock_diluent
                result["dilution_rows"] = alt_dilutions

    wb.close()
    return result


@app.get("/hplc/weights/{sample_id}", response_model=WeightExtractionResponse)
async def get_sample_weights(
    sample_id: str,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Search SharePoint Peptides folder for a sample ID and extract weight data
    + calibration curve from the associated Excel workbook.

    Scans each peptide subfolder's Raw Data directory for a folder matching
    the sample ID, then downloads and parses the lab Excel file.
    """
    import sharepoint as sp

    # 1. Search SharePoint for the sample folder
    try:
        sample_info = await sp.search_sample_folder(sample_id)
    except Exception as e:
        error_msg = str(e) or repr(e)
        print(f"[WARN] SharePoint search_sample_folder failed for '{sample_id}': {error_msg}")
        return WeightExtractionResponse(
            found=False,
            error=f"SharePoint search error: {error_msg}"
        )

    if not sample_info:
        return WeightExtractionResponse(
            found=False,
            error=f"No folder matching '{sample_id}' found in Peptides root"
        )

    folder_name = sample_info["name"]
    peptide_folder = sample_info["peptide_folder"]
    sample_path = sample_info["path"]

    # 2. List files in the sample folder (and 1290 subfolder)
    try:
        all_files = await sp.list_files_recursive(sample_path, extensions=[".xlsx"])
    except Exception as e:
        return WeightExtractionResponse(
            found=True,
            folder_name=folder_name,
            peptide_folder=peptide_folder,
            error=f"Error listing files: {e}"
        )

    # Filter to lab workbook files (not Agilent data exports)
    excel_candidates = []
    for f in all_files:
        name = f["name"]
        if name.startswith("~$"):
            continue
        if ".dx_" in name or "_PeakData" in name:
            continue
        excel_candidates.append(f)

    if not excel_candidates:
        return WeightExtractionResponse(
            found=True,
            folder_name=folder_name,
            peptide_folder=peptide_folder,
            error="No Excel file found in sample folder"
        )

    # Prefer _Samp_ or _Std_ files
    chosen = None
    for c in excel_candidates:
        if "_Samp_" in c["name"] or "_Std_" in c["name"]:
            chosen = c
            break
    if not chosen:
        chosen = excel_candidates[0]

    # 3. Download and parse the Excel file
    try:
        file_bytes, filename = await sp.download_file(chosen["id"])
    except Exception as e:
        return WeightExtractionResponse(
            found=True,
            folder_name=folder_name,
            peptide_folder=peptide_folder,
            excel_filename=chosen["name"],
            error=f"Error downloading Excel: {e}"
        )

    try:
        weights = _extract_weights_from_excel_bytes(file_bytes)
    except Exception as e:
        return WeightExtractionResponse(
            found=True,
            folder_name=folder_name,
            peptide_folder=peptide_folder,
            excel_filename=filename,
            error=f"Error parsing Excel: {e}"
        )

    # 4. Try to extract calibration curve from the same Excel file
    tech_cal = None
    try:
        cal_data = _parse_calibration_excel_bytes(file_bytes, filename)
        if cal_data and len(cal_data.get("concentrations", [])) >= 3:
            from calculations.calibration import calculate_calibration_curve
            regression = calculate_calibration_curve(
                cal_data["concentrations"], cal_data["areas"]
            )

            # Match against stored curves by area fingerprint
            tech_fp = tuple(sorted(round(a, 1) for a in cal_data["areas"]))
            matching_ids: list[int] = []

            # Find the peptide by folder name to scope the search
            peptide_row = db.execute(
                select(Peptide).where(
                    func.lower(Peptide.abbreviation) == func.lower(peptide_folder)
                )
            ).scalar_one_or_none()

            if peptide_row:
                stored_cals = db.execute(
                    select(CalibrationCurve.id, CalibrationCurve.standard_data, CalibrationCurve.slope, CalibrationCurve.intercept)
                    .where(CalibrationCurve.peptide_id == peptide_row.id)
                ).all()

                for row in stored_cals:
                    # Strategy 1: Area fingerprint match
                    if row.standard_data and "areas" in row.standard_data:
                        stored_fp = tuple(sorted(round(a, 1) for a in row.standard_data["areas"]))
                        if tech_fp == stored_fp:
                            matching_ids.append(row.id)
                            continue

                    # Strategy 2: Slope/intercept match (within 0.1% tolerance)
                    if row.slope and row.intercept:
                        slope_match = abs(row.slope - regression["slope"]) < abs(regression["slope"]) * 0.001
                        intercept_match = abs(row.intercept - regression["intercept"]) < max(abs(regression["intercept"]) * 0.001, 0.01)
                        if slope_match and intercept_match:
                            matching_ids.append(row.id)

            tech_cal = TechCalibrationData(
                concentrations=cal_data["concentrations"],
                areas=cal_data["areas"],
                slope=regression["slope"],
                intercept=regression["intercept"],
                r_squared=regression["r_squared"],
                n_points=regression["n_points"],
                matching_curve_ids=matching_ids,
            )
    except Exception as e:
        print(f"[WARN] Failed to extract calibration from {filename}: {e}")

    # Build analyte weight entries from per-sheet data
    analyte_entries = []
    for a in weights.get("analytes", []):
        analyte_entries.append(AnalyteWeights(
            sheet_name=a["sheet_name"],
            stock_vial_empty=a["stock_vial_empty"],
            stock_vial_with_diluent=a["stock_vial_with_diluent"],
            dilution_rows=[DilutionRow(**d) for d in a["dilution_rows"]],
        ))

    return WeightExtractionResponse(
        found=True,
        folder_name=folder_name,
        peptide_folder=peptide_folder,
        excel_filename=filename,
        stock_vial_empty=weights["stock_vial_empty"],
        stock_vial_with_diluent=weights["stock_vial_with_diluent"],
        dilution_rows=[DilutionRow(**d) for d in weights["dilution_rows"]],
        tech_calibration=tech_cal,
        analytes=analyte_entries,
    )


# --- Explorer Endpoints (Integration Service Database) ---

from integration_db import (
    fetch_orders,
    fetch_ingestions_for_order,
    fetch_attempts_for_order,
    fetch_coa_generations_for_order,
    fetch_sample_events_for_order,
    fetch_access_logs_for_order,
    test_connection,
    get_wordpress_host,
    get_integration_db,
)


class ExplorerOrderResponse(BaseModel):
    """Schema for order from Integration Service database."""
    id: str
    order_id: str
    order_number: str
    customer_id: Optional[int] = None  # WC customer id; null for guest orders (passthrough from IS)
    status: str
    samples_expected: int
    samples_delivered: int
    error_message: Optional[str] = None
    payload: Optional[dict] = None
    sample_results: Optional[dict] = None
    created_at: datetime
    updated_at: datetime
    completed_at: Optional[datetime] = None
    wp_order_status: Optional[str] = None


class ExplorerIngestionResponse(BaseModel):
    """Schema for ingestion from Integration Service database."""
    id: str
    sample_id: str
    coa_version: int
    order_ref: Optional[str] = None
    status: str
    s3_key: Optional[str] = None
    verification_code: Optional[str] = None
    error_message: Optional[str] = None
    created_at: datetime
    updated_at: datetime
    completed_at: Optional[datetime] = None
    processing_time_ms: Optional[int] = None


class ExplorerConnectionStatus(BaseModel):
    """Schema for database connection status."""
    connected: bool
    environment: Optional[str] = None
    database: Optional[str] = None
    host: Optional[str] = None
    wordpress_host: Optional[str] = None
    error: Optional[str] = None


class ExplorerAttemptResponse(BaseModel):
    """Schema for order submission attempt."""
    id: str
    attempt_number: int
    event_id: Optional[str] = None
    status: str
    error_message: Optional[str] = None
    samples_processed: Optional[dict] = None
    created_at: datetime


class ExplorerCOAGenerationResponse(BaseModel):
    """Schema for COA generation record."""
    id: str
    sample_id: str
    generation_number: int
    verification_code: str
    content_hash: str
    status: str
    anchor_status: str
    anchor_tx_hash: Optional[str] = None
    chromatogram_s3_key: Optional[str] = None
    chromatogram_5k_url: Optional[str] = None
    chromatogram_10k_url: Optional[str] = None
    published_at: Optional[datetime] = None
    superseded_at: Optional[datetime] = None
    created_at: datetime
    order_id: Optional[str] = None
    order_number: Optional[str] = None
    parent_generation_id: Optional[str] = None
    # 1-based vial number for per-vial HPLC COA children; None otherwise.
    vial_sequence: Optional[int] = None
    # True for the regular parent-services COA child (plain COA generated
    # alongside a variance primary); False otherwise.
    is_regular_coa: bool = False
    ingestion_status: Optional[str] = None


class ExplorerSampleEventResponse(BaseModel):
    """Schema for sample status event."""
    id: str
    sample_id: str
    transition: str
    new_status: str
    event_id: Optional[str] = None
    event_timestamp: Optional[int] = None
    wp_notified: bool
    wp_status_sent: Optional[str] = None
    wp_error: Optional[str] = None
    created_at: datetime


class ExplorerAccessLogResponse(BaseModel):
    """Schema for COA access log entry."""
    id: str
    sample_id: str
    coa_version: int
    action: str
    requester_ip: Optional[str] = None
    user_agent: Optional[str] = None
    requested_by: Optional[str] = None
    timestamp: datetime


class EnvironmentSwitchRequest(BaseModel):
    """Schema for environment switch request."""
    environment: str


class EnvironmentListResponse(BaseModel):
    """Schema for available environments response."""
    environments: list[str]
    current: str


@app.get("/explorer/environments", response_model=EnvironmentListResponse)
async def get_explorer_environments(_current_user=Depends(get_current_user)):
    """Get list of available database environments."""
    from integration_db import get_available_environments, get_environment
    return EnvironmentListResponse(
        environments=get_available_environments(),
        current=get_environment()
    )


@app.post("/explorer/environments", response_model=ExplorerConnectionStatus)
async def set_explorer_environment(request: EnvironmentSwitchRequest, _current_user=Depends(get_current_user)):
    """
    Switch to a different database environment.
    
    Returns the new connection status after switching.
    """
    from integration_db import set_environment
    try:
        set_environment(request.environment)
        result = test_connection()
        return ExplorerConnectionStatus(**result)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        return ExplorerConnectionStatus(connected=False, error=str(e))


@app.get("/explorer/status", response_model=ExplorerConnectionStatus)
def get_explorer_status(_current_user=Depends(get_current_user)):
    """Test connection to Integration Service database."""
    try:
        result = test_connection()
        return ExplorerConnectionStatus(**result)
    except Exception as e:
        return ExplorerConnectionStatus(connected=False, error=str(e))


@app.get("/explorer/orders", response_model=list[ExplorerOrderResponse])
async def get_explorer_orders(
    search: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    customer_id: Optional[int] = None,
    # UX revision (post-Phase 30): three independent search axes, AND-combined
    # at the IS SQL layer. We forward verbatim — the IS enforces per-axis
    # max_length=256 (T-30-02) and the per-axis SQL-safety pipeline (T-30-01).
    # None = "axis not requested"; empty string is forwarded as-is so the IS
    # sees the same absent-vs-empty semantics it would from a direct caller.
    search_order_number: Optional[str] = None,
    search_sample_id: Optional[str] = None,
    search_analyte: Optional[str] = None,
    sort: Optional[str] = None,
    _current_user=Depends(get_current_user),
):
    """
    Get orders from Integration Service database.

    Query params:
    - search: Filter by order_id or order_number (partial match)
    - limit: Max records to return (default 50)
    - offset: Pagination offset (default 0)
    - customer_id: When present, scopes the list to that WC customer (Phase 29).
      Local fetch_orders does not support this filter, so the request is
      proxied to the Integration Service which owns customer-scoped queries.
    - search_order_number / search_sample_id / search_analyte: UX-revision
      three-input AND search (Customer Detail → Customer Orders tab). Each is
      independently optional; the IS AND-combines whichever are set. Each is
      forwarded only when explicitly provided so an absent param stays absent.
    - sort: Phase 30 sort key (open_first | date_desc | date_asc); forwarded
      to the IS for customer-scoped requests.
    """
    # Phase 29: customer-scoped listing must round-trip to Integration Service.
    if customer_id is not None:
        import httpx as _httpx
        params: dict[str, str] = {
            "customer_id": str(customer_id),
            "limit": str(limit),
            "offset": str(offset),
        }
        # Forward UX-revision search axes + sort only when set so the IS sees
        # the same absent-vs-empty semantics it would from a direct caller.
        # Empty string ('') IS forwarded as-is (back-compat with debounce-flush).
        if search_order_number is not None:
            params["search_order_number"] = search_order_number
        if search_sample_id is not None:
            params["search_sample_id"] = search_sample_id
        if search_analyte is not None:
            params["search_analyte"] = search_analyte
        if sort is not None:
            params["sort"] = sort
        url = f"{os.environ.get('INTEGRATION_SERVICE_URL', 'http://host.docker.internal:8000')}/explorer/orders"
        api_key = os.environ.get("ACCU_MK1_API_KEY", "")
        try:
            async with _httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
                resp = await client.get(url, params=params, headers={"X-API-Key": api_key})
                resp.raise_for_status()
                return resp.json()
        except _httpx.HTTPStatusError as e:
            raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
        except Exception as e:
            raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")

    try:
        orders = fetch_orders(search=search, limit=limit, offset=offset)
        # Convert UUID to string for JSON serialization
        for order in orders:
            order['id'] = str(order['id'])
        return orders
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Failed to connect to Integration Service database: {e}"
        )


@app.get("/explorer/customers")
async def get_explorer_customers(
    search: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    include_test_emails: bool = False,
    _current_user=Depends(get_current_user),
):
    """
    Get WC customers with aggregate stats from Integration Service (Phase 29).

    Pure proxy — the Integration Service owns the aggregate query (`UNION` of
    registered customers + guest billing-email bucket) and the test-email
    filter. The backend forwards limit/offset/include_test_emails verbatim
    (matching the IS shape at integration-service/app/api/desktop.py) and
    returns the raw response: { customers: ExplorerCustomer[], total_count: int }.
    """
    import httpx as _httpx
    params: dict[str, str] = {
        "limit": str(limit),
        "offset": str(offset),
        "include_test_emails": "true" if include_test_emails else "false",
    }
    if search:
        params["search"] = search
    url = f"{os.environ.get('INTEGRATION_SERVICE_URL', 'http://host.docker.internal:8000')}/explorer/customers"
    api_key = os.environ.get("ACCU_MK1_API_KEY", "")
    try:
        async with _httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.get(url, params=params, headers={"X-API-Key": api_key})
            resp.raise_for_status()
            return resp.json()
    except _httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/customers/{customer_id}")
async def get_explorer_customer(customer_id: int, _current_user=Depends(get_current_user)):
    """Get a single WC customer (+ aggregate stats) by id from Integration Service.

    Pure proxy to IS /explorer/customers/{id}. Used by the customer-detail header
    on cold load (deep-link / refresh) when the customers-list cache is empty.
    """
    import httpx as _httpx
    url = f"{os.environ.get('INTEGRATION_SERVICE_URL', 'http://host.docker.internal:8000')}/explorer/customers/{customer_id}"
    api_key = os.environ.get("ACCU_MK1_API_KEY", "")
    try:
        async with _httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.get(url, headers={"X-API-Key": api_key})
            if resp.status_code == 404:
                raise HTTPException(status_code=404, detail=f"Customer {customer_id} not found")
            resp.raise_for_status()
            return resp.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/orders/{order_id}", response_model=ExplorerOrderResponse)
async def get_explorer_order(order_id: str, _current_user=Depends(get_current_user)):
    """Get a single order by WordPress order ID from Integration Service."""
    import httpx as _httpx
    url = f"{os.environ.get('INTEGRATION_SERVICE_URL', 'http://host.docker.internal:8000')}/explorer/orders/{order_id}"
    api_key = os.environ.get("ACCU_MK1_API_KEY", "")
    try:
        async with _httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.get(url, headers={"X-API-Key": api_key})
            if resp.status_code == 404:
                raise HTTPException(status_code=404, detail=f"Order {order_id} not found")
            resp.raise_for_status()
            return resp.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/orders/{order_id}/ingestions", response_model=list[ExplorerIngestionResponse])
def get_order_ingestions(order_id: str, _current_user=Depends(get_current_user)):
    """
    Get all ingestions for an order from Integration Service database.
    
    Args:
        order_id: The WordPress order ID (e.g., "12345")
    """
    try:
        ingestions = fetch_ingestions_for_order(order_id)
        # Convert UUID to string for JSON serialization
        for ing in ingestions:
            ing['id'] = str(ing['id'])
        return ingestions
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Failed to connect to Integration Service database: {e}"
        )


# ── Reports Endpoints (Published COA Results) ─────────────────────
# Read-optimized queries against the published_coa_results table
# in the Integration Service database.


class ReportsSummary(BaseModel):
    total_peptides: int
    total_coas: int
    conforming: int
    non_conforming: int


class PeptideCard(BaseModel):
    analyte_name: str
    is_blend: bool = False
    total_coas: int
    additional_coas: int
    conforming: int
    non_conforming: int
    most_recent_code: Optional[str] = None
    most_recent_sample: Optional[str] = None
    most_recent_status: Optional[str] = None
    most_recent_date: Optional[str] = None
    most_recent_lot: Optional[str] = None


class ReportsDashboard(BaseModel):
    summary: ReportsSummary
    peptides: list[PeptideCard]
    blends: list[PeptideCard] = []


class PurityTrendPoint(BaseModel):
    date: str
    purity_percent: float
    sample_id: str
    verification_code: str
    conforms: Optional[bool] = None


@app.get("/reports/dashboard", response_model=ReportsDashboard)
async def reports_dashboard(
    _current_user=Depends(get_current_user),
):
    """COA reporting dashboard — grouped by product (single peptide or blend)."""
    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                # Deduplicate to one row per COA — use blend overall row for blends, the single row for non-blends
                # This gives us product-level grouping
                cur.execute("""
                    WITH unique_coas AS (
                        SELECT DISTINCT ON (verification_code)
                            verification_code, product_name, is_blend, overall_status,
                            published_at, client_sample_id, lot_code, purity_percent, purity_conforms
                        FROM published_coa_results
                        WHERE (is_blend_overall = true OR is_blend = false)
                        ORDER BY verification_code, published_at DESC
                    )
                    SELECT
                        COUNT(DISTINCT product_name),
                        COUNT(*),
                        COUNT(*) FILTER (WHERE overall_status = 'PASSED'),
                        COUNT(*) FILTER (WHERE overall_status = 'FAILED')
                    FROM unique_coas
                """)
                total_products, total_coas, conforming, non_conforming = cur.fetchone()

                # Per-product cards
                cur.execute("""
                    WITH unique_coas AS (
                        SELECT DISTINCT ON (verification_code)
                            verification_code, product_name, is_blend, overall_status,
                            published_at, client_sample_id, lot_code
                        FROM published_coa_results
                        WHERE (is_blend_overall = true OR is_blend = false)
                        ORDER BY verification_code, published_at DESC
                    ),
                    product_stats AS (
                        SELECT
                            product_name,
                            bool_or(is_blend) as is_blend,
                            COUNT(*) as total_coas,
                            COUNT(*) FILTER (WHERE overall_status = 'PASSED') as conforming,
                            COUNT(*) FILTER (WHERE overall_status = 'FAILED') as non_conforming
                        FROM unique_coas
                        WHERE product_name IS NOT NULL
                        GROUP BY product_name
                    ),
                    most_recent AS (
                        SELECT DISTINCT ON (product_name)
                            product_name,
                            verification_code,
                            client_sample_id,
                            overall_status,
                            published_at,
                            lot_code
                        FROM unique_coas
                        WHERE product_name IS NOT NULL
                        ORDER BY product_name, published_at DESC
                    ),
                    additional_counts AS (
                        SELECT
                            r.product_name,
                            COUNT(DISTINCT child.id) as additional
                        FROM published_coa_results r
                        JOIN coa_generations cg ON cg.verification_code = r.verification_code
                        LEFT JOIN coa_generations child ON child.parent_generation_id = cg.id AND child.status = 'published'
                        WHERE (r.is_blend_overall = true OR r.is_blend = false) AND r.product_name IS NOT NULL AND child.id IS NOT NULL
                        GROUP BY r.product_name
                    )
                    SELECT
                        s.product_name,
                        s.is_blend,
                        s.total_coas,
                        COALESCE(ac.additional, 0) as additional_coas,
                        s.conforming,
                        s.non_conforming,
                        m.verification_code,
                        m.client_sample_id,
                        m.overall_status,
                        m.published_at,
                        m.lot_code
                    FROM product_stats s
                    LEFT JOIN most_recent m ON m.product_name = s.product_name
                    LEFT JOIN additional_counts ac ON ac.product_name = s.product_name
                    ORDER BY s.total_coas DESC
                """)
                products = []
                for row in cur.fetchall():
                    products.append(PeptideCard(
                        analyte_name=row[0],
                        is_blend=row[1],
                        total_coas=row[2],
                        additional_coas=row[3],
                        conforming=row[4],
                        non_conforming=row[5],
                        most_recent_code=row[6],
                        most_recent_sample=row[7],
                        most_recent_status=row[8],
                        most_recent_date=row[9].strftime("%Y-%m-%d") if row[9] else None,
                        most_recent_lot=row[10],
                    ))

                return ReportsDashboard(
                    summary=ReportsSummary(
                        total_peptides=total_products,
                        total_coas=total_coas,
                        conforming=conforming,
                        non_conforming=non_conforming,
                    ),
                    peptides=products,
                )
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Reports database error: {e}")


@app.get("/reports/purity-trend/{analyte_name}", response_model=list[PurityTrendPoint])
async def reports_purity_trend(
    analyte_name: str,
    is_blend: bool = False,
    _current_user=Depends(get_current_user),
):
    """Purity trend over time for a specific analyte or blend."""
    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                if is_blend:
                    # Blend: query by product_name using the blend overall rows
                    cur.execute("""
                        SELECT
                            published_at,
                            purity_percent,
                            sample_id,
                            verification_code,
                            purity_conforms
                        FROM published_coa_results
                        WHERE product_name = %s
                          AND is_blend_overall = true
                          AND purity_percent IS NOT NULL
                        ORDER BY published_at ASC
                    """, (analyte_name,))
                else:
                    cur.execute("""
                        SELECT
                            published_at,
                            purity_percent,
                            sample_id,
                            verification_code,
                            purity_conforms
                        FROM published_coa_results
                        WHERE analyte_name = %s
                          AND NOT is_blend_overall
                          AND purity_percent IS NOT NULL
                        ORDER BY published_at ASC
                    """, (analyte_name,))
                return [
                    PurityTrendPoint(
                        date=row[0].strftime("%Y-%m-%d") if row[0] else "",
                        purity_percent=row[1],
                        sample_id=row[2],
                        verification_code=row[3],
                        conforms=row[4],
                    )
                    for row in cur.fetchall()
                ]
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Reports database error: {e}")


class CheckInRecord(BaseModel):
    sample_id: str
    sample_uid: str
    date_received: str          # ISO 8601 UTC, trailing "Z"
    product_label: Optional[str] = None
    priority: str
    is_test_order: bool = False  # sample belongs to a TEST_EMAILS order (see inbox)


def _test_order_senaite_ids() -> set[str]:
    """SENAITE sample IDs that belong to a test order (billing email in TEST_EMAILS).

    Mirrors the /worksheets/inbox test-order definition: read order_submissions
    from the integration DB, map each order's sample_results → senaite_id, and
    flag those whose payload.billing.email is a known test email. Returns an
    empty set on any failure (graceful degradation — nothing flagged as test).
    """
    TEST_EMAILS = {"forrestp@outlook.com", "forrest@valenceanalytical.com"}
    test_ids: set[str] = set()

    def _as_dict(val):
        if isinstance(val, dict):
            return val
        if isinstance(val, str):
            try:
                parsed = json.loads(val)
                return parsed if isinstance(parsed, dict) else {}
            except (ValueError, TypeError):
                return {}
        return {}

    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT sample_results, payload FROM order_submissions WHERE sample_results IS NOT NULL"
                )
                for sample_results, payload in cur.fetchall():
                    billing = _as_dict(payload).get("billing", {})
                    email = (billing.get("email") or "").lower() if isinstance(billing, dict) else ""
                    if email not in TEST_EMAILS:
                        continue
                    for entry in _as_dict(sample_results).values():
                        if isinstance(entry, dict) and entry.get("senaite_id"):
                            test_ids.add(str(entry["senaite_id"]))
    except Exception:
        pass
    return test_ids


def _parse_day_bound(val: Optional[str], *, end: bool) -> Optional[datetime]:
    """Parse a YYYY-MM-DD string into a naive day boundary (start or end of day).

    Returns None on missing/invalid input so callers can skip the filter.
    date_received is stored naive (UTC), so bounds are naive to match.
    """
    if not val:
        return None
    from datetime import time as _time
    try:
        d = datetime.strptime(val[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None
    return datetime.combine(d, _time.max if end else _time.min)


@app.get("/reports/checkin-times", response_model=list[CheckInRecord])
async def reports_checkin_times(
    from_date: Optional[str] = Query(None, alias="from"),
    to_date: Optional[str] = Query(None, alias="to"),
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Sample check-in events sourced from worksheet_items.date_received.

    Returns raw UTC timestamps (one row per sample); time-of-day bucketing is done
    client-side in the browser's local timezone. worksheet_items holds one row per
    (sample, analysis), so a sample with multiple analyses produces several rows
    sharing one date_received — results are deduped by sample_uid (earliest
    date_received kept, product labels merged). Rows with a null date_received are
    excluded. Optional `from`/`to` are inclusive YYYY-MM-DD day bounds.
    """
    stmt = select(WorksheetItem).where(WorksheetItem.date_received.is_not(None))
    lo = _parse_day_bound(from_date, end=False)
    hi = _parse_day_bound(to_date, end=True)
    if lo is not None:
        stmt = stmt.where(WorksheetItem.date_received >= lo)
    if hi is not None:
        stmt = stmt.where(WorksheetItem.date_received <= hi)

    rows = db.execute(stmt.order_by(WorksheetItem.date_received.desc())).scalars().all()

    test_ids = _test_order_senaite_ids()

    by_uid: dict[str, dict] = {}
    for it in rows:
        names: list[str] = []
        if it.analyses_json:
            try:
                for a in json.loads(it.analyses_json):
                    pn = (a.get("peptide_name") or "").strip()
                    if pn and pn not in names:
                        names.append(pn)
            except (ValueError, TypeError):
                pass
        entry = by_uid.get(it.sample_uid)
        if entry is None:
            by_uid[it.sample_uid] = {
                "sample_id": it.sample_id,
                "sample_uid": it.sample_uid,
                "date_received": it.date_received,
                "priority": it.priority,
                "names": names,
            }
        else:
            if it.date_received < entry["date_received"]:
                entry["date_received"] = it.date_received
            for pn in names:
                if pn not in entry["names"]:
                    entry["names"].append(pn)

    records = [
        CheckInRecord(
            sample_id=e["sample_id"],
            sample_uid=e["sample_uid"],
            date_received=e["date_received"].isoformat() + "Z",
            product_label=", ".join(e["names"]) if e["names"] else None,
            priority=e["priority"],
            is_test_order=e["sample_id"] in test_ids,
        )
        for e in by_uid.values()
    ]
    records.sort(key=lambda r: r.date_received, reverse=True)
    return records


class TurnaroundSample(BaseModel):
    sample_id: str
    ordered_at: Optional[str] = None
    received_at: Optional[str] = None
    submitted_at: Optional[str] = None
    verified_at: Optional[str] = None
    published_at: Optional[str] = None
    is_test_order: bool = False


@app.get("/reports/turnaround", response_model=list[TurnaroundSample])
async def reports_turnaround(
    _current_user=Depends(get_current_user),
):
    """Per-sample SENAITE milestone timestamps for phase-turnaround (bottleneck) analysis.

    Pivots sample_status_events to first-occurrence milestone timestamps
    (receive / submit|partial_submit / verify|partial_verify / publish), left-joins
    the order for the Ordered milestone + test-order flag, and returns raw per-sample
    rows. Phase durations and percentiles are aggregated client-side. event_timestamp
    (Unix seconds — SENAITE's real time) is preferred over our created_at. Both
    source tables live in the integration DB.
    """
    sql = """
        WITH m AS (
            SELECT
                sample_id,
                -- Postgres has no MAX(uuid); a sample's events share one order, so
                -- take the first non-null order_submission_id.
                (array_agg(order_submission_id) FILTER (WHERE order_submission_id IS NOT NULL))[1] AS order_id,
                MIN(COALESCE(to_timestamp(event_timestamp), created_at))
                    FILTER (WHERE transition = 'receive') AS received_at,
                MIN(COALESCE(to_timestamp(event_timestamp), created_at))
                    FILTER (WHERE transition IN ('submit', 'partial_submit')) AS submitted_at,
                MIN(COALESCE(to_timestamp(event_timestamp), created_at))
                    FILTER (WHERE transition IN ('verify', 'partial_verify')) AS verified_at,
                MIN(COALESCE(to_timestamp(event_timestamp), created_at))
                    FILTER (WHERE transition = 'publish') AS published_at
            FROM sample_status_events
            GROUP BY sample_id
        )
        SELECT m.sample_id, os.created_at AS ordered_at,
               m.received_at, m.submitted_at, m.verified_at, m.published_at,
               (LOWER(os.payload->'billing'->>'email') IN
                  ('forrestp@outlook.com', 'forrest@valenceanalytical.com')) AS is_test_order
        FROM m
        LEFT JOIN order_submissions os ON os.id = m.order_id
    """

    from datetime import timezone as _tz

    def _iso(dt):
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=_tz.utc)
        return dt.astimezone(_tz.utc).isoformat().replace("+00:00", "Z")

    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Reports database error: {e}")

    return [
        TurnaroundSample(
            sample_id=sample_id,
            ordered_at=_iso(ordered_at),
            received_at=_iso(received_at),
            submitted_at=_iso(submitted_at),
            verified_at=_iso(verified_at),
            published_at=_iso(published_at),
            is_test_order=bool(is_test),
        )
        for (sample_id, ordered_at, received_at, submitted_at,
             verified_at, published_at, is_test) in rows
    ]


class ReportsSyncStatus(BaseModel):
    source_published: int
    source_verification_codes: int
    report_table_rows: int
    report_verification_codes: int
    missing_codes: list[str]
    orphaned_codes: list[str]
    in_sync: bool


@app.get("/reports/sync-status", response_model=ReportsSyncStatus)
async def reports_sync_status(
    _current_user=Depends(get_current_user),
):
    """Compare published_coa_results with coa_generations source table."""
    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                # Source: coa_generations
                cur.execute("SELECT count(*) FROM coa_generations WHERE status = 'published'")
                source_published = cur.fetchone()[0]
                cur.execute("SELECT count(DISTINCT verification_code) FROM coa_generations WHERE status = 'published'")
                source_codes = cur.fetchone()[0]

                # Report table
                cur.execute("SELECT count(*) FROM published_coa_results")
                report_rows = cur.fetchone()[0]
                cur.execute("SELECT count(DISTINCT verification_code) FROM published_coa_results")
                report_codes = cur.fetchone()[0]

                # Missing: in source but not in report table
                cur.execute("""
                    SELECT verification_code FROM coa_generations
                    WHERE status = 'published' AND coa_data IS NOT NULL
                      AND verification_code NOT IN (SELECT DISTINCT verification_code FROM published_coa_results)
                """)
                missing = [r[0] for r in cur.fetchall()]

                # Orphaned: in report table but no longer published in source
                cur.execute("""
                    SELECT DISTINCT verification_code FROM published_coa_results
                    WHERE verification_code NOT IN (
                        SELECT verification_code FROM coa_generations WHERE status = 'published'
                    )
                """)
                orphaned = [r[0] for r in cur.fetchall()]

                return ReportsSyncStatus(
                    source_published=source_published,
                    source_verification_codes=source_codes,
                    report_table_rows=report_rows,
                    report_verification_codes=report_codes,
                    missing_codes=missing,
                    orphaned_codes=orphaned,
                    in_sync=len(missing) == 0 and len(orphaned) == 0,
                )
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Reports sync check failed: {e}")


@app.post("/reports/resync")
async def reports_resync(
    _current_user=Depends(require_admin),
):
    """Re-sync published_coa_results: backfill missing codes, remove orphans."""
    import json as _json
    synced = 0
    removed = 0
    try:
        with get_integration_db() as conn:
            with conn.cursor() as cur:
                # Remove orphaned rows
                cur.execute("""
                    DELETE FROM published_coa_results
                    WHERE verification_code NOT IN (
                        SELECT verification_code FROM coa_generations WHERE status = 'published'
                    )
                """)
                removed = cur.rowcount

                # Find missing codes
                cur.execute("""
                    SELECT id, sample_id, verification_code, coa_data, published_at, created_at
                    FROM coa_generations
                    WHERE status = 'published' AND coa_data IS NOT NULL
                      AND verification_code NOT IN (SELECT DISTINCT verification_code FROM published_coa_results)
                """)
                missing_rows = cur.fetchall()

                def _parse_float(val):
                    if val is None:
                        return None
                    s = str(val).strip().replace("%", "").replace("mg/mL", "").replace("mg", "").replace("EU/mL", "").strip()
                    try:
                        return float(s)
                    except (ValueError, TypeError):
                        return None

                def _parse_date(val):
                    if not val:
                        return None
                    from datetime import datetime as _dt
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
                        try:
                            return _dt.strptime(val, fmt).date()
                        except ValueError:
                            continue
                    return None

                for gen_id, sample_id, vcode, coa_data, published_at, created_at in missing_rows:
                    if not isinstance(coa_data, dict):
                        coa_data = _json.loads(coa_data)
                    pub_at = published_at or created_at
                    sample = coa_data.get("sample", {})
                    client = coa_data.get("client", {})
                    product = coa_data.get("product", {})
                    results = coa_data.get("results", {})
                    overall_status = coa_data.get("overall_status", "UNKNOWN")
                    received_date = _parse_date(sample.get("received_date"))
                    product_components = product.get("components", [])
                    product_name = ", ".join(product_components) if product_components else product.get("name")
                    is_blend = len(product_components) > 1
                    addons = results.get("addons", [])
                    has_endo = any("endotoxin" in (a.get("test_name", "")).lower() for a in addons)
                    endo_conf = None
                    if has_endo:
                        e = next((a for a in addons if "endotoxin" in a.get("test_name", "").lower()), None)
                        if e: endo_conf = e.get("status", "").upper() == "CONFORMS"
                    has_ster = any("sterility" in (a.get("test_name", "")).lower() for a in addons)
                    ster_conf = None
                    if has_ster:
                        s = next((a for a in addons if "sterility" in a.get("test_name", "").lower()), None)
                        if s: ster_conf = s.get("status", "").upper() == "CONFORMS"

                    base = (gen_id, vcode, sample_id, pub_at, received_date,
                            client.get("name"), sample.get("name"), sample.get("batch_id"), None,
                            product_name, sample.get("matrix_type"), is_blend)

                    insert_sql = """INSERT INTO published_coa_results (
                        coa_generation_id, verification_code, sample_id, published_at, received_date,
                        client_name, client_sample_id, lot_code, order_number,
                        product_name, sample_type, is_blend,
                        analyte_name, is_blend_overall,
                        purity_percent, purity_conforms, purity_spec, identity_conforms,
                        quantity_value, quantity_unit, overall_status,
                        has_endotoxin, endotoxin_conforms, has_sterility, sterility_conforms
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

                    analytes = results.get("analytes", [])
                    if analytes:
                        for analyte in analytes:
                            name = analyte.get("name", "Unknown")
                            is_overall = name.lower() in ("peptide blend", "blend")
                            p = analyte.get("purity", {}); i = analyte.get("identity", {}); q = analyte.get("quantity", {})
                            cur.execute(insert_sql, base + (
                                name, is_overall, _parse_float(p.get("result")), p.get("conforms"), p.get("specification"),
                                i.get("conforms"), _parse_float(q.get("result")), q.get("unit"), overall_status,
                                has_endo, endo_conf, has_ster, ster_conf))
                    else:
                        p = results.get("purity", {}); i = results.get("identity", {}); q = results.get("quantity", {})
                        a_name = p.get("analyte") or (product_components[0] if product_components else "Unknown")
                        cur.execute(insert_sql, base + (
                            a_name, False, _parse_float(p.get("result")), p.get("conforms"), p.get("specification"),
                            i.get("conforms"), _parse_float(q.get("result")), q.get("unit"), overall_status,
                            has_endo, endo_conf, has_ster, ster_conf))
                    synced += 1

                conn.commit()
        return {"synced": synced, "removed": removed, "message": f"Synced {synced} missing, removed {removed} orphaned"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Resync failed: {e}")


# ── Integration Service HTTP Proxy ─────────────────────────────────
# These endpoints proxy through the Integration Service HTTP API
# rather than querying the database directly. This keeps the
# integration-service as the single source of truth for query logic.

import httpx
from httpx_shared import HTTPX_SSL_CONTEXT

INTEGRATION_SERVICE_URL = os.environ.get("INTEGRATION_SERVICE_URL", "http://host.docker.internal:8000")
INTEGRATION_SERVICE_API_KEY = os.environ.get("ACCU_MK1_API_KEY", "")
COA_BUILDER_URL = os.environ.get("COA_BUILDER_URL", "")


_VIAL_SHAPED_ID_RE = re.compile(r"^(?P<parent>.+)-S\d{2,}$")


def _worksheet_notify_target(db: Session, sample_id: str) -> str:
    """Resolve the sample id to notify the IS with when a worksheet item is
    added. Order-status mapping on the IS side (/explorer/worksheet-assigned)
    is keyed by PARENT AR ids — receive-webhook sample_status_events and
    order payload sample_results — so vial ids (…-SNN) must be translated or
    the notification no-ops (no_order_found). DB linkage wins; regex strip is
    the fallback for vial-shaped ids with no lims_sub_samples row."""
    if not sample_id:
        return sample_id
    m = _VIAL_SHAPED_ID_RE.match(sample_id)
    if not m:
        return sample_id
    parent_sid = db.execute(
        select(LimsSample.sample_id)
        .join(LimsSubSample, LimsSubSample.parent_sample_pk == LimsSample.id)
        .where(LimsSubSample.sample_id == sample_id)
    ).scalar_one_or_none()
    return parent_sid or m.group("parent")


async def _notify_worksheet_assigned(sample_id: str) -> None:
    """Fire-and-forget: tell Integration Service a sample was added to a worksheet."""
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/worksheet-assigned"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=10.0) as client:
            resp = await client.post(
                url,
                json={"sample_id": sample_id},
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            resp.raise_for_status()
            print(f"[INFO] worksheet_assigned notified for {sample_id}: {resp.json()}")
    except Exception as e:
        print(f"[WARN] worksheet_assigned notification failed for {sample_id}: {e}")


async def _proxy_explorer_get(path: str) -> list[dict]:
    """Proxy a GET request to the Integration Service explorer API."""
    url = f"{INTEGRATION_SERVICE_URL}/explorer{path}"
    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
        resp = await client.get(url, headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY})
        resp.raise_for_status()
        return resp.json()


@app.get("/explorer/orders/{order_id}/coa-generations", response_model=list[ExplorerCOAGenerationResponse])
async def get_order_coa_generations(order_id: str, _current_user=Depends(get_current_user)):
    """Get COA generation records for an order (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/orders/{order_id}/coa-generations")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/orders/{order_id}/attempts", response_model=list[ExplorerAttemptResponse])
async def get_order_attempts(order_id: str, _current_user=Depends(get_current_user)):
    """Get submission attempts for an order (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/orders/{order_id}/attempts")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/sample-events", response_model=list[ExplorerSampleEventResponse])
async def get_all_sample_events(limit: int = 200, _current_user=Depends(get_current_user)):
    """Get all sample status events across all orders (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/sample-events?limit={limit}")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/orders/{order_id}/sample-events", response_model=list[ExplorerSampleEventResponse])
async def get_order_sample_events(order_id: str, _current_user=Depends(get_current_user)):
    """Get sample status events for an order (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/orders/{order_id}/sample-events")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/orders/{order_id}/access-logs", response_model=list[ExplorerAccessLogResponse])
async def get_order_access_logs(order_id: str, _current_user=Depends(get_current_user)):
    """Get COA access logs for an order (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/orders/{order_id}/access-logs")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


# --- Order Box-Label Summary ---

class BoxLabelSummary(BaseModel):
    order_number: str
    order_date: Optional[str] = None
    counts: dict  # {"hplc": int, "endo": int, "ster": int}


def _fetch_order_submission_row(order_number: str) -> Optional[dict]:
    """The order_submissions row for a WP order number, or None.
    Frontend passes the SENAITE ClientOrderNumber (e.g. 'WP-3263'); the table
    stores the bare number ('3263') in both order_number and order_id. Strip a
    leading 'WP-' and match either column. Returns keys: order_number,
    created_at (datetime|None), sample_results (dict)."""
    from integration_db import get_integration_db
    from psycopg2.extras import RealDictCursor
    raw = (order_number or "").strip()
    stripped = raw[3:] if raw.upper().startswith("WP-") else raw
    with get_integration_db() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT order_number, order_id, created_at, sample_results
                  FROM order_submissions
                 WHERE order_number = %(raw)s OR order_number = %(stripped)s
                    OR order_id = %(raw)s OR order_id = %(stripped)s
                 ORDER BY created_at DESC
                 LIMIT 1
                """,
                {"raw": raw, "stripped": stripped},
            )
            row = cur.fetchone()
            if row is None:
                return None
            return {
                "order_number": raw,
                "created_at": row.get("created_at"),
                "sample_results": row.get("sample_results") or {},
            }


@app.get("/orders/{order_number}/box-label-summary", response_model=BoxLabelSummary)
def get_order_box_label_summary(order_number: str, _current_user=Depends(get_current_user)):
    row = _fetch_order_submission_row(order_number)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Order {order_number} not found")
    counts = {"hplc": 0, "endo": 0, "ster": 0}
    fetch_error = False
    for entry in (row["sample_results"] or {}).values():
        sid = entry.get("senaite_id") if isinstance(entry, dict) else None
        if not sid:
            continue
        try:
            services_resp = sub_service.fetch_sample_services(sid)
        except Exception:
            # fetch_sample_services raises on network/non-2xx and returns None
            # only on a legit 404. A raise means the count would be unreliable —
            # fail loud (below) rather than silently undercount the box label.
            fetch_error = True
            continue
        if not services_resp:
            continue  # legit 404 / unmapped sample → contributes 0
        # IS returns {"services": {...flags...}, ...}; derive_* wants the inner
        # flags dict (mirrors sub_samples.service.build_vial_plan).
        services = services_resp.get("services") or {}
        d = derive_base_demand(services)
        counts["hplc"] += d["hplc"]
        counts["endo"] += d["endo"]
        counts["ster"] += d["ster"]
    if fetch_error:
        # Don't return a silently-undercounted total (which the FE would print as
        # a misleading/blank box label); let the wizard's soft-fail engage.
        raise HTTPException(
            status_code=503,
            detail="Could not reach the analysis service for one or more samples; try again.",
        )
    created = row.get("created_at")
    order_date = created.date().isoformat() if created else None
    return BoxLabelSummary(order_number=row["order_number"], order_date=order_date, counts=counts)


def _fetch_order_submission_rows_batch(order_numbers: list[str]) -> dict[str, dict]:
    """Batched _fetch_order_submission_row: ONE IS-DB query for a page of
    order numbers. Returns {requested_number: row} (absent = not found);
    row keys mirror the single helper. Newest submission wins per order."""
    from integration_db import get_integration_db
    from psycopg2.extras import RealDictCursor

    wanted: dict[str, set[str]] = {}  # requested → its match forms
    for raw in order_numbers:
        raw = (raw or "").strip()
        if not raw:
            continue
        stripped = raw[3:] if raw.upper().startswith("WP-") else raw
        wanted[raw] = {raw, stripped}
    if not wanted:
        return {}
    all_forms = sorted({f for forms in wanted.values() for f in forms})
    with get_integration_db() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT order_number, order_id, created_at, sample_results
                  FROM order_submissions
                 WHERE order_number = ANY(%(forms)s) OR order_id = ANY(%(forms)s)
                 ORDER BY created_at DESC
                """,
                {"forms": all_forms},
            )
            rows = cur.fetchall()
    out: dict[str, dict] = {}
    for requested, forms in wanted.items():
        for row in rows:  # rows are newest-first; first hit wins
            if row.get("order_number") in forms or row.get("order_id") in forms:
                out[requested] = {
                    "order_number": requested,
                    "created_at": row.get("created_at"),
                    "sample_results": row.get("sample_results") or {},
                }
                break
    return out


class BoxLabelSummariesRequest(BaseModel):
    order_numbers: list[str]


class BoxLabelSummariesResponse(BaseModel):
    summaries: dict[str, BoxLabelSummary]  # keyed by REQUESTED order number
    errors: list[str] = []  # orders whose IS fetch failed (never undercounted)


@app.post("/orders/box-label-summaries", response_model=BoxLabelSummariesResponse)
def get_order_box_label_summaries(
    body: BoxLabelSummariesRequest,
    _current_user=Depends(get_current_user),
):
    """Batched box-label summaries: ONE request per receive-by-order PAGE.

    The per-row endpoint above holds this request's DB pool connection through
    a per-sample IS fan-out; ~50 concurrent row cells under HTTP/2 (no browser
    connection cap) exhausted the pool (QueuePool 30s timeout waves) and took
    get_current_user — and login — down with it (prod brownout 2026-07-09).
    Here a whole page costs one batched IS-DB order lookup + a bounded
    (8-thread) IS fan-out over UNIQUE sample ids, holding a single pool
    connection briefly. Per-order failure isolation mirrors the single
    endpoint's fail-loud rule: an order whose services fetch raises lands in
    `errors` (never a silent undercount); the rest resolve normally.
    """
    requested = [n.strip() for n in body.order_numbers if n and n.strip()]
    if len(requested) > 100:
        raise HTTPException(status_code=400, detail="At most 100 order_numbers per request")
    if not requested:
        return BoxLabelSummariesResponse(summaries={}, errors=[])

    rows = _fetch_order_submission_rows_batch(requested)

    order_sids: dict[str, list[str]] = {}
    for num, row in rows.items():
        sids: list[str] = []
        for entry in (row["sample_results"] or {}).values():
            sid = entry.get("senaite_id") if isinstance(entry, dict) else None
            if sid:
                sids.append(sid)
        order_sids[num] = sids
    unique_sids = sorted({s for sids in order_sids.values() for s in sids})

    services_by_sid: dict[str, Optional[dict]] = {}
    failed_sids: set[str] = set()
    if unique_sids:
        from concurrent.futures import ThreadPoolExecutor

        def _one(sid: str):
            try:
                return sid, sub_service.fetch_sample_services(sid), False
            except Exception:
                return sid, None, True

        # Bounded fan-out: 8 concurrent IS calls regardless of page size —
        # never a herd, and this handler's own threadpool slot is the only
        # anyio worker consumed.
        with ThreadPoolExecutor(max_workers=8) as ex:
            for sid, resp, failed in ex.map(_one, unique_sids):
                if failed:
                    failed_sids.add(sid)
                else:
                    services_by_sid[sid] = resp

    summaries: dict[str, BoxLabelSummary] = {}
    errors: list[str] = []
    for num, row in rows.items():
        sids = order_sids[num]
        if any(s in failed_sids for s in sids):
            errors.append(num)
            continue
        counts = {"hplc": 0, "endo": 0, "ster": 0}
        for sid in sids:
            resp = services_by_sid.get(sid)
            if not resp:
                continue  # legit 404 / unmapped sample → contributes 0
            services = resp.get("services") or {}
            d = derive_base_demand(services)
            counts["hplc"] += d["hplc"]
            counts["endo"] += d["endo"]
            counts["ster"] += d["ster"]
        created = row.get("created_at")
        summaries[num] = BoxLabelSummary(
            order_number=row["order_number"],
            order_date=created.date().isoformat() if created else None,
            counts=counts,
        )
    return BoxLabelSummariesResponse(summaries=summaries, errors=errors)


@app.get("/explorer/samples/{sample_id}/additional-coas")
async def get_sample_additional_coas(sample_id: str, _current_user=Depends(get_current_user)):
    """Get additional COA configs for a sample (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get(f"/samples/{sample_id}/additional-coas")
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 404:
            return []
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/analysis-services")
async def get_analysis_services(_current_user=Depends(get_current_user)):
    """List all active analysis services (proxied to Integration Service)."""
    try:
        return await _proxy_explorer_get("/analysis-services")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.post("/explorer/samples/{sample_id}/analyses")
async def add_sample_analysis(
    sample_id: str,
    body: dict,
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Add an analysis service to a sample.

    Native branch: if sample_id maps to a lims_sub_samples row whose
    external_lims_uid starts with 'mk1://', the analysis is created
    directly in Mk1 via add_analysis_to_native_vial.

    Non-native fallthrough: proxied to the Integration Service unchanged.
    """
    from sqlalchemy import select as _select
    from sqlalchemy.exc import IntegrityError as SQLIntegrityError
    from lims_analyses.service import (
        add_analysis_to_native_vial,
        BadRequestError as _BadRequestError,
        NotFoundError as _NotFoundError,
    )

    sub = db.execute(
        _select(LimsSubSample).where(
            LimsSubSample.sample_id == sample_id,
            LimsSubSample.external_lims_uid.like("mk1://%"),
        )
    ).scalar_one_or_none()

    if sub is not None:
        # Native branch
        senaite_service_uid = body.get("service_uid")
        try:
            add_analysis_to_native_vial(
                db,
                sub_sample_pk=sub.id,
                senaite_service_uid=senaite_service_uid,
                keyword=None,
                user_id=_current_user.id,
            )
        except _NotFoundError as e:
            raise HTTPException(status_code=404, detail=str(e))
        except _BadRequestError as e:
            raise HTTPException(status_code=409, detail=str(e))
        except SQLIntegrityError:
            # Unique-index collision (e.g. concurrent add). The index now
            # excludes retracted/rejected rows, so this is a true duplicate.
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail=f"analysis already exists on {sample_id}",
            )
        return {"success": True, "message": "Analysis added"}

    # Non-native: proxy to Integration Service
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/samples/{sample_id}/analyses"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=30.0) as client:
            resp = await client.post(url, json=body, headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY})
            resp.raise_for_status()
            result = resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")

    # ── Parent-add cascade (best-effort) ─────────────────────────────────
    # If the sample is a PARENT with assigned vials, re-run the idempotent
    # seeder per vial so the new service lands on the bench immediately
    # (cascade_parent_add_to_vials self-guards: sample_ids that aren't in
    # lims_samples — e.g. legacy SENAITE secondaries — fall out as no-ops).
    import logging as _logging
    _add_logger = _logging.getLogger(__name__)
    try:
        from lims_analyses.service import cascade_parent_add_to_vials
        _seeded = cascade_parent_add_to_vials(
            db, parent_sample_id=sample_id, user_id=_current_user.id,
        )
        if _seeded:
            _add_logger.info(
                "cascade_parent_add: parent=%s → seeded %s", sample_id, _seeded,
            )
    except Exception as _add_err:
        _add_logger.warning(
            "cascade_parent_add: unexpected error for sample=%s: %s",
            sample_id, _add_err,
        )

    # ── Parent analysis shadow mirror (A7 add, best-effort) ───────────────
    # A shadow row for an analysis with no result yet is correct — it
    # mirrors the line's existence. The IS proxy response carries no
    # keyword (just {success, message}), so resolve service_uid -> keyword
    # via the locally-synced analysis_services table (senaite_uid isn't
    # unique — see resolve_instrument_id's note — hence order_by + first()).
    # Unresolvable service_uid is the documented no-op: skip silently.
    _svc_uid = body.get("service_uid")
    if _svc_uid:
        _add_svc = db.execute(
            _select(AnalysisService).where(AnalysisService.senaite_uid == _svc_uid)
            .order_by(AnalysisService.id)
        ).scalars().first()
        if _add_svc is not None and _add_svc.keyword:
            from fastapi.concurrency import run_in_threadpool
            await run_in_threadpool(
                _mirror_parent_analysis_bg,
                sample_id=sample_id, keyword=_add_svc.keyword,
                mirror_review_state="unassigned",
            )

    return result


@app.get("/explorer/samples/{sample_id}/analyses/{keyword}/removal-impact")
async def get_analysis_removal_impact(
    sample_id: str,
    keyword: str,
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Classify the vial-tier rows a removal of {keyword} from parent
    {sample_id} would touch (pristine / worked_unverified / blocked). Drives
    the Manage Analyses retract-confirm modal. Thin wrapper over the tested
    classify_removal_impact."""
    from lims_analyses.service import classify_removal_impact
    return classify_removal_impact(db, parent_sample_id=sample_id, keyword=keyword)


@app.delete("/explorer/samples/{sample_id}/analyses/{keyword}")
async def remove_sample_analysis(
    sample_id: str,
    keyword: str,
    confirm_retract: bool = False,
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Remove an analysis from a sample.

    Tiered worked-row handling (parent samples with vials): verified/published
    vial rows BLOCK the remove (409 — invalidate/retest first); worked-but-
    unverified rows require confirm_retract=true (412 with the impact payload
    so the FE can show the modal) and are then audited-rejected; pristine rows
    fall through to the existing delete/cascade path below.

    Native branch: if sample_id maps to a lims_sub_samples row whose
    external_lims_uid starts with 'mk1://', the analysis is hard-deleted
    (if pristine) via delete_pristine_analysis.

    Non-native fallthrough: proxied to the Integration Service unchanged.
    """
    from sqlalchemy import select as _select
    from lims_analyses.service import (
        delete_pristine_analysis,
        classify_removal_impact,
        reject_vials_for_parent_keyword,
        BadRequestError as _BadRequestError,
        NotFoundError as _NotFoundError,
    )

    # ── Tiered worked-row guard (parent samples with vials) ──────────────────
    # Verified/published vial rows block; worked-unverified rows need an
    # explicit confirm and are audited-rejected before the delete proceeds.
    # Pristine rows are handled by the existing delete/cascade path below.
    _impact = classify_removal_impact(db, parent_sample_id=sample_id, keyword=keyword)
    if _impact["blocked"]:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Verified/published results exist on {len(_impact['blocked'])} "
                "vial(s) — invalidate or retest those first."
            ),
        )
    if _impact["worked_unverified"] and not confirm_retract:
        # 412: FE shows the retract-confirm modal, then re-submits with
        # confirm_retract=true. Payload is the full impact for the modal.
        raise HTTPException(status_code=412, detail=_impact)
    if _impact["worked_unverified"] and confirm_retract:
        reject_vials_for_parent_keyword(
            db, parent_sample_id=sample_id, keyword=keyword,
            user_id=_current_user.id,
        )

    sub = db.execute(
        _select(LimsSubSample).where(
            LimsSubSample.sample_id == sample_id,
            LimsSubSample.external_lims_uid.like("mk1://%"),
        )
    ).scalar_one_or_none()

    if sub is not None:
        # Native branch
        try:
            delete_pristine_analysis(
                db,
                sub_sample_pk=sub.id,
                keyword=keyword,
                user_id=_current_user.id,
            )
        except _NotFoundError as e:
            raise HTTPException(status_code=404, detail=str(e))
        except _BadRequestError as e:
            raise HTTPException(status_code=409, detail=str(e))
        return {"success": True, "message": "Analysis removed"}

    # Non-native: proxy to Integration Service
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/samples/{sample_id}/analyses/{keyword}"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=30.0) as client:
            resp = await client.delete(url, headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY})
            resp.raise_for_status()
            result = resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")

    # ── Parent-remove cascade (best-effort) ──────────────────────────────
    # If the sample is a PARENT with vials, hard-delete the PRISTINE vial
    # mirror rows of the removed service (rows with activity are skipped —
    # cascade_parent_remove_from_vials self-guards: sample_ids not in
    # lims_samples fall out as no-ops).
    import logging as _logging
    _rm_logger = _logging.getLogger(__name__)
    try:
        from lims_analyses.service import cascade_parent_remove_from_vials
        _removed = cascade_parent_remove_from_vials(
            db, parent_sample_id=sample_id, keyword=keyword,
            user_id=_current_user.id,
        )
        if _removed:
            _rm_logger.info(
                "cascade_parent_remove: parent=%s keyword=%s → removed %s",
                sample_id, keyword, _removed,
            )
    except Exception as _rm_err:
        _rm_logger.warning(
            "cascade_parent_remove: unexpected error for sample=%s keyword=%s: %s",
            sample_id, keyword, _rm_err,
        )

    # ── Parent analysis shadow mirror (A7 remove, best-effort) ───────────
    # Mark the keyword's live shadow row rejected — never deleted (audit
    # trail). Own session via _mirror_parent_analysis_bg, never the request
    # `db` used by the cascade above.
    from fastapi.concurrency import run_in_threadpool
    await run_in_threadpool(
        _mirror_parent_analysis_bg,
        sample_id=sample_id, keyword=keyword,
        mirror_review_state="rejected",
    )

    return result


class ReplaceAnalyteBody(BaseModel):
    """Replace the peptide on one analyte slot of a parent blend AR.

    senaite_uid is the parent AR's SENAITE uid (the FE has it as data.uid);
    old_peptide_id is the slot's current peptide (FE reads it from
    data.analytes) — passing it avoids a SENAITE slot round-trip and lets the
    orchestrator find the old per-substance vial rows.
    """
    new_peptide_id: int
    old_peptide_id: int
    senaite_uid: str
    # Strong confirm: authorizes retracting worked AND verified/promoted vial
    # results (the whole analyte is wrong). Published results still hard-block.
    force: bool = False


@app.post("/explorer/samples/{sample_id}/analytes/{slot}/replace")
async def replace_analyte(
    sample_id: str,
    slot: int,
    body: ReplaceAnalyteBody,
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Replace a wrong peptide-variant on slot {slot} of a parent blend.

    Order (gates first, then writes):
      1. offer-only gate — new peptide needs a full ID_/PUR_/QTY_ set (400)
      2. pre-write impact on the OLD peptide's vial rows — verified/published
         block (409); worked-unverified need confirm_retract (412 + impact)
      3. write Analyte{slot}Peptide = new identity title (SENAITE AR)
      4. reset the slot's COA display alias
      5. swap the parent AR identity service (remove ID_old, add ID_new) via
         the Integration-Service proxy — best-effort, no Mk1 cascade (the
         orchestrator is the sole authority on vial rows)
      6. re-mirror the slot across vials (replace_analyte_slot)
    """
    from sqlalchemy import select as _select
    from models import AnalysisService, LimsSample, Peptide, SampleAnalyteAlias
    from lims_analyses.service import (
        peptide_has_full_service_set,
        classify_slot_replacement_impact,
        presubsample_slot_blocked_keywords,
        replace_analyte_slot,
        BadRequestError as _BadRequestError,
        NotFoundError as _NotFoundError,
    )

    if slot < 1 or slot > 4:
        raise HTTPException(400, "slot must be between 1 and 4")

    new_pep = db.get(Peptide, body.new_peptide_id)
    if new_pep is None:
        raise HTTPException(404, f"peptide id={body.new_peptide_id} not found")

    # ── 1. offer-only gate ───────────────────────────────────────────────────
    if not peptide_has_full_service_set(db, peptide_id=body.new_peptide_id):
        raise HTTPException(
            400,
            f"{new_pep.name} has no full ID_/PUR_/QTY_ service set — set it up "
            "in Analysis Services first.",
        )

    new_id_svc = db.execute(
        _select(AnalysisService).where(
            AnalysisService.peptide_id == body.new_peptide_id,
            AnalysisService.keyword.like("ID%"),
        ).order_by(AnalysisService.keyword)
    ).scalars().first()
    old_id_kw = db.execute(
        _select(AnalysisService.keyword).where(
            AnalysisService.peptide_id == body.old_peptide_id,
            AnalysisService.keyword.like("ID%"),
        ).order_by(AnalysisService.keyword)
    ).scalars().first()
    if new_id_svc is None or not new_id_svc.title:
        raise HTTPException(400, f"{new_pep.name} has no Identity service title")

    # ── 2. pre-write impact gate ──────────────────────────────────────────────
    impact = classify_slot_replacement_impact(
        db, parent_sample_id=sample_id, old_peptide_id=body.old_peptide_id
    )
    # Published results can never be force-retracted here — invalidate via SENAITE.
    published = [b for b in impact["blocked"] if b.get("review_state") == "published"]
    if published:
        raise HTTPException(
            409,
            f"{len(published)} result(s) are on a published COA — invalidate or "
            "retest in SENAITE first.",
        )
    # Anything else that would be retracted (worked OR verified/promoted) needs
    # the strong confirm. 412 carries the impact so the FE shows the (escalated
    # when blocked rows are present) confirm modal, then re-posts with force=true.
    forceable = impact["worked_unverified"] or impact["blocked"]
    if forceable and not body.force:
        raise HTTPException(412, detail=impact)

    # ── 2b. pre-subsample SENAITE-results gate ────────────────────────────────
    # The vial-based impact gate above is blind to pre-subsample (pre-vial)
    # samples — they have no Mk1 LimsSample/vial rows, so their results live only
    # on the SENAITE AR. Guard the slot there before the destructive writes below
    # (step 5 removes the old identity service), so a replace can never strip a
    # worked identity or invalidate a verified/published result on an old sample.
    _is_presubsample = db.execute(
        _select(LimsSample.id).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none() is None
    if _is_presubsample:
        from lims_analyses.senaite_writeback import (
            list_parent_line_states,
            SenaiteWritebackError,
        )
        try:
            _senaite_states = list_parent_line_states(sample_id)
        except SenaiteWritebackError:
            # Fail closed: can't prove the slot is safe to replace → don't write.
            raise HTTPException(
                502,
                f"Could not read SENAITE results for {sample_id} to verify the "
                "slot is safe to replace — try again.",
            )
        _blocked_kw = presubsample_slot_blocked_keywords(
            _senaite_states, slot=slot, identity_keyword=old_id_kw
        )
        if _blocked_kw:
            raise HTTPException(
                409,
                f"Slot {slot} has verified/published result(s) in SENAITE "
                f"({', '.join(_blocked_kw)}) — invalidate or retest in SENAITE first.",
            )

    import logging as _logging
    _rep_logger = _logging.getLogger(__name__)

    # ── 3. write the slot's peptide (canonical source of truth) ───────────────
    field_result = await update_senaite_sample_fields(
        uid=body.senaite_uid,
        req=SenaiteFieldUpdateRequest(fields={f"Analyte{slot}Peptide": new_id_svc.title}),
        current_user=_current_user,
    )
    if not getattr(field_result, "success", False):
        raise HTTPException(
            502,
            f"SENAITE field write failed: {getattr(field_result, 'message', 'unknown')}",
        )

    # ── 4. reset the slot's COA display alias (was pegged to the old peptide) ──
    try:
        _alias = db.execute(
            _select(SampleAnalyteAlias).where(
                SampleAnalyteAlias.senaite_sample_id == sample_id,
                SampleAnalyteAlias.slot == slot,
            )
        ).scalar_one_or_none()
        if _alias is not None:
            db.delete(_alias)
            db.commit()
    except Exception as _alias_err:
        db.rollback()
        _rep_logger.warning("replace_analyte: alias reset failed for %s slot=%s: %s",
                            sample_id, slot, _alias_err)

    # ── 5. swap the parent AR identity service (direct IS proxy) ──────────────
    identity = {"removed": None, "added": None}
    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=30.0) as client:
        if old_id_kw:
            try:
                r = await client.delete(
                    f"{INTEGRATION_SERVICE_URL}/explorer/samples/{sample_id}/analyses/{old_id_kw}",
                    headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
                )
                r.raise_for_status()
                identity["removed"] = old_id_kw
            except Exception as _e:
                _rep_logger.warning("replace_analyte: remove %s failed: %s", old_id_kw, _e)
        if new_id_svc.senaite_uid:
            try:
                r = await client.post(
                    f"{INTEGRATION_SERVICE_URL}/explorer/samples/{sample_id}/analyses",
                    json={"service_uid": new_id_svc.senaite_uid},
                    headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
                )
                r.raise_for_status()
                identity["added"] = new_id_svc.keyword
            except Exception as _e:
                _rep_logger.warning("replace_analyte: add %s failed: %s", new_id_svc.keyword, _e)

    # ── 5b. parent analysis shadow mirror (best-effort) ───────────────────────
    # Mirror the OLD identity keyword as rejected and the NEW as unassigned —
    # gated per-op on the swap actually succeeding above (identity["removed"]
    # / identity["added"] are only set on a successful IS call each).
    from fastapi.concurrency import run_in_threadpool
    if identity["removed"]:
        await run_in_threadpool(
            _mirror_parent_analysis_bg,
            sample_id=sample_id, keyword=identity["removed"],
            mirror_review_state="rejected",
        )
    if identity["added"]:
        await run_in_threadpool(
            _mirror_parent_analysis_bg,
            sample_id=sample_id, keyword=identity["added"],
            mirror_review_state="unassigned",
        )

    # ── 6. re-mirror the slot across vials ────────────────────────────────────
    try:
        summary = replace_analyte_slot(
            db, parent_sample_id=sample_id, slot=slot,
            old_peptide_id=body.old_peptide_id, new_peptide_id=body.new_peptide_id,
            confirm_retract=body.force, force=body.force, user_id=_current_user.id,
        )
    except (_BadRequestError, _NotFoundError) as e:
        raise HTTPException(400, str(e))

    # ── 7. refresh the registry row (registry-owned analytes) ────────────────
    # lims_samples.analytes is the samples-list's authoritative analyte source
    # in Accu-Mk1 read mode, and this endpoint is the only Mk1-side mutation
    # of the slots. Re-read SENAITE truth (not in-memory state) so whatever
    # Replace actually landed is what the registry serves. Best-effort: a
    # failure never fails the replace — repair via the registry-debug refresh
    # or the backfill re-sweep. Commit the replace work FIRST so a refresh
    # error can't roll it back; run the sync SENAITE fetch in the threadpool
    # (this is an async-def handler — a blocking call would freeze the loop).
    db.commit()
    if not _is_presubsample:
        from starlette.concurrency import run_in_threadpool
        from sub_samples.service import _refresh_parent_from_senaite
        try:
            _row = db.execute(
                _select(LimsSample).where(LimsSample.sample_id == sample_id)
            ).scalar_one_or_none()
            if _row is not None:
                await run_in_threadpool(_refresh_parent_from_senaite, db, _row)
                db.commit()
        except Exception as _e:
            db.rollback()
            _rep_logger.warning(
                "replace_analyte: registry refresh failed for %s: %s", sample_id, _e
            )

    return {
        "success": True,
        "field_updated": f"Analyte{slot}Peptide",
        "new_peptide": new_pep.name,
        "identity": identity,
        **summary,
    }


@app.get("/peptides/with-service-set")
async def get_peptides_with_service_set(
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Peptide ids that have a complete per-substance HPLC service set
    (ID_/PUR_/QTY_). Drives the offer-only Replace picker — peptides not in
    this set are shown disabled ("set up in Analysis Services first")."""
    from models import AnalysisService

    rows = db.execute(
        select(AnalysisService.peptide_id, AnalysisService.keyword).where(
            AnalysisService.peptide_id.is_not(None)
        )
    ).all()
    by_pep: dict[int, set[str]] = {}
    for peptide_id, keyword in rows:
        if keyword and "_" in keyword:
            by_pep.setdefault(peptide_id, set()).add(keyword.split("_", 1)[0])
    eligible = [pid for pid, prefixes in by_pep.items()
                if {"ID", "PUR", "QTY"}.issubset(prefixes)]
    return {"peptide_ids": eligible}


# ── WooCommerce REST API Proxy ──────────────────────────────────────
# Fetches live order data directly from WooCommerce, including full
# financial breakdown (totals, discounts, coupons, shipping, tax).

WC_CONSUMER_KEY = os.environ.get("WC_CONSUMER_KEY", "")
WC_CONSUMER_SECRET = os.environ.get("WC_CONSUMER_SECRET", "")


def _get_wp_host() -> str:
    """Return WordPress host URL based on active environment."""
    env = os.environ.get("INTEGRATION_DB_ENV", "local").lower()
    if env == "production":
        return os.environ.get("WORDPRESS_PROD_HOST", "https://accumarklabs.com")
    return os.environ.get("WORDPRESS_LOCAL_HOST", "https://accumarklabs.local")


@app.get("/woo/orders/{order_id}")
async def get_woo_order(order_id: str, _current_user=Depends(get_current_user)):
    """Fetch a WooCommerce order directly from the WP REST API."""
    if not WC_CONSUMER_KEY or not WC_CONSUMER_SECRET:
        raise HTTPException(status_code=503, detail="WooCommerce API credentials not configured")
    url = f"{_get_wp_host()}/wp-json/wc/v3/orders/{order_id}"
    try:
        async with httpx.AsyncClient(timeout=15.0, verify=False) as client:
            resp = await client.get(url, auth=(WC_CONSUMER_KEY, WC_CONSUMER_SECRET))
            if resp.status_code == 404:
                raise HTTPException(status_code=404, detail=f"WooCommerce order {order_id} not found")
            resp.raise_for_status()
            return resp.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"WooCommerce unavailable: {e}")


class AdditionalCOAUpdateRequest(BaseModel):
    company_name: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    logo_url: Optional[str] = None
    chromatograph_background_url: Optional[str] = None


@app.patch("/explorer/additional-coas/{config_id}")
async def update_additional_coa_config(
    config_id: str,
    body: AdditionalCOAUpdateRequest,
    _current_user=Depends(get_current_user),
):
    """Update additional COA config branding (proxied to Integration Service)."""
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/additional-coas/{config_id}"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.patch(
                url,
                json=body.model_dump(exclude_unset=True),
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            resp.raise_for_status()
            return resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/coa-generations", response_model=list[ExplorerCOAGenerationResponse])
async def get_all_coa_generations(
    search: Optional[str] = None,
    status: Optional[str] = None,
    anchor_status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    _current_user=Depends(get_current_user),
):
    """Get all COA generations across all orders (proxied to Integration Service)."""
    try:
        params: dict[str, str] = {}
        if search:
            params["search"] = search
        if status:
            params["status"] = status
        if anchor_status:
            params["anchor_status"] = anchor_status
        params["limit"] = str(limit)
        params["offset"] = str(offset)
        url = f"{INTEGRATION_SERVICE_URL}/explorer/coa-generations"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.get(
                url,
                params=params,
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            resp.raise_for_status()
            return resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


class SignedURLRequest(BaseModel):
    sample_id: str
    version: int


class SignedURLResponse(BaseModel):
    url: str


@app.post("/explorer/signed-url/coa", response_model=SignedURLResponse)
async def get_coa_signed_url(
    body: SignedURLRequest,
    _current_user=Depends(get_current_user),
):
    """Get a presigned download URL for a COA PDF (proxied to Integration Service)."""
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/signed-url/coa"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.post(
                url,
                json=body.model_dump(),
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            resp.raise_for_status()
            return resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.post("/explorer/signed-url/chromatogram", response_model=SignedURLResponse)
async def get_chromatogram_signed_url(
    body: SignedURLRequest,
    _current_user=Depends(get_current_user),
):
    """Get a presigned download URL for a chromatogram image (proxied to Integration Service)."""
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/signed-url/chromatogram"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.post(
                url,
                json=body.model_dump(),
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
            resp.raise_for_status()
            return resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


@app.get("/explorer/chromatogram-lttb/{verification_code}/{resolution}")
async def get_chromatogram_lttb(
    verification_code: str,
    resolution: str,
    _current_user=Depends(get_current_user),
):
    """Proxy LTTB chromatogram JSON from Integration Service (avoids S3 CORS)."""
    if resolution not in ("5k", "10k"):
        raise HTTPException(status_code=400, detail="Resolution must be '5k' or '10k'")
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/chromatogram-lttb/{verification_code}/{resolution}"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=15.0) as client:
            resp = await client.get(url, headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY})
            resp.raise_for_status()
            return Response(
                content=resp.content,
                media_type="application/json",
                headers={"Cache-Control": "private, max-age=300"},
            )
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")


# ── COA Actions ────────────────────────────────────────────────────


class SampleCOAActionResponse(BaseModel):
    success: bool
    message: str
    verification_code: str | None = None
    warning: str | None = None


class AnalyteAliasSet(BaseModel):
    """Schema for setting a per-sample analyte display alias."""
    alias: str


class AnalyteAliasResponse(BaseModel):
    """Schema for a per-sample analyte display alias pick."""
    slot: int
    alias: str
    updated_at: datetime
    updated_by_email: Optional[str] = None

    class Config:
        from_attributes = True


def _load_sample_aliases(db: Session, sample_id: str) -> dict[int, str]:
    """Return {slot: alias} for a sample.  Used by COA-trigger endpoints to
    enrich the coabuilder request body with per-sample alias overrides."""
    rows = db.execute(
        select(SampleAnalyteAlias).where(
            SampleAnalyteAlias.senaite_sample_id == sample_id
        )
    ).scalars().all()
    return {r.slot: r.alias for r in rows}


def _build_coa_analyte_name_resolver(db: Session, sample_id: str, *, alias_map: dict[int, str]):
    """Build a keyword→display-name function for the COA block message.

    Gathers three name sources (no failure is fatal — each degrades):
      - Mk1 catalog titles (AnalysisService.keyword → title)
      - parent analyte-slot peptide names from SENAITE (best-effort)
      - per-sample customer aliases (passed in)

    Only invoked on the unresolved-block error path, so the one SENAITE
    slot read here is off the happy path.
    """
    from coa.block_summary import build_name_resolver
    from models import AnalysisService

    catalog_titles: dict[str, str] = {
        k: t for k, t in db.execute(
            select(AnalysisService.keyword, AnalysisService.title)
        ).all() if k and t
    }

    slot_names: dict[int, str] = {}
    try:
        from sub_samples import senaite as senaite_mod
        for slot, label in senaite_mod.fetch_parent_analyte_slots(sample_id).items():
            # Strip the "- Identity (HPLC)" service suffix → bare peptide name.
            stripped = re.sub(r"\s*-\s*[^-]+\([^)]+\)\s*$", "", str(label)).strip()
            slot_names[slot] = stripped or str(label)
    except Exception:
        pass

    return build_name_resolver(
        catalog_titles=catalog_titles, slot_names=slot_names, aliases=alias_map,
    )


@app.get(
    "/wizard/senaite/samples/{sample_id}/analyte-aliases",
    response_model=list[AnalyteAliasResponse],
)
async def get_sample_analyte_aliases(
    sample_id: str,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """List all per-slot COA display-alias picks for a SENAITE sample."""
    rows = db.execute(
        select(SampleAnalyteAlias)
        .where(SampleAnalyteAlias.senaite_sample_id == sample_id)
        .order_by(SampleAnalyteAlias.slot)
    ).scalars().all()
    return rows


@app.put(
    "/wizard/senaite/samples/{sample_id}/analyte-aliases/{slot}",
    response_model=AnalyteAliasResponse,
)
async def set_sample_analyte_alias(
    sample_id: str,
    slot: int,
    data: AnalyteAliasSet,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Set the COA display alias for one analyte slot (1-4) of a SENAITE sample.

    Alias is stored as a denormalized string so pruning a peptide's approved
    list later does not retroactively change this pick.  Caller (UI) is
    responsible for ensuring the alias is currently in the peptide's
    approved list; this endpoint does not validate against that list.
    """
    if slot < 1 or slot > 4:
        raise HTTPException(400, "slot must be between 1 and 4")
    if not data.alias or not data.alias.strip():
        raise HTTPException(400, "alias must be a non-empty string")

    alias = data.alias.strip()
    existing = db.execute(
        select(SampleAnalyteAlias).where(
            SampleAnalyteAlias.senaite_sample_id == sample_id,
            SampleAnalyteAlias.slot == slot,
        )
    ).scalar_one_or_none()

    if existing:
        existing.alias = alias
        existing.updated_by_user_id = current_user.id
        existing.updated_by_email = current_user.email
        row = existing
    else:
        row = SampleAnalyteAlias(
            senaite_sample_id=sample_id,
            slot=slot,
            alias=alias,
            updated_by_user_id=current_user.id,
            updated_by_email=current_user.email,
        )
        db.add(row)

    db.commit()
    db.refresh(row)
    return row


@app.delete(
    "/wizard/senaite/samples/{sample_id}/analyte-aliases/{slot}",
    status_code=204,
)
async def clear_sample_analyte_alias(
    sample_id: str,
    slot: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Clear the alias pick for one analyte slot.  Missing row is a no-op."""
    db.execute(
        delete(SampleAnalyteAlias).where(
            SampleAnalyteAlias.senaite_sample_id == sample_id,
            SampleAnalyteAlias.slot == slot,
        )
    )
    db.commit()


async def _parent_attachment_kinds(sample_id: str, auth) -> Optional[dict]:
    """Classify a sample's SENAITE attachments for the COA attachments gate.

    Returns {'has_image': bool, 'has_chromatogram': bool}, or None when the
    check can't be performed (SENAITE error / sample not found) — callers
    fail OPEN on None, matching the resolver pre-flight's error posture.

    Image = any attachment with an image/* content type. Chromatogram = any
    attachment typed "HPLC Graph" or with a .csv filename (same predicate the
    sample-details FE uses for its HPLC Graph cards).
    """
    import logging as _logging
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0), auth=auth, follow_redirects=True,
        ) as client:
            resp = await client.get(
                f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest",
                params={"getId": sample_id, "complete": "true"},
            )
            resp.raise_for_status()
            items = resp.json().get("items", [])
            if not items:
                return None
            refs = items[0].get("Attachment") or []
            has_image = False
            has_chromatogram = False
            for ref in refs:
                api_url = ref.get("api_url") if isinstance(ref, dict) else None
                if not api_url:
                    continue
                att_resp = await client.get(api_url)
                if att_resp.status_code >= 300:
                    continue
                att_data = att_resp.json()
                att_item = att_data["items"][0] if att_data.get("items") else att_data
                att_file = att_item.get("AttachmentFile") or {}
                filename = (att_file.get("filename") or "").lower()
                content_type = (att_file.get("content_type") or "").lower()
                att_type = att_item.get("AttachmentType") or att_item.get("getAttachmentType")
                if isinstance(att_type, dict):
                    att_type = att_type.get("title") or att_type.get("Title")
                if content_type.startswith("image/"):
                    has_image = True
                if att_type == "HPLC Graph" or filename.endswith(".csv"):
                    has_chromatogram = True
                if has_image and has_chromatogram:
                    break
            return {"has_image": has_image, "has_chromatogram": has_chromatogram}
    except Exception as e:
        _logging.getLogger(__name__).warning(
            "attachments-gate: could not classify attachments for %s: %s", sample_id, e,
        )
        return None


async def _maybe_emit_regular_coa_child(db, sample_id, parent_row, primary_data):
    """For a variance sample, generate the Regular parent-services COA as a child
    of the just-created variance primary: a SECOND COABuilder /process WITHOUT
    variance fields (so it renders the parent's promoted figures as a plain COA),
    tagged is_regular_coa. No-op for a non-variance sample (the primary already IS
    the regular COA). Best-effort — a failure must NOT fail the primary.
    """
    import logging as _logging
    _logger = _logging.getLogger(__name__)
    if not COA_BUILDER_URL or parent_row is None:
        return
    from coa.variance_series import build_variance_replicates
    if not build_variance_replicates(db, parent_row):
        return  # non-variance: the primary already IS the regular COA
    primary_gen_id = primary_data.get("generation_id")
    if not primary_gen_id:
        _logger.warning("regular COA child skipped for %s: no primary generation_id", sample_id)
        return
    body: dict = {"parent_generation_id": str(primary_gen_id), "is_regular_coa": True}
    alias_map = _load_sample_aliases(db, sample_id)
    if alias_map:
        body["analyte_display_names"] = {str(k): v for k, v in alias_map.items()}
    include_remarks = bool(parent_row.customer_remarks_include)
    body["include_lab_remarks"] = include_remarks
    if include_remarks and (parent_row.customer_remarks or "").strip():
        body["lab_remarks"] = parent_row.customer_remarks.strip()
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=120.0) as client:
            resp = await client.post(f"{COA_BUILDER_URL}/process/{sample_id}", json=body)
            resp.raise_for_status()
    except Exception as e:  # noqa: BLE001 — best-effort; must not fail the primary
        _logger.warning("regular COA child generation failed for %s: %s", sample_id, e)


@app.post("/wizard/senaite/samples/{sample_id}/generate-coa")
async def generate_sample_coa(
    sample_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Trigger Accumark COA generation for a SENAITE sample via COA Builder.

    Mirrors the SENAITE addon flow: call COA Builder, then immediately write
    the verification code back to the SENAITE sample.

    COA roll-up Phase 1: before invoking COABuilder, runs the source resolver
    over the parent + every linked sub-sample. The resolver:
      - auto-resolves analytes with a single reportable verified candidate
      - blocks (HTTP 422) when an analyte has >1 candidates with no actionable
        pin (or a stale pin / no candidates at all)
    After a successful generation, persists a per-generation manifest row
    per resolved analyte. Resolver runtime errors are non-fatal — they log
    and fall through so single-vial behavior is unchanged on a resolver bug.
    """
    import logging as _logging
    _logger = _logging.getLogger(__name__)

    if not COA_BUILDER_URL:
        return SampleCOAActionResponse(
            success=False,
            message="COA Builder not configured (COA_BUILDER_URL env var not set)",
        )

    # --- COA roll-up Phase 1: resolver pre-flight (parents only) ---
    # Sub-sample COAs aren't covered by the parent-level roll-up; their
    # generation path is independent and remains unchanged.
    from coa.source_resolver import resolve_sources, SenaiteAnalysesHttpReader
    from coa.manifest import write_generation_manifest

    is_sub = bool(re.search(r"-S\d{2,}$", sample_id))
    resolver_result = None
    if not is_sub:
        # ── COA pre-flight gates (parents only) ──────────────────────────────
        # Evaluate EVERY gate and accumulate the blockers, then raise ONCE — the
        # lab sees all blockers up front instead of one-at-a-time (clearing the
        # attachment gate only to then hit "variance not locked" was whack-a-mole).
        # Each gate keeps its own fail-open/fail-soft posture: a gate that can't
        # be evaluated simply contributes no blocker.
        _unresolved = None
        _missing_attach = None
        _variance_required = False

        if SENAITE_URL:
            try:
                reader = SenaiteAnalysesHttpReader(
                    base_url=SENAITE_URL,
                    auth=_get_senaite_auth(current_user),
                )
                resolver_result = await resolve_sources(sample_id, db, reader)
            except Exception as e:
                # Resolver failure is non-fatal in Phase 1; log and fall through.
                _logger.warning("COA resolver pre-flight failed for %s: %s", sample_id, e)
                resolver_result = None

            if resolver_result is not None:
                # Micro analytes (ENDO/STER/KF) never block: the lab finishes them
                # after the analytical COA and re-generates. Only NON-micro
                # unresolved analytes hold up generation.
                from coa.block_summary import (
                    build_name_resolver,
                    has_blocking_unresolved,
                    summarize_unresolved,
                )
                from lims_analyses.seeder import _micro_group_keywords

                micro_kw = _micro_group_keywords(db)
                if has_blocking_unresolved(resolver_result, micro_keywords=micro_kw):
                    name_for = _build_coa_analyte_name_resolver(db, sample_id, alias_map=_load_sample_aliases(db, sample_id))
                    _unresolved = summarize_unresolved(
                        resolver_result, micro_keywords=micro_kw, name_for=name_for,
                    )

            # --- Attachments gate: a parent COA requires a sample image on the
            # AR, and — when the sample carries analytical (non-micro) analytes —
            # a chromatogram. Both are attached from the sample page pickers
            # ("Select Vial Image" / "Select Vial Chromatogram") or the legacy
            # check-in/auto-fill flows. Fail-OPEN when SENAITE can't be read
            # (same posture as the resolver pre-flight: a flaky check must not
            # permanently block generation; COABuilder fails loudly anyway if
            # SENAITE is truly down).
            kinds = await _parent_attachment_kinds(sample_id, _get_senaite_auth(current_user))
            if kinds is not None:
                from lims_analyses.seeder import _micro_group_keywords as _micro_kws
                if resolver_result is not None:
                    _micro = _micro_kws(db)
                    needs_chromatogram = any(
                        d.analyte_keyword not in _micro for d in resolver_result.decisions
                    )
                else:
                    # Resolver unavailable — derive from the AR's active keywords;
                    # fail open (no chromatogram requirement) if that read fails too.
                    try:
                        from sub_samples.senaite import fetch_parent_analysis_keywords
                        _micro = _micro_kws(db)
                        needs_chromatogram = any(
                            k not in _micro
                            for k in fetch_parent_analysis_keywords(sample_id)
                        )
                    except Exception:
                        needs_chromatogram = False
                gate_missing = []
                if not kinds["has_image"]:
                    gate_missing.append({
                        "kind": "sample_image",
                        "message": "Sample image — attach one from the sample page "
                                   "(Select Vial Image, or upload a Sample Image attachment).",
                    })
                if needs_chromatogram and not kinds["has_chromatogram"]:
                    gate_missing.append({
                        "kind": "chromatogram",
                        "message": "Chromatogram — attach one from the sample page "
                                   "(Select Vial Chromatogram, after processing the HPLC prep).",
                    })
                if gate_missing:
                    _missing_attach = gate_missing

        # Variance-lock gate: a variance-purchased lot must have its variance set
        # LOCKED before generation, so the COA never certifies an incomplete
        # series (lock_variance_set enforces every in-set vial is signed off).
        # Applies to peptide + BW. No SENAITE_URL dependency. Fail-soft: skip if
        # services are unavailable.
        _pre_parent = db.execute(
            select(LimsSample).where(LimsSample.sample_id == sample_id)
        ).scalar_one_or_none()
        if _pre_parent is not None:
            try:
                from sub_samples.service import fetch_sample_services, variance_lock_required
                _vsvc = fetch_sample_services(sample_id)
                _variance_required = variance_lock_required(
                    (_vsvc or {}).get("services") or {}, _pre_parent.variance_locked_at
                )
            except HTTPException:
                raise
            except Exception:
                _variance_required = False

        from coa.preflight import collect_preflight_blockers, build_preflight_error
        _blockers = collect_preflight_blockers(
            unresolved=_unresolved,
            missing_attachments=_missing_attach,
            variance_locked_required=_variance_required,
        )
        if _blockers:
            raise HTTPException(status_code=422, detail=build_preflight_error(_blockers))

    # Enrich with per-sample analyte display alias picks so the COA renders
    # the customer-facing name (real name still drives conformance).
    alias_body: dict = {}
    alias_map = _load_sample_aliases(db, sample_id)
    if alias_map:
        alias_body["analyte_display_names"] = {str(k): v for k, v in alias_map.items()}

    # Variance replicate series (parent's assignment_kind='variance' vials).
    # Raw per-vial values; COABuilder prepends its own parent figure and renders
    # the comma-delimited series. Best-effort — a builder error must not block
    # generation. Parents only (sub-sample COAs have no variance children).
    # Tracks whether this generation delivered customer remarks — drives the
    # "Delivered on" timestamp stamped after a successful generation.
    # (The variance-lock gate above already blocked an unlocked variance lot.)
    _remarks_included = False
    if not is_sub:
        _parent_row = db.execute(
            select(LimsSample).where(LimsSample.sample_id == sample_id)
        ).scalar_one_or_none()
        if _parent_row is not None:
            # Customer-remarks snapshot + "Include with Publish?" gating. When
            # the flag is False we omit lab_remarks and tell COABuilder the
            # suppression was intentional so its non-conforming gate is skipped.
            # include_lab_remarks is ALWAYS sent. Set OUTSIDE the variance try
            # below so a variance-build error can't drop it.
            _include_remarks = bool(_parent_row.customer_remarks_include)
            alias_body["include_lab_remarks"] = _include_remarks
            _remarks_text = (_parent_row.customer_remarks or "").strip()
            if _include_remarks and _remarks_text:
                alias_body["lab_remarks"] = _remarks_text
                _remarks_included = True
            # Variance series (peptide replicates + BW/generic analyte series) —
            # best-effort; a builder error must not block generation. Shared
            # helper so regen-primary sends the identical series (parity).
            from coa.variance_series import process_variance_fields
            alias_body.update(process_variance_fields(db, _parent_row))

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=120.0) as client:
            resp = await client.post(
                f"{COA_BUILDER_URL}/process/{sample_id}",
                json=alias_body if alias_body else None,
            )
            resp.raise_for_status()
            data = resp.json()
    except httpx.TimeoutException:
        return SampleCOAActionResponse(success=False, message="COA Builder request timed out (PDF generation can take up to 2 minutes)")
    except httpx.HTTPStatusError as e:
        try:
            detail = e.response.json().get("detail", str(e.response.status_code))
        except Exception:
            detail = str(e.response.status_code)
        return SampleCOAActionResponse(success=False, message=f"COA Builder error: {detail}")
    except Exception as e:
        return SampleCOAActionResponse(success=False, message=f"COA generation failed: {e}")

    verification_code: str | None = data.get("verification_code")
    generation_number: int | None = data.get("generation_number")
    pdf_base64: str | None = data.get("pdf_base64")

    # Attach PDF + verification code to SENAITE via the custom addon endpoint.
    # This mirrors the full COAGeneratorView flow: saves VerificationCode field
    # and creates an ARReport child object so the PDF appears in SENAITE's
    # Reports tab.  Best-effort — generation already succeeded at this point.
    if SENAITE_URL and pdf_base64:
        try:
            attach_payload = {
                "sample_id": sample_id,
                "pdf_base64": pdf_base64,
                "verification_code": verification_code or "",
            }
            attach_url = f"{SENAITE_URL}/senaite/@@accumark-attach-coa"
            # Try the user's own SENAITE creds first (audit attribution); a
            # stale stored password makes SENAITE treat the call as anonymous
            # (404/401 without raising), so check status and retry once with
            # the service account.
            for attach_auth in (_get_senaite_auth(current_user), httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD)):
                async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                    timeout=httpx.Timeout(30.0, connect=5.0),
                    auth=attach_auth,
                    follow_redirects=True,
                ) as senaite_client:
                    resp = await senaite_client.post(attach_url, json=attach_payload)
                if resp.status_code < 300:
                    break
                _logger.warning(
                    "SENAITE COA attach HTTP %s for %s (auth=%s)",
                    resp.status_code, sample_id,
                    "user" if attach_auth is not None else "service",
                )
        except Exception as e:
            # Non-fatal — COA is generated; SENAITE attach is best-effort.
            _logger.warning("SENAITE COA attach failed for %s: %s", sample_id, e)

    # Variance sample: also emit the Regular parent-services COA as a child of the
    # just-created primary (best-effort; the helper no-ops for non-variance).
    if not is_sub:
        await _maybe_emit_regular_coa_child(db, sample_id, _parent_row, data)

    # Build a meaningful message from the COA Builder response
    warnings = data.get("warnings", [])
    if verification_code and generation_number:
        message = f"COA generation #{generation_number} complete — code: {verification_code}"
    elif verification_code:
        message = f"COA generated — code: {verification_code}"
    else:
        message = "COA generated but verification code failed — Integration Service error"

    # Surface partial failures: PDF was generated but S3/verification failed
    if not verification_code:
        warn_detail = "; ".join(warnings) if warnings else "Integration Service may be unavailable"
        return SampleCOAActionResponse(
            success=False,
            message=f"S3 upload failed: {warn_detail}",
        )

    if warnings:
        message += f" (warnings: {'; '.join(warnings)})"

    # NOTE (1.0.1): "Delivered to Customer" is stamped at PUBLISH time
    # (publish_sample_coa), NOT here. The lab writes the customer remark after
    # reviewing the generated COA, so generation no longer represents delivery —
    # the remark is captured and delivered when the COA is published.

    # --- COA roll-up Phase 1: write the per-generation manifest ---
    # Best-effort. COABuilder's response shape may or may not include the
    # integration-DB generation UUID; fall back to a generated UUID + log
    # warning so the manifest is still queryable by (parent, generation_number).
    # No is_blocked gate here: if we reached this point we passed the
    # non-micro block check, and write_generation_manifest already skips any
    # residual blocked decision (e.g. an unfinished micro analyte) defensively.
    if (
        resolver_result is not None
        and verification_code
        and generation_number
    ):
        import uuid as _uuid
        gen_id_str = data.get("generation_id")
        try:
            gen_id = _uuid.UUID(gen_id_str) if gen_id_str else _uuid.uuid4()
        except (TypeError, ValueError):
            _logger.warning(
                "COA generation_id from COABuilder is not a valid UUID (%r); "
                "manifest will be keyed by a generated UUID",
                gen_id_str,
            )
            gen_id = _uuid.uuid4()
        try:
            write_generation_manifest(
                db,
                generation_id=gen_id,
                generation_number=generation_number,
                result=resolver_result,
            )
        except Exception as e:
            _logger.warning(
                "COA manifest write failed for %s gen %s: %s",
                sample_id, generation_number, e,
            )

    return SampleCOAActionResponse(
        success=True,
        message=message,
        verification_code=verification_code,
    )


class GenerateVialCOAsRequest(BaseModel):
    # No parent_generation_id: the server DERIVES the sample's primary COA
    # itself (never trusts a client-supplied generation id). Only optional
    # remarks overrides are accepted.
    lab_remarks: Optional[str] = None
    include_lab_remarks: Optional[bool] = None


@app.post("/wizard/senaite/samples/{sample_id}/generate-vial-coas")
async def generate_vial_coas(
    sample_id: str,
    body: Optional[GenerateVialCOAsRequest] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Generate one per-vial HPLC COA for every reportable HPLC vial of a parent.

    The parent generation the vial COAs attach to is DERIVED server-side (the
    sample's current primary, non-superseded generation) — never supplied by the
    client, so a stale UI or a forged id can't graft vial COAs onto the wrong
    lineage. Idempotent: a vial that already has a live child COA is skipped, so
    re-running fills gaps instead of duplicating. Completeness-aware: success is
    true only when no vial failed.
    """
    import logging as _logging
    _logger = _logging.getLogger(__name__)

    def _resp(success, message, *, expected=0, generated=None, skipped=None, errors=None):
        return {
            "success": success, "message": message, "expected": expected,
            "generated": generated or [], "skipped": skipped or [], "errors": errors or [],
        }

    if not COA_BUILDER_URL:
        return _resp(False, "COA Builder not configured")
    if re.search(r"-S\d{2,}$", sample_id):
        return _resp(False, "Per-vial COAs are generated from the parent sample, not a sub-sample.")

    parent = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()
    if parent is None:
        return _resp(False, f"Parent sample {sample_id} not found in LIMS")

    # Derive + validate the parent COA generation server-side (do not trust the
    # client). fetch_primary_generation already scopes to this sample_id, a NULL
    # parent_generation_id (primary), and excludes superseded.
    from integration_db import fetch_primary_generation, fetch_existing_vial_sequences
    try:
        primary = fetch_primary_generation(sample_id)
    except Exception as e:  # noqa: BLE001
        _logger.warning("fetch_primary_generation failed for %s: %s", sample_id, e)
        return _resp(False, "Could not look up the parent COA (Integration Service unavailable).")
    if not primary:
        return _resp(False, "Generate the parent COA first — per-vial COAs attach to it.")
    parent_generation_id = str(primary["id"])

    from coa.variance_series import list_hplc_vials_with_figures
    vials = list_hplc_vials_with_figures(db, parent)
    if not vials:
        return _resp(False, "No HPLC vials with reportable results to generate COAs for.")

    # Idempotency: skip vials that already have a live child under this parent
    # (one child per parent_generation_id + vial_sequence).
    try:
        existing = fetch_existing_vial_sequences(parent_generation_id)
    except Exception as e:  # noqa: BLE001
        _logger.warning("fetch_existing_vial_sequences failed for %s: %s", sample_id, e)
        existing = set()

    # Customer-remarks snapshot — same source the parent COA uses, so a
    # non-conforming vial COA carries remarks (COABuilder 422s otherwise).
    include_remarks = body.include_lab_remarks if body else None
    lab_remarks = (body.lab_remarks or "").strip() if (body and body.lab_remarks) else ""
    if include_remarks is None:
        include_remarks = bool(parent.customer_remarks_include)
        if include_remarks and not lab_remarks:
            lab_remarks = (parent.customer_remarks or "").strip()

    generated: list[dict] = []
    skipped: list[int] = []
    errors: list[dict] = []
    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=120.0) as client:
        for vial_seq, figs in vials:
            if vial_seq in existing:
                skipped.append(vial_seq)
                continue
            vbody: dict = {
                "vial_figures": figs,
                "parent_generation_id": parent_generation_id,
                "vial_sequence": vial_seq,
                "include_lab_remarks": bool(include_remarks),
            }
            if include_remarks and lab_remarks:
                vbody["lab_remarks"] = lab_remarks
            try:
                resp = await client.post(f"{COA_BUILDER_URL}/process/{sample_id}", json=vbody)
                resp.raise_for_status()
                data = resp.json()
                generated.append({
                    "vial_sequence": vial_seq,
                    "verification_code": data.get("verification_code"),
                    "generation_id": data.get("generation_id"),
                })
            except httpx.HTTPStatusError as e:
                try:
                    detail = e.response.json().get("detail", str(e.response.status_code))
                except Exception:
                    detail = str(e.response.status_code)
                errors.append({"vial_sequence": vial_seq, "error": detail})
                _logger.warning("vial COA gen failed for %s vial %s: %s", sample_id, vial_seq, detail)
            except Exception as e:  # noqa: BLE001 — one vial failing must not abort the rest
                errors.append({"vial_sequence": vial_seq, "error": str(e)})
                _logger.warning("vial COA gen error for %s vial %s: %s", sample_id, vial_seq, e)

    g, s, f = len(generated), len(skipped), len(errors)
    parts = [f"Generated {g} per-vial COA(s)"]
    if s:
        parts.append(f"{s} already existed")
    if f:
        parts.append(f"{f} failed")
    # Success only when nothing failed (a fully-skipped re-run is still success).
    return {
        "success": f == 0, "message": "; ".join(parts), "expected": len(vials),
        "generated": generated, "skipped": skipped, "errors": errors,
    }


@app.post("/wizard/senaite/samples/{sample_id}/publish-coa")
async def publish_sample_coa(
    sample_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Publish the latest draft Accumark COA for a SENAITE sample.

    Order of operations:
    1. Reject sub-sample IDs (they inherit parent's order number, would clobber parent's COA on WP)
    2. Resolve SENAITE UID (fail fast before any state changes)
    3. Publish in Integration Service (marks generation published, publishes additional COAs)
    4. Write verification code to SENAITE
    5. Transition SENAITE sample to published workflow state
    """
    # 1. Sub-sample suffix matches `-S\d{2}` (canonical from sub_samples/senaite.py:89).
    if re.search(r"-S\d{2}$", sample_id):
        raise HTTPException(
            status_code=403,
            detail=(
                "Sub-sample COAs cannot be published to WordPress. "
                "Publish the parent sample's COA instead."
            ),
        )

    # 2. Resolve SENAITE UID upfront so we fail before touching integration service state
    senaite_uid: str | None = None
    if SENAITE_URL:
        try:
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=httpx.Timeout(15.0, connect=5.0),
                auth=_get_senaite_auth(current_user),
                follow_redirects=True,
            ) as client:
                search_url = (
                    f"{SENAITE_URL}/senaite/@@API/senaite/v1/search"
                    f"?portal_type=AnalysisRequest&getId={sample_id}&complete=true"
                )
                search_resp = await client.get(search_url)
                search_resp.raise_for_status()
                items = search_resp.json().get("items", [])
                if items:
                    senaite_uid = items[0].get("uid")
        except Exception as e:
            raise HTTPException(status_code=503, detail=f"SENAITE unreachable: {e}")

        if not senaite_uid:
            raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found in SENAITE")

    # Publish-time customer remark (1.0.1): the lab writes it AFTER reviewing the
    # generated COA, so capture it HERE and send to IS, which forwards it to the
    # WP COA email + order-page Lab Remarks button. publish-coa is parent-only
    # (sub-samples are rejected above), so the parent row carries the remark.
    _parent_row = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()
    # Snapshot the pre-publish status NOW, before any commit below can expire
    # `_parent_row` (expire_on_commit=True). The transition-log hook at the
    # bottom of this function passes `from_status` as a run_in_threadpool
    # kwarg, which is evaluated on the event loop at call time — touching an
    # expired ORM attribute there triggers a synchronous refresh SELECT on
    # the event loop that can raise and escape into the generic except below,
    # producing a false 502 AFTER the publish already succeeded.
    _pre_publish_status = _parent_row.status if _parent_row is not None else None
    _publish_body: dict = {}
    _deliver_remark = False
    if _parent_row is not None:
        _inc = bool(_parent_row.customer_remarks_include)
        _txt = (_parent_row.customer_remarks or "").strip()
        _publish_body["include_lab_remarks"] = _inc
        if _inc and _txt:
            _publish_body["lab_remarks"] = _txt
            _deliver_remark = True

    # 2. Publish in Integration Service (also publishes additional COAs)
    try:
        url = f"{INTEGRATION_SERVICE_URL}/explorer/samples/{sample_id}/publish-coa"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=30.0) as client:
            resp = await client.post(
                url,
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
                json=_publish_body or None,
            )
            resp.raise_for_status()
            data = resp.json()
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Integration Service unavailable: {e}")

    if not data.get("success"):
        return SampleCOAActionResponse(
            success=False, message=data.get("message", "Publish failed")
        )

    verification_code: str | None = data.get("verification_code")
    warning: str | None = None

    # Stamp "Delivered to Customer" now that the COA — carrying the current
    # remark — is published to WordPress. Only when a remark was actually
    # delivered (include on + non-empty). Best-effort; never fails the publish.
    if _deliver_remark and _parent_row is not None:
        try:
            from datetime import datetime as _dt
            _parent_row.customer_remarks_delivered_at = _dt.utcnow()
            db.commit()
        except Exception:
            db.rollback()
            _logger.warning("delivered_at stamp failed for %s", sample_id, exc_info=True)

    # 3 & 4. Write verification code and transition SENAITE workflow — guaranteed
    if senaite_uid:
        try:
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=httpx.Timeout(30.0, connect=5.0),
                auth=_get_senaite_auth(current_user),
                follow_redirects=True,
            ) as client:
                if verification_code:
                    code_resp = await client.post(
                        f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{senaite_uid}",
                        json={"VerificationCode": verification_code},
                    )
                    code_resp.raise_for_status()

                    # Dual-write mirror (registry slice 1): reflect the
                    # freshly-written verification code onto the local
                    # registry row. Best-effort — a mirror problem must
                    # never fail the publish.
                    try:
                        from sub_samples.service import apply_senaite_fields_to_row
                        if apply_senaite_fields_to_row(
                            db, senaite_uid, {"VerificationCode": verification_code}
                        ):
                            db.commit()
                    except Exception as mirror_err:
                        try:
                            db.rollback()
                        except Exception:
                            pass
                        logger.warning(
                            "registry.field_mirror_failed uid=%s err=%s",
                            senaite_uid, mirror_err,
                        )

                transition_resp = await client.post(
                    f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{senaite_uid}",
                    json={"transition": "publish"},
                )
                transition_resp.raise_for_status()

                # SENAITE returns 200 OK even when it silently rejects a
                # transition (e.g. sample still in `to_be_verified`, not
                # `verified`), AND returns 200 OK with an empty `items`
                # array when the transition is a no-op because the sample
                # is already in the target state (e.g. regen of an
                # already-published sample).  Verify by re-reading the AR
                # instead of relying on the transition response.
                #
                # Accepted terminal states:
                #   - published      : fully published, normal success
                #   - to_be_verified : lab-partial-publish flow.  Some
                #     analyses (e.g. sterility) take days longer than
                #     others; we issue a partial COA with current results
                #     and re-publish when final results come in.  The
                #     VerificationCode is already written to SENAITE above,
                #     and IS already marked the generation published, so
                #     the client-facing COA is live.  SENAITE's workflow
                #     label stays `to_be_verified` to correctly reflect
                #     that tests are still pending.
                #   - waiting_for_addon_results : same partial-publish
                #     flow as above, but for samples with add-on analyses
                #     (sterility, endotoxin) that run on a different
                #     timeline.  SENAITE's addon plugin parks the AR in
                #     this state until addon results return; the publish
                #     transition is accepted and the COA is live.
                items = transition_resp.json().get("items", [])
                actual_state = items[0].get("review_state", "") if items else ""
                accepted_states = {
                    "published",
                    "to_be_verified",
                    "waiting_for_addon_results",
                }
                # Pre-publish review states — the COA is live in our system
                # (IS marked it published, verification code is in SENAITE),
                # but SENAITE's workflow hasn't been advanced because the lab
                # tech hasn't run a verify transition. We surface this as a
                # warning rather than a hard error: the customer-facing COA
                # is published, but ops should advance the SENAITE state.
                pre_publish_states = {"ready_for_initial_review"}
                if actual_state not in accepted_states:
                    verify_resp = await client.get(
                        f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
                        f"?id={sample_id}&complete=true"
                    )
                    if verify_resp.status_code == 200:
                        verify_items = verify_resp.json().get("items", [])
                        if verify_items:
                            actual_state = verify_items[0].get("review_state", "")
                    if actual_state in pre_publish_states:
                        warning = (
                            f"Warning: Sample should not typically be published "
                            f"from the '{actual_state}' state. The COA is live "
                            f"but SENAITE's workflow remains at "
                            f"'{actual_state}' — verify the sample in SENAITE "
                            f"to advance it."
                        )
                    elif actual_state not in accepted_states:
                        raise HTTPException(
                            status_code=502,
                            detail=(
                                f"COA published in system but SENAITE silently "
                                f"rejected the 'publish' transition for {sample_id} "
                                f"(state is '{actual_state or 'unknown'}'). "
                                f"Verify the sample in SENAITE, then retry."
                            ),
                        )

                # ── Parent shadow mirror (A6 publish, best-effort) ────────────
                # mirror_review_state records SENAITE-side truth, so this is
                # gated on the AR having ACTUALLY reached 'published' — after
                # the reconciliation above (so the silent-rejection 502 path
                # can never leave a durable 'published' mirror behind), and
                # deliberately NOT on the broader accepted/pre-publish sets:
                # to_be_verified / waiting_for_addon_results mean the
                # customer-facing COA is live but SENAITE deferred the
                # workflow publish (partial-publish / addon flow) — the
                # SENAITE analyses are NOT published there, and mirroring
                # 'published' would be false. When publish lands later,
                # per-line A2/A3 hooks and/or the next publish call record
                # it. Own session via _mark_shadows_published_bg — never the
                # request `db`.
                if actual_state == "published":
                    from fastapi.concurrency import run_in_threadpool
                    await run_in_threadpool(
                        _mark_shadows_published_bg, sample_id=sample_id
                    )
                    # Task 3: native sample-transition log (own session,
                    # never-fail — see _record_sample_transition_bg).
                    await run_in_threadpool(
                        _record_sample_transition_bg,
                        sample_id=sample_id, verb="publish", to_status="published",
                        from_status=_pre_publish_status,
                        source="mk1",
                        actor_user_id=getattr(current_user, "id", None),
                    )
        except HTTPException:
            raise
        except Exception as e:
            # COA is published in our system — surface SENAITE failure clearly
            raise HTTPException(
                status_code=502,
                detail=f"COA published in system but SENAITE transition failed: {e}",
            )

    return SampleCOAActionResponse(
        success=True,
        message=data.get("message", "COA published"),
        verification_code=verification_code,
        warning=warning,
    )


@app.post("/wizard/senaite/samples/{sample_id}/regen-primary-coa")
async def regen_primary_coa(
    sample_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Regenerate ONLY the primary COA for a sample and republish it.

    Used for ops corrections where a sample's primary branding is wrong
    (wrong client on the mother cert) but its additional COAs have already
    been distributed with their own verification codes.  Calls COA Builder
    with skip_additional_coas=true so existing additional COAs keep their
    codes untouched, then publishes the new primary.
    """
    if not COA_BUILDER_URL:
        return SampleCOAActionResponse(
            success=False,
            message="COA Builder not configured",
        )

    alias_body: dict = {}
    alias_map = _load_sample_aliases(db, sample_id)
    if alias_map:
        alias_body["analyte_display_names"] = {str(k): v for k, v in alias_map.items()}

    # Variance series — MUST mirror generate_sample_coa, else regenerating strips
    # the variance series off the certified COA (the parent figure renders alone).
    # Same shared helper both paths use so they can't drift.
    _regen_parent = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()
    if _regen_parent is not None:
        from coa.variance_series import process_variance_fields
        alias_body.update(process_variance_fields(db, _regen_parent))

    # 1. Regenerate only the primary COA via COA Builder
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=120.0) as client:
            resp = await client.post(
                f"{COA_BUILDER_URL}/process/{sample_id}",
                params={"skip_additional_coas": "true"},
                json=alias_body if alias_body else None,
            )
            resp.raise_for_status()
            data = resp.json()
    except httpx.TimeoutException:
        return SampleCOAActionResponse(
            success=False,
            message="COA Builder timed out (PDF generation can take up to 2 minutes)",
        )
    except httpx.HTTPStatusError as e:
        try:
            detail = e.response.json().get("detail", str(e.response.status_code))
        except Exception:
            detail = str(e.response.status_code)
        return SampleCOAActionResponse(success=False, message=f"COA Builder error: {detail}")
    except Exception as e:
        return SampleCOAActionResponse(success=False, message=f"COA regeneration failed: {e}")

    verification_code: str | None = data.get("verification_code")
    pdf_base64: str | None = data.get("pdf_base64")

    if not verification_code:
        return SampleCOAActionResponse(
            success=False,
            message="Primary regenerated but no verification code returned",
        )

    # 2. Attach new PDF to SENAITE (best-effort — the generation already has a PDF in S3)
    if SENAITE_URL and pdf_base64:
        try:
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=httpx.Timeout(30.0, connect=5.0),
                auth=_get_senaite_auth(current_user),
                follow_redirects=True,
            ) as senaite_client:
                await senaite_client.post(
                    f"{SENAITE_URL}/senaite/@@accumark-attach-coa",
                    json={
                        "sample_id": sample_id,
                        "pdf_base64": pdf_base64,
                        "verification_code": verification_code,
                    },
                )
        except Exception as e:
            # Non-fatal — COA is in S3 already; SENAITE attach is best-effort.
            # The @@accumark-attach-coa addon is prod-only, so this 404s on dev.
            import logging as _logging
            _logging.getLogger(__name__).warning(
                "SENAITE COA attach failed for %s: %s", sample_id, e
            )

    # Mirror generate_sample_coa: refresh the Regular parent-services COA child
    # off the NEW primary BEFORE publishing, so the cascade publishes it. The
    # cascade's cross-primary supersede retires the old primary's regular child.
    # Best-effort; no-op for non-variance.
    await _maybe_emit_regular_coa_child(db, sample_id, _regen_parent, data)

    # 3. Publish the new primary — reuses publish_sample_coa's flow so
    # integration service marks the new generation as published,
    # supersedes the old primary, and _publish_additional_coas sees no
    # draft children (existing additionals keep their codes).
    return await publish_sample_coa(sample_id=sample_id, current_user=current_user, db=db)


@app.post("/wizard/senaite/additional-coas/{config_id}/regen-coa")
async def regen_additional_coa(
    config_id: str,
    current_user=Depends(get_current_user),
):
    """Regenerate ONLY one additional COA and republish it.

    Proxies to the integration service which orchestrates the single-
    additional regen against the existing published primary.  Produces
    a new verification code for this one additional COA; the primary
    and other additional COAs are untouched.
    """
    url = f"{INTEGRATION_SERVICE_URL}/explorer/additional-coas/{config_id}/regenerate"
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, timeout=120.0) as client:
            resp = await client.post(
                url,
                headers={"X-API-Key": INTEGRATION_SERVICE_API_KEY},
            )
    except Exception as e:
        raise HTTPException(
            status_code=503, detail=f"Integration Service unavailable: {e}"
        )

    if resp.status_code != 200:
        try:
            detail = resp.json().get("detail", resp.text[:200])
        except Exception:
            detail = resp.text[:200]
        return SampleCOAActionResponse(
            success=False,
            message=f"Regen failed ({resp.status_code}): {detail}",
        )

    data = resp.json()
    return SampleCOAActionResponse(
        success=bool(data.get("success")),
        message=data.get("message", "Additional COA regenerated"),
        verification_code=data.get("new_verification_code"),
    )



# ── SharePoint Integration ─────────────────────────────────────────

import sharepoint as sp


@app.get("/sharepoint/status")
async def sharepoint_status(_current_user=Depends(get_current_user)):
    """Test SharePoint connection and return site info."""
    return await sp.verify_connection()


@app.get("/sharepoint/browse")
async def sharepoint_browse(
    path: str = "",
    root: str = "lims",
    _current_user=Depends(get_current_user),
):
    """
    Browse folders/files in SharePoint.

    Args:
        path: Relative path within the root (empty = root itself)
        root: Which root — 'lims' (LIMS CSVs) or 'peptides'
    """
    try:
        if root == "lims":
            items, truncated = await sp.list_lims_folder(path, with_truncation=True)
        else:
            items, truncated = await sp.list_folder(path, with_truncation=True)
        # `truncated` is True when the folder was too large to list in full
        # (bounded crawl). The client can surface a "showing first N — open a
        # subfolder" hint rather than treating the listing as complete.
        return {"path": path, "root": root, "items": items, "truncated": truncated}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint error: {e}")


@app.get("/sharepoint/sample/{sample_id}/files")
async def sharepoint_sample_files(
    sample_id: str,
    _current_user=Depends(get_current_user),
):
    """
    Find a sample folder and list all CSV/Excel files within it.
    Searches the Peptides root for a sample ID match.
    """
    try:
        result = await sp.get_sample_files(sample_id)
        if not result:
            raise HTTPException(status_code=404, detail=f"Sample '{sample_id}' not found")
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint error: {e}")


@app.get("/sharepoint/download/{item_id}")
async def sharepoint_download(
    item_id: str,
    _current_user=Depends(get_current_user),
):
    """
    Download a file from SharePoint by its item ID.
    Returns the raw file content.
    """
    from fastapi.responses import Response

    try:
        content, filename = await sp.download_file(item_id)

        # Determine content type
        if filename.lower().endswith(".csv"):
            media_type = "text/csv"
        elif filename.lower().endswith(".xlsx"):
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            media_type = "application/octet-stream"

        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint download error: {e}")


@app.get("/sharepoint/folder-by-id/{folder_id}/chrom-files")
async def sharepoint_folder_chrom_files(
    folder_id: str,
    _current_user=Depends(get_current_user),
):
    """
    List chromatogram CSV files (dx_dad1a) in a SharePoint folder given its item ID.
    Used by the HPLC flyout to self-discover chrom files from the scan match folder
    when the original scan didn't include them.
    """
    try:
        items = await sp.list_folder_by_id(folder_id)
        chrom = [
            {"id": it["id"], "name": it["name"], "type": "file", "size": it.get("size", 0)}
            for it in items
            if it["type"] == "file"
            and it["name"].lower().endswith(".csv")
            and "dx_dad1a" in it["name"].lower()
        ]
        return {"files": chrom}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint error: {e}")


@app.post("/sharepoint/download-batch")
async def sharepoint_download_batch(
    file_ids: list[str],
    _current_user=Depends(get_current_user),
):
    """
    Download multiple files from SharePoint and return their contents.
    Used to fetch all CSVs for HPLC analysis in one request.
    """
    try:
        results = []
        for item_id in file_ids:
            content, filename = await sp.download_file(item_id)
            results.append({
                "id": item_id,
                "filename": filename,
                "content": content.decode("utf-8", errors="replace"),
            })
        return {"files": results}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint batch download error: {e}")


# ─── Wizard Session Endpoints ──────────────────────────────────────────────────

# --- Pydantic schemas ---

VALID_STEP_KEYS = {
    "stock_vial_empty_mg",
    "stock_vial_with_peptide_mg",  # Standards only: vial+cap+peptide aliquot, before diluent
    "stock_vial_loaded_mg",
    "dil_vial_empty_mg",
    "dil_vial_with_diluent_mg",
    "dil_vial_final_mg",
}


class WizardSessionCreate(BaseModel):
    """Schema for creating a new wizard session."""
    peptide_id: int
    sample_id_label: Optional[str] = None
    declared_weight_mg: Optional[float] = None  # mg; must be > 0 and < 5000 if provided
    target_conc_ug_ml: Optional[float] = None
    target_total_vol_ul: Optional[float] = None
    vial_params: Optional[dict] = None  # Multi-vial: {"1": {declared_weight_mg, target_conc, target_vol}, ...}
    # Phase 09: Standard prep metadata
    is_standard: Optional[bool] = None
    manufacturer: Optional[str] = None
    standard_notes: Optional[str] = None
    instrument_name: Optional[str] = None
    instrument_id: Optional[int] = None
    lims_sub_sample_pk: Optional[int] = None


class WizardSessionUpdate(BaseModel):
    """Schema for updating session fields (PATCH). All fields optional."""
    sample_id_label: Optional[str] = None
    declared_weight_mg: Optional[float] = None
    target_conc_ug_ml: Optional[float] = None
    target_total_vol_ul: Optional[float] = None
    peak_area: Optional[float] = None
    vial_params: Optional[dict] = None
    # Phase 09: Standard prep metadata
    is_standard: Optional[bool] = None
    manufacturer: Optional[str] = None
    standard_notes: Optional[str] = None
    instrument_name: Optional[str] = None
    instrument_id: Optional[int] = None


class WizardMeasurementCreate(BaseModel):
    """Schema for recording a weight measurement."""
    step_key: str  # Must be one of VALID_STEP_KEYS
    weight_mg: float  # Raw balance reading in milligrams
    source: str = "manual"  # 'manual' | 'scale'
    vial_number: int = 1  # Which vial this measurement belongs to (multi-vial blends)


class WizardMeasurementResponse(BaseModel):
    """Schema for measurement response."""
    id: int
    session_id: int
    step_key: str
    weight_mg: float
    source: str
    vial_number: int = 1
    is_current: bool
    recorded_at: datetime

    class Config:
        from_attributes = True


class WizardSessionResponse(BaseModel):
    """
    Full session response including current measurements and calculated values.
    Calculations are recalculated on demand — never stored in DB.
    Decimal values are converted to float at this boundary.
    """
    id: int
    peptide_id: int
    calibration_curve_id: Optional[int]
    status: str
    sample_id_label: Optional[str]
    declared_weight_mg: Optional[float]
    target_conc_ug_ml: Optional[float]
    target_total_vol_ul: Optional[float]
    peak_area: Optional[float]
    created_at: datetime
    updated_at: datetime
    completed_at: Optional[datetime]
    measurements: list[WizardMeasurementResponse] = []
    calculations: Optional[dict] = None  # Populated by _build_session_response()
    vial_params: Optional[dict] = None  # Per-vial target params (multi-vial blends)
    vial_calculations: Optional[dict] = None  # Per-vial calculations keyed by vial number
    # Phase 09: Standard prep metadata
    is_standard: bool = False
    manufacturer: Optional[str] = None
    standard_notes: Optional[str] = None
    instrument_name: Optional[str] = None
    instrument_id: Optional[int] = None
    lims_sub_sample_pk: Optional[int] = None

    class Config:
        from_attributes = True


class WizardSessionListItem(BaseModel):
    """Lightweight session entry for list view."""
    id: int
    peptide_id: int
    status: str
    sample_id_label: Optional[str]
    declared_weight_mg: Optional[float]
    created_at: datetime
    updated_at: datetime
    completed_at: Optional[datetime]

    class Config:
        from_attributes = True


# --- Helper: build session response with inline calculations ---

def _build_session_response(session: WizardSession, db: Session) -> WizardSessionResponse:
    """
    Build a WizardSessionResponse from an ORM session object.
    Loads current measurements and triggers calculation.
    Decimal arithmetic happens inside calculations/wizard.py.
    float() conversion happens here at the response boundary.
    """
    from decimal import Decimal

    # Collect current measurements keyed by step_key (vial 1 / single-vial backward compat)
    current = {
        m.step_key: m.weight_mg
        for m in session.measurements
        if m.is_current and m.vial_number == 1
    }

    calcs: dict = {}

    # Stage 1: Stock Prep
    # Production: needs declared_weight + empty + loaded
    # Standard:   needs empty + with_peptide + loaded (declared not required)
    stock_empty = current.get("stock_vial_empty_mg")
    stock_with_peptide = current.get("stock_vial_with_peptide_mg")
    stock_loaded = current.get("stock_vial_loaded_mg")
    declared = session.declared_weight_mg
    is_standard = session.is_standard

    can_calc = (
        (is_standard and all(v is not None for v in [stock_empty, stock_with_peptide, stock_loaded])) or
        (not is_standard and all(v is not None for v in [declared, stock_empty, stock_loaded]))
    )

    stock_conc_d = None  # Decimal — used in subsequent stages

    if can_calc:
        try:
            from calculations.wizard import calc_stock_prep
            _dd = session.calibration_curve.diluent_density if session.calibration_curve else 997.1
            density = Decimal(str(_dd))
            sp = calc_stock_prep(
                Decimal(str(declared)) if declared is not None else None,
                Decimal(str(stock_empty)),
                Decimal(str(stock_loaded)),
                density,
                Decimal(str(stock_with_peptide)) if stock_with_peptide is not None else None,
            )
            stock_conc_d = sp["stock_conc_ug_ml"]
            calcs["diluent_added_ml"] = float(sp["total_diluent_added_ml"])
            calcs["stock_conc_ug_ml"] = float(sp["stock_conc_ug_ml"])
            if is_standard:
                calcs["actual_peptide_mg"] = float(sp["actual_peptide_mg"])
        except Exception:
            pass  # Partial session — skip this stage

    # Stage 2: Required Volumes — requires Stage 1 + target params
    if stock_conc_d is not None and session.target_conc_ug_ml and session.target_total_vol_ul:
        try:
            from calculations.wizard import calc_required_volumes
            rv = calc_required_volumes(
                stock_conc_d,
                Decimal(str(session.target_conc_ug_ml)),
                Decimal(str(session.target_total_vol_ul)),
            )
            calcs["required_stock_vol_ul"] = float(rv["required_stock_vol_ul"])
            calcs["required_diluent_vol_ul"] = float(rv["required_diluent_vol_ul"])
        except Exception:
            pass

    # Stage 3: Actual Dilution — requires Stage 1 + 3 dilution vial weights
    dil_empty = current.get("dil_vial_empty_mg")
    dil_diluent = current.get("dil_vial_with_diluent_mg")
    dil_final = current.get("dil_vial_final_mg")

    actual_conc_d = None
    actual_total_d = None
    actual_stock_d = None

    if stock_conc_d is not None and all(v is not None for v in [dil_empty, dil_diluent, dil_final]):
        try:
            from calculations.wizard import calc_actual_dilution
            _dd2 = session.calibration_curve.diluent_density if session.calibration_curve else 997.1
            density = Decimal(str(_dd2))
            ad = calc_actual_dilution(
                stock_conc_d,
                Decimal(str(dil_empty)),
                Decimal(str(dil_diluent)),
                Decimal(str(dil_final)),
                density,
            )
            actual_conc_d = ad["actual_conc_ug_ml"]
            actual_total_d = ad["actual_total_vol_ul"]
            actual_stock_d = ad["actual_stock_vol_ul"]
            calcs["actual_diluent_vol_ul"] = float(ad["actual_diluent_vol_ul"])
            calcs["actual_stock_vol_ul"] = float(ad["actual_stock_vol_ul"])
            calcs["actual_total_vol_ul"] = float(ad["actual_total_vol_ul"])
            calcs["actual_conc_ug_ml"] = float(ad["actual_conc_ug_ml"])
        except Exception:
            pass

    # Stage 4: Results — requires Stage 3 + peak_area + calibration curve
    if actual_conc_d is not None and actual_total_d is not None and actual_stock_d is not None and session.peak_area and session.calibration_curve_id:
        try:
            from calculations.wizard import calc_results
            cal = db.execute(
                select(CalibrationCurve).where(CalibrationCurve.id == session.calibration_curve_id)
            ).scalar_one_or_none()
            if cal:
                res = calc_results(
                    Decimal(str(cal.slope)),
                    Decimal(str(cal.intercept)),
                    Decimal(str(session.peak_area)),
                    actual_conc_d,
                    actual_total_d,
                    actual_stock_d,
                )
                calcs["determined_conc_ug_ml"] = float(res["determined_conc_ug_ml"])
                calcs["peptide_mass_mg"] = float(res["peptide_mass_mg"])
                calcs["purity_pct"] = float(res["purity_pct"])
                calcs["dilution_factor"] = float(res["dilution_factor"])
        except Exception:
            pass

    # Per-analyte calculations for single-vial sessions
    # Check vial_params["1"] for analyte_params (or if only 1 vial in multi-vial)
    _sv_aparams = None
    if session.vial_params and len(session.vial_params) == 1:
        _sv_vp = list(session.vial_params.values())[0]
        _sv_aparams = _sv_vp.get("analyte_params") if isinstance(_sv_vp, dict) else None
    elif session.vial_params and "1" in session.vial_params:
        _sv_vp = session.vial_params["1"]
        _sv_aparams = _sv_vp.get("analyte_params") if isinstance(_sv_vp, dict) else None

    if _sv_aparams and isinstance(_sv_aparams, dict) and "diluent_added_ml" in calcs:
        from calculations.wizard import calc_stock_conc_per_analyte, calc_actual_conc_per_analyte
        sv_diluent_ml = Decimal(str(calcs["diluent_added_ml"]))
        sv_analyte_calcs: dict = {}
        for a_key, a_params in _sv_aparams.items():
            ac: dict = {}
            a_decl = a_params.get("declared_weight_mg")
            a_sc_d = None
            if a_decl and sv_diluent_ml:
                try:
                    a_sc_d = calc_stock_conc_per_analyte(Decimal(str(a_decl)), sv_diluent_ml)
                    ac["stock_conc_ug_ml"] = float(a_sc_d)
                except Exception:
                    pass
            a_tc = a_params.get("target_conc_ug_ml")
            a_tv = a_params.get("target_total_vol_ul")
            if a_sc_d is not None and a_tc and a_tv:
                try:
                    from calculations.wizard import calc_required_volumes as _crv
                    arv = _crv(a_sc_d, Decimal(str(a_tc)), Decimal(str(a_tv)))
                    ac["required_stock_vol_ul"] = float(arv["required_stock_vol_ul"])
                    ac["required_diluent_vol_ul"] = float(arv["required_diluent_vol_ul"])
                except Exception:
                    pass
            if a_sc_d is not None and "actual_stock_vol_ul" in calcs and "actual_total_vol_ul" in calcs:
                try:
                    a_act = calc_actual_conc_per_analyte(
                        a_sc_d, Decimal(str(calcs["actual_stock_vol_ul"])),
                        Decimal(str(calcs["actual_total_vol_ul"])),
                    )
                    ac["actual_conc_ug_ml"] = float(a_act)
                except Exception:
                    pass
            if ac:
                sv_analyte_calcs[a_key] = ac
        if sv_analyte_calcs:
            calcs["analyte_calculations"] = sv_analyte_calcs

    # Build current measurements list (only is_current=True)
    current_measurements = [m for m in session.measurements if m.is_current]

    # Per-vial calculations (multi-vial blends)
    vial_calcs: dict | None = None
    # Run vial calculations when: multiple vials OR any vial has analyte_params (blend in single vial)
    _has_analyte_params = session.vial_params and any(
        v.get("analyte_params") for v in session.vial_params.values()
    ) if session.vial_params else False
    if session.vial_params and (len(session.vial_params) > 1 or _has_analyte_params):
        from calculations.wizard import (
            calc_stock_prep, calc_required_volumes, calc_actual_dilution,
            calc_stock_conc_per_analyte, calc_actual_conc_per_analyte,
        )
        vial_calcs = {}
        _dd = session.calibration_curve.diluent_density if session.calibration_curve else 997.1
        density = Decimal(str(_dd))

        for vial_key, vparams in sorted(session.vial_params.items(), key=lambda x: int(x[0])):
            vn = int(vial_key)
            vcurrent = {
                m.step_key: m.weight_mg
                for m in session.measurements
                if m.is_current and m.vial_number == vn
            }
            vc: dict = {}
            v_declared = vparams.get("declared_weight_mg")
            v_stock_empty = vcurrent.get("stock_vial_empty_mg")
            v_stock_with_peptide = vcurrent.get("stock_vial_with_peptide_mg")
            v_stock_loaded = vcurrent.get("stock_vial_loaded_mg")
            v_stock_conc_d = None
            v_diluent_ml_d = None  # shared diluent volume for per-analyte calcs

            # Standard serial dilutions: vials 2+ use previous vial's actual_conc as stock
            if is_standard and vn > 1:
                prev_vc = vial_calcs.get(str(vn - 1), {})
                prev_actual = prev_vc.get("actual_conc_ug_ml")
                if prev_actual is not None:
                    v_stock_conc_d = Decimal(str(prev_actual))
            else:
                v_can_calc = (
                    (is_standard and all(v is not None for v in [v_stock_empty, v_stock_with_peptide, v_stock_loaded])) or
                    (not is_standard and all(v is not None for v in [v_declared, v_stock_empty, v_stock_loaded]))
                )
                if v_can_calc:
                    try:
                        sp = calc_stock_prep(
                            Decimal(str(v_declared)) if v_declared is not None else None,
                            Decimal(str(v_stock_empty)),
                            Decimal(str(v_stock_loaded)),
                            density,
                            Decimal(str(v_stock_with_peptide)) if v_stock_with_peptide is not None else None,
                        )
                        v_stock_conc_d = sp["stock_conc_ug_ml"]
                        v_diluent_ml_d = sp["total_diluent_added_ml"]
                        vc["diluent_added_ml"] = float(sp["total_diluent_added_ml"])
                        vc["stock_conc_ug_ml"] = float(sp["stock_conc_ug_ml"])
                        if is_standard:
                            vc["actual_peptide_mg"] = float(sp["actual_peptide_mg"])
                    except Exception:
                        pass

            v_target_conc = vparams.get("target_conc_ug_ml")
            v_target_vol = vparams.get("target_total_vol_ul")
            if v_stock_conc_d is not None and v_target_conc and v_target_vol:
                try:
                    rv = calc_required_volumes(
                        v_stock_conc_d, Decimal(str(v_target_conc)), Decimal(str(v_target_vol)),
                    )
                    vc["required_stock_vol_ul"] = float(rv["required_stock_vol_ul"])
                    vc["required_diluent_vol_ul"] = float(rv["required_diluent_vol_ul"])
                except Exception:
                    pass

            v_dil_empty = vcurrent.get("dil_vial_empty_mg")
            v_dil_diluent = vcurrent.get("dil_vial_with_diluent_mg")
            v_dil_final = vcurrent.get("dil_vial_final_mg")
            if v_stock_conc_d is not None and all(v is not None for v in [v_dil_empty, v_dil_diluent, v_dil_final]):
                try:
                    ad = calc_actual_dilution(
                        v_stock_conc_d, Decimal(str(v_dil_empty)),
                        Decimal(str(v_dil_diluent)), Decimal(str(v_dil_final)), density,
                    )
                    vc["actual_diluent_vol_ul"] = float(ad["actual_diluent_vol_ul"])
                    vc["actual_stock_vol_ul"] = float(ad["actual_stock_vol_ul"])
                    vc["actual_total_vol_ul"] = float(ad["actual_total_vol_ul"])
                    vc["actual_conc_ug_ml"] = float(ad["actual_conc_ug_ml"])
                except Exception:
                    pass

            # --- Per-analyte calculations (when analyte_params provided) ---
            analyte_params = vparams.get("analyte_params")
            if analyte_params and isinstance(analyte_params, dict) and v_diluent_ml_d is not None:
                analyte_calcs: dict = {}

                for analyte_key, aparams in analyte_params.items():
                    ac: dict = {}
                    a_declared = aparams.get("declared_weight_mg")

                    # Per-analyte stock concentration
                    a_stock_conc_d = None
                    if a_declared and v_diluent_ml_d:
                        try:
                            a_stock_conc_d = calc_stock_conc_per_analyte(
                                Decimal(str(a_declared)), v_diluent_ml_d,
                            )
                            ac["stock_conc_ug_ml"] = float(a_stock_conc_d)
                        except Exception:
                            pass

                    # Per-analyte required volumes (reference — not physically independent)
                    a_target_conc = aparams.get("target_conc_ug_ml")
                    a_target_vol = aparams.get("target_total_vol_ul")
                    if a_stock_conc_d is not None and a_target_conc and a_target_vol:
                        try:
                            arv = calc_required_volumes(
                                a_stock_conc_d, Decimal(str(a_target_conc)),
                                Decimal(str(a_target_vol)),
                            )
                            ac["required_stock_vol_ul"] = float(arv["required_stock_vol_ul"])
                            ac["required_diluent_vol_ul"] = float(arv["required_diluent_vol_ul"])
                        except Exception:
                            pass

                    # Per-analyte actual concentration (shared dilution factor)
                    if a_stock_conc_d is not None and "actual_stock_vol_ul" in vc and "actual_total_vol_ul" in vc:
                        try:
                            a_actual = calc_actual_conc_per_analyte(
                                a_stock_conc_d,
                                Decimal(str(vc["actual_stock_vol_ul"])),
                                Decimal(str(vc["actual_total_vol_ul"])),
                            )
                            ac["actual_conc_ug_ml"] = float(a_actual)
                        except Exception:
                            pass

                    if ac:
                        analyte_calcs[analyte_key] = ac

                if analyte_calcs:
                    vc["analyte_calculations"] = analyte_calcs

                    # For blends: override vial-level required volumes with the MAX
                    # across per-analyte requirements (all analytes share one pipette)
                    max_stock = max(
                        (ac.get("required_stock_vol_ul", 0) for ac in analyte_calcs.values()),
                        default=0,
                    )
                    if max_stock > 0:
                        vc["required_stock_vol_ul"] = max_stock
                        if v_target_vol:
                            vc["required_diluent_vol_ul"] = float(
                                Decimal(str(v_target_vol)) - Decimal(str(max_stock))
                            )

            if vc:
                vial_calcs[vial_key] = vc

        if not vial_calcs:
            vial_calcs = None

    return WizardSessionResponse(
        id=session.id,
        peptide_id=session.peptide_id,
        calibration_curve_id=session.calibration_curve_id,
        status=session.status,
        sample_id_label=session.sample_id_label,
        declared_weight_mg=session.declared_weight_mg,
        target_conc_ug_ml=session.target_conc_ug_ml,
        target_total_vol_ul=session.target_total_vol_ul,
        peak_area=session.peak_area,
        created_at=session.created_at,
        updated_at=session.updated_at,
        completed_at=session.completed_at,
        measurements=[
            WizardMeasurementResponse.model_validate(m) for m in current_measurements
        ],
        calculations=calcs if calcs else None,
        vial_params=session.vial_params,
        vial_calculations=vial_calcs,
        is_standard=session.is_standard,
        manufacturer=session.manufacturer,
        standard_notes=session.standard_notes,
        instrument_name=session.instrument_name,
        instrument_id=session.instrument_id,
        lims_sub_sample_pk=session.lims_sub_sample_pk,
    )


# --- Endpoints ---

@app.post("/wizard/sessions", response_model=WizardSessionResponse, status_code=201)
async def create_wizard_session(
    data: WizardSessionCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Start a new analysis wizard session (SESS-01).
    Auto-resolves the active calibration curve for the peptide.
    Returns 400 if no active calibration curve exists.
    Returns 404 if peptide not found.
    """
    peptide = db.execute(select(Peptide).where(Peptide.id == data.peptide_id)).scalar_one_or_none()
    if not peptide:
        raise HTTPException(status_code=404, detail=f"Peptide {data.peptide_id} not found")

    cal = None
    # Standards are creating the curve — skip calibration requirement
    if not data.is_standard:
        if peptide.is_blend:
            # For blends, find an active calibration from any component peptide
            component_ids = [
                row.component_id for row in db.execute(
                    select(blend_components.c.component_id)
                    .where(blend_components.c.blend_id == peptide.id)
                    .order_by(blend_components.c.display_order)
                ).all()
            ]
            if component_ids:
                blend_cal_query = (
                    select(CalibrationCurve)
                    .where(CalibrationCurve.peptide_id.in_(component_ids))
                    .where(CalibrationCurve.is_active == True)
                )
                if data.instrument_id:
                    blend_cal_query = blend_cal_query.where(CalibrationCurve.instrument_id == data.instrument_id)
                cal = db.execute(
                    blend_cal_query.order_by(desc(CalibrationCurve.created_at)).limit(1)
                ).scalar_one_or_none()
            if not cal:
                if data.instrument_id:
                    inst = db.execute(select(Instrument).where(Instrument.id == data.instrument_id)).scalar_one_or_none()
                    inst_name = inst.name if inst else f"ID {data.instrument_id}"
                    raise HTTPException(
                        status_code=400,
                        detail=f"No active calibration curves found for any component peptide in this blend on instrument '{inst_name}'. Star a curve for this instrument first."
                    )
                raise HTTPException(
                    status_code=400,
                    detail="No active calibration curves found for any component peptide in this blend."
                )
        else:
            # Look up active curve matching the session's instrument
            cal_query = (
                select(CalibrationCurve)
                .where(CalibrationCurve.peptide_id == data.peptide_id)
                .where(CalibrationCurve.is_active == True)
            )
            if data.instrument_id:
                cal_query = cal_query.where(CalibrationCurve.instrument_id == data.instrument_id)
            cal = db.execute(
                cal_query.order_by(desc(CalibrationCurve.created_at)).limit(1)
            ).scalar_one_or_none()
            if not cal:
                if data.instrument_id:
                    inst = db.execute(select(Instrument).where(Instrument.id == data.instrument_id)).scalar_one_or_none()
                    inst_name = inst.name if inst else f"ID {data.instrument_id}"
                    raise HTTPException(
                        status_code=400,
                        detail=f"No active calibration curve found for peptide {data.peptide_id} on instrument '{inst_name}'. Star a curve for this instrument first."
                    )
                raise HTTPException(
                    status_code=400,
                    detail=f"No active calibration curve found for peptide {data.peptide_id}. Activate a calibration curve before starting a session."
                )

    if data.declared_weight_mg is not None and not (0 < data.declared_weight_mg < 5000):
        raise HTTPException(status_code=422, detail="declared_weight_mg must be between 0 and 5000 mg")

    # Resolve instrument_id → instrument_name if only ID was provided
    resolved_inst_name = data.instrument_name
    resolved_inst_id = data.instrument_id
    if resolved_inst_id and not resolved_inst_name:
        inst = db.execute(select(Instrument).where(Instrument.id == resolved_inst_id)).scalar_one_or_none()
        if inst:
            resolved_inst_name = inst.name

    session = WizardSession(
        peptide_id=data.peptide_id,
        calibration_curve_id=cal.id if cal else None,
        sample_id_label=data.sample_id_label,
        declared_weight_mg=data.declared_weight_mg,
        target_conc_ug_ml=data.target_conc_ug_ml,
        target_total_vol_ul=data.target_total_vol_ul,
        vial_params=data.vial_params,
        is_standard=data.is_standard if data.is_standard is not None else False,
        manufacturer=data.manufacturer,
        standard_notes=data.standard_notes,
        instrument_name=resolved_inst_name,
        instrument_id=resolved_inst_id,
        lims_sub_sample_pk=data.lims_sub_sample_pk,
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return _build_session_response(session, db)


@app.get("/wizard/sessions", response_model=list[WizardSessionListItem])
async def list_wizard_sessions(
    status: Optional[str] = None,
    peptide_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    List wizard sessions with optional filtering.
    Returns lightweight list items (no measurements, no calculations).
    """
    stmt = select(WizardSession).order_by(desc(WizardSession.created_at))
    if status:
        stmt = stmt.where(WizardSession.status == status)
    if peptide_id:
        stmt = stmt.where(WizardSession.peptide_id == peptide_id)
    stmt = stmt.offset(offset).limit(limit)
    sessions = db.execute(stmt).scalars().all()
    return sessions


@app.get("/wizard/sessions/{session_id}", response_model=WizardSessionResponse)
async def get_wizard_session(
    session_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Get a wizard session with all current measurements and recalculated values.
    Used for resuming an in-progress session (SESS-02).
    """
    session = db.execute(
        select(WizardSession).where(WizardSession.id == session_id)
    ).scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    return _build_session_response(session, db)


@app.patch("/wizard/sessions/{session_id}", response_model=WizardSessionResponse)
async def update_wizard_session(
    session_id: int,
    data: WizardSessionUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Update session fields (target params, peak_area, sample label, declared weight).
    Returns updated session with recalculated values.
    """
    session = db.execute(
        select(WizardSession).where(WizardSession.id == session_id)
    ).scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    if session.status == "completed":
        raise HTTPException(status_code=400, detail="Cannot update a completed session")

    from sqlalchemy.orm.attributes import flag_modified

    update_data = data.model_dump(exclude_unset=True)

    # Resolve instrument_id → instrument_name if only ID was provided
    if "instrument_id" in update_data and update_data["instrument_id"] and "instrument_name" not in update_data:
        inst = db.execute(select(Instrument).where(Instrument.id == update_data["instrument_id"])).scalar_one_or_none()
        if inst:
            update_data["instrument_name"] = inst.name

    if "declared_weight_mg" in update_data and update_data["declared_weight_mg"] is not None:
        if not (0 < update_data["declared_weight_mg"] < 5000):
            raise HTTPException(status_code=422, detail="declared_weight_mg must be between 0 and 5000 mg")

    if "vial_params" in update_data and update_data["vial_params"] is not None:
        session.vial_params = update_data.pop("vial_params")
        flag_modified(session, "vial_params")

    for field, value in update_data.items():
        setattr(session, field, value)

    db.commit()
    db.refresh(session)

    # Sync standard metadata to the linked sample_prep row (if any)
    std_meta_fields = {"instrument_name", "instrument_id", "manufacturer", "standard_notes"}
    if std_meta_fields & set(update_data.keys()):
        from mk1_db import get_mk1_db, update_sample_prep as _update_sp
        with get_mk1_db() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM sample_preps WHERE wizard_session_id = %s", [session_id])
                sp_row = cur.fetchone()
        if sp_row:
            sync_data = {k: update_data[k] for k in std_meta_fields if k in update_data}
            _update_sp(sp_row[0], sync_data)

    return _build_session_response(session, db)


@app.post("/wizard/sessions/{session_id}/measurements", response_model=WizardSessionResponse, status_code=201)
async def record_measurement(
    session_id: int,
    data: WizardMeasurementCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Record a weight for a wizard step. If a measurement for this step already exists,
    mark the old record as is_current=False (audit trail preserved) and insert a new one.
    Returns updated session with recalculated values.
    """
    session = db.execute(
        select(WizardSession).where(WizardSession.id == session_id)
    ).scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    if session.status == "completed":
        raise HTTPException(status_code=400, detail="Cannot add measurements to a completed session")

    if data.step_key not in VALID_STEP_KEYS:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid step_key '{data.step_key}'. Must be one of: {sorted(VALID_STEP_KEYS)}"
        )
    if data.source not in ("manual", "scale"):
        raise HTTPException(status_code=422, detail="source must be 'manual' or 'scale'")

    # Mark existing current measurement for this step+vial as superseded
    old = db.execute(
        select(WizardMeasurement)
        .where(WizardMeasurement.session_id == session_id)
        .where(WizardMeasurement.step_key == data.step_key)
        .where(WizardMeasurement.vial_number == data.vial_number)
        .where(WizardMeasurement.is_current == True)
    ).scalar_one_or_none()
    if old:
        old.is_current = False

    # Insert new measurement
    new_m = WizardMeasurement(
        session_id=session_id,
        step_key=data.step_key,
        weight_mg=data.weight_mg,
        source=data.source,
        vial_number=data.vial_number,
        is_current=True,
    )
    db.add(new_m)
    db.commit()
    db.refresh(session)

    # Auto-sync: if a sample prep already exists for this session, update it
    # so vial_data stays current without requiring a manual re-save
    try:
        from mk1_db import get_mk1_db as _sync_db, update_sample_prep as _sync_update
        from psycopg2.extras import RealDictCursor as _sync_RDC
        with _sync_db() as _sconn:
            with _sconn.cursor(cursor_factory=_sync_RDC) as _scur:
                _scur.execute(
                    "SELECT id FROM sample_preps WHERE wizard_session_id = %s",
                    [session_id],
                )
                _existing_prep = _scur.fetchone()
        if _existing_prep:
            # Re-run the full save logic by calling the save endpoint internally
            response = _build_session_response(session, db)
            calcs = response.calculations or {}
            current = {
                m.step_key: m.weight_mg
                for m in session.measurements
                if m.is_current and m.vial_number == 1
            }
            update_data = {
                "stock_vial_empty_mg": current.get("stock_vial_empty_mg"),
                "stock_vial_loaded_mg": current.get("stock_vial_loaded_mg"),
                "stock_conc_ug_ml": calcs.get("stock_conc_ug_ml"),
                "required_diluent_vol_ul": calcs.get("required_diluent_vol_ul"),
                "required_stock_vol_ul": calcs.get("required_stock_vol_ul"),
                "dil_vial_empty_mg": current.get("dil_vial_empty_mg"),
                "dil_vial_with_diluent_mg": current.get("dil_vial_with_diluent_mg"),
                "dil_vial_final_mg": current.get("dil_vial_final_mg"),
                "actual_conc_ug_ml": calcs.get("actual_conc_ug_ml"),
                "actual_diluent_vol_ul": calcs.get("actual_diluent_vol_ul"),
                "actual_stock_vol_ul": calcs.get("actual_stock_vol_ul"),
                "actual_total_vol_ul": calcs.get("actual_total_vol_ul"),
            }
            # Rebuild vial_data for blends
            if response.vial_calculations:
                peptide = session.peptide
                if peptide and peptide.is_blend:
                    comp_rows = db.execute(
                        select(Peptide, blend_components.c.vial_number)
                        .join(blend_components, blend_components.c.component_id == Peptide.id)
                        .where(blend_components.c.blend_id == peptide.id)
                        .order_by(blend_components.c.display_order)
                    ).all()
                    vial_data_list = []
                    for vial_key, vc in response.vial_calculations.items():
                        vn = int(vial_key)
                        vp = (session.vial_params or {}).get(vial_key, {})
                        v_current = {
                            m.step_key: m.weight_mg
                            for m in session.measurements
                            if m.is_current and m.vial_number == vn
                        }
                        vd_entry = {
                            "vial_number": vn,
                            "component_ids": [c.id for c, cvn in comp_rows if (cvn or 1) == vn],
                            "component_abbreviations": [c.abbreviation for c, cvn in comp_rows if (cvn or 1) == vn],
                            "declared_weight_mg": vp.get("declared_weight_mg"),
                            "target_conc_ug_ml": vp.get("target_conc_ug_ml"),
                            "target_total_vol_ul": vp.get("target_total_vol_ul"),
                            "stock_vial_empty_mg": v_current.get("stock_vial_empty_mg"),
                            "stock_vial_loaded_mg": v_current.get("stock_vial_loaded_mg"),
                            "stock_conc_ug_ml": vc.get("stock_conc_ug_ml"),
                            "required_diluent_vol_ul": vc.get("required_diluent_vol_ul"),
                            "required_stock_vol_ul": vc.get("required_stock_vol_ul"),
                            "dil_vial_empty_mg": v_current.get("dil_vial_empty_mg"),
                            "dil_vial_with_diluent_mg": v_current.get("dil_vial_with_diluent_mg"),
                            "dil_vial_final_mg": v_current.get("dil_vial_final_mg"),
                            "actual_conc_ug_ml": vc.get("actual_conc_ug_ml"),
                            "actual_diluent_vol_ul": vc.get("actual_diluent_vol_ul"),
                            "actual_stock_vol_ul": vc.get("actual_stock_vol_ul"),
                            "actual_total_vol_ul": vc.get("actual_total_vol_ul"),
                        }
                        vial_data_list.append(vd_entry)
                    update_data["vial_data"] = json.dumps(vial_data_list)
            _sync_update(_existing_prep["id"], update_data)
            print(f"[INFO] Auto-synced sample prep {_existing_prep['id']} from session {session_id}")
    except Exception as e:
        print(f"[WARN] Auto-sync sample prep failed for session {session_id}: {e}")

    return _build_session_response(session, db)


@app.post("/wizard/sessions/{session_id}/complete", response_model=WizardSessionResponse)
async def complete_wizard_session(
    session_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Mark a wizard session as complete (SESS-03).
    Sets status='completed' and records completed_at timestamp.
    Returns 400 if already completed.
    """
    session = db.execute(
        select(WizardSession).where(WizardSession.id == session_id)
    ).scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    if session.status == "completed":
        raise HTTPException(status_code=400, detail="Session is already completed")

    session.status = "completed"
    session.completed_at = datetime.utcnow()
    db.commit()
    db.refresh(session)
    return _build_session_response(session, db)


# ─── Sample Preps API ──────────────────────────────────────────────────────────

class SamplePrepCreateRequest(BaseModel):
    """Save a completed wizard session as a sample prep record in the integration DB."""
    wizard_session_id: int
    notes: Optional[str] = None


class SamplePrepUpdateRequest(BaseModel):
    """PATCH fields on an existing sample prep."""
    senaite_sample_id: Optional[str] = None
    declared_weight_mg: Optional[float] = None
    target_conc_ug_ml: Optional[float] = None
    target_total_vol_ul: Optional[float] = None
    stock_vial_empty_mg: Optional[float] = None
    stock_vial_loaded_mg: Optional[float] = None
    stock_conc_ug_ml: Optional[float] = None
    required_diluent_vol_ul: Optional[float] = None
    required_stock_vol_ul: Optional[float] = None
    dil_vial_empty_mg: Optional[float] = None
    dil_vial_with_diluent_mg: Optional[float] = None
    dil_vial_final_mg: Optional[float] = None
    actual_conc_ug_ml: Optional[float] = None
    actual_diluent_vol_ul: Optional[float] = None
    actual_stock_vol_ul: Optional[float] = None
    actual_total_vol_ul: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None


@app.post("/sample-preps", status_code=201)
async def create_sample_prep_endpoint(
    body: SamplePrepCreateRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Save a wizard session as a sample prep into the Integration-Services Postgres DB.
    Loads the wizard session by id, pulls all measurements + derived calculations,
    and writes a flat record to sample_preps.
    """
    from mk1_db import ensure_sample_preps_table, create_sample_prep

    # Load wizard session
    session = db.execute(
        select(WizardSession).where(WizardSession.id == body.wizard_session_id)
    ).scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail=f"Wizard session {body.wizard_session_id} not found")

    # Build inline calculations (reuse existing helper)
    response = _build_session_response(session, db)
    calcs = response.calculations or {}

    # Current measurements keyed by step_key (vial 1 for backward compat)
    current = {
        m.step_key: m.weight_mg
        for m in session.measurements
        if m.is_current and m.vial_number == 1
    }

    data = {
        "wizard_session_id": session.id,
        "peptide_id": session.peptide_id,
        "peptide_name": session.peptide.name if session.peptide else None,
        "peptide_abbreviation": session.peptide.abbreviation if session.peptide else None,
        "senaite_sample_id": session.sample_id_label,
        "lims_sub_sample_pk": session.lims_sub_sample_pk,
        "declared_weight_mg": float(session.declared_weight_mg) if session.declared_weight_mg else None,
        "target_conc_ug_ml": float(session.target_conc_ug_ml) if session.target_conc_ug_ml else None,
        "target_total_vol_ul": float(session.target_total_vol_ul) if session.target_total_vol_ul else None,
        # Step 2 measurements
        "stock_vial_empty_mg": current.get("stock_vial_empty_mg"),
        "stock_vial_loaded_mg": current.get("stock_vial_loaded_mg"),
        # Step 2 derived
        "stock_conc_ug_ml": calcs.get("stock_conc_ug_ml"),
        "required_diluent_vol_ul": calcs.get("required_diluent_vol_ul"),
        "required_stock_vol_ul": calcs.get("required_stock_vol_ul"),
        # Step 3 measurements
        "dil_vial_empty_mg": current.get("dil_vial_empty_mg"),
        "dil_vial_with_diluent_mg": current.get("dil_vial_with_diluent_mg"),
        "dil_vial_final_mg": current.get("dil_vial_final_mg"),
        # Step 3 derived
        "actual_conc_ug_ml": calcs.get("actual_conc_ug_ml"),
        "actual_diluent_vol_ul": calcs.get("actual_diluent_vol_ul"),
        "actual_stock_vol_ul": calcs.get("actual_stock_vol_ul"),
        "actual_total_vol_ul": calcs.get("actual_total_vol_ul"),
        "status": "awaiting_hplc",
        "notes": body.notes,
        # Phase 09: Standard prep metadata
        "is_standard": session.is_standard,
        "manufacturer": session.manufacturer,
        "standard_notes": session.standard_notes,
        "instrument_name": session.instrument_name or (
            session.instrument_obj.name if session.instrument_id and session.instrument_obj else None
        ),
        "instrument_id": session.instrument_id,
        # User tracking
        "created_by_user_id": current_user.id,
        "created_by_email": current_user.email,
        "updated_by_user_id": current_user.id,
        "updated_by_email": current_user.email,
    }

    # Standard multi-vial support: store per-vial actual concentrations
    if session.is_standard and response.vial_calculations and len(response.vial_calculations) > 1:
        vial_data_list = []
        for vial_key, vc in response.vial_calculations.items():
            vn = int(vial_key)
            vp = (session.vial_params or {}).get(vial_key, {})
            v_current = {
                m.step_key: m.weight_mg
                for m in session.measurements
                if m.is_current and m.vial_number == vn
            }
            vial_data_list.append({
                "vial_number": vn,
                "target_conc_ug_ml": vp.get("target_conc_ug_ml"),
                "target_total_vol_ul": vp.get("target_total_vol_ul"),
                "stock_vial_empty_mg": v_current.get("stock_vial_empty_mg"),
                "stock_vial_loaded_mg": v_current.get("stock_vial_loaded_mg"),
                "stock_conc_ug_ml": vc.get("stock_conc_ug_ml"),
                "required_diluent_vol_ul": vc.get("required_diluent_vol_ul"),
                "required_stock_vol_ul": vc.get("required_stock_vol_ul"),
                "dil_vial_empty_mg": v_current.get("dil_vial_empty_mg"),
                "dil_vial_with_diluent_mg": v_current.get("dil_vial_with_diluent_mg"),
                "dil_vial_final_mg": v_current.get("dil_vial_final_mg"),
                "actual_conc_ug_ml": vc.get("actual_conc_ug_ml"),
                "actual_diluent_vol_ul": vc.get("actual_diluent_vol_ul"),
                "actual_stock_vol_ul": vc.get("actual_stock_vol_ul"),
                "actual_total_vol_ul": vc.get("actual_total_vol_ul"),
            })
        data["vial_data"] = json.dumps(vial_data_list)

    # Blend support: include component info so HPLC flyout can load per-component curves
    peptide = session.peptide
    if peptide and peptide.is_blend:
        data["is_blend"] = True
        comp_rows = db.execute(
            select(Peptide, blend_components.c.vial_number)
            .join(blend_components, blend_components.c.component_id == Peptide.id)
            .where(blend_components.c.blend_id == peptide.id)
            .order_by(blend_components.c.display_order)
        ).all()
        data["components_json"] = json.dumps([
            {"id": c.id, "name": c.name, "abbreviation": c.abbreviation, "vial_number": vn or 1,
             "hplc_aliases": c.hplc_aliases}
            for c, vn in comp_rows
        ])

        # Build vial_data for multi-vial blends
        if response.vial_calculations:
            vial_data_list = []
            for vial_key, vc in response.vial_calculations.items():
                vn = int(vial_key)
                vp = (session.vial_params or {}).get(vial_key, {})
                v_current = {
                    m.step_key: m.weight_mg
                    for m in session.measurements
                    if m.is_current and m.vial_number == vn
                }
                # Build per-analyte data snapshot for this vial
                vial_analyte_data = None
                a_calcs = vc.get("analyte_calculations")
                a_params = vp.get("analyte_params") if isinstance(vp, dict) else None
                if a_calcs and a_params:
                    vial_analyte_data = []
                    for a_key, a_calc in a_calcs.items():
                        ap = a_params.get(a_key, {})
                        # Resolve component_id from abbreviation
                        comp_match = next((c for c, cvn in comp_rows if c.abbreviation == a_key), None)
                        vial_analyte_data.append({
                            "component_id": comp_match.id if comp_match else None,
                            "abbreviation": a_key,
                            "declared_weight_mg": ap.get("declared_weight_mg"),
                            "target_conc_ug_ml": ap.get("target_conc_ug_ml"),
                            "target_total_vol_ul": ap.get("target_total_vol_ul"),
                            "stock_conc_ug_ml": a_calc.get("stock_conc_ug_ml"),
                            "required_stock_vol_ul": a_calc.get("required_stock_vol_ul"),
                            "required_diluent_vol_ul": a_calc.get("required_diluent_vol_ul"),
                            "actual_conc_ug_ml": a_calc.get("actual_conc_ug_ml"),
                        })

                vd_entry = {
                    "vial_number": vn,
                    "component_ids": [c.id for c, cvn in comp_rows if (cvn or 1) == vn],
                    "component_abbreviations": [c.abbreviation for c, cvn in comp_rows if (cvn or 1) == vn],
                    "declared_weight_mg": vp.get("declared_weight_mg"),
                    "target_conc_ug_ml": vp.get("target_conc_ug_ml"),
                    "target_total_vol_ul": vp.get("target_total_vol_ul"),
                    "stock_vial_empty_mg": v_current.get("stock_vial_empty_mg"),
                    "stock_vial_loaded_mg": v_current.get("stock_vial_loaded_mg"),
                    "stock_conc_ug_ml": vc.get("stock_conc_ug_ml"),
                    "required_diluent_vol_ul": vc.get("required_diluent_vol_ul"),
                    "required_stock_vol_ul": vc.get("required_stock_vol_ul"),
                    "dil_vial_empty_mg": v_current.get("dil_vial_empty_mg"),
                    "dil_vial_with_diluent_mg": v_current.get("dil_vial_with_diluent_mg"),
                    "dil_vial_final_mg": v_current.get("dil_vial_final_mg"),
                    "actual_conc_ug_ml": vc.get("actual_conc_ug_ml"),
                    "actual_diluent_vol_ul": vc.get("actual_diluent_vol_ul"),
                    "actual_stock_vol_ul": vc.get("actual_stock_vol_ul"),
                    "actual_total_vol_ul": vc.get("actual_total_vol_ul"),
                }
                if vial_analyte_data:
                    vd_entry["analyte_data"] = vial_analyte_data
                vial_data_list.append(vd_entry)
            data["vial_data"] = json.dumps(vial_data_list)

    try:
        ensure_sample_preps_table()
        # Idempotency: if a sample prep already exists for this session, update it
        from mk1_db import get_mk1_db as _get_mk1_db, update_sample_prep as _update_sp
        from psycopg2.extras import RealDictCursor as _RDC
        existing_id = None
        with _get_mk1_db() as _conn:
            with _conn.cursor(cursor_factory=_RDC) as _cur:
                _cur.execute(
                    "SELECT id FROM sample_preps WHERE wizard_session_id = %s",
                    [body.wizard_session_id],
                )
                _existing = _cur.fetchone()
                if _existing:
                    existing_id = _existing["id"]
        if existing_id:
            row = _update_sp(existing_id, data)
        else:
            row = create_sample_prep(data)
        # Stamp the prep's bench assignment onto the vial's AR rows: the
        # wizard already knows instrument (picked at Step 1) and method (the
        # peptide's method chain), so apply both to the vial's unassigned
        # HPLC rows now — instead of method staying manual and instrument
        # arriving only at results-bridge time. Best-effort: a stamp failure
        # must never lose the prep (already saved above).
        if session.lims_sub_sample_pk is not None:
            try:
                from lims_analyses.prep_bridge import stamp_prep_assignment
                _method_id = None
                if session.peptide and session.peptide.methods:
                    _method_id = session.peptide.methods[0].id
                stamp_prep_assignment(
                    db,
                    lims_sub_sample_pk=session.lims_sub_sample_pk,
                    instrument_id=session.instrument_id,
                    method_id=_method_id,
                    user_id=current_user.id,
                )
            except Exception:
                db.rollback()
                logger.exception(
                    "prep assignment stamp failed for sub_pk=%s",
                    session.lims_sub_sample_pk,
                )
        # Serialize datetime fields
        for k in ("created_at", "updated_at"):
            if row.get(k) and hasattr(row[k], "isoformat"):
                row[k] = row[k].isoformat()
        return row
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save sample prep: {e}")


@app.get("/sample-preps")
async def list_sample_preps_endpoint(
    search: Optional[str] = None,
    is_standard: Optional[bool] = None,
    limit: int = 100,
    offset: int = 0,
    _current_user=Depends(get_current_user),
):
    """List sample preps from the integration DB (newest first)."""
    from mk1_db import ensure_sample_preps_table, list_sample_preps

    try:
        ensure_sample_preps_table()
        rows = list_sample_preps(search=search, is_standard=is_standard, limit=limit, offset=offset)
        for row in rows:
            for k in ("created_at", "updated_at"):
                if row.get(k) and hasattr(row[k], "isoformat"):
                    row[k] = row[k].isoformat()
        return rows
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list sample preps: {e}")


@app.get("/sample-preps/scan-hplc")
async def scan_sample_preps_hplc(_current_user=Depends(get_current_user)):
    """
    Scan the LIMS SharePoint folder for HPLC data matching each sample prep.
    Streams Server-Sent Events: log, progress, match, done, error events.
    Must be defined before /sample-preps/{id} to avoid route shadowing.
    """
    import json as _json
    from starlette.responses import StreamingResponse as _SR
    from mk1_db import ensure_sample_preps_table, list_sample_preps as _list_preps

    async def _generate():
        def ev(etype: str, data: dict) -> str:
            return f"event: {etype}\ndata: {_json.dumps(data)}\n\n"

        try:
            yield ev("log", {"msg": "Initialising sample preps...", "level": "dim"})
            ensure_sample_preps_table()

            yield ev("log", {"msg": "Fetching sample preps...", "level": "info"})
            all_preps = _list_preps(limit=500)
            preps = [p for p in all_preps if p.get("senaite_sample_id")]

            if not preps:
                yield ev("log", {"msg": "No preps with SENAITE IDs found.", "level": "warn"})
                yield ev("done", {"matches": []})
                return

            yield ev("log", {"msg": f"{len(preps)} prep(s) with SENAITE IDs to scan", "level": "info"})
            yield ev("progress", {"current": 0, "total": len(preps)})

            # List LIMS SharePoint root
            yield ev("log", {"msg": "Listing LIMS SharePoint root folder...", "level": "info"})
            try:
                root_items = await sp.list_lims_folder("")
            except Exception as sp_err:
                yield ev("error", {"msg": f"SharePoint error: {sp_err}"})
                return

            root_folders = [item for item in root_items if item["type"] == "folder"]
            yield ev("log", {"msg": f"{len(root_folders)} folder(s) found in LIMS root", "level": "dim"})

            matches = []
            for i, prep in enumerate(preps):
                sid = prep["senaite_sample_id"]
                yield ev("progress", {"current": i, "total": len(preps)})

                # Match folders prefixed by the sample ID (e.g. "P-0248 Peptide Name")
                matching = [
                    f for f in root_folders
                    if f["name"] == sid
                    or f["name"].startswith(sid + " ")
                    or f["name"].startswith(sid + "_")
                ]

                if not matching:
                    yield ev("log", {"msg": f"{sid}: no folder found", "level": "dim"})
                    continue

                folder = matching[0]
                yield ev("log", {"msg": f"{sid}: found '{folder['name']}', checking for CSVs...", "level": "info"})

                try:
                    all_csvs = await sp.list_files_recursive(
                        folder["name"], extensions=[".csv"], root="lims",
                    )
                except Exception:
                    yield ev("log", {"msg": f"{sid}: could not list folder contents", "level": "warn"})
                    continue

                peak_files = [
                    c for c in all_csvs
                    if "_PeakData" in c["name"] and c["name"].endswith(".csv")
                ]
                chrom_files = [
                    c for c in all_csvs
                    if c["name"].lower().endswith(".csv")
                    and "dx_dad1a" in c["name"].lower()
                ]

                if peak_files:
                    match_data = {
                        "prep_id": prep["id"],
                        "senaite_sample_id": sid,
                        "folder_name": folder["name"],
                        "folder_id": folder.get("id"),
                        "folder_web_url": folder.get("web_url"),
                        "peak_files": peak_files,
                        "chrom_files": chrom_files,
                    }
                    matches.append(match_data)
                    yield ev("match", match_data)
                    yield ev("log", {
                        "msg": f"{sid}: ✓ {len(peak_files)} PeakData, {len(chrom_files)} chromatogram file(s)",
                        "level": "success",
                    })
                else:
                    yield ev("log", {"msg": f"{sid}: folder found but no PeakData CSVs", "level": "warn"})

            yield ev("progress", {"current": len(preps), "total": len(preps)})
            total_matches = len(matches)
            yield ev("log", {
                "msg": f"Scan complete — {total_matches} match(es) found",
                "level": "success" if total_matches else "warn",
            })
            yield ev("done", {"matches": matches})

        except Exception as exc:
            yield ev("error", {"msg": f"Scan failed: {exc}"})

    return _SR(
        _generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/sample-preps/hplc-folder-match")
async def hplc_folder_match(
    folder_path: str,
    _current_user=Depends(get_current_user),
):
    """Manual HPLC-data override: run the scan's per-folder matching (recursive
    CSV listing + PeakData/chromatogram filters) against an arbitrary LIMS
    folder, so the bench can pin any folder's data to a prep — e.g. on test
    stacks where no folder matches the prep's sample id.

    Returns the same peak_files/chrom_files shapes the scan emits, plus the
    folder's id/web_url (resolved from the parent listing) so the flyout's
    chromatogram fallback keeps working.

    Must be defined before /sample-preps/{sample_prep_id} to avoid shadowing.
    """
    folder_path = folder_path.strip().strip("/")
    if not folder_path:
        raise HTTPException(status_code=400, detail="folder_path is required")
    try:
        all_csvs = await sp.list_files_recursive(
            folder_path, extensions=[".csv"], root="lims",
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"SharePoint error: {e}")

    # Same filters as the scan above — keep in lockstep.
    peak_files = [
        c for c in all_csvs
        if "_PeakData" in c["name"] and c["name"].endswith(".csv")
    ]
    chrom_files = [
        c for c in all_csvs
        if c["name"].lower().endswith(".csv") and "dx_dad1a" in c["name"].lower()
    ]

    # Resolve the folder's own item (id / web_url) from its parent's listing.
    folder_id = None
    folder_web_url = None
    parent_path, _, folder_name = folder_path.rpartition("/")
    try:
        siblings = await sp.list_lims_folder(parent_path)
        me = next(
            (i for i in siblings if i["type"] == "folder" and i["name"] == folder_name),
            None,
        )
        if me:
            folder_id = me.get("id")
            folder_web_url = me.get("web_url")
    except Exception:
        pass  # best-effort — the match works without id/web_url

    return {
        "folder_name": folder_name,
        "folder_path": folder_path,
        "folder_id": folder_id,
        "folder_web_url": folder_web_url,
        "peak_files": peak_files,
        "chrom_files": chrom_files,
    }


@app.get("/sample-preps/{sample_prep_id}")
async def get_sample_prep_endpoint(
    sample_prep_id: int,
    _current_user=Depends(get_current_user),
):
    """Fetch a single sample prep by id."""
    from mk1_db import ensure_sample_preps_table, get_sample_prep

    try:
        ensure_sample_preps_table()
        row = get_sample_prep(sample_prep_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Sample prep {sample_prep_id} not found")
        for k in ("created_at", "updated_at"):
            if row.get(k) and hasattr(row[k], "isoformat"):
                row[k] = row[k].isoformat()
        return row
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch sample prep: {e}")


@app.patch("/sample-preps/{sample_prep_id}")
async def update_sample_prep_endpoint(
    sample_prep_id: int,
    body: SamplePrepUpdateRequest,
    current_user=Depends(get_current_user),
):
    """Update fields on a sample prep."""
    from mk1_db import ensure_sample_preps_table, update_sample_prep

    try:
        ensure_sample_preps_table()
        update_data = body.model_dump(exclude_none=True)
        update_data["updated_by_user_id"] = current_user.id
        update_data["updated_by_email"] = current_user.email
        # Backfill created_by if missing (pre-tracking records)
        from mk1_db import get_sample_prep as _get_sp
        existing = _get_sp(sample_prep_id)
        if existing and not existing.get("created_by_user_id"):
            update_data["created_by_user_id"] = current_user.id
            update_data["created_by_email"] = current_user.email
        row = update_sample_prep(sample_prep_id, update_data)
        if not row:
            raise HTTPException(status_code=404, detail=f"Sample prep {sample_prep_id} not found")
        for k in ("created_at", "updated_at"):
            if row.get(k) and hasattr(row[k], "isoformat"):
                row[k] = row[k].isoformat()
        return row
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update sample prep: {e}")



@app.delete("/sample-preps/{sample_prep_id}", status_code=204)
async def delete_sample_prep_endpoint(
    sample_prep_id: int,
    _current_user=Depends(get_current_user),
):
    """Permanently delete a sample prep record."""
    from mk1_db import ensure_sample_preps_table, delete_sample_prep

    try:
        ensure_sample_preps_table()
        deleted = delete_sample_prep(sample_prep_id)
        if not deleted:
            raise HTTPException(status_code=404, detail=f"Sample prep {sample_prep_id} not found")
        return Response(status_code=204)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete sample prep: {e}")


# ─── SENAITE Integration (Phase 5) ────────────────────────────────────────────

# -- SENAITE Integration -----------------------------------------------
SENAITE_URL = os.environ.get("SENAITE_URL")          # None = disabled
SENAITE_USER = os.environ.get("SENAITE_USER", "")
SENAITE_PASSWORD = os.environ.get("SENAITE_PASSWORD", "")
SENAITE_TIMEOUT = httpx.Timeout(30.0, connect=10.0)
# Host used when constructing browser-facing SENAITE deep-links returned to
# the frontend.  Each dev stack publishes SENAITE on a different port so this
# lets the stack override the host without touching SENAITE_URL (which is the
# internal container-to-container address).  Falls back to None (relative path
# only) when unset, preserving previous behaviour.
SENAITE_PUBLIC_URL = os.environ.get("SENAITE_PUBLIC_URL", "").rstrip("/") or None

# ── Senaite per-user credential helpers ──────────────────────────────────────

import base64
import hashlib
from cryptography.fernet import Fernet
from auth import SECRET_KEY as _JWT_SECRET

def _get_fernet() -> Fernet:
    """Derive a Fernet key from JWT_SECRET for encrypting Senaite passwords."""
    key_bytes = hashlib.sha256(_JWT_SECRET.encode()).digest()
    return Fernet(base64.urlsafe_b64encode(key_bytes))

def _encrypt_senaite_password(plain: str) -> str:
    return _get_fernet().encrypt(plain.encode()).decode()

def _decrypt_senaite_password(encrypted: str) -> str:
    return _get_fernet().decrypt(encrypted.encode()).decode()

def _get_senaite_auth(user=None) -> httpx.BasicAuth:
    """Return BasicAuth for Senaite — user's own credentials if available, else admin."""
    if user and getattr(user, "senaite_password_encrypted", None):
        try:
            pwd = _decrypt_senaite_password(user.senaite_password_encrypted)
            return httpx.BasicAuth(user.email, pwd)
        except Exception:
            pass  # Fall through to admin
    return httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD)

def _user_to_read(user) -> UserRead:
    """Convert User model to UserRead schema with senaite_configured."""
    return UserRead(
        id=user.id,
        email=user.email,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
        senaite_configured=user.senaite_password_encrypted is not None,
        first_name=user.first_name,
        last_name=user.last_name,
    )


class SenaiteAnalyte(BaseModel):
    raw_name: str
    slot_number: int  # 1-4, corresponding to Analyte1..Analyte4 in SENAITE
    matched_peptide_id: Optional[int] = None
    matched_peptide_name: Optional[str] = None
    declared_quantity: Optional[float] = None  # per-analyte declared qty (mg)


class SenaiteCOAInfo(BaseModel):
    company_logo_url: Optional[str] = None
    chromatograph_background_url: Optional[str] = None
    company_name: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    verification_code: Optional[str] = None


class SenaiteRemark(BaseModel):
    content: str  # HTML string from SENAITE
    user_id: Optional[str] = None
    created: Optional[str] = None


class SenaiteAnalysis(BaseModel):
    uid: Optional[str] = None
    keyword: Optional[str] = None
    title: str
    result: Optional[str] = None
    result_options: list[dict] = []  # [{value: str, label: str}] for selection-type analyses
    unit: Optional[str] = None
    method: Optional[str] = None
    method_uid: Optional[str] = None  # UID for editing
    method_options: list[dict] = []   # [{uid: str, title: str}] allowed methods for this analysis
    instrument: Optional[str] = None
    instrument_uid: Optional[str] = None  # UID for editing
    instrument_options: list[dict] = []   # [{uid: str, title: str}] allowed instruments for this analysis
    analyst: Optional[str] = None
    due_date: Optional[str] = None
    review_state: Optional[str] = None
    sort_key: Optional[float] = None
    captured: Optional[str] = None
    retested: bool = False
    # Mk1-local enrichment: which service_group this analysis belongs to
    # (resolved from analysis_services table by keyword). Drives the
    # per-vial "primary analysis" highlight on the sample detail page.
    service_group_id: Optional[int] = None
    service_group_name: Optional[str] = None


class SenaiteAttachment(BaseModel):
    uid: str
    filename: str
    content_type: Optional[str] = None
    attachment_type: Optional[str] = None  # e.g. "Sample Image", "HPLC Graph"
    download_url: Optional[str] = None  # proxied through our backend


# Cache SENAITE download URLs for attachment proxy (uid -> {download_url, content_type, filename})
_attachment_download_cache: dict[str, dict[str, str]] = {}

# Cache SENAITE download URLs for ARReport PDF proxy (uid -> download_url)
_report_download_cache: dict[str, str] = {}


class SenaitePublishedCOA(BaseModel):
    report_uid: str
    filename: str
    file_size_bytes: Optional[int] = None
    published_date: Optional[str] = None
    published_by: Optional[str] = None
    download_url: str  # proxied through our backend


class SenaiteLookupResult(BaseModel):
    sample_id: str
    sample_uid: Optional[str] = None
    client: Optional[str] = None
    contact: Optional[str] = None
    sample_type: Optional[str] = None
    date_received: Optional[str] = None
    date_sampled: Optional[str] = None
    profiles: list[str] = []
    client_order_number: Optional[str] = None
    client_sample_id: Optional[str] = None
    client_lot: Optional[str] = None
    review_state: Optional[str] = None
    declared_weight_mg: Optional[float] = None
    analytes: list[SenaiteAnalyte]
    coa: SenaiteCOAInfo = SenaiteCOAInfo()
    remarks: list[SenaiteRemark] = []
    analyses: list[SenaiteAnalysis] = []
    attachments: list[SenaiteAttachment] = []
    published_coa: Optional[SenaitePublishedCOA] = None
    senaite_url: Optional[str] = None  # e.g. "/clients/client-8/PB-0057"
    cached_at: Optional[str] = None  # ISO timestamp when this result was cached


class RegistrySampleReadResult(SenaiteLookupResult):
    """SenaiteLookupResult with basic-info overlaid from the Accu-Mk1 registry.
    field_sources records, per overlay field, whether the value shown came from
    the registry ('mk1') or fell back to SENAITE ('senaite')."""
    read_source: str = "mk1"
    registry_missing: bool = False
    field_sources: dict[str, str] = {}


class SenaiteStatusResponse(BaseModel):
    enabled: bool


def _senaite_path(item: dict) -> Optional[str]:
    """Build the browser-facing SENAITE URL for a sample item.

    Extracts the SENAITE-relative path (e.g. '/clients/client-8/PB-0057') and,
    when SENAITE_PUBLIC_URL is configured, prepends it to form an absolute URL
    (e.g. 'http://localhost:5538/clients/client-8/PB-0057').  When
    SENAITE_PUBLIC_URL is not set the relative path is returned unchanged —
    the frontend is expected to prepend its own senaiteBaseUrl in that case.
    """
    raw = item.get("path") or ""
    if raw.startswith("/senaite/"):
        path = raw[len("/senaite"):]  # strip '/senaite' prefix, keep leading slash
    else:
        path = raw or None
    if SENAITE_PUBLIC_URL and path:
        return f"{SENAITE_PUBLIC_URL}{path}"
    return path


def _strip_method_suffix(name: str) -> str:
    """Strip trailing ' - Method (Type)' suffixes from SENAITE analyte names.

    Example: 'BPC-157 - Identity (HPLC)' -> 'BPC-157'
    """
    import re
    return re.sub(r'\s*-\s*[^-]+\([^)]+\)\s*$', '', name).strip()


def _fuzzy_match_peptide(stripped_name: str, peptides: list) -> Optional[tuple]:
    """Case-insensitive match of stripped analyte name against local peptides.

    Priority order to avoid false positives on blend names:
      1. Exact normalized match (ignoring hyphens and spaces)
      2. Substring match — only against non-blend peptides (no '+' in name)
      3. Abbreviation exact match

    Returns (peptide.id, peptide.name) if a match is found, else None.
    """
    needle = stripped_name.lower()
    needle_norm = needle.replace("-", "").replace(" ", "")

    # Pass 1: exact normalized match (handles BPC-157 ↔ BPC157 etc.)
    for peptide in peptides:
        hay_norm = peptide.name.lower().replace("-", "").replace(" ", "")
        if needle_norm == hay_norm:
            return (peptide.id, peptide.name)

    # Pass 2: substring match — skip blend names (containing '+') to prevent
    # "Semaglutide" matching "Cagrilinitide + Semaglutide"
    for peptide in peptides:
        if "+" in peptide.name:
            continue
        hay = peptide.name.lower()
        hay_norm = hay.replace("-", "").replace(" ", "")
        if needle in hay or needle_norm in hay_norm:
            return (peptide.id, peptide.name)

    # Pass 3: abbreviation exact match
    for peptide in peptides:
        if peptide.abbreviation and needle == peptide.abbreviation.lower():
            return (peptide.id, peptide.name)

    return None


async def _resolve_instrument_from_senaite(sample_id: str) -> Optional[str]:
    """
    Look up the instrument model for a sample via SENAITE.

    Strategy:
      1. Query the search endpoint for Analysis objects by getRequestID
      2. Check each analysis for instrument title containing "1290" or "1260"

    Returns:
        "1290", "1260", or None if SENAITE is unavailable or no instrument found.
    """
    if SENAITE_URL is None:
        print(f"[INST] SENAITE_URL is None — skipping instrument lookup for {sample_id}")
        return None

    sid = sample_id.strip().upper()
    try:
        # Use the search endpoint (proven to work for catalog queries)
        search_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/search"
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
        ) as client:
            resp = await client.get(search_url, params={
                "portal_type": "Analysis",
                "getRequestID": sid,
                "limit": "20",
            })
            resp.raise_for_status()
            items = resp.json().get("items", [])
            print(f"[INST] {sid}: search returned {len(items)} analyses")

            for an in items:
                # Try getInstrumentTitle first (catalog metadata)
                title = an.get("getInstrumentTitle") or ""
                # Fall back to nested Instrument object
                if not title:
                    inst_obj = an.get("Instrument")
                    if isinstance(inst_obj, dict):
                        title = inst_obj.get("title") or inst_obj.get("Title") or ""
                # Fall back to InstrumentTitle string field
                if not title:
                    title = an.get("InstrumentTitle") or ""
                if title:
                    print(f"[INST] {sid}: found instrument title '{title}'")
                    if "1290" in title:
                        return "1290"
                    if "1260" in title:
                        return "1260"

            # If search returned analyses but none had instruments, log it
            if items:
                print(f"[INST] {sid}: {len(items)} analyses found but none had instrument titles")
            else:
                print(f"[INST] {sid}: no analyses found via search endpoint")

    except Exception as e:
        print(f"[INST] {sid}: SENAITE query failed: {e}")
    return None


async def _fetch_senaite_sample(sample_id: str) -> dict:
    """Fetch a sample from SENAITE by ID using the AnalysisRequest API.

    Calls GET {SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest?id={id}&complete=yes
    with HTTP Basic auth. Returns the full parsed JSON response dict.
    Raises httpx exceptions on network/HTTP errors.
    """
    url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
    print(f"[INFO] _fetch_senaite_sample: GET {url}?id={sample_id}&complete=yes")
    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
        timeout=SENAITE_TIMEOUT,
        auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
    ) as client:
        resp = await client.get(url, params={"id": sample_id, "complete": "yes"})
        print(f"[INFO] _fetch_senaite_sample: status={resp.status_code}")
        resp.raise_for_status()
        return resp.json()


@app.get("/wizard/senaite/status", response_model=SenaiteStatusResponse)
async def get_senaite_status(_current_user=Depends(get_current_user)):
    """Return whether SENAITE integration is enabled (SENAITE_URL env var is set)."""
    return SenaiteStatusResponse(enabled=SENAITE_URL is not None)


@app.get("/wizard/senaite/raw-fields/{sample_id}")
async def get_senaite_raw_fields(
    sample_id: str,
    _current_user=Depends(get_current_user),
):
    """
    Return the raw SENAITE API fields for a sample — useful for diagnosing what
    Analyte1Peptide, SampleType, Profiles, etc. actually contain.
    """
    if SENAITE_URL is None:
        raise HTTPException(status_code=503, detail="SENAITE not configured")
    sample_id = sample_id.strip().upper()
    data = await _fetch_senaite_sample(sample_id)
    if data.get("count", 0) == 0:
        raise HTTPException(status_code=404, detail=f"Sample {sample_id} not found")
    item = data["items"][0]
    keys = [
        "id", "title", "SampleType", "getSampleTypeTitle",
        "Analyte1Peptide", "Analyte2Peptide", "Analyte3Peptide", "Analyte4Peptide",
        "Analyte1DeclaredQuantity", "Analyte2DeclaredQuantity",
        "Analyte3DeclaredQuantity", "Analyte4DeclaredQuantity",
        "DeclaredTotalQuantity", "getProfilesTitleStr", "ProfilesTitleStr",
        "review_state", "getClientTitle",
    ]
    return {k: item.get(k) for k in keys}


# ── Senaite lookup cache (shared across all users) ─────────────────
_senaite_lookup_cache: dict[str, tuple[float, SenaiteLookupResult]] = {}  # id → (timestamp, result)
_SENAITE_LOOKUP_TTL = 15 * 60  # 15 minutes


@app.delete("/wizard/senaite/lookup-cache")
async def clear_senaite_lookup_cache(
    _current_user=Depends(get_current_user),
):
    """Clear the server-side Senaite lookup cache so next lookups fetch fresh data."""
    count = len(_senaite_lookup_cache)
    _senaite_lookup_cache.clear()
    return {"cleared": count}


@app.get("/wizard/senaite/lookup", response_model=SenaiteLookupResult)
async def lookup_senaite_sample(
    id: str,
    no_cache: bool = True,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Look up a sample in SENAITE by ID and return structured analyte data.

    By default (no_cache=true) always fetches fresh from SENAITE.
    Pass ?no_cache=false to use the 15-minute server-side cache
    (only the Order Status page should do this, to avoid hammering Zope).

    Returns:
        SenaiteLookupResult with sample_id, declared_weight_mg, and analytes list.

    Raises:
        503 if SENAITE is not configured or is unreachable/timed out.
        404 if the sample ID does not exist in SENAITE.
    """
    if SENAITE_URL is None:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    # SENAITE sample IDs are always uppercase (e.g. PB-0056) — normalize
    id = id.strip().upper()

    # Check server-side cache (skipped when no_cache=true)
    import time as _time
    if not no_cache:
        cached = _senaite_lookup_cache.get(id)
        if cached:
            ts, result = cached
            if _time.time() - ts < _SENAITE_LOOKUP_TTL:
                return result

    try:
        data = await _fetch_senaite_sample(id)

        if data.get("count", 0) == 0:
            # Distinguish "sample not found" from "credentials/permissions failure".
            # If a sanity query (no ID filter) also returns 0, SENAITE auth is broken.
            url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            ) as client:
                sanity = await client.get(url, params={"limit": 1})
                sanity.raise_for_status()
                if sanity.json().get("count", 0) == 0:
                    raise HTTPException(
                        status_code=503,
                        detail="SENAITE is currently unavailable \u2014 use manual entry",
                    )
            raise HTTPException(status_code=404, detail=f"Sample {id} not found in SENAITE")

        item = data["items"][0]

        # Parse sample_id
        sample_id = item["id"]

        # Parse declared_weight_mg — DeclaredTotalQuantity is a decimal string or null
        declared_weight_mg: Optional[float] = None
        raw_qty = item.get("DeclaredTotalQuantity")
        if raw_qty is not None and str(raw_qty).strip() != "":
            try:
                declared_weight_mg = float(raw_qty)
            except (ValueError, TypeError):
                declared_weight_mg = None

        # Parse analytes from Analyte1Peptide through Analyte4Peptide
        all_peptides = db.query(Peptide).all()
        analytes: list[SenaiteAnalyte] = []
        for slot, key in enumerate(("Analyte1Peptide", "Analyte2Peptide", "Analyte3Peptide", "Analyte4Peptide"), start=1):
            raw_name = item.get(key)
            if raw_name is None or str(raw_name).strip() == "":
                continue
            stripped = _strip_method_suffix(str(raw_name))
            match = _fuzzy_match_peptide(stripped, all_peptides)
            # Parse per-analyte declared quantity
            qty_key = f"Analyte{slot}DeclaredQuantity"
            raw_analyte_qty = item.get(qty_key)
            analyte_declared_qty = None
            if raw_analyte_qty is not None and str(raw_analyte_qty).strip() != "":
                try:
                    analyte_declared_qty = float(raw_analyte_qty)
                except (ValueError, TypeError):
                    analyte_declared_qty = None
            analytes.append(SenaiteAnalyte(
                raw_name=raw_name,
                slot_number=slot,
                matched_peptide_id=match[0] if match else None,
                matched_peptide_name=match[1] if match else None,
                declared_quantity=analyte_declared_qty,
            ))

        # Parse profiles
        profiles: list[str] = []
        profiles_str = item.get("getProfilesTitleStr") or item.get("ProfilesTitleStr")
        if profiles_str:
            profiles = [p.strip() for p in str(profiles_str).split(",") if p.strip()]

        # Resolve image URLs — prepend WordPress host for relative paths
        def resolve_wp_url(raw: str | None) -> str | None:
            if not raw:
                return None
            if raw.startswith("http://") or raw.startswith("https://"):
                return raw
            return get_wordpress_host().rstrip("/") + "/" + raw.lstrip("/")

        # Parse COA info
        coa = SenaiteCOAInfo(
            company_logo_url=resolve_wp_url(item.get("CompanyLogoUrl")),
            chromatograph_background_url=resolve_wp_url(item.get("ChromatographBackgroundUrl")),
            company_name=item.get("CoaCompanyName") or None,
            email=item.get("CoaEmail") or None,
            website=item.get("CoaWebsite") or None,
            address=item.get("CoaAddress") or None,
            verification_code=item.get("VerificationCode") or None,
        )

        # Parse remarks (list of {content, user_id, created, ...})
        senaite_remarks: list[SenaiteRemark] = []
        raw_remarks = item.get("Remarks")
        if isinstance(raw_remarks, list):
            for r in raw_remarks:
                if isinstance(r, dict) and r.get("content"):
                    senaite_remarks.append(SenaiteRemark(
                        content=r["content"],
                        user_id=r.get("user_id") or None,
                        created=r.get("created") or None,
                    ))

        # Fetch analyses (lab test results) for this sample.
        # Use getRequestID (the sample ID string) — this is the only reliable
        # filter on SENAITE's Analysis endpoint (getRequestUID is ignored).
        sample_uid = item.get("uid") or item.get("UID") or ""
        senaite_analyses: list[SenaiteAnalysis] = []
        try:
            analysis_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/Analysis"
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            ) as client:
                an_resp = await client.get(analysis_url, params={
                    "getRequestID": sample_id,
                    "complete": "yes",
                    "limit": "100",
                })
                an_resp.raise_for_status()
                an_data = an_resp.json()

                _inst_uid_to_indices: dict[str, list[int]] = {}

                _svc_uid_to_indices: dict[str, list[int]] = {}

                # Task 7 passive drift observer: a parallel raw-line capture
                # (NOT the SenaiteAnalysis pydantic objects below, which
                # collapse RetestOf to a bare bool). This endpoint returns
                # EVERY analysis line including retest-superseded ones (no
                # review_state filter) — same shape/reason
                # sub_samples/senaite.py's fetch_parent_analyses documents —
                # so `select_current_lines` (same one the backfill and the
                # registry-debug hook use) reduces this to one current line
                # per keyword before anything is handed to the observer.
                _observer_raw_lines: list[dict] = []

                for an_item in an_data.get("items", []):
                    # Result: prefer formatted string for selection-type results,
                    # then fall back to raw numeric Result
                    raw_result = an_item.get("Result") or an_item.get("getResult") or an_item.get("result")
                    result_str: Optional[str] = None
                    # Build result_options for selection-type analyses
                    raw_options = an_item.get("ResultOptions") or an_item.get("getResultOptions") or []
                    parsed_options: list[dict] = []
                    if raw_options and isinstance(raw_options, list):
                        for opt in raw_options:
                            if isinstance(opt, dict) and opt.get("ResultValue") is not None:
                                parsed_options.append({
                                    "value": str(opt["ResultValue"]),
                                    "label": str(opt.get("ResultText", opt["ResultValue"])),
                                })
                    if raw_result not in (None, ""):
                        # For selection-type analyses, store the raw numeric value so
                        # it can round-trip back to SENAITE correctly on save.
                        # The frontend maps value → label for display.
                        result_str = str(raw_result)

                    # Sort key: numeric priority for display ordering
                    raw_sort_key = an_item.get("SortKey") or an_item.get("getSortKey")
                    sort_key_val: Optional[float] = None
                    if raw_sort_key is not None:
                        try:
                            sort_key_val = float(raw_sort_key)
                        except (ValueError, TypeError):
                            sort_key_val = None

                    # Captured: when the result was entered
                    captured = an_item.get("ResultCaptureDate") or an_item.get("getResultCaptureDate") or None

                    # Retested: RetestOf is a dict — non-empty means this IS a retest
                    retest_of = an_item.get("RetestOf") or {}
                    retested_val = bool(retest_of) if isinstance(retest_of, dict) else False

                    # Method: getMethodTitle is a string, Method is an object ref
                    method_title = an_item.get("getMethodTitle") or None
                    method_uid_val = None
                    method_obj = an_item.get("Method")
                    if isinstance(method_obj, dict):
                        method_uid_val = method_obj.get("uid") or None
                        if not method_title and method_obj.get("title"):
                            method_title = method_obj["title"]

                    # Instrument: try getInstrumentTitle first, then Instrument object ref
                    instrument_title = an_item.get("getInstrumentTitle") or None
                    instrument_uid_val = None
                    instrument_uid = None  # used for deferred title resolution
                    instrument_obj = an_item.get("Instrument")
                    if isinstance(instrument_obj, dict):
                        instrument_uid_val = instrument_obj.get("uid") or None
                        if not instrument_title:
                            instrument_title = instrument_obj.get("title") or instrument_obj.get("Title") or None
                            if not instrument_title and instrument_uid_val:
                                instrument_uid = instrument_uid_val

                    # Analyst: Analyst field is often None; getSubmittedBy has the user
                    analyst = an_item.get("Analyst") or an_item.get("getSubmittedBy") or None

                    # SENAITE shows "Manual" for submitted analyses with no method/instrument
                    an_review_state = an_item.get("review_state") or ""
                    has_result = an_review_state in ("verified", "published", "to_be_verified")
                    if not method_title and has_result:
                        method_title = "Manual"
                    # Delay instrument "Manual" fallback — resolve UIDs first
                    if not instrument_title and not instrument_uid and has_result:
                        instrument_title = "Manual"

                    senaite_analyses.append(SenaiteAnalysis(
                        uid=an_item.get("uid") or an_item.get("UID") or None,
                        keyword=an_item.get("Keyword") or an_item.get("getKeyword") or None,
                        title=an_item.get("title") or an_item.get("Title") or str(an_item.get("id", "")),
                        result=result_str,
                        result_options=parsed_options,
                        unit=an_item.get("Unit") or an_item.get("getUnit") or None,
                        method=method_title,
                        method_uid=method_uid_val,
                        instrument=instrument_title,
                        instrument_uid=instrument_uid_val,
                        analyst=analyst,
                        due_date=an_item.get("getDueDate") or an_item.get("DueDate") or None,
                        review_state=an_item.get("review_state") or None,
                        sort_key=sort_key_val,
                        captured=str(captured) if captured else None,
                        retested=retested_val,
                    ))
                    # Task 7 raw-line capture (see comment above the list init):
                    # uid/retest_of_uid/created mirror sub_samples/senaite.py's
                    # fetch_parent_analyses projection exactly, so the same
                    # `select_current_lines` reduces this list too.
                    _observer_raw_lines.append({
                        "uid": an_item.get("uid") or an_item.get("UID") or None,
                        "keyword": an_item.get("Keyword") or an_item.get("getKeyword") or None,
                        "review_state": an_item.get("review_state") or None,
                        "result": result_str,
                        "retest_of_uid": (
                            an_item.get("getRetestOfUID")
                            or (an_item.get("RetestOf") or {}).get("uid")
                            or None
                        ),
                        "created": (
                            an_item.get("created") or an_item.get("creation_date")
                            or an_item.get("DateCreated") or an_item.get("getDateCreated")
                        ),
                    })
                    # Track indices that need instrument UID resolution
                    if instrument_uid:
                        if instrument_uid not in _inst_uid_to_indices:
                            _inst_uid_to_indices[instrument_uid] = []
                        _inst_uid_to_indices[instrument_uid].append(len(senaite_analyses) - 1)
                    # Track service UID → analysis indices for per-analysis method/instrument options
                    svc_uid = an_item.get("getServiceUID") or (an_item.get("AnalysisService") or {}).get("uid") or None
                    if svc_uid:
                        if svc_uid not in _svc_uid_to_indices:
                            _svc_uid_to_indices[svc_uid] = []
                        _svc_uid_to_indices[svc_uid].append(len(senaite_analyses) - 1)
            # Resolve instrument UIDs → titles via batch API call
            if _inst_uid_to_indices:
                try:
                    inst_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/Instrument"
                    uid_filter = "|".join(_inst_uid_to_indices.keys())
                    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                        timeout=SENAITE_TIMEOUT,
                        auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
                    ) as inst_client:
                        inst_resp = await inst_client.get(inst_url, params={
                            "UID": uid_filter,
                            "limit": "50",
                        })
                        inst_resp.raise_for_status()
                        inst_data = inst_resp.json()
                        uid_to_title: dict[str, str] = {}
                        for inst_item in inst_data.get("items", []):
                            uid = inst_item.get("uid") or inst_item.get("UID") or ""
                            title = inst_item.get("title") or inst_item.get("Title") or ""
                            if uid and title:
                                uid_to_title[uid] = title
                        # Apply resolved titles to analyses
                        for uid, indices in _inst_uid_to_indices.items():
                            resolved = uid_to_title.get(uid)
                            for idx in indices:
                                if resolved:
                                    senaite_analyses[idx].instrument = resolved
                                elif not senaite_analyses[idx].instrument:
                                    senaite_analyses[idx].instrument = "Manual"
                except Exception as inst_exc:
                    print(f"[WARN] Failed to resolve instrument UIDs: {inst_exc}")
                    # Fall back to "Manual" for unresolved
                    for uid, indices in _inst_uid_to_indices.items():
                        for idx in indices:
                            if not senaite_analyses[idx].instrument:
                                senaite_analyses[idx].instrument = "Manual"

            # Fetch AnalysisServices individually by UID to get per-analysis allowed methods/instruments
            if _svc_uid_to_indices:
                async def _fetch_one_service(client: httpx.AsyncClient, svc_uid: str) -> tuple[str, dict]:
                    url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/analysisservice/{svc_uid}"
                    resp = await client.get(url)
                    resp.raise_for_status()
                    data = resp.json()
                    items = data.get("items")
                    svc_item = items[0] if items else data
                    return svc_uid, svc_item

                try:
                    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                        timeout=SENAITE_TIMEOUT,
                        auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
                    ) as svc_client:
                        svc_results = await asyncio.gather(
                            *[_fetch_one_service(svc_client, uid) for uid in _svc_uid_to_indices],
                            return_exceptions=True,
                        )

                    # Collect UIDs per service (titles not included by SENAITE in nested objects)
                    # svc_uid → (method_uids, instrument_uids)
                    _svc_method_uids: dict[str, list[str]] = {}
                    _svc_instr_uids: dict[str, list[str]] = {}
                    for result in svc_results:
                        if isinstance(result, Exception):
                            print(f"[WARN] Failed to fetch one AnalysisService: {result}")
                            continue
                        svc_uid, svc_item = result
                        if svc_uid not in _svc_uid_to_indices:
                            continue
                        m_list = svc_item.get("Methods")
                        m_uids = [m["uid"] for m in m_list if isinstance(m, dict) and m.get("uid")] if isinstance(m_list, list) else []
                        i_list = svc_item.get("Instruments")
                        i_uids = [i["uid"] for i in i_list if isinstance(i, dict) and i.get("uid")] if isinstance(i_list, list) else []
                        _svc_method_uids[svc_uid] = m_uids
                        _svc_instr_uids[svc_uid] = i_uids

                    # Batch-resolve method titles
                    all_m_uids = list({uid for uids in _svc_method_uids.values() for uid in uids})
                    all_i_uids = list({uid for uids in _svc_instr_uids.values() for uid in uids})
                    method_uid_to_title: dict[str, str] = {}
                    instr_uid_to_title: dict[str, str] = {}

                    async def _fetch_title(client: httpx.AsyncClient, kind: str, uid: str) -> tuple[str, str, str]:
                        # kind = "method" or "instrument"
                        resp = await client.get(f"{SENAITE_URL}/senaite/@@API/senaite/v1/{kind}/{uid}")
                        resp.raise_for_status()
                        data = resp.json()
                        items = data.get("items")
                        obj = items[0] if items else data
                        title = obj.get("title") or obj.get("Title") or ""
                        return kind, uid, title

                    async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                        timeout=SENAITE_TIMEOUT,
                        auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
                    ) as title_client:
                        fetch_tasks = (
                            [_fetch_title(title_client, "method", u) for u in all_m_uids] +
                            [_fetch_title(title_client, "instrument", u) for u in all_i_uids]
                        )
                        title_results = await asyncio.gather(*fetch_tasks, return_exceptions=True)
                        for tr in title_results:
                            if isinstance(tr, Exception):
                                print(f"[WARN] Failed to fetch title: {tr}")
                                continue
                            kind, uid, title = tr
                            if uid and title:
                                if kind == "method":
                                    method_uid_to_title[uid] = title
                                else:
                                    instr_uid_to_title[uid] = title
                    # Apply resolved titles to each analysis
                    for svc_uid, m_uids in _svc_method_uids.items():
                        parsed_methods = [
                            {"uid": u, "title": method_uid_to_title[u]}
                            for u in m_uids if u in method_uid_to_title
                        ]
                        parsed_instruments = [
                            {"uid": u, "title": instr_uid_to_title[u]}
                            for u in _svc_instr_uids.get(svc_uid, []) if u in instr_uid_to_title
                        ]
                        for idx in _svc_uid_to_indices[svc_uid]:
                            senaite_analyses[idx].method_options = parsed_methods
                            senaite_analyses[idx].instrument_options = parsed_instruments
                except Exception as svc_exc:
                    print(f"[WARN] Failed to fetch AnalysisService options: {svc_exc}")

            # Sort by sort_key, then title, then non-retested first to match SENAITE
            senaite_analyses.sort(key=lambda a: (
                a.sort_key if a.sort_key is not None else float("inf"),
                a.title.lower(),
                a.retested,  # False (0) before True (1)
            ))
            print(f"[INFO] Fetched {len(senaite_analyses)} analyses for sample {sample_id}")

            # Passive drift observer (Task 7): schedule ONLY on this success
            # path (never in the except branch below) — zero additional
            # SENAITE load. Deliberately NOT the SenaiteAnalysis objects
            # above (`senaite_analyses`) — this endpoint returns every
            # analysis line including retest-superseded ones, and the
            # observer has no dedup of its own, so feeding it a raw
            # multi-line-per-keyword list can heal a shadow to a stale
            # (superseded) state. `select_current_lines` (the same reducer
            # the backfill script and the registry-debug hook use) collapses
            # `_observer_raw_lines` to one current line per keyword first.
            # Own session, never raises; scheduled via run_in_threadpool
            # since this route is `async def`.
            from fastapi.concurrency import run_in_threadpool
            from lims_analyses.parent_mirror import select_current_lines
            _observer_current = select_current_lines(_observer_raw_lines)
            await run_in_threadpool(
                _observe_parent_analyses_bg,
                sample_id=sample_id,
                observed=[
                    {"keyword": kw, "review_state": ln.get("review_state"), "result": ln.get("result")}
                    for kw, ln in _observer_current.items()
                ],
            )
        except Exception as exc:
            print(f"[WARN] Failed to fetch analyses for {sample_id}: {exc}")

        # Fetch sample-level attachments
        senaite_attachments: list[SenaiteAttachment] = []
        try:
            raw_attachments = item.get("Attachment") or []
            if isinstance(raw_attachments, list) and raw_attachments:
                async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                    timeout=SENAITE_TIMEOUT,
                    auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
                ) as att_client:
                    for att_ref in raw_attachments:
                        att_api_url = att_ref.get("api_url")
                        att_uid = att_ref.get("uid")
                        if not att_api_url or not att_uid:
                            continue
                        try:
                            att_resp = await att_client.get(att_api_url)
                            att_resp.raise_for_status()
                            att_data = att_resp.json()
                            att_item = att_data["items"][0] if "items" in att_data and att_data["items"] else att_data
                            att_file = att_item.get("AttachmentFile") or {}
                            filename = att_file.get("filename") or ""
                            content_type = att_file.get("content_type") or ""
                            att_type_title = att_item.get("AttachmentType") or att_item.get("getAttachmentType") or None
                            if isinstance(att_type_title, dict):
                                att_type_title = att_type_title.get("title") or att_type_title.get("Title") or None
                            # Cache the SENAITE download URL for the proxy endpoint
                            senaite_dl_url = att_file.get("download") or ""
                            if senaite_dl_url:
                                _attachment_download_cache[att_uid] = {
                                    "download_url": senaite_dl_url,
                                    "content_type": content_type or "application/octet-stream",
                                    "filename": filename or "attachment",
                                }
                            senaite_attachments.append(SenaiteAttachment(
                                uid=att_uid,
                                filename=filename,
                                content_type=content_type,
                                attachment_type=att_type_title,
                                download_url=f"/wizard/senaite/attachment/{att_uid}",
                            ))
                        except Exception as att_exc:
                            print(f"[WARN] Failed to fetch attachment {att_uid}: {att_exc}")
                print(f"[INFO] Fetched {len(senaite_attachments)} attachments for sample {sample_id}")
        except Exception as exc:
            print(f"[WARN] Failed to fetch attachments for {sample_id}: {exc}")

        # Fetch published COA ARReport (PDF attached by Accumark COA Builder or SENAITE).
        # Use path-constrained catalog search to only fetch ARReports under this specific
        # sample, avoiding global scans that break on stale catalog entries (e.g. P-0028).
        published_coa_report: Optional[SenaitePublishedCOA] = None
        sample_path = item.get("path") or ""  # e.g. /senaite/clients/client-8/PB-0061
        try:
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            ) as report_client:
                sample_reports = []
                # Strategy 1: Path-constrained catalog search (avoids stale brains)
                search_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/search"
                try:
                    r_resp = await report_client.get(search_url, params={
                        "portal_type": "ARReport",
                        "path": sample_path,
                        "depth": "1",
                        "complete": "yes",
                        "limit": "10",
                    })
                    r_resp.raise_for_status()
                    r_data = r_resp.json()
                    sample_reports = r_data.get("items", [])
                except Exception as search_exc:
                    print(f"[WARN] ARReport search failed, trying direct traversal: {search_exc}")
                    # Strategy 2: Direct traversal into sample folder
                    # The sample path is e.g. /senaite/clients/client-8/P-0233
                    # Strip leading /senaite to get the API-relative path
                    rel_path = sample_path.lstrip("/")
                    if rel_path.startswith("senaite/"):
                        rel_path = rel_path[len("senaite/"):]
                    traverse_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/{rel_path}"
                    try:
                        t_resp = await report_client.get(traverse_url, params={"complete": "yes"})
                        t_resp.raise_for_status()
                        t_data = t_resp.json()
                        # Check if the response contains child items of type ARReport
                        for t_item in t_data.get("items", []):
                            if t_item.get("portal_type") == "ARReport":
                                sample_reports.append(t_item)
                    except Exception as trav_exc:
                        print(f"[WARN] ARReport traversal also failed: {trav_exc}")
                if sample_reports:
                    # Pick the most recently created one
                    sample_reports.sort(key=lambda r: r.get("created", ""), reverse=True)
                    r = sample_reports[0]
                    r_uid = r.get("uid") or r.get("UID") or ""
                    pdf_info = r.get("Pdf") or {}
                    pdf_filename = (pdf_info.get("filename") if isinstance(pdf_info, dict) else None) or f"{sample_id}_COA.pdf"
                    pub_date = r.get("created") or None
                    pub_by = r.get("Creator") or r.get("creator") or None
                    pdf_dl_url = pdf_info.get("download") if isinstance(pdf_info, dict) else None
                    if r_uid and pdf_dl_url:
                        _report_download_cache[r_uid] = pdf_dl_url
                    # Get file size via streaming GET headers (HEAD returns wrong Content-Length)
                    pdf_size: Optional[int] = None
                    if pdf_dl_url:
                        try:
                            async with report_client.stream("GET", pdf_dl_url) as size_resp:
                                cl = size_resp.headers.get("content-length")
                                if cl and cl.isdigit():
                                    pdf_size = int(cl)
                        except Exception:
                            pass
                    if r_uid:
                        published_coa_report = SenaitePublishedCOA(
                            report_uid=r_uid,
                            filename=pdf_filename,
                            file_size_bytes=pdf_size,
                            published_date=str(pub_date) if pub_date else None,
                            published_by=str(pub_by) if pub_by else None,
                            download_url=f"/wizard/senaite/report/{r_uid}",
                        )
            print(f"[INFO] ARReport for {sample_id} (path={sample_path}): {'found' if published_coa_report else 'none'}")
        except Exception as exc:
            print(f"[WARN] Failed to fetch ARReport for {sample_id}: {exc}")

        # Enrich each analysis with its service_group_id (Mk1-local, joined by
        # keyword from the analysis_services table). Drives the "primary
        # analysis for this vial" highlight on the sample detail page based on
        # the sample's assignment_role.
        try:
            keywords_seen = {a.keyword for a in senaite_analyses if a.keyword}
            if keywords_seen:
                rows = db.execute(
                    select(
                        AnalysisService.keyword,
                        ServiceGroup.id,
                        ServiceGroup.name,
                    )
                    .join(
                        service_group_members,
                        service_group_members.c.analysis_service_id == AnalysisService.id,
                    )
                    .join(
                        ServiceGroup,
                        ServiceGroup.id == service_group_members.c.service_group_id,
                    )
                    .where(AnalysisService.keyword.in_(keywords_seen))
                ).all()
                kw_to_group = {row[0]: (row[1], row[2]) for row in rows}
                for a in senaite_analyses:
                    if a.keyword and a.keyword in kw_to_group:
                        gid, gname = kw_to_group[a.keyword]
                        a.service_group_id = gid
                        a.service_group_name = gname
        except Exception as e:
            # Best-effort: a service-group lookup failure should never break
            # the SENAITE lookup. Just log and proceed without enrichment.
            print(f"[WARN] service_group enrichment failed for {sample_id}: {e}")

        from datetime import datetime, timezone
        now_iso = datetime.now(timezone.utc).isoformat()
        result = SenaiteLookupResult(
            sample_id=sample_id,
            sample_uid=sample_uid or None,
            client=item.get("getClientTitle") or item.get("ClientTitle") or None,
            contact=item.get("ContactFullName") or None,
            sample_type=item.get("SampleTypeTitle") or item.get("getSampleTypeTitle") or None,
            date_received=item.get("DateReceived") or item.get("getDateReceived") or None,
            date_sampled=item.get("DateSampled") or item.get("getDateSampled") or None,
            profiles=profiles,
            client_order_number=item.get("ClientOrderNumber") or item.get("getClientOrderNumber") or None,
            client_sample_id=item.get("ClientSampleID") or item.get("getClientSampleID") or None,
            client_lot=str(item["ClientLot"]) if item.get("ClientLot") is not None else None,
            review_state=item.get("review_state") or None,
            declared_weight_mg=declared_weight_mg,
            analytes=analytes,
            coa=coa,
            remarks=senaite_remarks,
            analyses=senaite_analyses,
            attachments=senaite_attachments,
            published_coa=published_coa_report,
            senaite_url=_senaite_path(item),
            cached_at=now_iso,
        )
        _senaite_lookup_cache[id] = (_time.time(), result)
        return result

    except HTTPException:
        raise
    except httpx.TimeoutException as exc:
        print(f"[WARN] SENAITE lookup timeout for {id}: {exc}")
        raise HTTPException(status_code=503, detail="SENAITE is currently unavailable \u2014 use manual entry")
    except httpx.HTTPStatusError as exc:
        print(f"[WARN] SENAITE lookup HTTP error for {id}: {exc}")
        raise HTTPException(status_code=503, detail="SENAITE is currently unavailable \u2014 use manual entry")
    except Exception as exc:
        print(f"[WARN] SENAITE lookup error for {id}: {type(exc).__name__} {exc}")
        raise HTTPException(status_code=503, detail="SENAITE is currently unavailable \u2014 use manual entry")


@app.get("/wizard/senaite/attachment/{uid}")
async def get_senaite_attachment(
    uid: str,
    _current_user=Depends(get_current_user),
):
    """Proxy an attachment file from SENAITE, streaming it with correct content type."""
    if SENAITE_URL is None:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    from starlette.responses import StreamingResponse

    try:
        cached = _attachment_download_cache.get(uid)

        if cached:
            # Use cached download URL from the lookup
            download_url = cached["download_url"]
            content_type = cached["content_type"]
            filename = cached["filename"]
        else:
            # Fallback: fetch attachment metadata directly via api_url pattern
            att_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/Attachment/{uid}"
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            ) as client:
                meta_resp = await client.get(att_url)
                meta_resp.raise_for_status()
                meta_data = meta_resp.json()
                # Response could be {items: [...]} or direct object
                att_item = meta_data
                if "items" in meta_data and meta_data["items"]:
                    att_item = meta_data["items"][0]
                att_file = att_item.get("AttachmentFile") or {}
                download_url = att_file.get("download")
                content_type = att_file.get("content_type") or "application/octet-stream"
                filename = att_file.get("filename") or "attachment"

        if not download_url:
            raise HTTPException(status_code=404, detail="Attachment has no download URL")

        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
        ) as client:
            file_resp = await client.get(download_url)
            file_resp.raise_for_status()

            return StreamingResponse(
                iter([file_resp.content]),
                media_type=content_type,
                headers={"Content-Disposition": f'inline; filename="{filename}"'},
            )
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[WARN] Failed to proxy attachment {uid}: {exc}")
        raise HTTPException(status_code=503, detail="Failed to retrieve attachment")


@app.get("/wizard/senaite/report/{uid}")
async def get_senaite_report(
    uid: str,
    _current_user=Depends(get_current_user),
):
    """Proxy an ARReport PDF from SENAITE."""
    if SENAITE_URL is None:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    from starlette.responses import StreamingResponse

    try:
        download_url = _report_download_cache.get(uid)
        if not download_url:
            # Fallback: fetch ARReport metadata directly
            report_api_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/ARReport/{uid}"
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
            ) as client:
                meta_resp = await client.get(report_api_url, params={"complete": "yes"})
                meta_resp.raise_for_status()
                meta_data = meta_resp.json()
                r_item = meta_data
                if "items" in meta_data and meta_data["items"]:
                    r_item = meta_data["items"][0]
                pdf_info = r_item.get("Pdf") or {}
                download_url = pdf_info.get("download") if isinstance(pdf_info, dict) else None

        if not download_url:
            raise HTTPException(status_code=404, detail="Report PDF not found")

        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
        ) as client:
            file_resp = await client.get(download_url)
            file_resp.raise_for_status()
            return StreamingResponse(
                iter([file_resp.content]),
                media_type="application/pdf",
                headers={"Content-Disposition": f'inline; filename="COA.pdf"'},
            )
    except HTTPException:
        raise
    except Exception as exc:
        print(f"[WARN] Failed to proxy report {uid}: {exc}")
        raise HTTPException(status_code=503, detail="Failed to retrieve report")


class SenaiteUploadAttachmentResponse(BaseModel):
    success: bool
    message: str


@app.post(
    "/wizard/senaite/samples/{sample_uid}/attachments",
    response_model=SenaiteUploadAttachmentResponse,
)
async def upload_senaite_attachment(
    sample_uid: str,
    file: UploadFile,
    attachment_type: str = Form(...),  # "HPLC Graph" or "Sample Image"
    current_user=Depends(get_current_user),
):
    """Upload a file attachment to a SENAITE sample.

    Uses the Plone @@attachments_view/add form endpoint — same mechanism as
    the intake wizard image upload. The attachment_type name is matched against
    the options rendered in the sample page HTML to resolve its UID.
    """
    if SENAITE_URL is None:
        return SenaiteUploadAttachmentResponse(success=False, message="SENAITE not configured")

    try:
        file_bytes = await file.read()
        filename = file.filename or "attachment"
        content_type = file.content_type or "application/octet-stream"

        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(60.0, connect=10.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            # Step 1: Resolve the sample's Plone URL via REST API
            api_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/analysisrequest/{sample_uid}"
            meta_resp = await client.get(api_url)
            meta_resp.raise_for_status()
            meta_data = meta_resp.json()
            items = meta_data.get("items", [])
            if not items:
                return SenaiteUploadAttachmentResponse(
                    success=False, message="Sample not found in SENAITE"
                )
            sample_url = items[0].get("url") or items[0].get("absolute_url") or None
            if not sample_url:
                return SenaiteUploadAttachmentResponse(
                    success=False, message="Could not resolve sample URL"
                )

            # Step 2: GET sample page to extract CSRF authenticator + attachment type UID
            page_resp = await client.get(sample_url)
            page_html = page_resp.text

            auth_match = re.search(r'name="_authenticator"\s+value="([^"]+)"', page_html)
            authenticator = auth_match.group(1) if auth_match else ""

            # Find the UID for the requested attachment type name
            type_pattern = re.compile(
                r'<option\s+value="([^"]+)"[^>]*>\s*' + re.escape(attachment_type) + r'\s*</option>',
                re.IGNORECASE,
            )
            type_match = type_pattern.search(page_html)
            attachment_type_uid = type_match.group(1) if type_match else ""

            # Step 3: POST to @@attachments_view/add
            form_url = f"{sample_url}/@@attachments_view/add"
            form_data = {
                "submitted": "1",
                "_authenticator": authenticator,
                "AttachmentType": attachment_type_uid,
                "Analysis": "",  # empty = "Attach to Sample"
                "AttachmentKeys": "",
                "RenderInReport:boolean": "True",
                "RenderInReport:boolean:default": "False",
                "addARAttachment": "Add Attachment",
            }
            files = {
                "AttachmentFile_file": (filename, file_bytes, content_type),
            }
            headers = {
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Referer": sample_url,
            }
            att_resp = await client.post(
                form_url,
                data=form_data,
                files=files,
                headers=headers,
            )
            if att_resp.status_code in (200, 301, 302):
                return SenaiteUploadAttachmentResponse(success=True, message="Attachment uploaded")
            else:
                return SenaiteUploadAttachmentResponse(
                    success=False,
                    message=f"SENAITE returned {att_resp.status_code}",
                )

    except httpx.TimeoutException:
        return SenaiteUploadAttachmentResponse(success=False, message="SENAITE request timed out")
    except Exception as e:
        print(f"[WARN] Failed to upload attachment to sample {sample_uid}: {e}")
        return SenaiteUploadAttachmentResponse(success=False, message=str(e))


class SenaiteSampleItem(BaseModel):
    uid: str
    id: str
    title: str
    client_id: Optional[str] = None
    client_order_number: Optional[str] = None
    date_created: Optional[str] = None
    date_received: Optional[str] = None
    date_sampled: Optional[str] = None
    review_state: str
    sample_type: Optional[str] = None
    contact: Optional[str] = None
    verification_code: Optional[str] = None
    analytes: list[str] = []


class SenaiteSamplesResponse(BaseModel):
    items: list[SenaiteSampleItem]
    total: int
    b_start: int


@app.get("/senaite/samples", response_model=SenaiteSamplesResponse)
async def list_senaite_samples(
    review_state: Optional[str] = None,
    limit: int = 50,
    b_start: int = 0,
    search: Optional[str] = None,
    search_field: Optional[str] = None,
    include_sub_samples: bool = False,
    slim: bool = False,
    _current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    List AnalysisRequests from SENAITE with optional review_state filter.

    Query params:
    - review_state: Comma-separated state(s) e.g. "sample_received,to_be_verified"
    - limit: Max results (default 50)
    - b_start: Pagination offset (default 0)
    - include_sub_samples: when False (default), secondary ARs whose ID
      matches the <parent>-S\\d{2} convention are filtered out so the
      list shows parent samples only. The receive wizard surfaces
      sub-samples under their parent rather than as standalone rows.
      Note: pages can be slightly shorter than `limit` when many
      sub-samples are interleaved in the SENAITE result; the caller
      should bump `limit` if a denser display is desired.
    - slim: when True, skip SENAITE's complete=yes hydration and serve
      catalog brains only — review_state/id/uid are live, but analytes and
      verification_code come back empty (brains don't carry the custom
      Analyte{N}Peptide/VerificationCode schema fields; spike-verified
      2026-07-08). Used by the mk1-read-mode list refresh, which merges
      review_state only. SENAITE-mode callers must NOT pass it.
    """
    if SENAITE_URL is None:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
    base_params: dict = {"sort_on": "created", "sort_order": "descending"}
    if not slim:
        # Full hydration wakes every object in Zope — the expensive mode.
        base_params["complete"] = "yes"

    # SENAITE secondary AR ID convention. Used to drop sub-sample rows
    # from the parent listing unless include_sub_samples=True.
    _SUB_SAMPLE_RE = re.compile(r"-S\d{2}$")
    def _is_visible(it: dict) -> bool:
        if include_sub_samples:
            return True
        return not _SUB_SAMPLE_RE.search(str(it.get("id", "")))

    # SENAITE supports review_state:list for multiple states
    states = [s.strip() for s in review_state.split(",") if s.strip()] if review_state else []

    def _add_state_params(params: dict, states_list: list[str]) -> dict:
        """Add review_state filter to params dict."""
        if len(states_list) == 1:
            return {**params, "review_state": states_list[0]}
        return params

    def _build_state_url(base_url: str, params: dict, states_list: list[str]) -> str | None:
        """Build URL with multivalue review_state:list if needed."""
        if len(states_list) > 1:
            base_qs = "&".join(f"{k}={v}" for k, v in params.items())
            state_qs = "&".join(f"review_state:list={s}" for s in states_list)
            return f"{base_url}?{base_qs}&{state_qs}"
        return None

    def _extract_contact(item: dict) -> Optional[str]:
        contact = item.get("contact")
        if not contact:
            return None
        if isinstance(contact, dict):
            return contact.get("title") or contact.get("id")
        return str(contact)

    def _extract_analytes(it: dict) -> list[str]:
        """Extract analyte peptide names from Analyte1Peptide..Analyte4Peptide fields."""
        analytes = []
        for i in range(1, 5):
            val = it.get(f"Analyte{i}Peptide")
            if val and str(val).strip():
                analytes.append(str(val).strip())
        return analytes

    def _item_to_model(it: dict) -> SenaiteSampleItem:
        return SenaiteSampleItem(
            uid=str(it.get("uid", "")),
            id=str(it.get("id", "")),
            title=str(it.get("title", "")),
            client_id=it.get("getClientTitle") or it.get("ClientID") or it.get("getClientID") or None,
            client_order_number=it.get("getClientOrderNumber") or it.get("ClientOrderNumber") or None,
            date_created=it.get("created") or it.get("creation_date") or it.get("DateCreated") or it.get("getDateCreated") or None,
            date_received=it.get("getDateReceived") or it.get("DateReceived") or None,
            date_sampled=it.get("getDateSampled") or it.get("DateSampled") or None,
            review_state=str(it.get("review_state", "")),
            sample_type=it.get("getSampleTypeTitle") or it.get("SampleTypeTitle") or it.get("SampleType") or None,
            contact=_extract_contact(it),
            verification_code=it.get("VerificationCode") or it.get("getVerificationCode") or None,
            analytes=_extract_analytes(it),
        )

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=httpx.BasicAuth(SENAITE_USER, SENAITE_PASSWORD),
        ) as client:
            if search:
                # Multi-strategy search: fire parallel queries across different indexes
                # to cover sample ID, WP order number, and exact ID matches.
                import asyncio
                search_term = search.strip()

                async def _query(extra_params: dict) -> list[dict]:
                    """Run a single SENAITE search query and return items."""
                    params = {**base_params, "limit": limit, **extra_params}
                    params = _add_state_params(params, states)
                    multi_url = _build_state_url(url, params, states)
                    try:
                        if multi_url:
                            resp = await client.get(multi_url)
                        else:
                            resp = await client.get(url, params=params)
                        resp.raise_for_status()
                        return resp.json().get("items", [])
                    except Exception as exc:
                        print(f"[DEBUG] Search strategy failed ({extra_params}): {exc}")
                        return []

                # SENAITE catalog limitations (tested Mar 2026):
                # - getId: exact match only, no wildcards. Works perfectly and instantly.
                # - getClientOrderNumber: BROKEN — returns all samples regardless of value.
                # - SearchableText: tokenizes on hyphens, "P-0085" matches everything with "P".
                #
                # Strategy:
                # - search_field=None (default): getId exact match for sample ID search
                # - search_field=verification_code: Postgres lookup → sample IDs → getId
                # - search_field=order_number: Postgres lookup → sample IDs → getId

                if search_field == "verification_code":
                    from integration_db import search_sample_ids_by_verification_code
                    sample_ids = search_sample_ids_by_verification_code(search_term, limit=50)
                elif search_field == "order_number":
                    from integration_db import search_sample_ids_by_order_number
                    sample_ids = search_sample_ids_by_order_number(search_term, limit=50)
                else:
                    # Default: direct getId lookup (sample ID search)
                    sample_ids = [search_term]

                # Fetch each matching sample from SENAITE via getId
                all_items: list[dict] = []
                for sid in sample_ids:
                    items_for_id = await _query({"getId": sid})
                    all_items.extend(items_for_id)

                # Deduplicate by UID
                seen_uids: set[str] = set()
                deduped: list[dict] = []
                for it in all_items:
                    uid = str(it.get("uid", ""))
                    if uid and uid not in seen_uids:
                        seen_uids.add(uid)
                        deduped.append(it)

                # Sort by creation date descending
                deduped.sort(key=lambda x: x.get("created", "") or "", reverse=True)

                # Search bypasses the sub-sample filter. If the operator
                # explicitly searches by a verification code, order number,
                # or a sub-sample ID directly, they should find that record
                # — the parents-only filter is for browsing, not lookups.
                items = [_item_to_model(it) for it in deduped]

                # Mk1-native vials (Model D) have no SENAITE AR, so getId
                # can't find them. When an ID search comes back without the
                # exact term, resolve it against lims_sub_samples and
                # synthesize the row — direct vial-id lookup keeps working
                # post-cutover. (Legacy dual-written vials still resolve via
                # SENAITE above; the found-ids guard prevents duplicates.)
                if search_field is None:
                    found_ids = {it.id.upper() for it in items}
                    if search_term.upper() not in found_ids:
                        _sub = db.execute(
                            select(LimsSubSample).where(
                                func.upper(LimsSubSample.sample_id) == search_term.upper()
                            )
                        ).scalar_one_or_none()
                        if _sub is not None:
                            _parent = db.get(LimsSample, _sub.parent_sample_pk)
                            items.append(SenaiteSampleItem(
                                uid=_sub.external_lims_uid or f"mk1-sub-{_sub.id}",
                                id=_sub.sample_id,
                                title=_sub.sample_id,
                                date_created=_sub.created_at.isoformat() if _sub.created_at else None,
                                date_received=_sub.received_at.isoformat() if _sub.received_at else None,
                                # A vial with received_at was physically checked
                                # in (mirrors buildNativeSubSampleLookup).
                                review_state="sample_received" if _sub.received_at
                                             else (_parent.status if _parent and _parent.status else "registered"),
                                sample_type="Sub-sample",
                                analytes=[_parent.peptide_name] if _parent and _parent.peptide_name else [],
                            ))

                return SenaiteSamplesResponse(
                    items=items,
                    total=len(items),
                    b_start=0,
                )

            else:
                # Normal paginated listing (no search). When filtering
                # sub-samples server-side, over-fetch from SENAITE so each
                # page yields close to `limit` parents. b_start is in
                # user-page-units; translate to SENAITE-row-units by the
                # same factor so consecutive pages don't overlap.
                fetch_factor = 1 if include_sub_samples else 2
                params = {
                    **base_params,
                    "limit": limit * fetch_factor,
                    "b_start": b_start * fetch_factor,
                }
                params = _add_state_params(params, states)
                multi_url = _build_state_url(url, params, states)
                if multi_url:
                    resp = await client.get(multi_url)
                else:
                    resp = await client.get(url, params=params)

            resp.raise_for_status()
            data = resp.json()

            raw = data.get("items", [])
            visible = [it for it in raw if _is_visible(it)]
            # Cap at the user's requested `limit` so each page renders the
            # same density even when over-fetch found more parents than
            # expected. Anything trimmed here is picked up on the next page
            # (b_start advances by limit*fetch_factor on the SENAITE side).
            items = [_item_to_model(it) for it in visible[:limit]]

            # Estimate the parent total. SENAITE's count covers the
            # unfiltered set; scale by the parent ratio observed in the
            # current fetch so the frontend's totalPages stays roughly
            # right after filtering. Falls back to raw_total on an empty
            # page to avoid div-by-zero.
            raw_total = data.get("count") or data.get("total") or len(raw)
            if not include_sub_samples and raw:
                ratio = len(visible) / len(raw) if len(raw) else 1.0
                total = int(raw_total * ratio)
            else:
                total = raw_total

            return SenaiteSamplesResponse(
                items=items,
                total=total,
                b_start=b_start,
            )

    except HTTPException:
        raise
    except httpx.TimeoutException:
        raise HTTPException(status_code=503, detail="SENAITE is currently unavailable")
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=503, detail=f"SENAITE returned {e.response.status_code}")
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"SENAITE error: {e}")


class SenaiteReceiveSampleRequest(BaseModel):
    sample_uid: str
    sample_id: str
    image_base64: Optional[str] = None  # data URL or raw base64 (optional)
    remarks: Optional[str] = None


class SenaiteReceiveSampleResponse(BaseModel):
    success: bool
    message: str
    senaite_response: Optional[dict] = None


@app.post(
    "/wizard/senaite/receive-sample",
    response_model=SenaiteReceiveSampleResponse,
)
async def receive_senaite_sample(
    req: SenaiteReceiveSampleRequest,
    current_user=Depends(get_current_user),
):
    """Check-in / receive a sample in SENAITE.

    Performs up to three actions in sequence:
      1. Upload sample image attachment (if image provided).
      2. Add remarks to the sample (if remarks provided).
      3. Transition the sample to 'received' state.
    """
    if SENAITE_URL is None:
        return SenaiteReceiveSampleResponse(
            success=False, message="SENAITE not configured"
        )

    import base64
    import re

    # Decode image if provided
    image_bytes = None
    if req.image_base64:
        image_data = req.image_base64
        if image_data.startswith("data:"):
            image_data = image_data.split(",", 1)[1]
        try:
            image_bytes = base64.b64decode(image_data)
        except Exception as e:
            return SenaiteReceiveSampleResponse(
                success=False, message=f"Invalid base64 image data: {e}"
            )

    steps_done = []

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            # Fetch sample via JSON API to get physical path
            api_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
            sample_resp = await client.get(
                api_url, params={"UID": req.sample_uid, "limit": 1}
            )
            sample_resp.raise_for_status()
            sample_data = sample_resp.json()

            if sample_data.get("count", 0) == 0:
                return SenaiteReceiveSampleResponse(
                    success=False,
                    message=f"Sample {req.sample_id} not found in SENAITE",
                    senaite_response=sample_data,
                )

            sample_item = sample_data["items"][0]
            current_state = sample_item.get("review_state", "")
            sample_path = sample_item.get("path", "")
            if not sample_path:
                return SenaiteReceiveSampleResponse(
                    success=False,
                    message="Could not determine sample path in SENAITE",
                )
            sample_url = f"{SENAITE_URL}{sample_path}"

            # GET sample page for CSRF token (needed for attachment + workflow)
            page_resp = await client.get(sample_url)
            page_html = page_resp.text

            auth_match = re.search(
                r'name="_authenticator"\s+value="([^"]+)"', page_html
            )
            authenticator = auth_match.group(1) if auth_match else ""

            # --- Step 1: Upload image attachment (optional) ---
            if image_bytes:
                type_match = re.search(
                    r'<option\s+value="([^"]+)"[^>]*>\s*Sample Image\s*</option>',
                    page_html,
                )
                attachment_type_uid = (
                    type_match.group(1) if type_match else ""
                )

                filename = f"{req.sample_id}-sample-image.png"
                form_url = f"{sample_url}/@@attachments_view/add"
                form_data = {
                    "submitted": "1",
                    "_authenticator": authenticator,
                    "AttachmentType": attachment_type_uid,
                    "Analysis": "",  # empty = "Attach to Sample"
                    "AttachmentKeys": "",
                    "RenderInReport:boolean": "True",
                    "RenderInReport:boolean:default": "False",
                    "addARAttachment": "Add Attachment",
                }
                files = {
                    "AttachmentFile_file": (
                        filename,
                        image_bytes,
                        "image/png",
                    ),
                }
                headers = {
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Referer": sample_url,
                }
                att_resp = await client.post(
                    form_url,
                    data=form_data,
                    files=files,
                    headers=headers,
                )
                if att_resp.status_code in (200, 301, 302):
                    steps_done.append("image_uploaded")
                else:
                    return SenaiteReceiveSampleResponse(
                        success=False,
                        message=f"Image upload failed: SENAITE returned {att_resp.status_code}",
                        senaite_response={"steps_done": steps_done},
                    )

            # --- Step 2: Add remarks (optional) ---
            if req.remarks and req.remarks.strip():
                update_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{req.sample_uid}"
                remark_resp = await client.post(
                    update_url, json={"Remarks": req.remarks.strip()}
                )
                if remark_resp.status_code == 200:
                    steps_done.append("remarks_added")
                else:
                    return SenaiteReceiveSampleResponse(
                        success=False,
                        message=f"Remarks update failed: SENAITE returned {remark_resp.status_code}",
                        senaite_response={"steps_done": steps_done},
                    )

            # --- Step 3: Transition to 'received' ---
            # Only attempt if sample is still in 'sample_due' state.
            # Samples already received or further along don't need this.
            if current_state == "sample_due":
                # Always re-fetch CSRF right before workflow transition
                # (prior steps may have rotated the token)
                page_resp2 = await client.get(sample_url)
                auth_match2 = re.search(
                    r'name="_authenticator"\s+value="([^"]+)"',
                    page_resp2.text,
                )
                authenticator = (
                    auth_match2.group(1) if auth_match2 else authenticator
                )

                wf_url = f"{sample_url}/workflow_action"
                wf_data = {
                    "workflow_action": "receive",
                    "_authenticator": authenticator,
                }
                headers = {
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Referer": sample_url,
                }
                wf_resp = await client.post(
                    wf_url, data=wf_data, headers=headers
                )
                if wf_resp.status_code not in (200, 301, 302):
                    return SenaiteReceiveSampleResponse(
                        success=False,
                        message=f"Workflow transition failed: SENAITE returned {wf_resp.status_code}",
                        senaite_response={"steps_done": steps_done},
                    )

                # Verify the transition actually took effect
                verify_resp = await client.get(
                    api_url, params={"UID": req.sample_uid, "limit": 1}
                )
                verify_data = verify_resp.json()
                new_state = (
                    verify_data["items"][0].get("review_state", "")
                    if verify_data.get("count", 0) > 0
                    else ""
                )
                if new_state == "sample_received":
                    steps_done.append("received")
                    # Task 3: native sample-transition log (own session,
                    # never-fail — see _record_sample_transition_bg).
                    from fastapi.concurrency import run_in_threadpool
                    await run_in_threadpool(
                        _record_sample_transition_bg,
                        sample_id=req.sample_id, verb="receive", to_status="sample_received",
                        from_status="sample_due", source="mk1",
                        actor_user_id=getattr(current_user, "id", None),
                    )
                else:
                    return SenaiteReceiveSampleResponse(
                        success=False,
                        message=f"Workflow transition did not take effect (state is still '{new_state}')",
                        senaite_response={"steps_done": steps_done},
                    )
            else:
                steps_done.append(f"already_{current_state}")
                return SenaiteReceiveSampleResponse(
                    success=True,
                    message=f"Sample {req.sample_id} is already '{current_state}' — image/remarks added but no state change needed",
                    senaite_response={"steps_done": steps_done},
                )

        return SenaiteReceiveSampleResponse(
            success=True,
            message=f"Sample {req.sample_id} received successfully",
            senaite_response={"steps_done": steps_done},
        )

    except httpx.TimeoutException:
        return SenaiteReceiveSampleResponse(
            success=False,
            message="SENAITE request timed out",
            senaite_response={"steps_done": steps_done},
        )
    except httpx.HTTPStatusError as e:
        return SenaiteReceiveSampleResponse(
            success=False,
            message=f"SENAITE returned {e.response.status_code}",
            senaite_response={"steps_done": steps_done},
        )
    except Exception as e:
        return SenaiteReceiveSampleResponse(
            success=False,
            message=f"Receive error: {e}",
            senaite_response={"steps_done": steps_done},
        )


# --- SENAITE field update endpoint ---


class SenaiteFieldUpdateRequest(BaseModel):
    fields: dict  # e.g. {"ClientOrderNumber": "WP-1234", "ClientLot": "LOT-5"}


class SenaiteFieldUpdateResponse(BaseModel):
    success: bool
    message: str
    updated_fields: Optional[list] = None


@app.post(
    "/wizard/senaite/samples/{uid}/update",
    response_model=SenaiteFieldUpdateResponse,
)
async def update_senaite_sample_fields(
    uid: str,
    req: SenaiteFieldUpdateRequest,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Update one or more fields on a SENAITE sample via the JSON API."""
    if SENAITE_URL is None:
        return SenaiteFieldUpdateResponse(
            success=False, message="SENAITE not configured"
        )

    if not req.fields:
        return SenaiteFieldUpdateResponse(
            success=False, message="No fields provided"
        )

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            update_url = (
                f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{uid}"
            )
            senaite_fields = {
                k: str(v) if v is not None else ""
                for k, v in req.fields.items()
            }

            # Strategy: try JSON body first (required for extension fields
            # like CompanyLogoUrl, ChromatographBackgroundUrl which are
            # silently ignored when sent form-encoded).
            #
            # If SENAITE returns 400 (e.g. isDecimal validator rejects
            # unicode strings in Python 2), fall back to form-encoded
            # which sends Python 2 str values that pass the validator.
            try:
                resp = await client.post(update_url, json=senaite_fields)
                resp.raise_for_status()
            except httpx.HTTPStatusError as json_err:
                if json_err.response.status_code == 400:
                    # Fallback: form-encoded for isDecimal-type fields
                    resp = await client.post(
                        update_url, data=senaite_fields
                    )
                    resp.raise_for_status()
                else:
                    raise

            # Dual-write mirror (registry slice 1): reflect the accepted
            # SENAITE edit onto the local registry row. Best-effort — a
            # mirror problem must never fail the user's edit.
            try:
                from sub_samples.service import apply_senaite_fields_to_row
                if apply_senaite_fields_to_row(db, uid, req.fields):
                    db.commit()
            except Exception as mirror_err:
                try:
                    db.rollback()
                except Exception:
                    pass
                logger.warning(
                    "registry.field_mirror_failed uid=%s err=%s", uid, mirror_err
                )

            return SenaiteFieldUpdateResponse(
                success=True,
                message=f"Updated {len(req.fields)} field(s)",
                updated_fields=list(req.fields.keys()),
            )

    except httpx.TimeoutException:
        return SenaiteFieldUpdateResponse(
            success=False, message="SENAITE request timed out"
        )
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:200]
        except Exception:
            pass
        return SenaiteFieldUpdateResponse(
            success=False,
            message=f"SENAITE returned {e.response.status_code}: {detail}".strip(),
        )
    except Exception as e:
        return SenaiteFieldUpdateResponse(
            success=False, message=f"Update error: {e}"
        )


# --- Analysis result and transition endpoints ---


def _mirror_parent_analysis_bg(**kwargs) -> None:
    """Best-effort parent-analysis shadow mirror on its own short-lived session
    (never holds the request DB across the SENAITE HTTP call). Never raises.

    Accepts optional `method_uid`/`instrument_uid` (raw SENAITE uids) — popped
    and resolved to Mk1 ids on THIS function's own session (never on the event
    loop) via resolve_method_id/resolve_instrument_id, then passed through to
    mirror_parent_analysis as method_id/instrument_id, but only when resolved
    non-None (an unresolvable uid must never overwrite an existing value with
    None). Existing callers (A1/A2/A3) never pass these kwargs — pop() with a
    None default plus the resolvers' own None-safety makes this a pure,
    backward-compatible extension.

    `SessionLocal()`, the resolver/helper imports, and the kwargs.pop() calls
    all live INSIDE the try (not just the mirror call) so that even a
    pathological construction/import failure is caught here rather than
    escaping to the caller — this function must never raise. `db` is set to
    None before the try so the finally block can guard `db.close()` for the
    case where SessionLocal() itself never returned a session.
    """
    db = None
    try:
        from database import SessionLocal
        from lims_analyses.parent_mirror import (
            mirror_parent_analysis, resolve_instrument_id, resolve_method_id,
        )
        method_uid = kwargs.pop("method_uid", None)
        instrument_uid = kwargs.pop("instrument_uid", None)
        db = SessionLocal()
        method_id = resolve_method_id(db, method_uid)
        instrument_id = resolve_instrument_id(db, instrument_uid)
        if method_id is not None:
            kwargs["method_id"] = method_id
        if instrument_id is not None:
            kwargs["instrument_id"] = instrument_id
        if mirror_parent_analysis(db, **kwargs):
            db.commit()
    except Exception as mirror_err:  # noqa: BLE001
        if db is not None:
            try:
                db.rollback()
            except Exception:
                pass
        logger.warning("registry.analysis_mirror_failed kw=%s err=%s",
                       kwargs.get("keyword"), mirror_err)
    finally:
        if db is not None:
            db.close()


def _mark_shadows_published_bg(sample_id: str) -> None:
    """Best-effort A6 sibling of _mirror_parent_analysis_bg: on its own
    short-lived session, flip every LIVE parent shadow row to
    mirror_review_state='published' (mark_parent_shadows_published). Never
    holds the request `db` across the SENAITE HTTP calls in publish_sample_coa;
    never raises — a mirror failure must never fail the publish.

    `SessionLocal()` and the helper import live INSIDE the try, same
    hardening rationale as `_mirror_parent_analysis_bg`: `db` starts as None
    so `finally` can guard `db.close()` if construction itself failed.
    """
    db = None
    try:
        from database import SessionLocal
        from lims_analyses.parent_mirror import mark_parent_shadows_published
        db = SessionLocal()
        if mark_parent_shadows_published(db, sample_id=sample_id):
            db.commit()
    except Exception as mirror_err:  # noqa: BLE001
        if db is not None:
            try:
                db.rollback()
            except Exception:
                pass
        logger.warning("registry.publish_shadow_mark_failed sample_id=%s err=%s",
                       sample_id, mirror_err)
    finally:
        if db is not None:
            db.close()


def _record_sample_transition_bg(**kwargs) -> None:
    """Best-effort native sample-transition log write (Task 3) on its own
    short-lived session — never holds the request `db` across the SENAITE
    HTTP calls at the two call sites (publish, receive). Never raises: a
    log-write failure must never fail or delay-fail the endpoint it's
    scheduled from.

    `SessionLocal()` and the recorder import live INSIDE the try, same
    hardening rationale as `_mirror_parent_analysis_bg`: `db` starts as None
    so `finally` can guard `db.close()` if construction itself failed.
    """
    db = None
    try:
        from database import SessionLocal
        from workflow.sample_log import record_sample_transition
        db = SessionLocal()
        if record_sample_transition(db, **kwargs):
            db.commit()
    except Exception as log_err:  # noqa: BLE001
        if db is not None:
            try:
                db.rollback()
            except Exception:
                pass
        logger.warning("workflow.sample_log_failed sample_id=%s err=%s",
                       kwargs.get("sample_id"), log_err)
    finally:
        if db is not None:
            db.close()


def _observe_parent_analyses_bg(sample_id: str, observed: list[dict]) -> None:
    """Best-effort passive analysis drift observer (Task 7) on its own
    short-lived session — never holds the request/route `db` across the
    SENAITE fetch either hook site already made for display. Never raises: a
    healing failure must never affect the page that triggered it.

    Commits unconditionally on success (not gated on `observe_parent_
    analyses`'s return count): result-only healing writes no transition row
    but still mutates `result_value` and must persist — gating the commit on
    the count would silently drop that write.

    `SessionLocal()` and the observer import live INSIDE the try, same
    hardening rationale as `_mirror_parent_analysis_bg`: `db` starts as None
    so `finally` can guard `db.close()` if construction itself failed.
    """
    db = None
    try:
        from database import SessionLocal
        from workflow.observer import observe_parent_analyses
        db = SessionLocal()
        observe_parent_analyses(db, sample_id=sample_id, observed=observed)
        db.commit()
    except Exception as observer_err:  # noqa: BLE001
        if db is not None:
            try:
                db.rollback()
            except Exception:
                pass
        logger.warning("workflow.observer_failed sample_id=%s err=%s",
                       sample_id, observer_err)
    finally:
        if db is not None:
            db.close()


class AnalysisResultRequest(BaseModel):
    result: str  # The result value to set


class AnalysisResultResponse(BaseModel):
    success: bool
    message: str
    new_review_state: Optional[str] = None
    keyword: Optional[str] = None


@app.post(
    "/wizard/senaite/analyses/{uid}/result",
    response_model=AnalysisResultResponse,
)
async def set_analysis_result(
    uid: str,
    req: AnalysisResultRequest,
    current_user=Depends(get_current_user),
):
    """Set the Result value on a SENAITE analysis.

    Proxies to SENAITE REST API: POST /update/{uid} with {"Result": value}.
    Does NOT trigger a workflow transition — that is a separate explicit action.
    """
    if SENAITE_URL is None:
        return AnalysisResultResponse(
            success=False, message="SENAITE not configured"
        )

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            update_url = (
                f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{uid}"
            )
            resp = await client.post(update_url, json={"Result": req.result})
            resp.raise_for_status()
            data = resp.json()
            items = data.get("items", [])
            if not items:
                return AnalysisResultResponse(
                    success=False,
                    message="SENAITE returned no items — update may have failed",
                )
            item = items[0]

            from fastapi.concurrency import run_in_threadpool
            _sid = item.get("getRequestID") or item.get("RequestID")
            _kw = item.get("Keyword")
            if _sid and _kw:
                await run_in_threadpool(
                    _mirror_parent_analysis_bg,
                    sample_id=_sid, keyword=_kw,
                    mirror_review_state=item.get("review_state"),
                    result_value=req.result,
                )

            return AnalysisResultResponse(
                success=True,
                message="Result updated",
                new_review_state=item.get("review_state", ""),
                keyword=item.get("Keyword", ""),
            )

    except httpx.TimeoutException:
        return AnalysisResultResponse(
            success=False, message="SENAITE request timed out"
        )
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:200]
        except Exception:
            pass
        return AnalysisResultResponse(
            success=False,
            message=f"SENAITE returned {e.response.status_code}: {detail}".strip(),
        )
    except Exception as e:
        return AnalysisResultResponse(
            success=False, message=f"Update error: {e}"
        )


class AnalysisMethodInstrumentRequest(BaseModel):
    method_uid: Optional[str] = None
    instrument_uid: Optional[str] = None


@app.post(
    "/wizard/senaite/analyses/{uid}/method-instrument",
    response_model=AnalysisResultResponse,
)
async def set_analysis_method_instrument(
    uid: str,
    req: AnalysisMethodInstrumentRequest,
    current_user=Depends(get_current_user),
):
    """Set Method and/or Instrument on a SENAITE analysis.

    Proxies to SENAITE REST API: POST /update/{uid} with {Method: uid, Instrument: uid}.
    Pass an empty string to clear a field, or omit to leave it unchanged.
    """
    if SENAITE_URL is None:
        return AnalysisResultResponse(success=False, message="SENAITE not configured")

    payload: dict = {}
    if req.method_uid is not None:
        payload["Method"] = req.method_uid
    if req.instrument_uid is not None:
        payload["Instrument"] = req.instrument_uid
    if not payload:
        return AnalysisResultResponse(success=False, message="No fields to update")

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            update_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{uid}"
            resp = await client.post(update_url, json=payload)
            resp.raise_for_status()
            data = resp.json()
            items = data.get("items", [])
            if not items:
                return AnalysisResultResponse(
                    success=False,
                    message="SENAITE returned no items — update may have failed",
                )
            item = items[0]

            from fastapi.concurrency import run_in_threadpool
            _sid = item.get("getRequestID") or item.get("RequestID")
            _kw = item.get("Keyword")
            if _sid and _kw:
                await run_in_threadpool(
                    _mirror_parent_analysis_bg,
                    sample_id=_sid, keyword=_kw,
                    mirror_review_state=item.get("review_state"),
                    method_uid=req.method_uid, instrument_uid=req.instrument_uid,
                )

            return AnalysisResultResponse(
                success=True,
                message="Updated",
                new_review_state=item.get("review_state", ""),
                keyword=item.get("Keyword", ""),
            )

    except httpx.TimeoutException:
        return AnalysisResultResponse(success=False, message="SENAITE request timed out")
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:200]
        except Exception:
            pass
        return AnalysisResultResponse(
            success=False,
            message=f"SENAITE returned {e.response.status_code}: {detail}".strip(),
        )
    except Exception as e:
        return AnalysisResultResponse(success=False, message=f"Update error: {e}")


class AnalysisTransitionRequest(BaseModel):
    transition: str  # "submit", "verify", "retract", "reject"


# After a successful transition, SENAITE should move to this review_state.
# If the actual state differs, SENAITE silently rejected the transition (DATA-04).
# NOTE: "retract"'s "unassigned" entry is DEAD for validation purposes —
# retract is retire-and-replace, not an in-place flip (see the dedicated
# early-return branch in transition_analysis), so its real post-state
# ('retracted') is checked there instead of via this dict/DATA-04. The key
# stays here only so the membership gate a few lines below still accepts
# "retract" as a valid transition name.
EXPECTED_POST_STATES: dict[str, str] = {
    "submit": "to_be_verified",
    "verify": "verified",
    "retract": "unassigned",
    "reject": "rejected",
    "retest": "verified",
}


@app.post(
    "/wizard/senaite/analyses/{uid}/transition",
    response_model=AnalysisResultResponse,
)
async def transition_analysis(
    uid: str,
    req: AnalysisTransitionRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Trigger a workflow transition on a SENAITE analysis.

    Proxies to SENAITE REST API: POST /update/{uid} with {"transition": action}.
    Validates post-transition review_state against EXPECTED_POST_STATES to catch
    silent rejections (SENAITE returns 200 OK even when transitions are skipped).

    When the transition is 'retest', cascades the retest to source vial-tier
    analyses via cascade_parent_retest_to_sources (best-effort, wrapped in
    try/except — cascade failure never fails the SENAITE transition).
    """
    if SENAITE_URL is None:
        return AnalysisResultResponse(
            success=False, message="SENAITE not configured"
        )

    if req.transition not in EXPECTED_POST_STATES:
        return AnalysisResultResponse(
            success=False,
            message=f"Invalid transition: {req.transition}. "
            f"Must be one of: {', '.join(EXPECTED_POST_STATES.keys())}",
        )

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            update_url = (
                f"{SENAITE_URL}/senaite/@@API/senaite/v1/update/{uid}"
            )
            resp = await client.post(
                update_url, json={"transition": req.transition}
            )
            resp.raise_for_status()

            # ── Retract: retire-and-replace, not an in-place state flip ──────
            # SENAITE's retract transition retires the ORIGINAL analysis line
            # (review_state -> 'retracted') and spawns a brand-new, DIFFERENT
            # analysis object born 'unassigned' carrying the old Result, with
            # retest_of_uid pointing at the original (live evidence: BW-0002
            # FILL-NET-CONTENT — original 800582b9 -> retracted result=30;
            # new copy 736bb35a -> unassigned result=30
            # retest_of_uid=800582b9). Two bugs this exposes in the generic
            # path below:
            #   1. SENAITE's /update response for retract comes back with an
            #      EMPTY items list even on success — the items-empty guard
            #      a few lines down would report a false failure.
            #   2. EXPECTED_POST_STATES["retract"] == "unassigned" models the
            #      in-place flip that doesn't exist; the original's true
            #      post-state is 'retracted' — the unassigned thing is a
            #      DIFFERENT object, so DATA-04 below can't validate retract.
            # Handled as a self-contained early-return branch so every other
            # verb (submit/verify/reject/retest) stays byte-identical to
            # before: whether or not the POST's items came back, re-fetch the
            # target uid's current state (same GET-by-uid shape as the
            # retest/reject cascade's getRequestID fallback further below) and
            # use its review_state as the sole success criterion.
            if req.transition == "retract":
                _fetch_url = (
                    f"{SENAITE_URL}/senaite/@@API/senaite/v1/Analysis/{uid}"
                )
                _fetch_resp = await client.get(_fetch_url)
                _retract_item: dict = {}
                if _fetch_resp.status_code == 200:
                    _fetch_items = _fetch_resp.json().get("items", [])
                    if _fetch_items:
                        _retract_item = _fetch_items[0]

                _retracted_state = _retract_item.get("review_state", "")
                _retract_keyword = _retract_item.get("Keyword", "")

                if _retracted_state != "retracted":
                    return AnalysisResultResponse(
                        success=False,
                        message=(
                            "Transition 'retract' was silently rejected by "
                            f"SENAITE. Expected state 'retracted' but got "
                            f"'{_retracted_state or 'unknown'}'."
                        ),
                        new_review_state=_retracted_state,
                        keyword=_retract_keyword,
                    )

                # Confirmed retracted — fire the chained shadow mirror:
                # the OLD live shadow row is stamped retracted+retested, and
                # a NEW live row is born unassigned carrying the re-fetched
                # Result (SENAITE's copy). Best-effort, own session, same
                # posture as the A2/A3 mirror block below.
                _retract_sid = (
                    _retract_item.get("getRequestID")
                    or _retract_item.get("RequestID")
                )
                if _retract_sid and _retract_keyword:
                    from fastapi.concurrency import run_in_threadpool
                    await run_in_threadpool(
                        _mirror_parent_analysis_bg,
                        sample_id=_retract_sid, keyword=_retract_keyword,
                        mirror_review_state="unassigned",
                        result_value=(_retract_item.get("Result") or None),
                        is_retest=True,
                        old_mirror_review_state="retracted",
                    )

                return AnalysisResultResponse(
                    success=True,
                    message="Transition 'retract' completed",
                    new_review_state=_retracted_state,
                    keyword=_retract_keyword,
                )
            # ── end retract special-case ──────────────────────────────────────

            data = resp.json()
            items = data.get("items", [])
            if not items:
                return AnalysisResultResponse(
                    success=False,
                    message="SENAITE returned no items — transition may have failed",
                )
            item = items[0]
            actual_state = item.get("review_state", "")
            keyword = item.get("Keyword", "")
            expected_state = EXPECTED_POST_STATES[req.transition]

            # DATA-04: Detect silent rejection by comparing actual vs expected state
            if actual_state != expected_state:
                return AnalysisResultResponse(
                    success=False,
                    message=(
                        f"Transition '{req.transition}' was silently rejected "
                        f"by SENAITE. Expected state '{expected_state}' but "
                        f"got '{actual_state}'."
                    ),
                    new_review_state=actual_state,
                    keyword=keyword,
                )

            # ── Parent retest/reject cascade (best-effort) ───────────────────
            # retest: find the Mk1 parent-tier row and cascade the retest to
            # all source vial-tier analyses so the bench sees the work requests.
            # reject: the service was removed from the offering — reject the
            # UNPOPULATED vial-tier mirror rows of that keyword across the
            # family (rows with results are never touched).
            #
            # sample_id resolution:
            #   1. Try item["getRequestID"] — available when the update endpoint
            #      returns a full catalog-aware item (usually the case).
            #   2. Fallback: fetch the Analysis object by uid and read
            #      getRequestID from there (one extra GET, always reliable).
            #   3. If still absent, log a warning and skip — the cascade is
            #      best-effort and must not fail the SENAITE transition.
            if req.transition in ("retest", "reject"):
                import logging as _logging
                _cascade_logger = _logging.getLogger(__name__)
                _cascade_tag = f"cascade_parent_{req.transition}"
                try:
                    from lims_analyses.service import (
                        cascade_parent_reject_to_vials,
                        cascade_parent_retest_to_sources,
                    )

                    _parent_sample_id: Optional[str] = (
                        item.get("getRequestID") or item.get("RequestID") or None
                    )
                    if not _parent_sample_id:
                        # Fallback: re-fetch the analysis object for its request ID
                        try:
                            _fetch_url = (
                                f"{SENAITE_URL}/senaite/@@API/senaite/v1/Analysis/{uid}"
                            )
                            _fetch_resp = await client.get(_fetch_url)
                            if _fetch_resp.status_code == 200:
                                _fetch_data = _fetch_resp.json()
                                _fetch_items = _fetch_data.get("items", [])
                                if _fetch_items:
                                    _parent_sample_id = (
                                        _fetch_items[0].get("getRequestID")
                                        or _fetch_items[0].get("RequestID")
                                        or None
                                    )
                        except Exception as _fetch_err:
                            _cascade_logger.warning(
                                "%s: fallback fetch for uid=%s failed: %s",
                                _cascade_tag, uid, _fetch_err,
                            )

                    if _parent_sample_id and keyword:
                        _user_id = getattr(current_user, "id", None)
                        if req.transition == "retest":
                            _row_ids = cascade_parent_retest_to_sources(
                                db,
                                parent_sample_id=_parent_sample_id,
                                keyword=keyword,
                                user_id=_user_id,
                            )
                            _verb = "created vial retest rows"
                        else:
                            _row_ids = cascade_parent_reject_to_vials(
                                db,
                                parent_sample_id=_parent_sample_id,
                                keyword=keyword,
                                user_id=_user_id,
                            )
                            _verb = "rejected vial mirror rows"
                        if _row_ids:
                            _cascade_logger.info(
                                "%s: parent=%s keyword=%s → %s %s",
                                _cascade_tag, _parent_sample_id, keyword,
                                _verb, _row_ids,
                            )
                    else:
                        _cascade_logger.warning(
                            "%s: could not resolve parent_sample_id "
                            "for uid=%s keyword=%r — cascade skipped",
                            _cascade_tag, uid, keyword,
                        )
                except Exception as _cascade_err:
                    _cascade_logger.warning(
                        "%s: unexpected error for uid=%s: %s",
                        _cascade_tag, uid, _cascade_err,
                    )
            # ── end parent retest/reject cascade ──────────────────────────────

            # ── Parent analysis shadow mirror (A2/A3, best-effort) ───────────
            # Only reached after the silent-rejection check above passes, i.e.
            # SENAITE actually applied the transition. Uses its own SessionLocal
            # via _mirror_parent_analysis_bg — never the request's `db` (that
            # session is reserved for the retest/reject cascade above).
            from fastapi.concurrency import run_in_threadpool
            _sid = item.get("getRequestID") or item.get("RequestID")
            if _sid and keyword:
                _is_retest = req.transition == "retest"
                # For retest, `actual_state` ("verified", per
                # EXPECTED_POST_STATES["retest"]) describes the OLD SENAITE
                # analysis line SENAITE echoed back — not the NEW retest
                # analysis object spawned under the hood, which is born
                # unassigned. The mirror helper's is_retest branch creates
                # that new row and stamps it with whatever mirror_review_state
                # is passed here, so passing actual_state would mislabel the
                # brand-new row as already verified. The old row keeps its own
                # (correct) mirror_review_state — is_retest only marks it
                # retested, it doesn't touch that field.
                #
                # Live registry-inspect UAT showed SENAITE's retest
                # transition also copies the OLD line's Result onto the new
                # retest analysis under the hood (new line born unassigned
                # but already carrying the old Result) — so `item["Result"]`
                # here is that carried-over value, not a real new result.
                # Mirror it onto the new row for the same reason as the
                # state above: leaving result_value=None would drift from
                # SENAITE truth. `or None` guards against mirroring an
                # empty-string Result. Non-retest transitions are unchanged —
                # they never pass result_value (state-only mirror).
                _mirror_state = "unassigned" if _is_retest else actual_state
                _retest_result = (item.get("Result") or None) if _is_retest else None
                await run_in_threadpool(
                    _mirror_parent_analysis_bg,
                    sample_id=_sid, keyword=keyword,
                    mirror_review_state=_mirror_state,
                    result_value=_retest_result,
                    is_retest=_is_retest,
                )

            return AnalysisResultResponse(
                success=True,
                message=f"Transition '{req.transition}' completed",
                new_review_state=actual_state,
                keyword=keyword,
            )

    except httpx.TimeoutException:
        return AnalysisResultResponse(
            success=False, message="SENAITE request timed out"
        )
    except httpx.HTTPStatusError as e:
        detail = ""
        try:
            detail = e.response.text[:200]
        except Exception:
            pass
        return AnalysisResultResponse(
            success=False,
            message=f"SENAITE returned {e.response.status_code}: {detail}".strip(),
        )
    except Exception as e:
        return AnalysisResultResponse(
            success=False, message=f"Transition error: {e}"
        )


# --- Scale endpoints (Phase 2) ---

@app.get("/scale/status")
async def get_scale_status(
    request: Request,
    _current_user=Depends(get_current_user),
):
    """
    Get scale connection status.

    Returns:
        disabled     — SCALE_HOST not configured; manual-entry mode
        connected    — SCALE_HOST set and balance is reachable
        disconnected — SCALE_HOST set but balance is unreachable
    """
    bridge = getattr(request.app.state, 'scale_bridge', None)
    if bridge is None:
        return {"status": "disabled", "host": None, "port": None}
    return {
        "status": "connected" if bridge.connected else "disconnected",
        "host": bridge.host,
        "port": bridge.port,
    }


@app.get("/scale/weight/stream")
async def stream_scale_weight(
    request: Request,
    _current_user=Depends(get_current_user),
):
    """
    Stream live weight readings from the balance via SSE at 4 Hz.

    Yields 'weight' events with value/unit/stable fields.
    Yields 'error' events on ConnectionError or ValueError (bridge may reconnect — does not stop stream).
    Returns 503 when SCALE_HOST is not configured.
    """
    from starlette.responses import StreamingResponse
    import asyncio

    bridge = getattr(request.app.state, 'scale_bridge', None)
    if bridge is None:
        raise HTTPException(status_code=503, detail="Scale not configured (SCALE_HOST not set)")

    async def event_generator():
        def send_event(event_type: str, data: dict) -> str:
            payload = json.dumps(data)
            return f"event: {event_type}\ndata: {payload}\n\n"

        try:
            while True:
                if await request.is_disconnected():
                    break

                try:
                    reading = await bridge.read_weight()
                    yield send_event("weight", {
                        "value": reading["value"],
                        "unit": reading["unit"],
                        "stable": reading["stable"],
                    })
                except (ConnectionError, ValueError) as e:
                    yield send_event("error", {"message": str(e)})

                await asyncio.sleep(0.25)
        except asyncio.CancelledError:
            pass

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ─── Service Groups ───────────────────────────────────────────────────────────

@app.get("/service-groups", response_model=list[ServiceGroupResponse])
async def get_service_groups(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Return all service groups ordered by sort_order, name."""
    groups = db.execute(
        select(ServiceGroup)
        .options(joinedload(ServiceGroup.analysis_services))
        .order_by(ServiceGroup.sort_order, ServiceGroup.name)
    ).scalars().unique().all()

    result = []
    for group in groups:
        resp = ServiceGroupResponse(
            id=group.id,
            name=group.name,
            description=group.description,
            color=group.color,
            sort_order=group.sort_order,
            is_default=group.is_default,
            sla_tier_id=group.sla_tier_id,
            member_count=len(group.analysis_services),
            member_ids=[s.id for s in group.analysis_services],
            created_at=group.created_at,
            updated_at=group.updated_at,
        )
        result.append(resp)
    return result


@app.post("/service-groups", response_model=ServiceGroupResponse, status_code=201)
async def create_service_group(
    data: ServiceGroupCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Create a new service group."""
    existing = db.execute(
        select(ServiceGroup).where(ServiceGroup.name == data.name)
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(400, f"Service group '{data.name}' already exists")

    group = ServiceGroup(**data.model_dump())
    if group.is_default:
        db.execute(
            select(ServiceGroup).where(ServiceGroup.is_default == True)  # noqa: E712
        )
        db.query(ServiceGroup).filter(ServiceGroup.is_default == True).update({"is_default": False})  # noqa: E712
    db.add(group)
    db.commit()
    db.refresh(group)
    return ServiceGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        sort_order=group.sort_order,
        is_default=group.is_default,
        sla_tier_id=group.sla_tier_id,
        member_count=0,
        created_at=group.created_at,
        updated_at=group.updated_at,
    )


@app.put("/service-groups/{group_id}", response_model=ServiceGroupResponse)
async def update_service_group(
    group_id: int,
    data: ServiceGroupUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Update an existing service group."""
    group = db.execute(
        select(ServiceGroup)
        .options(joinedload(ServiceGroup.analysis_services))
        .where(ServiceGroup.id == group_id)
    ).unique().scalar_one_or_none()
    if not group:
        raise HTTPException(404, f"Service group {group_id} not found")

    update_data = data.model_dump(exclude_unset=True)
    if update_data.get("is_default"):
        db.query(ServiceGroup).filter(
            ServiceGroup.is_default == True, ServiceGroup.id != group_id  # noqa: E712
        ).update({"is_default": False})
    for field, value in update_data.items():
        setattr(group, field, value)

    db.commit()
    db.refresh(group)
    return ServiceGroupResponse(
        id=group.id,
        name=group.name,
        description=group.description,
        color=group.color,
        sort_order=group.sort_order,
        is_default=group.is_default,
        sla_tier_id=group.sla_tier_id,
        member_count=len(group.analysis_services),
        member_ids=[s.id for s in group.analysis_services],
        created_at=group.created_at,
        updated_at=group.updated_at,
    )


@app.delete("/service-groups/{group_id}")
async def delete_service_group(
    group_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Delete a service group. Membership rows cascade-delete."""
    group = db.execute(
        select(ServiceGroup).where(ServiceGroup.id == group_id)
    ).scalar_one_or_none()
    if not group:
        raise HTTPException(404, f"Service group {group_id} not found")

    db.delete(group)
    db.commit()
    return {"message": f"Service group '{group.name}' deleted"}


@app.get("/service-groups/{group_id}/members", response_model=list[int])
async def get_service_group_members(
    group_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Return the list of analysis_service IDs currently in the group."""
    group = db.execute(
        select(ServiceGroup).where(ServiceGroup.id == group_id)
    ).scalar_one_or_none()
    if not group:
        raise HTTPException(404, f"Service group {group_id} not found")

    rows = db.execute(
        select(service_group_members.c.analysis_service_id).where(
            service_group_members.c.service_group_id == group_id
        )
    ).all()
    return [row.analysis_service_id for row in rows]


@app.put("/service-groups/{group_id}/members")
async def set_service_group_members(
    group_id: int,
    req: ServiceGroupMembersRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Replace the full membership set for a service group."""
    group = db.execute(
        select(ServiceGroup)
        .options(joinedload(ServiceGroup.analysis_services))
        .where(ServiceGroup.id == group_id)
    ).unique().scalar_one_or_none()
    if not group:
        raise HTTPException(404, f"Service group {group_id} not found")

    services = db.execute(
        select(AnalysisService).where(AnalysisService.id.in_(req.analysis_service_ids))
    ).scalars().all()

    group.analysis_services = list(services)
    db.commit()
    return {"count": len(services)}


# ─── SLA tiers (sub-project A, revised to tiers) ──────────────────────────────


def _demote_other_default_tiers(db: Session, keep_id: Optional[int] = None) -> None:
    """Clear is_default on every tier except keep_id, flushing before the caller
    inserts/updates the promoted row (the partial unique index is immediate)."""
    q = db.query(SlaTier).filter(SlaTier.is_default == True)  # noqa: E712
    if keep_id is not None:
        q = q.filter(SlaTier.id != keep_id)
    q.update({"is_default": False})
    db.flush()


@app.get("/sla-tiers", response_model=list[SlaTierResponse])
async def list_sla_tiers(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """All SLA tiers, default first. Consumed by the settings UI (C) and, cached,
    by D2 (which resolves client-side)."""
    return db.execute(
        select(SlaTier).order_by(SlaTier.is_default.desc(), SlaTier.name)
    ).scalars().all()


@app.post("/sla-tiers", response_model=SlaTierResponse, status_code=201)
async def create_sla_tier(
    data: SlaTierCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Create a tier. Setting is_default demotes any existing default."""
    tier = SlaTier(**data.model_dump())
    if tier.is_default:
        _demote_other_default_tiers(db)
    db.add(tier)
    db.commit()
    db.refresh(tier)
    return tier


@app.put("/sla-tiers/{tier_id}", response_model=SlaTierResponse)
async def update_sla_tier(
    tier_id: int,
    data: SlaTierUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Update a tier. Promoting demotes the rest; demoting the only default is
    rejected (it's the 24h backstop for unmatched samples)."""
    tier = db.get(SlaTier, tier_id)
    if not tier:
        raise HTTPException(404, f"SLA tier {tier_id} not found")
    update_data = data.model_dump(exclude_unset=True)
    if "is_default" in update_data:
        if update_data["is_default"]:
            _demote_other_default_tiers(db, keep_id=tier_id)
        elif tier.is_default:
            raise HTTPException(
                409,
                "Cannot unset the only default SLA tier; set another as default instead",
            )
    for field, value in update_data.items():
        setattr(tier, field, value)
    db.commit()
    db.refresh(tier)
    return tier


@app.delete("/sla-tiers/{tier_id}")
async def delete_sla_tier(
    tier_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Delete a tier. The default cannot be deleted. Groups referencing it have
    sla_tier_id set NULL (FK); priority overrides referencing it cascade-delete."""
    tier = db.get(SlaTier, tier_id)
    if not tier:
        raise HTTPException(404, f"SLA tier {tier_id} not found")
    if tier.is_default:
        raise HTTPException(409, "Cannot delete the default SLA tier; promote another first")
    db.delete(tier)
    db.commit()
    return {"message": f"SLA tier {tier_id} deleted"}


@app.get("/sla-priority-tiers", response_model=list[SlaPriorityTierResponse])
async def list_sla_priority_tiers(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """The sparse priority -> tier override map (only overriding priorities)."""
    return db.execute(select(SlaPriorityTier)).scalars().all()


@app.put("/sla-priority-tiers/{priority}", response_model=SlaPriorityTierResponse)
async def set_sla_priority_tier(
    priority: SlaPriority,
    data: SlaPriorityTierSet,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Upsert a priority -> tier override.

    The override is identified by (priority, service_group_id). Omit
    service_group_id to upsert the global override; supply it to scope the
    override to a single service group.
    """
    if not db.get(SlaTier, data.sla_tier_id):
        raise HTTPException(404, f"SLA tier {data.sla_tier_id} not found")
    if data.service_group_id is not None and not db.get(
        ServiceGroup, data.service_group_id
    ):
        raise HTTPException(
            404, f"Service group {data.service_group_id} not found"
        )
    # SQL NULL semantics — `=` against NULL never matches, so the global-row
    # branch must use IS NULL explicitly. Both code paths return at most one
    # row courtesy of the two partial unique indexes from _run_migrations.
    q = select(SlaPriorityTier).where(SlaPriorityTier.priority == priority)
    if data.service_group_id is None:
        q = q.where(SlaPriorityTier.service_group_id.is_(None))
    else:
        q = q.where(SlaPriorityTier.service_group_id == data.service_group_id)
    row = db.execute(q).scalar_one_or_none()
    if row:
        row.sla_tier_id = data.sla_tier_id
    else:
        row = SlaPriorityTier(
            priority=priority,
            sla_tier_id=data.sla_tier_id,
            service_group_id=data.service_group_id,
        )
        db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.delete("/sla-priority-tiers/{priority}")
async def delete_sla_priority_tier(
    priority: SlaPriority,
    service_group_id: int | None = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Remove a priority override.

    Without `service_group_id`, removes the global (NULL group) override.
    With `service_group_id`, removes the override scoped to that group only.
    """
    q = select(SlaPriorityTier).where(SlaPriorityTier.priority == priority)
    if service_group_id is None:
        q = q.where(SlaPriorityTier.service_group_id.is_(None))
    else:
        q = q.where(SlaPriorityTier.service_group_id == service_group_id)
    row = db.execute(q).scalar_one_or_none()
    if not row:
        scope = "global" if service_group_id is None else f"group_id={service_group_id}"
        raise HTTPException(404, f"No override for priority '{priority}' ({scope})")
    db.delete(row)
    db.commit()
    scope = "global" if service_group_id is None else f"group_id={service_group_id}"
    return {"message": f"Priority override '{priority}' ({scope}) removed"}


# ── Business-hours config (sub-project B) ──────────────────────────────────

@app.get("/business-hours-config", response_model=BusinessHoursConfigResponse)
async def get_business_hours_config(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """The singleton business-hours schedule. Read-only for non-admins (UI-gated)."""
    cfg = db.get(BusinessHoursConfig, 1)
    if not cfg:
        raise HTTPException(500, "Business-hours config not initialized")
    return cfg


@app.put("/business-hours-config", response_model=BusinessHoursConfigResponse)
async def update_business_hours_config(
    data: BusinessHoursConfigUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Update the schedule. Validates IANA timezone, close>open, working_days ⊆ 0..6."""
    try:
        ZoneInfo(data.timezone)
    except Exception:
        raise HTTPException(422, f"Unknown timezone: {data.timezone}")
    if data.close_time <= data.open_time:
        raise HTTPException(422, "close_time must be after open_time")
    if not data.working_days or any(d < 0 or d > 6 for d in data.working_days):
        raise HTTPException(422, "working_days must be a non-empty subset of 0..6")
    cfg = db.get(BusinessHoursConfig, 1)
    if not cfg:
        raise HTTPException(500, "Business-hours config not initialized")
    cfg.open_time = data.open_time
    cfg.close_time = data.close_time
    cfg.timezone = data.timezone
    cfg.working_days = sorted(set(data.working_days))
    db.commit()
    db.refresh(cfg)
    return cfg


# ── Lab holidays (sub-project B) ───────────────────────────────────────────

@app.get("/lab-holidays", response_model=list[LabHolidayResponse])
async def list_lab_holidays(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """All stored closures for `year` (defaults to current), federal + custom, ordered by date."""
    y = year if year is not None else date.today().year
    return db.execute(
        select(LabHoliday)
        .where(extract("year", LabHoliday.holiday_date) == y)
        .order_by(LabHoliday.holiday_date)
    ).scalars().all()


@app.post("/lab-holidays", response_model=LabHolidayResponse, status_code=201)
async def create_lab_holiday(
    data: LabHolidayCreate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Add a custom closure (source='custom'). 409 if a closure already exists on that date."""
    existing = db.execute(
        select(LabHoliday).where(LabHoliday.holiday_date == data.holiday_date)
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(409, f"A closure already exists on {data.holiday_date}")
    row = LabHoliday(holiday_date=data.holiday_date, name=data.name, source="custom")
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.delete("/lab-holidays/{holiday_date}")
async def delete_lab_holiday(
    holiday_date: date,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Remove any closure (federal or custom). Deleting a federal row = the lab works that day."""
    row = db.execute(
        select(LabHoliday).where(LabHoliday.holiday_date == holiday_date)
    ).scalar_one_or_none()
    if not row:
        raise HTTPException(404, f"No closure on {holiday_date}")
    db.delete(row)
    db.commit()
    return {"message": f"Closure on {holiday_date} removed"}


@app.post("/lab-holidays/generate-federal")
async def generate_federal_holidays_endpoint(
    year: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Insert any missing federal closures for `year`. Primary use: extend
    coverage into a new year. Caveat: this re-adds ANY missing federal date for
    that year — including ones the lab previously deleted — because it's a
    deliberate, user-triggered action. (Startup seeding does NOT do this; it is
    first-boot-only, so deletions survive restarts.)"""
    from database import seed_federal_holidays

    added = seed_federal_holidays(db.connection(), year)
    db.commit()
    return {"year": year, "added": added}


# ─── SLA Status Batch ─────────────────────────────────────────────────────────

@app.post("/sla/status", response_model=SlaStatusResponse)
async def compute_sla_statuses(
    req: SlaStatusRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Batch SLA status for a page of rows (the render endpoint).

    Loads the schedule + holiday set ONCE, then maps over items — O(items) with
    O(1) DB reads. `now` is server time; the response is a snapshot. `key` is
    opaque and echoed (the client correlates by key, not array order). `status`
    is null iff `received_at` is null.
    """
    now = datetime.utcnow()  # naive UTC, codebase convention
    cfg = db.get(BusinessHoursConfig, 1)
    schedule = BusinessSchedule.from_orm(cfg) if cfg else None
    holiday_dates = {r[0] for r in db.execute(select(LabHoliday.holiday_date)).all()}
    is_holiday = lambda d: d in holiday_dates  # noqa: E731

    results: list[SlaStatusResultItem] = []
    for item in req.items:
        recv = item.received_at
        if recv is None:
            results.append(SlaStatusResultItem(key=item.key, status=None))
            continue
        # Normalize to naive UTC (an offset-aware ISO string is converted).
        if recv.tzinfo is not None:
            recv = recv.astimezone(timezone.utc).replace(tzinfo=None)
        # Per-item "now": published samples send `now_override` (their
        # publication date) so elapsed = (published - received). Live samples
        # leave it null and get the request-time wall clock.
        item_now = item.now_override or now
        if item_now.tzinfo is not None:
            item_now = item_now.astimezone(timezone.utc).replace(tzinfo=None)
        if item.business_hours_only and schedule is not None:
            elapsed = compute_business_minutes(recv, item_now, schedule, is_holiday)
        else:
            elapsed = (item_now - recv).total_seconds() / 60.0
        results.append(
            SlaStatusResultItem(key=item.key, status=sla_status_dict(item.target_minutes, elapsed))
        )
    return SlaStatusResponse(items=results)


# ─── SENAITE Analyst Proxy ────────────────────────────────────────────────────

@app.get("/senaite/analysts")
async def get_senaite_analysts(
    current_user=Depends(get_current_user),
):
    """Proxy to SENAITE LabContact — returns list of {username, fullname}."""
    if not SENAITE_URL:
        raise HTTPException(503, "SENAITE URL not configured")

    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=httpx.Timeout(30.0, connect=5.0),
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/LabContact"
            resp = await client.get(url, params={"complete": "yes", "limit": 200})
            resp.raise_for_status()
            data = resp.json()
            items = data.get("items", [])
            return [
                {
                    "uid": i.get("uid"),
                    "username": i.get("Username") or i.get("getUsername"),
                    "fullname": i.get("getFullname", i.get("title", "")),
                }
                for i in items
            ]
    except httpx.TimeoutException:
        raise HTTPException(504, "SENAITE request timed out")
    except httpx.HTTPStatusError as e:
        raise HTTPException(502, f"SENAITE returned {e.response.status_code}")
    except Exception as e:
        raise HTTPException(500, f"Analyst fetch error: {e}")



# NOTE: SENAITE Analyst field is read-only on Analysis objects — it only gets set
# when an analysis is added to a SENAITE Worksheet. Since AccuMark replaces SENAITE
# worksheets, analyst assignment lives in AccuMark's local worksheet_items table.
# The GET /senaite/analysts endpoint above remains useful for populating analyst


# ─── Worksheets Inbox ─────────────────────────────────────────────────────────

class InboxAnalysisItem(BaseModel):
    uid: Optional[str] = None
    title: str
    keyword: Optional[str] = None
    peptide_name: Optional[str] = None
    method: Optional[str] = None
    review_state: Optional[str] = None
    # Service group context — surfaced per-analysis in the flat list so the
    # frontend (or AddSamplesModal) can regroup by service group when needed.
    group_id: int
    group_name: str
    group_color: str


class InboxVialItem(BaseModel):
    """One inbox card == one vial (parent AR or sub-sample AR).

    Replaces the old InboxSampleItem/InboxServiceGroupSection nesting. The
    role-filtered analyses are flat and carry their own group metadata.
    Vial position fields let the frontend stack same-family vials visually.
    """
    uid: str
    sample_id: str
    is_parent: bool
    parent_sample_id: str
    assignment_role: Optional[str] = None
    # 'core' | 'variance' | None — explicit per-vial variance marker set at
    # check-in (lims_sub_samples.assignment_kind). Parents are always None.
    assignment_kind: Optional[str] = None
    vial_sequence: int          # 0 for parent, 1+ for subs (lims_sub_samples.vial_sequence)
    vial_total: int             # family size: legacy = parent + subs; container = subs only
    # Container family (parent is a pure depository): vial position label is
    # vial_sequence itself, not vial_sequence + 1.
    container_mode: bool = False
    title: str
    client_id: Optional[str] = None
    client_order_number: Optional[str] = None
    date_received: Optional[str] = None
    review_state: str
    priority: str = "normal"
    analyses: list[InboxAnalysisItem] = []
    assignment_summary: str = ""  # e.g., "1/1 assigned" — vial-level


class InboxResponse(BaseModel):
    items: list[InboxVialItem]
    total: int
    filter_role: Optional[str] = None  # echo of the query param so the frontend can confirm


# Role → service_group_name set. Hardcoded — the lab has had Analytics +
# Microbiology for years and a 2-entry mapping doesn't deserve a table.
ROLE_TO_GROUP_NAMES: dict[str, set[str]] = {
    "hplc": {"Analytics"},
    "microbiology": {"Microbiology"},
}
VALID_INBOX_ROLES = set(ROLE_TO_GROUP_NAMES.keys())

# Role-set membership for the assignment_role column. Microbiology covers
# both 'ster' and 'endo' (collapsed into one filter chip per spec Q1).
ROLE_TO_VIAL_ROLES: dict[str, set[str]] = {
    "hplc": {"hplc"},
    "microbiology": {"ster", "endo"},
}


class PriorityUpdate(BaseModel):
    priority: str  # "normal" | "high" | "expedited"


class BulkInboxUpdate(BaseModel):
    sample_uids: list[str]
    priority: Optional[str] = None
    service_group_id: Optional[int] = None  # required when setting analyst or instrument
    analyst_id: Optional[int] = None
    instrument_uid: Optional[str] = None

    @validator("service_group_id", pre=True, always=True)
    def zero_to_none(cls, v):
        return None if v == 0 else v


class WorksheetCreate(BaseModel):
    title: str
    sample_uids: list[str]
    notes: Optional[str] = None


# Server-side SENAITE inbox cache — shared across all users, 30-minute TTL
_inbox_senaite_cache: dict[str, list] = {}
_inbox_senaite_cache_time: float = 0
_INBOX_CACHE_TTL_SECONDS = 30 * 60  # 30 minutes


# ── Phase 3.5: Mk1-sourced inbox analyses for sub-samples ───────────────────


# Color fallback when the ServiceGroup join misses (rare — only if a vial's
# analysis_service has no service_group_members row). Mirrors the FE's role
# palette.
_INBOX_ROLE_COLOR_FALLBACK = {
    "hplc": "sky",
    "endo": "violet",
    "ster": "violet",
    "xtra": "zinc",
}


def _inbox_family_sizes(parent_rows, sub_rows) -> dict[int, int]:
    """Family size per parent pk for the inbox's "Vial K / N". Legacy parents
    count as a vial themselves (parent IS vial 1); container parents are pure
    report depositories and count 0 (container-parent design). Sub rows can
    arrive via two query predicates — dedup on (parent_pk, vial_sequence)."""
    sizes: dict[int, int] = {
        r.id: (0 if r.container_mode else 1) for r in parent_rows
    }
    seen: set[tuple[int, int]] = set()
    for s in sub_rows:
        key = (s.parent_sample_pk, s.vial_sequence)
        if key in seen:
            continue
        seen.add(key)
        sizes[s.parent_sample_pk] = sizes.get(s.parent_sample_pk, 0) + 1
    return sizes


def _fetch_mk1_inbox_analyses_for_sub_sample(
    db: Session,
    sub_sample_pk: int,
    role: Optional[str],
    keyword_to_peptide: dict,
) -> list["InboxAnalysisItem"]:
    """Build the per-vial inbox analysis list from Mk1 lims_analyses.

    Returns the same InboxAnalysisItem shape as the existing SENAITE-derived
    builder. UIDs carry the 'mk1:' prefix so any downstream write-path
    dispatches to the Mk1 endpoints (Phase 3 adapter).

    Filtering: evaluate the live row per (vial, service) (retested=False) and
    drop dead + terminal-done review_states, so a finished vial does not return
    to the inbox once its open-worksheet claim is removed.

    Returns an empty list if the vial has no Mk1 rows — caller falls back
    to the SENAITE path.
    """
    from models import LimsAnalysis  # local import; not at module top

    rows = db.execute(
        select(LimsAnalysis, AnalysisService, ServiceGroup)
        .join(AnalysisService, AnalysisService.id == LimsAnalysis.analysis_service_id)
        .outerjoin(
            service_group_members,
            service_group_members.c.analysis_service_id == AnalysisService.id,
        )
        .outerjoin(ServiceGroup, ServiceGroup.id == service_group_members.c.service_group_id)
        .where(LimsAnalysis.lims_sub_sample_pk == sub_sample_pk)
        # Current/live row per (vial, service) is retested=False — NOT
        # retest_of_id IS NULL, which selects the superseded original after a
        # retest (see architecture_retest_current_row_idiom). Reading the
        # original masked finished retests, bouncing done vials into the inbox.
        .where(LimsAnalysis.retested.is_(False))
    ).all()

    # Drop dead rows (rejected/retracted) AND terminal done-states. A vial whose
    # live result is promoted (rolled up to parent), verified/published, or
    # variance_verified (vial-tier terminal; parent uses the mean model) is
    # finished. Without these, completing a worksheet — which removes the only
    # open-worksheet claim keeping the vial out — bounced finished Micro/variance
    # vials back into the inbox (2026-06-25 incident). Core HPLC vials were masked
    # by hide_prepped; Micro/variance vials have no prep, so nothing else hid them.
    EXCLUDED_STATES = {
        "rejected", "retracted",
        "promoted", "verified", "published", "variance_verified",
    }
    out: list[InboxAnalysisItem] = []
    for la, svc, sg in rows:
        if la.review_state in EXCLUDED_STATES:
            continue
        if sg is not None:
            grp_id = sg.id
            grp_name = sg.name or ""
            grp_color = getattr(sg, "color", None) or _INBOX_ROLE_COLOR_FALLBACK.get(role or "", "zinc")
        else:
            grp_id = 0
            grp_name = ""
            grp_color = _INBOX_ROLE_COLOR_FALLBACK.get(role or "", "zinc")

        out.append(InboxAnalysisItem(
            uid=f"mk1:{la.id}",
            title=la.title or la.keyword or "",
            keyword=la.keyword,
            peptide_name=keyword_to_peptide.get(la.keyword or "") if keyword_to_peptide else None,
            method=None,                  # Mk1 vial method not yet wired
            review_state=la.review_state,
            group_id=grp_id,
            group_name=grp_name,
            group_color=grp_color,
        ))
    return out


# Column set for the inbox's lims_sub_samples fetch. Single source of truth:
# every attribute that _build_native_vial_inbox_items or the vial_meta loop
# reads off a sub row MUST be here — these are column-limited Rows, not ORM
# objects, so a missing column is a runtime AttributeError the full-ORM test
# fixtures won't catch (tests select through this same tuple to stay honest).
INBOX_SUB_SAMPLE_COLUMNS = (
    LimsSubSample.parent_sample_pk,
    LimsSubSample.external_lims_uid,
    LimsSubSample.sample_id,
    LimsSubSample.assignment_role,
    LimsSubSample.assignment_kind,    # variance badge passthrough
    LimsSubSample.vial_sequence,
    LimsSubSample.id,                 # Phase 3.5: needed for Mk1 inbox source
    LimsSubSample.received_at,        # native-vial rows: own check-in date
)


def _build_native_vial_inbox_items(
    db: Session,
    *,
    parent_item: dict,
    parent_sample_id: str,
    native_subs: list,
    family_size: int,
    allowed_vial_roles: set,
    assigned_pairs: set,
    assigned_uids_for_null_group: set,
    hide_prepped: bool,
    prepped_sub_pks: set,
    prepped_senaite_ids: set,
    priority_map: dict,
    order_priority: Optional[str],
    assignment_map: dict,
    keyword_to_peptide: dict,
    container_mode: bool = True,
) -> "list[InboxVialItem]":
    """Inbox rows for the Mk1-native vials of one family.

    Called whenever a parent has Mk1-native vials. For a CONTAINER parent the
    caller also suppresses the parent's own row (depository); for a
    non-container (legacy / already-received) parent the parent row is kept
    (it is vial 1) and these native vials are added alongside it — the
    after-the-fact-add case. `container_mode` is stamped onto each emitted
    item so the UI shapes it correctly. Spec
    2026-06-11-vial-level-worksheets-inbox-design.md. Mirrors the per-vial
    filters of the SENAITE loop in get_worksheets_inbox: role, open-worksheet
    claims, prepped, and no-analyses-left. Row identity is the vial's own
    external_lims_uid (mk1://…) — already what worksheet_analyst.stamp_for_item
    resolves and unique per vial.

    Order-level priority persists to sample_priorities exactly like step 4b
    does for parents, so the worksheet add endpoints (which read
    SamplePriority by uid) see it too.
    """
    out: list[InboxVialItem] = []
    priorities_dirty = False
    for sub in native_subs:
        uid = sub.external_lims_uid
        role = sub.assignment_role
        if role not in allowed_vial_roles:
            continue
        if uid in assigned_uids_for_null_group:
            continue
        if hide_prepped and (
            sub.id in prepped_sub_pks or sub.sample_id in prepped_senaite_ids
        ):
            continue

        analyses = _fetch_mk1_inbox_analyses_for_sub_sample(
            db, sub.id, role, keyword_to_peptide,
        )
        analyses = [a for a in analyses if (uid, a.group_id) not in assigned_pairs]
        if not analyses:
            continue
        analyses.sort(key=lambda a: (a.group_name.lower(), a.title.lower()))

        prio = priority_map.get(uid)
        if prio is None and order_priority in ("high", "expedited"):
            existing = db.execute(
                select(SamplePriority).where(SamplePriority.sample_uid == uid)
            ).scalar_one_or_none()
            if existing is None:
                db.add(SamplePriority(sample_uid=uid, priority=order_priority))
                priorities_dirty = True
            elif existing.priority == "normal":
                existing.priority = order_priority
                priorities_dirty = True
            prio = order_priority

        unique_groups = {a.group_id for a in analyses}
        assigned_count = 0
        for gid in unique_groups:
            assignment = assignment_map.get((uid, gid))
            if assignment and assignment.assigned_analyst_id:
                assigned_count += 1
        summary = (
            f"{assigned_count}/{len(unique_groups)} assigned"
            if (unique_groups and assigned_count > 0)
            else ""
        )

        received_at = getattr(sub, "received_at", None)
        date_received = (
            received_at.isoformat()
            if received_at
            else (parent_item.get("getDateReceived") or parent_item.get("DateReceived") or None)
        )

        out.append(InboxVialItem(
            uid=uid,
            sample_id=sub.sample_id,
            is_parent=False,
            parent_sample_id=parent_sample_id,
            assignment_role=role,
            assignment_kind=sub.assignment_kind,
            vial_sequence=sub.vial_sequence,
            vial_total=family_size,
            container_mode=container_mode,
            title=str(parent_item.get("title", "")),
            client_id=parent_item.get("getClientTitle") or parent_item.get("ClientID") or None,
            client_order_number=parent_item.get("getClientOrderNumber") or parent_item.get("ClientOrderNumber") or None,
            date_received=date_received,
            review_state=str(parent_item.get("review_state", "sample_received")),
            priority=prio or "normal",
            assignment_summary=summary,
            analyses=analyses,
        ))
    if priorities_dirty:
        db.commit()
    return out


@app.get("/worksheets/inbox", response_model=InboxResponse)
async def get_worksheets_inbox(
    hide_test_orders: bool = True,
    hide_prepped: bool = True,
    force_refresh: bool = False,
    role: Optional[str] = None,
    show_xtra: bool = False,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Return inbox items (one per VIAL — parent AR or sub-sample AR) ready for worksheet
    assignment. Each vial carries its role-filtered analyses, vial-position context, and
    parent linkage. See `docs/superpowers/specs/2026-06-02-worksheet-vial-inbox-redesign.md`.

    Query params:
      role         — 'hplc' | 'microbiology' | omitted. Omitted means all roles (used by
                     AddSamplesModal, which adds across both benches). 400 on invalid value.
      show_xtra    — when True, append XTRA-role vials to the active filter's results.
      hide_test_*  — existing behavior.
      force_refresh — bypass the 30-min SENAITE cache.
    """
    global _inbox_senaite_cache, _inbox_senaite_cache_time

    # Validate role (None == "all roles", legal)
    if role is not None and role not in VALID_INBOX_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role: {role!r}. Expected one of {sorted(VALID_INBOX_ROLES)} or omit.",
        )

    if not SENAITE_URL:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    # Resolve role → allowed service_group IDs. None means "no filter; pass all groups".
    allowed_group_ids: Optional[set[int]] = None
    if role is not None:
        group_names = ROLE_TO_GROUP_NAMES[role]
        allowed_group_ids = {
            r[0] for r in db.execute(
                select(ServiceGroup.id).where(ServiceGroup.name.in_(group_names))
            ).all()
        }

    # Resolve allowed vial assignment_role values. NULL roles always excluded (auto-
    # assign on /vial-plan is the cure for those). XTRA gated by show_xtra.
    if role is None:
        # No bench filter: all known roles. XTRA still gated by the toggle.
        allowed_vial_roles: set[str] = {"hplc", "ster", "endo"}
        if show_xtra:
            allowed_vial_roles.add("xtra")
    else:
        allowed_vial_roles = set(ROLE_TO_VIAL_ROLES[role])
        if show_xtra:
            allowed_vial_roles.add("xtra")

    # Step 1: Fetch sample_received samples from SENAITE (with cache)
    import time as _time
    now = _time.time()
    cache_age = now - _inbox_senaite_cache_time
    cache_valid = not force_refresh and _inbox_senaite_cache.get("items") is not None and cache_age < _INBOX_CACHE_TTL_SECONDS

    if cache_valid:
        senaite_data = _inbox_senaite_cache
    else:
        senaite_url = f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest"
        params = {
            "review_state": "sample_received",
            "complete": "yes",
            "limit": 200,
            "sort_on": "created",
            "sort_order": "descending",
        }

        try:
            async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
                timeout=SENAITE_TIMEOUT,
                auth=_get_senaite_auth(current_user),
                follow_redirects=True,
            ) as client:
                resp = await client.get(senaite_url, params=params)
                resp.raise_for_status()
                senaite_data = resp.json()
                # Update cache
                _inbox_senaite_cache = senaite_data
                _inbox_senaite_cache_time = now
        except httpx.TimeoutException:
            # If cache exists but expired, serve stale data rather than error
            if _inbox_senaite_cache.get("items") is not None:
                senaite_data = _inbox_senaite_cache
            else:
                raise HTTPException(status_code=504, detail="SENAITE request timed out")
        except httpx.HTTPStatusError as e:
            if _inbox_senaite_cache.get("items") is not None:
                senaite_data = _inbox_senaite_cache
            else:
                raise HTTPException(status_code=502, detail=f"SENAITE returned {e.response.status_code}")
        except Exception as e:
            if _inbox_senaite_cache.get("items") is not None:
                senaite_data = _inbox_senaite_cache
            else:
                raise HTTPException(status_code=500, detail=f"SENAITE fetch error: {e}")

    senaite_items = senaite_data.get("items", [])

    # Step 1b: Filter to only samples linked to tracked orders in integration DB
    # Also build senaite_id → order priority map for auto-priority
    TEST_EMAILS = ["forrestp@outlook.com", "forrest@valenceanalytical.com"]
    order_priority_map: dict[str, str] = {}  # senaite_id → priority from order payload
    try:
        from integration_db import get_integration_db
        from psycopg2.extras import RealDictCursor
        linked_senaite_ids: set[str] = set()
        with get_integration_db() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT sample_results, payload FROM order_submissions WHERE sample_results IS NOT NULL"
                )
                for row in cur.fetchall():
                    payload = row["payload"] if isinstance(row.get("payload"), dict) else {}
                    # Check if this is a test order
                    if hide_test_orders:
                        billing = payload.get("billing", {})
                        email = (billing.get("email") or "").lower() if isinstance(billing, dict) else ""
                        if email in TEST_EMAILS:
                            continue
                    # Extract order-level priority (sent by WP, optional)
                    order_priority = payload.get("priority")
                    sr = row["sample_results"]
                    if isinstance(sr, dict):
                        for entry in sr.values():
                            if isinstance(entry, dict) and entry.get("senaite_id"):
                                sid = entry["senaite_id"]
                                linked_senaite_ids.add(sid)
                                if order_priority and order_priority in ("high", "expedited"):
                                    order_priority_map[sid] = order_priority
        # Extend with sub-samples of any linked parent. Sub-samples are created
        # post-order by the Receive Wizard and never appear in order_submissions,
        # so without this step they'd be dropped from the inbox entirely. One
        # SQL roundtrip pulls them all.
        if linked_senaite_ids:
            sub_rows = db.execute(
                select(LimsSubSample.sample_id)
                .join(LimsSample, LimsSubSample.parent_sample_pk == LimsSample.id)
                .where(LimsSample.sample_id.in_(linked_senaite_ids))
            ).scalars().all()
            linked_senaite_ids.update(sub_rows)
        # Filter SENAITE items to only those with a linked order (now includes subs)
        senaite_items = [
            it for it in senaite_items
            if str(it.get("id", "")) in linked_senaite_ids
        ]
    except Exception:
        pass  # If integration DB is unavailable, show all samples (graceful degradation)

    # Step 2: Build set of (sample_uid, service_group_id) pairs already in open worksheets
    # Only exclude specific service groups, not the entire sample — a sample can have
    # Microbiology in a worksheet while Core HPLC is still available in the inbox.
    open_worksheet_pairs = db.execute(
        select(WorksheetItem.sample_uid, WorksheetItem.service_group_id)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(Worksheet.status == "open")
    ).all()
    assigned_pairs: set[tuple[str, int | None]] = {
        (row.sample_uid, row.service_group_id) for row in open_worksheet_pairs
    }
    # Also track fully-assigned samples (all groups in worksheets) for backward compat
    # We'll filter at the group level later in step 6, not at the sample level here
    assigned_uids_for_null_group: set[str] = set()
    # Samples with service_group_id=None in a worksheet are fully claimed
    for uid, gid in assigned_pairs:
        if gid is None:
            assigned_uids_for_null_group.add(uid)

    # Step 2b: Load SENAITE sample IDs that already have a sample prep
    prepped_senaite_ids: set[str] = set()
    prepped_sub_pks: set[int] = set()
    if hide_prepped:
        try:
            from mk1_db import ensure_sample_preps_table, get_mk1_db
            ensure_sample_preps_table()
            with get_mk1_db() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT DISTINCT senaite_sample_id FROM sample_preps WHERE senaite_sample_id IS NOT NULL")
                    prepped_senaite_ids = {row[0] for row in cur.fetchall()}
                    # Vial-scoped preps (post prep-cutover) tag the vial pk —
                    # the native-vial rows filter on this, not the senaite id.
                    cur.execute("SELECT DISTINCT lims_sub_sample_pk FROM sample_preps WHERE lims_sub_sample_pk IS NOT NULL")
                    prepped_sub_pks = {row[0] for row in cur.fetchall()}
        except Exception:
            pass  # If mk1 DB is unavailable, show all samples

    # Filter out samples that already have preps (by SENAITE sample ID e.g. "P-0085")
    filtered_items = [
        it for it in senaite_items
        if not hide_prepped or str(it.get("id", "")) not in prepped_senaite_ids
    ]

    # Step 3: Build keyword → service group map
    group_rows = db.execute(
        select(
            AnalysisService.keyword,
            ServiceGroup.id,
            ServiceGroup.name,
            ServiceGroup.color,
        )
        .join(service_group_members, AnalysisService.id == service_group_members.c.analysis_service_id)
        .join(ServiceGroup, ServiceGroup.id == service_group_members.c.service_group_id)
        .where(AnalysisService.keyword.isnot(None))
    ).all()
    keyword_to_group: dict[str, tuple[int, str, str]] = {
        row.keyword: (row.id, row.name, row.color) for row in group_rows
    }

    # Step 3b: Build keyword → local enrichment map (peptide name + method)
    # AnalysisService.keyword → (peptide_name, method_name)
    # For identity services (ID_*), peptide_name is on the AnalysisService directly.
    # For slot services (ANALYTE-N-*), peptide_name is None — resolved per-sample from AnalyteNPeptide.
    # Method comes from the peptide → peptide_methods → HplcMethod chain.
    svc_rows = db.execute(
        select(AnalysisService.keyword, AnalysisService.peptide_name)
        .where(AnalysisService.keyword.isnot(None))
    ).all()
    keyword_to_peptide: dict[str, str | None] = {
        row.keyword: row.peptide_name for row in svc_rows
    }

    # Build peptide_name → method_name map from peptide_methods
    from sqlalchemy.orm import joinedload as _jl
    peptide_rows = db.execute(
        select(Peptide)
        .options(_jl(Peptide.methods))
        .where(Peptide.active == True)  # noqa: E712
    ).unique().scalars().all()
    peptide_to_method: dict[str, str] = {}
    for pep in peptide_rows:
        if pep.methods:
            # Use first active method as the display method
            peptide_to_method[pep.name] = pep.methods[0].name

    # Default group for unmatched analyses
    default_group_row = db.execute(
        select(ServiceGroup).where(ServiceGroup.is_default == True)  # noqa: E712
    ).scalar_one_or_none()
    default_group = (
        (default_group_row.id, default_group_row.name, default_group_row.color)
        if default_group_row
        else (0, "Other", "gray")
    )

    # Step 4: Load local priorities for these samples
    uids = [str(it.get("uid", "")) for it in filtered_items if it.get("uid")]
    priority_rows = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid.in_(uids))
    ).scalars().all()
    priority_map: dict[str, str] = {row.sample_uid: row.priority for row in priority_rows}

    # Step 4b: Apply order-level priority for samples without a manual override
    # WP can send priority on the order payload — auto-set for samples still at "normal"
    for it in filtered_items:
        uid = str(it.get("uid", ""))
        senaite_id = str(it.get("id", ""))
        if uid and senaite_id and uid not in priority_map and senaite_id in order_priority_map:
            order_prio = order_priority_map[senaite_id]
            # Persist so it shows up immediately and survives page reloads
            existing = db.execute(
                select(SamplePriority).where(SamplePriority.sample_uid == uid)
            ).scalar_one_or_none()
            if not existing:
                db.add(SamplePriority(sample_uid=uid, priority=order_prio))
                priority_map[uid] = order_prio
            # If existing and still "normal", upgrade to order priority
            elif existing.priority == "normal":
                existing.priority = order_prio
                priority_map[uid] = order_prio
    if order_priority_map:
        db.commit()

    # Step 4c: Load vial metadata (assignment_role, parent linkage, vial_sequence)
    # per item.uid. Parents come from lims_samples; sub-samples from lims_sub_samples.
    # The structure: vial_meta_by_uid[external_lims_uid] = dict(...).
    # vial_total for each family lets the frontend render "vial K of N" — derived
    # from a parent + its lims_sub_samples count.
    vial_meta_by_uid: dict[str, dict] = {}
    # Bulk-fetch parents that match the inbox items
    parent_rows = db.execute(
        select(
            LimsSample.id,
            LimsSample.external_lims_uid,
            LimsSample.sample_id,
            LimsSample.assignment_role,
            LimsSample.container_mode,
        ).where(LimsSample.external_lims_uid.in_(uids))
    ).all()
    parent_id_to_sample_id: dict[int, str] = {r.id: r.sample_id for r in parent_rows}
    parent_container_mode: dict[int, bool] = {r.id: r.container_mode for r in parent_rows}
    for r in parent_rows:
        vial_meta_by_uid[r.external_lims_uid] = {
            "sample_id": r.sample_id,
            "is_parent": True,
            "parent_sample_id": r.sample_id,
            "parent_lims_id": r.id,
            "assignment_role": r.assignment_role,
            "vial_sequence": 0,
            "container_mode": r.container_mode,
        }
    # Sub-samples: include both subs whose UID is in this fetch (direct hit on the
    # inbox set) AND any sub of a parent we just loaded (to compute vial_total).
    parent_ids = list(parent_id_to_sample_id.keys())
    sub_rows = db.execute(
        select(*INBOX_SUB_SAMPLE_COLUMNS).where(
            (LimsSubSample.external_lims_uid.in_(uids))
            | (LimsSubSample.parent_sample_pk.in_(parent_ids) if parent_ids else False)
        )
    ).all()

    # Sub-samples whose PARENT isn't already in parent_id_to_sample_id
    # (parent has moved past sample_received, but its sub is still inbox-eligible).
    # Without this lookup, parent_sample_id would come back empty on those subs.
    missing_parent_ids = {
        r.parent_sample_pk for r in sub_rows
        if r.parent_sample_pk not in parent_id_to_sample_id
        and r.external_lims_uid in uids
    }
    all_parent_rows = list(parent_rows)
    if missing_parent_ids:
        extra_parents = db.execute(
            select(LimsSample.id, LimsSample.sample_id, LimsSample.container_mode).where(
                LimsSample.id.in_(missing_parent_ids)
            )
        ).all()
        for r in extra_parents:
            parent_id_to_sample_id[r.id] = r.sample_id
            parent_container_mode[r.id] = r.container_mode
        all_parent_rows += list(extra_parents)
    family_sizes = _inbox_family_sizes(all_parent_rows, sub_rows)
    # Native vials (mk1:// uid — no SENAITE AR) per parent pk, for the
    # container-parent suppression branch in step 7. AR-backed subs are NOT
    # collected here; they arrive through their own SENAITE loop items.
    native_subs_by_parent: dict[int, list] = {}
    for r in sub_rows:
        if r.external_lims_uid and r.external_lims_uid.startswith("mk1://"):
            native_subs_by_parent.setdefault(r.parent_sample_pk, []).append(r)
    for subs in native_subs_by_parent.values():
        subs.sort(key=lambda r: r.vial_sequence)
    for r in sub_rows:
        # Only inbox items keyed by external_lims_uid get vial_meta entries
        # (family-size counting lives in _inbox_family_sizes).
        if r.external_lims_uid in uids and r.external_lims_uid not in vial_meta_by_uid:
            parent_sid = parent_id_to_sample_id.get(r.parent_sample_pk, "")
            vial_meta_by_uid[r.external_lims_uid] = {
                "sample_id": r.sample_id,
                "is_parent": False,
                "parent_sample_id": parent_sid,
                "parent_lims_id": r.parent_sample_pk,
                "sub_sample_pk": r.id,    # Phase 3.5: needed for Mk1 inbox source
                "assignment_role": r.assignment_role,
                "assignment_kind": r.assignment_kind,
                "vial_sequence": r.vial_sequence,
                "container_mode": parent_container_mode.get(r.parent_sample_pk, False),
            }

    # Step 5: Load per-group worksheet_item assignments (analyst + instrument)
    # Key is (sample_uid, service_group_id) → WorksheetItem
    item_rows = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid.in_(uids),
            Worksheet.status == "staging",
        )
    ).scalars().all()
    assignment_map: dict[tuple[str, int | None], WorksheetItem] = {
        (row.sample_uid, row.service_group_id): row for row in item_rows
    }

    # Collect analyst IDs to load emails
    analyst_ids = {row.assigned_analyst_id for row in item_rows if row.assigned_analyst_id}
    analyst_map: dict[int, str] = {}
    if analyst_ids:
        analyst_users = db.execute(
            select(User.id, User.email).where(User.id.in_(analyst_ids))
        ).all()
        analyst_map = {row.id: row.email for row in analyst_users}

    # Step 6: Batch-fetch analyses for all samples via the Analysis endpoint
    # The AnalysisRequest endpoint only returns analysis references ({url, uid, api_url}),
    # not full analysis objects. We need the Analysis endpoint with getRequestID filter
    # to get title, keyword, method, review_state — same approach as sample details page.
    sample_id_list = [str(it.get("id", "")) for it in filtered_items if it.get("id")]
    analyses_by_sample: dict[str, list[dict]] = {sid: [] for sid in sample_id_list}

    if sample_id_list:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            # Fetch analyses for each sample — SENAITE doesn't support multi-ID filter
            # so we do per-sample fetches, but with a shared client connection
            for sid in sample_id_list:
                try:
                    an_resp = await client.get(
                        f"{SENAITE_URL}/senaite/@@API/senaite/v1/Analysis",
                        params={"getRequestID": sid, "complete": "yes", "limit": "50"},
                    )
                    an_resp.raise_for_status()
                    an_data = an_resp.json()
                    analyses_by_sample[sid] = an_data.get("items", [])
                except Exception as fetch_err:
                    print(f"[WARN] Failed to fetch analyses for {sid}: {fetch_err}")

    # Step 7: Build vial-level inbox items.
    # Per-vial filters applied in order:
    #   * vial_meta exists in lims_samples / lims_sub_samples (else skip)
    #   * vial.assignment_role ∈ allowed_vial_roles (else skip)
    #   * not fully claimed by a null-group worksheet item (else skip)
    #   * after analysis filtering, at least one analysis remains (else skip)
    result_items: list[InboxVialItem] = []

    EXCLUDED_STATES = {"rejected", "retracted", "cancelled"}

    for it in filtered_items:
        uid = str(it.get("uid", ""))
        sample_id = str(it.get("id", ""))

        # Look up vial metadata (parent or sub) loaded in step 4c
        vial_meta = vial_meta_by_uid.get(uid)
        if vial_meta is None:
            # Legacy fallback: SENAITE knows this sample but Mk1 has no
            # lims_samples / lims_sub_samples row for it. This is the dominant
            # shape on production deploys where lims_samples didn't exist
            # before the sub-samples feature shipped. Without this fallback the
            # inbox would be empty until a backfill populates the table.
            #
            # Treat a parent-shaped id ("BW-0009", "P-0140") as a parent vial
            # with the migration-default 'hplc' role. Sub-sample-shaped ids
            # ("...-S01") are skipped — sub-samples shouldn't exist without a
            # lims_sub_samples row, and fabricating parent linkage for them
            # would be wrong.
            if re.match(r"^.+-S\d{2,}$", sample_id):
                continue
            vial_meta = {
                "sample_id": sample_id,
                "is_parent": True,
                "parent_sample_id": sample_id,
                "parent_lims_id": None,
                "assignment_role": "hplc",
                "vial_sequence": 0,
            }

        # Native (mk1://) vials emit whenever a parent has them, regardless of
        # container_mode. A CONTAINER parent is a pure depository, so its own
        # row is additionally suppressed (depository branch below — unchanged).
        # A non-container (legacy / already-received) parent is itself vial 1,
        # so its native vials are added here AND its own row falls through to
        # emit too (the after-the-fact-add case). AR-backed vials always arrive
        # via their own SENAITE loop items, never here. Spec:
        # docs/superpowers/specs/2026-06-11-vial-level-worksheets-inbox-design.md
        if vial_meta.get("is_parent") and native_subs_by_parent.get(
            vial_meta.get("parent_lims_id")
        ):
            result_items.extend(_build_native_vial_inbox_items(
                db,
                parent_item=it,
                parent_sample_id=vial_meta["parent_sample_id"],
                native_subs=native_subs_by_parent.get(vial_meta["parent_lims_id"], []),
                family_size=family_sizes.get(vial_meta["parent_lims_id"], 1),
                allowed_vial_roles=allowed_vial_roles,
                assigned_pairs=assigned_pairs,
                assigned_uids_for_null_group=assigned_uids_for_null_group,
                hide_prepped=hide_prepped,
                prepped_sub_pks=prepped_sub_pks,
                prepped_senaite_ids=prepped_senaite_ids,
                priority_map=priority_map,
                order_priority=order_priority_map.get(sample_id),
                assignment_map=assignment_map,
                keyword_to_peptide=keyword_to_peptide,
                container_mode=bool(vial_meta.get("container_mode")),
            ))

        # Depository suppression: a CONTAINER parent with any vials is a pure
        # depository — suppress its own row. Gated on container_mode (NOT on
        # whether the vials are native), so an AR-only container family stays
        # suppressed exactly as before. A zero-vial container family keeps its
        # row — its only inbox handle until the Receive Wizard registers vials.
        if (
            vial_meta.get("is_parent")
            and vial_meta.get("container_mode")
            and family_sizes.get(vial_meta.get("parent_lims_id"), 0) > 0
        ):
            continue

        vial_role = vial_meta["assignment_role"]
        if vial_role not in allowed_vial_roles:
            continue

        if uid in assigned_uids_for_null_group:
            # Sample with null service_group_id in an open worksheet — fully claimed
            continue

        raw_analyses = analyses_by_sample.get(sample_id, [])

        # Build slot → peptide name map for "Analyte N" title renaming
        analyte_name_map: dict[int, str] = {}
        for slot, key in enumerate(
            ("Analyte1Peptide", "Analyte2Peptide", "Analyte3Peptide", "Analyte4Peptide"),
            start=1,
        ):
            raw_name = it.get(key)
            if raw_name and str(raw_name).strip():
                stripped = re.sub(r"\s*-\s*[^-]+\([^)]+\)\s*$", "", str(raw_name)).strip()
                analyte_name_map[slot] = stripped

        # Dedup analyses: skip excluded states, prefer retests over originals on shared keywords.
        seen_keywords: dict[str, dict] = {}
        for analysis in raw_analyses:
            if not isinstance(analysis, dict):
                continue
            a_state = analysis.get("review_state") or analysis.get("getReviewState") or ""
            if a_state in EXCLUDED_STATES:
                continue
            kw = analysis.get("getKeyword") or analysis.get("keyword") or ""
            if kw and kw in seen_keywords:
                is_retest = bool(analysis.get("RetestOf") or analysis.get("getRetestOf"))
                if is_retest:
                    seen_keywords[kw] = analysis
            else:
                seen_keywords[kw] = analysis
        deduped_analyses = list(seen_keywords.values())
        for analysis in raw_analyses:
            if not isinstance(analysis, dict):
                continue
            a_state = analysis.get("review_state") or analysis.get("getReviewState") or ""
            if a_state in EXCLUDED_STATES:
                continue
            kw = analysis.get("getKeyword") or analysis.get("keyword") or ""
            if not kw:
                deduped_analyses.append(analysis)

        # Build flat InboxAnalysisItem list with role + assigned-group filters applied inline.
        flat_analyses: list[InboxAnalysisItem] = []
        for analysis in deduped_analyses:
            if not isinstance(analysis, dict):
                continue
            keyword = analysis.get("getKeyword") or analysis.get("keyword") or ""
            title = analysis.get("title") or analysis.get("getTitle") or keyword or ""

            analyte_match = re.match(r"^Analyte\s+(\d)\s*(.*)", title, re.IGNORECASE)
            if analyte_match:
                slot_num = int(analyte_match.group(1))
                suffix = analyte_match.group(2) or ""
                peptide_name = analyte_name_map.get(slot_num)
                if peptide_name:
                    title = f"{peptide_name} {suffix}".strip()

            a_uid = analysis.get("uid") or analysis.get("UID")
            review_state = analysis.get("review_state") or analysis.get("getReviewState")

            resolved_peptide: Optional[str] = keyword_to_peptide.get(keyword)
            if not resolved_peptide and analyte_match:
                slot_num = int(analyte_match.group(1))
                resolved_peptide = analyte_name_map.get(slot_num)

            method: Optional[str] = None
            if resolved_peptide:
                method = peptide_to_method.get(resolved_peptide)
            if not method:
                method = analysis.get("getMethodTitle") or None
                if not method:
                    method_obj = analysis.get("Method") or analysis.get("getMethod")
                    if isinstance(method_obj, dict):
                        method = method_obj.get("title")

            group_id, group_name, group_color = keyword_to_group.get(keyword, default_group)

            # Role → allowed_group_ids filter (None == pass all)
            if allowed_group_ids is not None and group_id not in allowed_group_ids:
                continue
            # Already on an open worksheet for this (vial, group) — drop the analysis
            if (uid, group_id) in assigned_pairs:
                continue

            flat_analyses.append(
                InboxAnalysisItem(
                    uid=str(a_uid) if a_uid else None,
                    title=str(title),
                    keyword=str(keyword) if keyword else None,
                    peptide_name=resolved_peptide,
                    method=str(method) if method else None,
                    review_state=str(review_state) if review_state else None,
                    group_id=group_id,
                    group_name=group_name,
                    group_color=group_color,
                )
            )

        # Phase 3.5 (mk1-native-analyses): for sub-sample vials with seeded
        # Mk1 lims_analyses rows, use Mk1 as the source of truth for the
        # inbox view. Pre-Phase-2 sub-samples (no Mk1 rows) keep the
        # SENAITE-derived flat_analyses built above. Parent vials are
        # untouched. Role-filtering already happened in Phase 2's seeder
        # (vial only has rows matching its role) so no extra filter needed.
        if vial_meta is not None and not vial_meta.get("is_parent"):
            sub_pk = vial_meta.get("sub_sample_pk")
            if sub_pk:
                mk1_rows = _fetch_mk1_inbox_analyses_for_sub_sample(
                    db, sub_pk, vial_meta.get("assignment_role"),
                    keyword_to_peptide,
                )
                if mk1_rows:
                    flat_analyses = mk1_rows

        if not flat_analyses:
            # No analyses survive filtering — hide this vial
            continue

        # Sort by group then title for stable rendering
        flat_analyses.sort(key=lambda a: (a.group_name.lower(), a.title.lower()))

        # Assignment summary: count unique groups present that have an analyst on the staging item
        unique_groups = {a.group_id for a in flat_analyses}
        assigned_count = 0
        for gid in unique_groups:
            assignment = assignment_map.get((uid, gid))
            if assignment and assignment.assigned_analyst_id:
                assigned_count += 1
        total_groups = len(unique_groups)
        summary = (
            f"{assigned_count}/{total_groups} assigned"
            if (total_groups > 0 and assigned_count > 0)
            else ""
        )

        family_size = family_sizes.get(vial_meta["parent_lims_id"], 1)

        result_items.append(
            InboxVialItem(
                uid=uid,
                sample_id=sample_id,
                is_parent=vial_meta["is_parent"],
                parent_sample_id=vial_meta["parent_sample_id"],
                assignment_role=vial_role,
                assignment_kind=vial_meta.get("assignment_kind"),
                vial_sequence=vial_meta["vial_sequence"],
                vial_total=family_size,
                container_mode=vial_meta.get("container_mode", False),
                title=str(it.get("title", "")),
                client_id=it.get("getClientTitle") or it.get("ClientID") or None,
                client_order_number=it.get("getClientOrderNumber") or it.get("ClientOrderNumber") or None,
                date_received=it.get("getDateReceived") or it.get("DateReceived") or None,
                review_state=str(it.get("review_state", "")),
                priority=priority_map.get(uid, "normal"),
                assignment_summary=summary,
                analyses=flat_analyses,
            )
        )

    # Visual grouping is the frontend's job; here we just give it a stable sort:
    # by parent_sample_id (so same-family vials are adjacent), then is_parent first,
    # then vial_sequence ascending within the family.
    result_items.sort(key=lambda v: (v.parent_sample_id, not v.is_parent, v.vial_sequence))

    return InboxResponse(items=result_items, total=len(result_items), filter_role=role)


@app.put("/worksheets/inbox/{sample_uid}/priority")
async def update_inbox_priority(
    sample_uid: str,
    data: PriorityUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Upsert priority for a received sample. Persists in sample_priorities table."""
    valid_priorities = {"normal", "high", "expedited"}
    if data.priority not in valid_priorities:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid priority '{data.priority}'. Must be one of: {', '.join(sorted(valid_priorities))}",
        )

    existing = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid == sample_uid)
    ).scalar_one_or_none()

    if existing:
        existing.priority = data.priority
        existing.updated_at = datetime.utcnow()
    else:
        db.add(SamplePriority(sample_uid=sample_uid, priority=data.priority))

    db.commit()
    return {"sample_uid": sample_uid, "priority": data.priority}


class InboxPriorityUpdate(BaseModel):
    sample_uid: str
    priority: str


@app.put("/worksheets/inbox/priority")
async def update_inbox_priority_by_body(
    data: InboxPriorityUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Upsert priority for a received sample; sample_uid travels in the body.

    Body-based sibling of /worksheets/inbox/{sample_uid}/priority. Native
    `mk1://<hex>` UIDs can't ride in a path segment — the nginx proxy mangles
    the slashes into extra path segments -> 404 (see remove_worksheet_item_by_id).
    """
    valid_priorities = {"normal", "high", "expedited"}
    if data.priority not in valid_priorities:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid priority '{data.priority}'. Must be one of: {', '.join(sorted(valid_priorities))}",
        )

    existing = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid == data.sample_uid)
    ).scalar_one_or_none()
    if existing:
        existing.priority = data.priority
        existing.updated_at = datetime.utcnow()
    else:
        db.add(SamplePriority(sample_uid=data.sample_uid, priority=data.priority))

    db.commit()
    return {"sample_uid": data.sample_uid, "priority": data.priority}


# ── D2: bulk per-sample priority lookup ────────────────────────────────────

@app.post("/sample-priorities/lookup", response_model=SamplePriorityLookupResponse)
async def lookup_sample_priorities(
    req: SamplePriorityLookupRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Sparse bulk read of sample_priorities for the order-list SLA cell.

    Returns only rows that exist; absent UIDs are omitted (the client treats
    absence as the default 'normal', matching the tier-resolution model).
    Hard cap 500 UIDs per request — a sanity bound that more than covers the
    visible-orders page at tens-to-low-hundreds of samples.
    """
    rows = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid.in_(req.sample_uids))
    ).scalars().all()
    return SamplePriorityLookupResponse(
        items=[SamplePriorityResponseItem.model_validate(r, from_attributes=True) for r in rows]
    )


@app.get("/worksheets/users")
async def get_worksheets_users(
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Return active users for analyst assignment. Accessible to all authenticated users (not admin-only).

    LEFT JOINs slack_dm_prefs so Slack-linked users carry their profile photo
    (avatar_url); others come back null and keep the initials fallback. This
    directory is SHARED with the worksheets UI, so worksheet avatars get photos
    from the same field.
    """
    from models import SlackDmPrefs
    users = db.execute(
        select(
            User.id, User.email, User.first_name, User.last_name,
            SlackDmPrefs.slack_avatar_url,
        )
        .outerjoin(SlackDmPrefs, SlackDmPrefs.user_id == User.id)
        .where(User.is_active == True)  # noqa: E712
        .order_by(User.email)
    ).all()
    return [
        {"id": row.id, "email": row.email, "first_name": row.first_name,
         "last_name": row.last_name, "avatar_url": row.slack_avatar_url}
        for row in users
    ]


@app.put("/worksheets/inbox/bulk")
async def bulk_update_inbox(
    data: BulkInboxUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """
    Bulk update priority and/or analyst/instrument assignments for multiple inbox samples.
    Priority upserts go to sample_priorities. Analyst/instrument go to worksheet_items
    as orphan pre-assignments (picked up when worksheet is created).
    """
    if not data.sample_uids:
        raise HTTPException(status_code=400, detail="sample_uids must not be empty")

    # Upsert priorities
    if data.priority is not None:
        valid_priorities = {"normal", "high", "expedited"}
        if data.priority not in valid_priorities:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid priority '{data.priority}'. Must be one of: {', '.join(sorted(valid_priorities))}",
            )
        existing_priorities = db.execute(
            select(SamplePriority).where(SamplePriority.sample_uid.in_(data.sample_uids))
        ).scalars().all()
        existing_map = {row.sample_uid: row for row in existing_priorities}
        for uid in data.sample_uids:
            if uid in existing_map:
                existing_map[uid].priority = data.priority
                existing_map[uid].updated_at = datetime.utcnow()
            else:
                db.add(SamplePriority(sample_uid=uid, priority=data.priority))

    # Upsert analyst/instrument per service group as staging worksheet_items
    if data.analyst_id is not None or data.instrument_uid is not None:
        if data.service_group_id is None:
            raise HTTPException(
                status_code=400,
                detail="service_group_id is required when setting analyst or instrument",
            )

        # Find existing staging items keyed by (sample_uid, service_group_id)
        existing_items = db.execute(
            select(WorksheetItem)
            .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
            .where(
                WorksheetItem.sample_uid.in_(data.sample_uids),
                WorksheetItem.service_group_id == data.service_group_id,
                Worksheet.status == "staging",
            )
        ).scalars().all()
        existing_item_map = {row.sample_uid: row for row in existing_items}

        # Get or create staging worksheet
        missing_uids = [uid for uid in data.sample_uids if uid not in existing_item_map]
        staging_ws = None
        if missing_uids:
            staging_ws = db.execute(
                select(Worksheet).where(
                    Worksheet.title == "__inbox_staging__",
                    Worksheet.status == "staging",
                )
            ).scalar_one_or_none()
            if not staging_ws:
                staging_ws = Worksheet(
                    title="__inbox_staging__",
                    status="staging",
                    created_by=_current_user.id,
                )
                db.add(staging_ws)
                db.flush()

        for uid in data.sample_uids:
            if uid in existing_item_map:
                item = existing_item_map[uid]
                if data.analyst_id is not None:
                    item.assigned_analyst_id = data.analyst_id
                if data.instrument_uid is not None:
                    item.instrument_uid = data.instrument_uid
            else:
                db.add(WorksheetItem(
                    worksheet_id=staging_ws.id,
                    sample_uid=uid,
                    sample_id=uid,
                    service_group_id=data.service_group_id,
                    assigned_analyst_id=data.analyst_id,
                    instrument_uid=data.instrument_uid,
                ))

    db.commit()
    return {"updated": len(data.sample_uids)}


@app.post("/worksheets", status_code=201)
async def create_worksheet(
    data: WorksheetCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Create a new worksheet from selected inbox samples.

    Performs a stale data guard: verifies each sample is still in sample_received
    state in SENAITE before creating the worksheet. Returns HTTP 409 if any samples
    have changed state.
    """
    if not data.sample_uids:
        raise HTTPException(status_code=400, detail="sample_uids must not be empty")

    if not SENAITE_URL:
        raise HTTPException(status_code=503, detail="SENAITE not configured")

    # Stale data guard: verify all samples still in sample_received state (INBX-10)
    stale_uids: list[str] = []
    try:
        async with httpx.AsyncClient(verify=HTTPX_SSL_CONTEXT, 
            timeout=SENAITE_TIMEOUT,
            auth=_get_senaite_auth(current_user),
            follow_redirects=True,
        ) as client:
            for uid in data.sample_uids:
                try:
                    verify_resp = await client.get(
                        f"{SENAITE_URL}/senaite/@@API/senaite/v1/AnalysisRequest/{uid}",
                        params={"complete": "no"},
                    )
                    verify_resp.raise_for_status()
                    sample_data = verify_resp.json()
                    # Handle both direct item and items-list responses
                    item_data = sample_data
                    if "items" in sample_data and sample_data["items"]:
                        item_data = sample_data["items"][0]
                    current_state = item_data.get("review_state", "")
                    if current_state != "sample_received":
                        stale_uids.append(uid)
                except Exception as verify_err:
                    print(f"[WARN] Could not verify state for {uid}: {verify_err}")
                    stale_uids.append(uid)
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="SENAITE request timed out during state verification")

    if stale_uids:
        raise HTTPException(
            status_code=409,
            detail={
                "stale_uids": stale_uids,
                "message": f"{len(stale_uids)} samples have changed state",
            },
        )

    # Load priorities for these samples
    priority_rows = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid.in_(data.sample_uids))
    ).scalars().all()
    priority_map: dict[str, str] = {row.sample_uid: row.priority for row in priority_rows}

    # Load sample IDs from any existing pre-assignment worksheet_items
    existing_items = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid.in_(data.sample_uids),
            Worksheet.status == "open",
        )
    ).scalars().all()
    existing_item_map: dict[str, WorksheetItem] = {row.sample_uid: row for row in existing_items}

    # Create the worksheet
    ws = Worksheet(
        title=data.title,
        status="open",
        notes=data.notes,
        created_by=current_user.id,
    )
    db.add(ws)
    db.flush()  # Get ID before adding items

    # Create worksheet items
    for uid in data.sample_uids:
        existing = existing_item_map.get(uid)
        db.add(WorksheetItem(
            worksheet_id=ws.id,
            sample_uid=uid,
            sample_id=existing.sample_id if existing else uid,
            priority=priority_map.get(uid, "normal"),
            assigned_analyst_id=existing.assigned_analyst_id if existing else None,
            instrument_uid=existing.instrument_uid if existing else None,
        ))

    db.commit()
    db.refresh(ws)

    # Notify integration service for each sample — order status → analyzing
    items = db.execute(
        select(WorksheetItem.sample_id).where(WorksheetItem.worksheet_id == ws.id)
    ).scalars().all()
    for sid in items:
        if sid:
            await _notify_worksheet_assigned(_worksheet_notify_target(db, sid))

    return {
        "id": ws.id,
        "title": ws.title,
        "status": ws.status,
        "item_count": len(data.sample_uids),
    }


@app.get("/worksheets")
def list_worksheets(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """List worksheets with summary. Excludes staging worksheets.

    Deliberately sync (`def`, not `async def`): the body is ~2.5s of pure
    synchronous DB work with zero awaits, and the page fires it twice — as
    `async def` it ran ON the event loop and froze every other request
    behind it (a 32ms flag GET measured 5.3s during two in-flight calls,
    prod probe 2026-07-07). Sync endpoints run in the threadpool. Keep it
    `def` unless the body becomes genuinely async end-to-end.
    """
    query = (
        select(Worksheet)
        .where(Worksheet.status != "staging")
        .order_by(Worksheet.created_at.desc())
    )
    if status:
        query = query.where(Worksheet.status == status)

    worksheets = db.execute(query).scalars().all()
    result = []
    for ws in worksheets:
        items = db.execute(
            select(WorksheetItem)
            .where(WorksheetItem.worksheet_id == ws.id)
            .order_by(WorksheetItem.sort_order, WorksheetItem.id)
        ).scalars().all()

        # Resolve service group names and peptide IDs for display
        group_ids = {it.service_group_id for it in items if it.service_group_id}
        group_name_map: dict[int, str] = {}
        group_peptide_map: dict[int, int | None] = {}
        if group_ids:
            groups = db.execute(
                select(ServiceGroup.id, ServiceGroup.name, ServiceGroup.color).where(ServiceGroup.id.in_(group_ids))
            ).all()
            group_name_map = {g.id: g.name for g in groups}
            group_color_map: dict[int, str] = {g.id: g.color for g in groups}
            # Resolve peptide_id and analyses per group
            group_analyses_map: dict[int, list[dict]] = {}
            for gid in group_ids:
                svcs = db.execute(
                    select(AnalysisService)
                    .join(service_group_members, service_group_members.c.analysis_service_id == AnalysisService.id)
                    .where(service_group_members.c.service_group_id == gid)
                ).scalars().all()
                analyses = []
                first_peptide_id = None
                for s in svcs:
                    if s.peptide_id and not first_peptide_id:
                        first_peptide_id = s.peptide_id
                    # Extract method name from methods JSON if available
                    method_name = None
                    if s.methods and isinstance(s.methods, list) and len(s.methods) > 0:
                        method_name = s.methods[0].get("title") if isinstance(s.methods[0], dict) else None
                    analyses.append({
                        "title": s.title,
                        "keyword": s.keyword,
                        "peptide_name": s.peptide_name,
                        "method": method_name,
                    })
                group_peptide_map[gid] = first_peptide_id
                group_analyses_map[gid] = analyses

        # Resolve assigned analyst email
        analyst_email = None
        if ws.assigned_analyst_id:
            analyst_user = db.execute(
                select(User.email).where(User.id == ws.assigned_analyst_id)
            ).scalar_one_or_none()
            analyst_email = analyst_user

        # Pre-resolve instrument senaite_uid → instrument.id for method lookup
        instrument_uids = {it.instrument_uid for it in items if it.instrument_uid}
        inst_uid_to_id: dict[str, int] = {}
        if instrument_uids:
            inst_rows = db.execute(
                select(Instrument.senaite_uid, Instrument.id).where(Instrument.senaite_uid.in_(instrument_uids))
            ).all()
            inst_uid_to_id = {r.senaite_uid: r.id for r in inst_rows if r.senaite_uid}

        # Resolve per-item analyst emails
        item_analyst_ids = {it.assigned_analyst_id for it in items if it.assigned_analyst_id}
        item_analyst_email_map: dict[int, str] = {}
        if item_analyst_ids:
            item_analyst_users = db.execute(
                select(User.id, User.email).where(User.id.in_(item_analyst_ids))
            ).all()
            item_analyst_email_map = {u.id: u.email for u in item_analyst_users}

        # Resolve each item's sub-sample (vial) pk so a worksheet "Start Prep"
        # can tag the wizard session as vial-scoped. Join on sample_id
        # (P-XXXX-SNN) — naturally null for parent-sample ids (P-XXXX), which
        # have no lims_sub_samples row. Additive; parents stay unaffected.
        item_sample_ids = {it.sample_id for it in items if it.sample_id}
        sub_sample_pk_map: dict[str, int] = {}
        sub_kind_map: dict[str, Optional[str]] = {}
        sub_box_id_map: dict[str, Optional[int]] = {}
        if item_sample_ids:
            sub_rows = db.execute(
                select(
                    LimsSubSample.sample_id,
                    LimsSubSample.id,
                    LimsSubSample.assignment_kind,  # variance badge passthrough
                    LimsSubSample.box_id,  # current physical box, if any
                ).where(
                    LimsSubSample.sample_id.in_(item_sample_ids)
                )
            ).all()
            sub_sample_pk_map = {r.sample_id: r.id for r in sub_rows}
            sub_kind_map = {r.sample_id: r.assignment_kind for r in sub_rows}
            sub_box_id_map = {r.sample_id: r.box_id for r in sub_rows}

        # Resolve current box labels for boxed vials so techs know which
        # physical box to grab. None for parent-sample items / unboxed vials.
        box_label_map: dict[int, str] = {}
        boxed_ids = {b for b in sub_box_id_map.values() if b}
        if boxed_ids:
            box_rows = db.execute(
                select(LimsBox).where(LimsBox.id.in_(boxed_ids))
            ).scalars().all()
            box_label_map = {b.id: box_label_code(b) for b in box_rows}

        def _resolve_method(it_instrument_uid: str | None, it_service_group_id: int | None) -> str | None:
            """Resolve HPLC method name from instrument + peptide (via service group)."""
            if not it_instrument_uid or not it_service_group_id:
                return None
            inst_id = inst_uid_to_id.get(it_instrument_uid)
            peptide_id = group_peptide_map.get(it_service_group_id)
            if not inst_id or not peptide_id:
                return None
            method = db.execute(
                select(HplcMethod.name)
                .join(peptide_methods, peptide_methods.c.method_id == HplcMethod.id)
                .join(instrument_methods, instrument_methods.c.method_id == HplcMethod.id)
                .where(peptide_methods.c.peptide_id == peptide_id)
                .where(instrument_methods.c.instrument_id == inst_id)
                .limit(1)
            ).scalar_one_or_none()
            return method

        result.append({
            "id": ws.id,
            "title": ws.title,
            "status": ws.status,
            "notes": ws.notes,
            "assigned_analyst": ws.assigned_analyst_id,
            "assigned_analyst_email": analyst_email,
            "item_count": len(items),
            "created_at": (ws.created_at.isoformat() + "Z") if ws.created_at else None,
            "completed_at": (ws.completed_at.isoformat() + "Z") if ws.completed_at else None,
            "items": [
                {
                    "id": it.id,
                    "sample_id": it.sample_id,
                    "sample_uid": it.sample_uid,
                    "service_group_id": it.service_group_id,
                    "group_name": group_name_map.get(it.service_group_id, "—") if it.service_group_id else "—",
                    "group_color": group_color_map.get(it.service_group_id, "zinc") if it.service_group_id else "zinc",
                    "priority": it.priority,
                    "added_at": (it.added_at.isoformat() + "Z") if it.added_at else None,
                    "date_received": (it.date_received.isoformat() + "Z") if it.date_received else None,
                    "instrument_uid": it.instrument_uid,
                    "assigned_analyst_id": it.assigned_analyst_id,
                    "assigned_analyst_email": item_analyst_email_map.get(it.assigned_analyst_id) if it.assigned_analyst_id else None,
                    "notes": it.notes,
                    "peptide_id": group_peptide_map.get(it.service_group_id) if it.service_group_id else None,
                    "method_name": _resolve_method(it.instrument_uid, it.service_group_id),
                    "lims_sub_sample_pk": sub_sample_pk_map.get(it.sample_id),
                    # 'core' | 'variance' | None — None for parent-sample ids
                    # (no lims_sub_samples row), same join as lims_sub_sample_pk
                    "assignment_kind": sub_kind_map.get(it.sample_id),
                    "box_id": sub_box_id_map.get(it.sample_id),
                    "box_label": box_label_map.get(sub_box_id_map.get(it.sample_id)),
                    "analyses": json.loads(it.analyses_json) if it.analyses_json else (group_analyses_map.get(it.service_group_id, []) if it.service_group_id else []),
                    "prep_status": it.prep_status,
                }
                for it in items
            ],
        })
    return result


class WorksheetUpdate(BaseModel):
    title: Optional[str] = None
    assigned_analyst: Optional[int] = None
    notes: Optional[str] = None


@app.put("/worksheets/{worksheet_id}")
async def update_worksheet(
    worksheet_id: int,
    data: WorksheetUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Update worksheet title and/or assigned analyst."""
    ws = db.execute(
        select(Worksheet).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    if not ws:
        raise HTTPException(404, "Worksheet not found")

    if data.title is not None:
        ws.title = data.title
    if data.assigned_analyst is not None:
        # null = not-provided (PATCH semantics, handled by the `is not None`
        # gate); a FE-sent 0 means UNASSIGN — coerce to None so we clear the
        # stamp rather than FK-exploding on user_id=0.
        ws.assigned_analyst_id = data.assigned_analyst or None
        # Also reassign all items in this worksheet to the new analyst
        items = db.execute(
            select(WorksheetItem).where(WorksheetItem.worksheet_id == worksheet_id)
        ).scalars().all()
        for item in items:
            item.assigned_analyst_id = data.assigned_analyst or None
        # Analyst-from-worksheet: re-stamp vial-tier analyses to the new analyst.
        # Best-effort: stamping failures must not break the worksheet update.
        from lims_analyses.worksheet_analyst import restamp_for_worksheet
        import logging as _logging
        try:
            restamp_for_worksheet(
                db, worksheet=ws, acting_user_id=getattr(_current_user, "id", None)
            )
        except Exception:
            _logging.getLogger(__name__).warning(
                "analyst restamp failed during worksheet update", exc_info=True
            )
    if data.notes is not None:
        ws.notes = data.notes

    db.commit()
    return {"status": "updated"}


class AddToWorksheetAnalysis(BaseModel):
    title: str
    keyword: Optional[str] = None
    peptide_name: Optional[str] = None
    method: Optional[str] = None


class AddToWorksheetRequest(BaseModel):
    sample_uid: str
    sample_id: str
    service_group_id: int | None = None
    date_received: Optional[str] = None
    analyses: Optional[list[AddToWorksheetAnalysis]] = None

    @validator("service_group_id", pre=True, always=True)
    def zero_to_none(cls, v):
        return None if v == 0 else v


@app.post("/worksheets/{worksheet_id}/add-group")
async def add_group_to_worksheet(
    worksheet_id: int,
    data: AddToWorksheetRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Add a service group (from a sample) to a worksheet. Used by drag-and-drop."""
    ws = db.execute(
        select(Worksheet).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    if not ws:
        raise HTTPException(404, "Worksheet not found")

    # Check if this sample+group is already in ANY open worksheet (collision guard)
    gid = data.service_group_id
    gid_filter = WorksheetItem.service_group_id.is_(None) if gid is None else (WorksheetItem.service_group_id == gid)
    existing_anywhere = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid == data.sample_uid,
            gid_filter,
            Worksheet.status == "open",
        )
    ).scalar_one_or_none()
    if existing_anywhere:
        owner_ws = db.execute(
            select(Worksheet.title).where(Worksheet.id == existing_anywhere.worksheet_id)
        ).scalar_one_or_none()
        if existing_anywhere.worksheet_id == worksheet_id:
            return {"status": "already_exists", "item_id": existing_anywhere.id}
        raise HTTPException(
            409,
            detail=f"Sample {data.sample_id} is already in worksheet \"{owner_ws or 'unknown'}\"",
        )

    # Pick up any staging pre-assignments for this sample+group
    staging_item = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid == data.sample_uid,
            gid_filter,
            Worksheet.status == "staging",
        )
    ).scalar_one_or_none()

    # Look up actual priority from sample_priorities
    sample_priority = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid == data.sample_uid)
    ).scalar_one_or_none()
    priority = sample_priority.priority if sample_priority else "normal"

    # If worksheet has an assigned analyst, use that (overrides card's tech)
    analyst_id = ws.assigned_analyst_id
    if not analyst_id and staging_item:
        analyst_id = staging_item.assigned_analyst_id

    item = WorksheetItem(
        worksheet_id=worksheet_id,
        sample_uid=data.sample_uid,
        sample_id=data.sample_id,
        service_group_id=data.service_group_id,
        assigned_analyst_id=analyst_id,
        instrument_uid=staging_item.instrument_uid if staging_item else None,
        priority=priority,
        date_received=datetime.fromisoformat(data.date_received.replace("Z", "+00:00")) if data.date_received else None,
        analyses_json=json.dumps([a.model_dump() for a in data.analyses]) if data.analyses else None,
    )
    db.add(item)

    # Analyst-from-worksheet (spec 2026-06-07): stamp vial-tier analyses.
    # No-ops for parent-AR uids (resolver matches lims_sub_samples only).
    # Best-effort: stamping failures must not break the worksheet operation.
    from lims_analyses.worksheet_analyst import stamp_for_item
    import logging as _logging
    try:
        stamp_for_item(
            db,
            sample_uid=data.sample_uid,
            service_group_id=data.service_group_id,
            analyst_user_id=analyst_id,
            acting_user_id=getattr(_current_user, "id", None),
            worksheet_id=worksheet_id,
            worksheet_title=ws.title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp failed during add-group-to-worksheet", exc_info=True
        )

    # Remove staging item if picked up
    if staging_item:
        db.delete(staging_item)

    db.commit()

    # Notify integration service — order status → analyzing. Vial items
    # notify with the PARENT id (the IS can only map parent ARs to orders).
    await _notify_worksheet_assigned(_worksheet_notify_target(db, data.sample_id))

    return {"status": "added", "item_id": item.id}


@app.post("/worksheets/create-from-drop")
async def create_worksheet_from_drop(
    data: AddToWorksheetRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Create a new worksheet from a drag-and-drop action."""
    # Collision guard: check if sample+group is already in any open worksheet
    gid = data.service_group_id
    gid_filter = WorksheetItem.service_group_id.is_(None) if gid is None else (WorksheetItem.service_group_id == gid)
    existing_anywhere = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid == data.sample_uid,
            gid_filter,
            Worksheet.status == "open",
        )
    ).scalar_one_or_none()
    if existing_anywhere:
        owner_ws = db.execute(
            select(Worksheet.title).where(Worksheet.id == existing_anywhere.worksheet_id)
        ).scalar_one_or_none()
        raise HTTPException(
            409,
            detail=f"Sample {data.sample_id} is already in worksheet \"{owner_ws or 'unknown'}\"",
        )

    from datetime import datetime as _dt
    title = f"WS-{_dt.utcnow().strftime('%Y-%m-%d')}-{db.query(Worksheet).filter(Worksheet.status != 'staging').count() + 1:03d}"

    ws = Worksheet(
        title=title,
        status="open",
        created_by=current_user.id,
    )
    db.add(ws)
    db.flush()

    # Look up actual priority from sample_priorities
    sample_priority = db.execute(
        select(SamplePriority).where(SamplePriority.sample_uid == data.sample_uid)
    ).scalar_one_or_none()
    priority = sample_priority.priority if sample_priority else "normal"

    # Pick up staging pre-assignments
    gid = data.service_group_id
    gid_filter = WorksheetItem.service_group_id.is_(None) if gid is None else (WorksheetItem.service_group_id == gid)
    staging_item = db.execute(
        select(WorksheetItem)
        .join(Worksheet, WorksheetItem.worksheet_id == Worksheet.id)
        .where(
            WorksheetItem.sample_uid == data.sample_uid,
            gid_filter,
            Worksheet.status == "staging",
        )
    ).scalar_one_or_none()

    item = WorksheetItem(
        worksheet_id=ws.id,
        sample_uid=data.sample_uid,
        sample_id=data.sample_id,
        service_group_id=gid,
        assigned_analyst_id=staging_item.assigned_analyst_id if staging_item else None,
        instrument_uid=staging_item.instrument_uid if staging_item else None,
        priority=priority,
        date_received=datetime.fromisoformat(data.date_received.replace("Z", "+00:00")) if data.date_received else None,
        analyses_json=json.dumps([a.model_dump() for a in data.analyses]) if data.analyses else None,
    )
    db.add(item)

    # Analyst-from-worksheet: stamp vial-tier analyses (analyst comes from the
    # staging pre-assignment via item.assigned_analyst_id). Best-effort.
    from lims_analyses.worksheet_analyst import stamp_for_item
    import logging as _logging
    try:
        stamp_for_item(
            db,
            sample_uid=data.sample_uid,
            service_group_id=gid,
            analyst_user_id=item.assigned_analyst_id,
            acting_user_id=getattr(current_user, "id", None),
            worksheet_id=ws.id,
            worksheet_title=ws.title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp failed during create-worksheet-from-drop", exc_info=True
        )

    if staging_item:
        db.delete(staging_item)

    db.commit()

    # Notify integration service — order status → analyzing. Vial items
    # notify with the PARENT id (the IS can only map parent ARs to orders).
    await _notify_worksheet_assigned(_worksheet_notify_target(db, data.sample_id))

    return {"id": ws.id, "title": ws.title, "status": ws.status, "item_count": 1}


@app.delete("/worksheets/{worksheet_id}")
async def delete_worksheet(
    worksheet_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Delete a worksheet. Items are removed (analyses return to inbox on next poll)."""
    ws = db.execute(
        select(Worksheet).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    if not ws:
        raise HTTPException(404, "Worksheet not found")

    # Analyst-from-worksheet: deleting a worksheet returns analyses to the
    # inbox — clear their stamps, like per-item removal.
    from lims_analyses.worksheet_analyst import clear_for_item
    import logging as _logging
    acting_id = getattr(_current_user, "id", None)
    ws_items = db.execute(
        select(WorksheetItem).where(WorksheetItem.worksheet_id == worksheet_id)
    ).scalars().all()
    for ws_item in ws_items:
        try:
            clear_for_item(
                db, sample_uid=ws_item.sample_uid,
                service_group_id=ws_item.service_group_id,
                acting_user_id=acting_id, worksheet_id=worksheet_id,
                worksheet_title=ws.title,
            )
        except Exception:
            # Stamps are best-effort relative to worksheet deletion (module
            # caller contract) — the delete must complete regardless.
            _logging.getLogger(__name__).warning(
                "analyst stamp clear failed during worksheet delete", exc_info=True
            )

    # Delete items first (CASCADE should handle it, but be explicit)
    db.execute(
        WorksheetItem.__table__.delete().where(WorksheetItem.worksheet_id == worksheet_id)
    )
    db.delete(ws)
    db.commit()
    return {"status": "deleted"}


@app.delete("/worksheets/{worksheet_id}/items/{sample_uid}/{service_group_id}")
async def remove_worksheet_item(
    worksheet_id: int,
    sample_uid: str,
    service_group_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Remove a single service group item from a worksheet. Analysis returns to inbox."""
    gid = None if service_group_id == 0 else service_group_id
    gid_filter = WorksheetItem.service_group_id.is_(None) if gid is None else (WorksheetItem.service_group_id == gid)
    item = db.execute(
        select(WorksheetItem).where(
            WorksheetItem.worksheet_id == worksheet_id,
            WorksheetItem.sample_uid == sample_uid,
            gid_filter,
        )
    ).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    # Analyst-from-worksheet: clear vial-tier stamps; analysis returns to inbox.
    # Best-effort: stamping failures must not break item removal.
    from lims_analyses.worksheet_analyst import clear_for_item
    import logging as _logging
    ws_title = db.execute(
        select(Worksheet.title).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    try:
        clear_for_item(
            db,
            sample_uid=sample_uid,
            service_group_id=gid,
            acting_user_id=getattr(_current_user, "id", None),
            worksheet_id=worksheet_id,
            worksheet_title=ws_title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp clear failed during remove-worksheet-item", exc_info=True
        )

    db.delete(item)
    db.commit()
    return {"status": "removed"}


@app.delete("/worksheets/{worksheet_id}/items/{item_id}")
async def remove_worksheet_item_by_id(
    worksheet_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Remove a worksheet item by its integer id. Analysis returns to inbox.

    Preferred over the {sample_uid}/{service_group_id} sibling route above:
    Mk1-native UIDs are `mk1://<hex>`, and a slash-bearing UID placed in a path
    segment gets mangled by the nginx proxy (encoded `://` -> `%3A%2F%2F` ->
    decoded + slash-merged -> extra path segments -> no route match -> 404). The
    integer id carries no slashes. Mirrors PATCH .../items/{item_id}.
    """
    item = db.execute(
        select(WorksheetItem).where(
            WorksheetItem.id == item_id,
            WorksheetItem.worksheet_id == worksheet_id,
        )
    ).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")

    # Analyst-from-worksheet: clear vial-tier stamps; analysis returns to inbox.
    # Best-effort: stamping failures must not break item removal.
    from lims_analyses.worksheet_analyst import clear_for_item
    import logging as _logging
    ws_title = db.execute(
        select(Worksheet.title).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    try:
        clear_for_item(
            db,
            sample_uid=item.sample_uid,
            service_group_id=item.service_group_id,
            acting_user_id=getattr(_current_user, "id", None),
            worksheet_id=worksheet_id,
            worksheet_title=ws_title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp clear failed during remove-worksheet-item-by-id", exc_info=True
        )

    db.delete(item)
    db.commit()
    return {"status": "removed"}


class ReassignRequest(BaseModel):
    target_worksheet_id: int


@app.post("/worksheets/{worksheet_id}/complete")
async def complete_worksheet(
    worksheet_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Transition a worksheet from open to completed. Records who completed it and when."""
    ws = db.execute(select(Worksheet).where(Worksheet.id == worksheet_id)).scalar_one_or_none()
    if not ws:
        raise HTTPException(404, "Worksheet not found")
    if ws.status != "open":
        raise HTTPException(400, f"Worksheet is already {ws.status}")
    ws.status = "completed"
    ws.completed_by = current_user.id
    ws.completed_at = datetime.utcnow()
    db.commit()
    return {"status": "completed", "completed_by": current_user.email, "completed_at": ws.completed_at.isoformat()}


@app.post("/worksheets/{worksheet_id}/items/{sample_uid}/{service_group_id}/reassign")
async def reassign_worksheet_item(
    worksheet_id: int,
    sample_uid: str,
    service_group_id: int,
    data: ReassignRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Move a worksheet item to a different (open) worksheet."""
    gid = None if service_group_id == 0 else service_group_id
    gid_filter = WorksheetItem.service_group_id.is_(None) if gid is None else (WorksheetItem.service_group_id == gid)
    item = db.execute(
        select(WorksheetItem).where(
            WorksheetItem.worksheet_id == worksheet_id,
            WorksheetItem.sample_uid == sample_uid,
            gid_filter,
        )
    ).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    target = db.execute(
        select(Worksheet).where(Worksheet.id == data.target_worksheet_id, Worksheet.status == "open")
    ).scalar_one_or_none()
    if not target:
        raise HTTPException(404, "Target worksheet not found or not open")
    # Analyst-from-worksheet: reassign = remove from source + add to target.
    # Best-effort: stamping failures must not break the reassign. The
    # item.worksheet_id move stays OUTSIDE the try so the reassign always
    # happens; only the stamp side-effects are guarded.
    from lims_analyses.worksheet_analyst import clear_for_item, stamp_for_item
    import logging as _logging
    acting_id = getattr(_current_user, "id", None)
    src_ws_title = db.execute(
        select(Worksheet.title).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    item.worksheet_id = data.target_worksheet_id
    # Target's worksheet-level analyst wins; else keep the item's own.
    if target.assigned_analyst_id:
        item.assigned_analyst_id = target.assigned_analyst_id
    try:
        clear_for_item(
            db, sample_uid=sample_uid, service_group_id=gid,
            acting_user_id=acting_id, worksheet_id=worksheet_id,
            worksheet_title=src_ws_title,
        )
        stamp_for_item(
            db, sample_uid=sample_uid, service_group_id=gid,
            analyst_user_id=target.assigned_analyst_id or item.assigned_analyst_id,
            acting_user_id=acting_id,
            worksheet_id=target.id, worksheet_title=target.title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp failed during reassign-worksheet-item", exc_info=True
        )
    db.commit()
    return {"status": "reassigned", "target_worksheet_id": data.target_worksheet_id}


@app.post("/worksheets/{worksheet_id}/items/{item_id}/reassign")
async def reassign_worksheet_item_by_id(
    worksheet_id: int,
    item_id: int,
    data: ReassignRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Move a worksheet item to a different (open) worksheet, keyed by integer id.

    By-id sibling of the {sample_uid}/{service_group_id} reassign route — see
    remove_worksheet_item_by_id for why native `mk1://` UIDs can't ride in a
    path segment.
    """
    item = db.execute(
        select(WorksheetItem).where(
            WorksheetItem.id == item_id,
            WorksheetItem.worksheet_id == worksheet_id,
        )
    ).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Item not found")
    target = db.execute(
        select(Worksheet).where(Worksheet.id == data.target_worksheet_id, Worksheet.status == "open")
    ).scalar_one_or_none()
    if not target:
        raise HTTPException(404, "Target worksheet not found or not open")
    # Mirror the {sample_uid} reassign: item.worksheet_id move stays OUTSIDE the
    # try so the reassign always happens; only the stamp side-effects are guarded.
    from lims_analyses.worksheet_analyst import clear_for_item, stamp_for_item
    import logging as _logging
    acting_id = getattr(_current_user, "id", None)
    src_ws_title = db.execute(
        select(Worksheet.title).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    sample_uid = item.sample_uid
    gid = item.service_group_id
    item.worksheet_id = data.target_worksheet_id
    if target.assigned_analyst_id:
        item.assigned_analyst_id = target.assigned_analyst_id
    try:
        clear_for_item(
            db, sample_uid=sample_uid, service_group_id=gid,
            acting_user_id=acting_id, worksheet_id=worksheet_id,
            worksheet_title=src_ws_title,
        )
        stamp_for_item(
            db, sample_uid=sample_uid, service_group_id=gid,
            analyst_user_id=target.assigned_analyst_id or item.assigned_analyst_id,
            acting_user_id=acting_id,
            worksheet_id=target.id, worksheet_title=target.title,
        )
    except Exception:
        _logging.getLogger(__name__).warning(
            "analyst stamp failed during reassign-worksheet-item-by-id", exc_info=True
        )
    db.commit()
    return {"status": "reassigned", "target_worksheet_id": data.target_worksheet_id}


class WorksheetItemUpdate(BaseModel):
    instrument_uid: Optional[str] = None
    prep_status: Optional[str] = None


@app.patch("/worksheets/{worksheet_id}/items/{item_id}")
async def update_worksheet_item(
    worksheet_id: int,
    item_id: int,
    data: WorksheetItemUpdate,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Update a worksheet item's instrument assignment. Auto-resolves method when instrument changes."""
    item = db.execute(
        select(WorksheetItem).where(
            WorksheetItem.id == item_id,
            WorksheetItem.worksheet_id == worksheet_id,
        )
    ).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Worksheet item not found")

    resolved_method = None
    if data.instrument_uid is not None:
        item.instrument_uid = data.instrument_uid if data.instrument_uid else None

        # Auto-resolve method: find instrument → get its hplc_method → match peptide
        if data.instrument_uid and item.service_group_id:
            # Get peptide_id from service group's first analysis service
            svc_peptide = db.execute(
                select(AnalysisService.peptide_id)
                .join(service_group_members, service_group_members.c.analysis_service_id == AnalysisService.id)
                .where(service_group_members.c.service_group_id == item.service_group_id)
                .where(AnalysisService.peptide_id.isnot(None))
                .limit(1)
            ).scalar_one_or_none()

            if svc_peptide:
                # Find instrument by senaite_uid
                inst = db.execute(
                    select(Instrument).where(Instrument.senaite_uid == data.instrument_uid)
                ).scalar_one_or_none()
                if inst:
                    # Find method for this peptide + instrument (M2M via instrument_methods)
                    method = db.execute(
                        select(HplcMethod)
                        .join(peptide_methods, peptide_methods.c.method_id == HplcMethod.id)
                        .join(instrument_methods, instrument_methods.c.method_id == HplcMethod.id)
                        .where(peptide_methods.c.peptide_id == svc_peptide)
                        .where(instrument_methods.c.instrument_id == inst.id)
                        .limit(1)
                    ).scalar_one_or_none()
                    if method:
                        resolved_method = method.name
                        # Update analyses_json with method
                        if item.analyses_json:
                            analyses = json.loads(item.analyses_json)
                            for a in analyses:
                                if not a.get("method"):
                                    a["method"] = resolved_method
                            item.analyses_json = json.dumps(analyses)

    if data.prep_status is not None:
        allowed = {"ready", "in_progress", "complete"}
        if data.prep_status in allowed:
            item.prep_status = data.prep_status

    db.commit()
    return {"status": "updated", "item_id": item_id, "resolved_method": resolved_method}


class ReorderRequest(BaseModel):
    item_ids: list[int]  # WorksheetItem IDs in desired order


@app.put("/worksheets/{worksheet_id}/reorder")
async def reorder_worksheet_items(
    worksheet_id: int,
    data: ReorderRequest,
    db: Session = Depends(get_db),
    _current_user=Depends(get_current_user),
):
    """Reorder items within a worksheet by setting sort_order."""
    ws = db.execute(
        select(Worksheet).where(Worksheet.id == worksheet_id)
    ).scalar_one_or_none()
    if not ws:
        raise HTTPException(404, "Worksheet not found")
    for idx, item_id in enumerate(data.item_ids):
        item = db.execute(
            select(WorksheetItem).where(
                WorksheetItem.id == item_id,
                WorksheetItem.worksheet_id == worksheet_id,
            )
        ).scalar_one_or_none()
        if item:
            item.sort_order = idx
    db.commit()
    return {"status": "reordered", "count": len(data.item_ids)}


# ── Variance payload (integration-service bridge) ────────────────────
# Called server-to-server by integration-service when it regenerates an
# additional (re-branded) COA on an already-published sample. /process-additional
# re-fetches SENAITE bare, which drops the variance series and could certify a
# re-branded COA with a verdict that disagrees with the primary. This returns the
# same {variance_replicates, variance_analytes} the primary /process flow builds
# (see main.py ~9233), so the additional COA renders an identical series.

@app.get("/samples/{sample_id}/variance-payload")
def get_sample_variance_payload(
    sample_id: str,
    db: Session = Depends(get_db),
    _: None = Depends(require_internal_service_token),
):
    """Variance replicate payload for a sample, for S2S consumers.

    Internal service token required (X-Service-Token). 404 if the parent
    LimsSample does not exist; otherwise 200 with possibly-empty dicts (a sample
    that never bought variance simply yields {}, {} — the caller proceeds bare).
    """
    parent = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()
    if parent is None:
        raise HTTPException(status_code=404, detail=f"sample {sample_id} not found")

    from coa.variance_series import (
        build_variance_replicates,
        build_variance_analyte_series,
    )
    return {
        "variance_replicates": build_variance_replicates(db, parent) or {},
        "variance_analytes": build_variance_analyte_series(db, parent) or {},
    }


# ── Registry creation signal (integration-service bridge) ────────────
# Called server-to-server by integration-service immediately after it creates
# a SENAITE AR (dual-write slice 1, 2026-07-06 spec). Idempotent upsert into
# lims_samples + native-id mint; see sub_samples.service.upsert_sample_from_signal.

class RegistrySampleSignal(BaseModel):
    """IS -> Mk1 creation signal (dual-write slice 1). meta is a
    SENAITE-shaped field dict (same keys as a complete=true AR payload)."""
    sample_id: Optional[str] = None
    senaite_uid: Optional[str] = None
    meta: dict


class RegistrySampleSignalResponse(BaseModel):
    sample_id: str
    native_id: Optional[str]


@app.post("/s2s/lims-samples", response_model=RegistrySampleSignalResponse)
def s2s_upsert_lims_sample(
    req: RegistrySampleSignal,
    db: Session = Depends(get_db),
    _: None = Depends(require_internal_service_token),
):
    """Server-to-server registry upsert, called by the Integration Service
    immediately after it creates a SENAITE AR (or, for future SENAITE-free
    lines, with no SENAITE id at all). Idempotent."""
    from sub_samples.service import upsert_sample_from_signal
    row = upsert_sample_from_signal(db, req.sample_id, req.senaite_uid, req.meta)
    db.commit()
    return RegistrySampleSignalResponse(sample_id=row.sample_id, native_id=row.native_id)


# ── Registry debug (admin diagnostic) ─────────────────────────────────
# Non-mutating registry-vs-SENAITE compare for the admin debug panel
# (2026-07-07-sample-registry-debug-panel-design.md). Reads the raw
# LimsSample row directly — never ensure_sample_row / list_sub_samples /
# _reconcile_from_senaite — so drift is observable instead of auto-healed.
# Exception (Task 7, workflow state system): the analyses column's SENAITE
# fetch now ALSO feeds the passive drift observer, which heals the parent's
# shadow mirror rows in place — the LimsSample row itself stays untouched.

def _registry_origin(row) -> str:
    if row.external_lims_system == "mk1":
        return "native"
    if row.native_id:
        return "creation-signal"
    return "lazy-or-backfill"


def _build_analysis_debug_rows(db: Session, row: LimsSample, sample_id: str) -> dict:
    """Registry-debug panel's analyses column (Task 10, 2026-07-07-sample-
    registry-debug-panel-design.md amendment): per-keyword compare of current
    SENAITE parent-analysis lines vs native `lims_analyses` rows (live shadow
    + current canonical) for this parent. Read-only for the registry row and
    the comparison itself — EXCEPT for the passive drift observer (Task 7)
    scheduled LAST, after `result` is built: it heals the parent's live
    shadow row(s) in place (+ logs a `transition_kind='observed'` row) using
    the deduped current SENAITE lines this function already fetched for
    display, at zero additional SENAITE load. Scheduling it last (not right
    after the fetch) means THIS view still renders the true pre-heal drift
    for the admin to see — the shadow heals for the NEXT view. Deliberate,
    documented exception to the panel's otherwise non-mutating posture — see
    `workflow.observer`.

    Independent try/except around the SENAITE analyses-catalog fetch: this is
    a SEPARATE SENAITE call from the basic-info `fetch_parent_metadata` above
    (own failure mode), so a failure here must not blank the field diff, and
    a `senaite_error` on the basic-info side must not blank this section. On
    failure, short-circuits to an empty rows/summary (same posture as the
    `meta is None` early-return above it) rather than rendering shadow/
    canonical rows as misleading "shadow_only" — we genuinely don't know
    whether a current SENAITE line exists, so showing none beats showing a
    wrong signal."""
    from lims_analyses.parent_mirror import build_analysis_sync_rows, select_current_lines
    from models import LimsAnalysis

    try:
        items = senaite.fetch_parent_analyses(sample_id)
    except Exception as e:
        return {"rows": [], "summary": None, "error": str(e)}

    current = select_current_lines(items)  # same selection as the backfill
    senaite_map = {
        kw: {"review_state": line.get("review_state"), "result": line.get("result")}
        for kw, line in current.items()
    }

    native_rows = db.execute(
        select(LimsAnalysis).where(LimsAnalysis.lims_sample_pk == row.id)
    ).scalars().all()

    def _is_live_canonical(r) -> bool:
        # Mirrors the DB's own `uq_lims_analyses_parent_service_root` partial
        # unique index definition (database.py) exactly: at most one
        # canonical row per (parent, keyword) may have retest_of_id IS NULL
        # AND review_state NOT IN ('retracted', 'rejected') at a time.
        return r.retest_of_id is None and r.review_state not in ("retracted", "rejected")

    # Newest-wins per keyword, same "prefer live, else fallback to newest"
    # idiom as parent_mirror.py's _existing_shadow / resolve_instrument_id.
    shadow_best: dict = {}
    canonical_best: dict = {}
    for r in native_rows:
        if r.provenance == "shadow" and not r.retested:
            cur = shadow_best.get(r.keyword)
            if cur is None or r.id > cur.id:
                shadow_best[r.keyword] = r
        elif r.provenance == "canonical":
            cur = canonical_best.get(r.keyword)
            if cur is None:
                canonical_best[r.keyword] = r
            else:
                r_live, cur_live = _is_live_canonical(r), _is_live_canonical(cur)
                if (r_live and not cur_live) or (r_live == cur_live and r.id > cur.id):
                    canonical_best[r.keyword] = r

    shadow_map = {
        kw: {"mirror_review_state": r.mirror_review_state, "result": r.result_value, "title": r.title}
        for kw, r in shadow_best.items()
    }
    canonical_map = {
        kw: {"review_state": r.review_state, "result": r.result_value, "title": r.title}
        for kw, r in canonical_best.items()
    }

    # Resolve a display title for SENAITE-only keywords (no native row to
    # source one from) with a single batched lookup.
    missing_title_kws = [
        kw for kw in senaite_map if kw not in shadow_map and kw not in canonical_map
    ]
    title_lookup: dict = {}
    if missing_title_kws:
        svc_rows = db.execute(
            select(AnalysisService.keyword, AnalysisService.title)
            .where(AnalysisService.keyword.in_(missing_title_kws))
        ).all()
        title_lookup = {kw: title for kw, title in svc_rows}
    for kw, entry in senaite_map.items():
        entry["title"] = (
            shadow_map.get(kw, {}).get("title")
            or canonical_map.get(kw, {}).get("title")
            or title_lookup.get(kw) or kw
        )

    result = build_analysis_sync_rows(senaite_map, shadow_map, canonical_map)
    result["error"] = None

    # Passive drift observer (Task 7): scheduled LAST, after `result` (and
    # the `native_rows` snapshot it's built from) is already captured —
    # THIS view still shows the true pre-heal drift (the panel's diagnostic
    # purpose), and the shadow row heals for the NEXT view. `current`
    # (`select_current_lines(items)` above) is passed rather than raw
    # `items`: the raw SENAITE Analysis fetch returns every line including
    # retest-superseded ones (no review_state filter), and the observer's
    # per-keyword loop has no dedup of its own — feeding it a superseded
    # line for a keyword that already has a newer current line would heal
    # the shadow to the WRONG (stale) state. `current` is already reduced to
    # one (newest, non-superseded) line per keyword, same guarantee the
    # backfill script relies on. This function runs on a plain `def` route
    # (FastAPI's threadpool), so the observer is called inline rather than
    # via `await run_in_threadpool(...)`; it's already never-raise (own
    # session, commits/rolls back internally) so a healing failure can't
    # blank this read-only panel.
    _observe_parent_analyses_bg(
        sample_id=sample_id,
        observed=[
            {"keyword": kw, "review_state": ln.get("review_state"), "result": ln.get("result")}
            for kw, ln in current.items()
        ],
    )
    return result


def _build_sample_transitions(db: Session, row: LimsSample) -> dict:
    """Registry-debug panel's recent-transitions tail (Task 8): the last 5
    `lims_sample_transitions` rows for this parent, newest first. Pure DB
    read, no SENAITE I/O — but still wrapped in its own try/except with its
    own error surface (`transitions.error`), same independent-failure
    posture as `_build_analysis_debug_rows`'s SENAITE fetch: a failure here
    must not blank the rest of the payload, and must not be blanked by a
    basic-info or analyses failure elsewhere."""
    from models import LimsSampleTransition

    try:
        rows = db.execute(
            select(LimsSampleTransition)
            .where(LimsSampleTransition.lims_sample_pk == row.id)
            .order_by(LimsSampleTransition.occurred_at.desc(), LimsSampleTransition.id.desc())
            .limit(5)
        ).scalars().all()
    except Exception as e:
        return {"rows": [], "error": str(e)}

    return {
        "rows": [
            {
                "verb": r.verb, "from_status": r.from_status, "to_status": r.to_status,
                "source": r.source, "occurred_at": r.occurred_at.isoformat(),
            }
            for r in rows
        ],
        "error": None,
    }


def _build_registry_debug_response(db: Session, sample_id: str) -> dict:
    """Assemble the registry-debug payload. Basic-info half is read-only;
    analyses half schedules the passive drift observer (Task 7) which heals
    shadow rows + audits transitions asynchronously. Display-view only;
    never mutates lims_analyses row count."""
    row = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()

    if row is None:
        return {
            "sample_id": sample_id,
            "load": {"exists": False, "native_id": None, "external_lims_system": None,
                     "last_synced_at": None, "age_seconds": None, "reconcile_due": None},
            "linkage": None, "origin": None, "container": None,
            "fields": [], "summary": None, "vials": None,
            "verdict": None, "senaite_error": None, "raw": None,
            "analyses": None, "transitions": None,
        }

    age = None
    reconcile_due = None
    if row.last_synced_at:
        age = int((datetime.utcnow() - row.last_synced_at).total_seconds())
        reconcile_due = age > 300  # CACHE_FRESHNESS = 5 min
    load = {
        "exists": True, "native_id": row.native_id,
        "external_lims_system": row.external_lims_system,
        "last_synced_at": row.last_synced_at.isoformat() if row.last_synced_at else None,
        "age_seconds": age, "reconcile_due": reconcile_due,
    }
    container = {"container_mode": row.container_mode, "assignment_role": row.assignment_role}

    # Independent of the basic-info meta fetch below — its own try/except,
    # its own error surface (analyses.error), never blanked by nor blanking
    # the field diff.
    analyses = _build_analysis_debug_rows(db, row, sample_id)

    # Recent-transitions tail (Task 8): same independent-failure posture —
    # own try/except, own error surface, never blanked by nor blanking
    # anything else in this payload.
    transitions = _build_sample_transitions(db, row)

    meta = None
    senaite_error = None
    try:
        meta = senaite.fetch_parent_metadata(sample_id)
    except Exception as e:
        senaite_error = str(e)

    if meta is None:
        return {
            "sample_id": sample_id, "load": load,
            "linkage": {"registry_uid": row.external_lims_uid, "senaite_uid": None,
                        "status": "senaite_missing"},
            "origin": _registry_origin(row), "container": container,
            "fields": [], "summary": None, "vials": None, "verdict": None,
            "senaite_error": senaite_error,
            "raw": {"registry": _row_to_dict(row), "senaite": None},
            "analyses": analyses, "transitions": transitions,
        }

    diff = diff_registry_vs_senaite(row, meta)
    senaite_uid = meta.get("uid")
    linkage_status = ("match" if row.external_lims_uid == senaite_uid
                      else "senaite_missing" if not senaite_uid else "mismatch")

    vials = None
    try:
        local_ct = db.execute(
            select(func.count()).select_from(LimsSubSample)
            .where(LimsSubSample.parent_sample_pk == row.id)
        ).scalar_one()
        senaite_ct = len(senaite.fetch_secondaries(sample_id))
        vstatus = ("in_sync" if local_ct == senaite_ct
                   else "local_extra" if local_ct > senaite_ct else "senaite_extra")
        vials = {"local": local_ct, "senaite": senaite_ct, "status": vstatus}
    except Exception:
        vials = None

    return {
        "sample_id": sample_id, "load": load,
        "linkage": {"registry_uid": row.external_lims_uid, "senaite_uid": senaite_uid,
                    "status": linkage_status},
        "origin": _registry_origin(row), "container": container,
        "fields": diff["fields"], "summary": diff["summary"], "vials": vials,
        "verdict": {"linkage_ok": linkage_status == "match",
                    "vials_ok": (vials or {}).get("status") == "in_sync" if vials else None,
                    "drift": diff["summary"]["drift"],
                    "registry_null": diff["summary"]["registry_null"]},
        "senaite_error": None,
        "raw": {"registry": _row_to_dict(row), "senaite": meta},
        "analyses": analyses, "transitions": transitions,
    }


def _row_to_dict(row) -> dict:
    """Registry row → JSON-safe dict for the raw panel."""
    out = {}
    for col in row.__table__.columns:
        v = getattr(row, col.name)
        out[col.name] = v.isoformat() if isinstance(v, datetime) else v
    return out


@app.get("/debug/sample-registry/{sample_id}")
def get_sample_registry_debug(
    sample_id: str,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin registry diagnostic — read-only for basic info; viewing may
    heal analysis-shadow drift via the passive observer (audit-logged as
    transition_kind='observed').

    Plain `def` (not `async def`): the SENAITE calls behind this (fetch_parent_
    analyses / fetch_parent_metadata / fetch_secondaries, all `requests`-based)
    are blocking. FastAPI runs sync `def` route handlers in the threadpool so
    they no longer stall the event loop while holding this request's `db`."""
    return _build_registry_debug_response(db, sample_id)


@app.post("/debug/sample-registry/{sample_id}/refresh")
def refresh_sample_registry_debug(
    sample_id: str,
    admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin action: force a SENAITE reconcile of the registry row, then
    return the re-diffed debug payload so drift can be watched resolving.
    Distinct POST verb because it mutates.

    Plain `def` for the same reason as the GET sibling: `_refresh_parent_
    from_senaite` and `_build_registry_debug_response` both make blocking
    SENAITE calls; threadpool execution keeps them off the event loop."""
    from sub_samples.service import _refresh_parent_from_senaite
    row = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id)
    ).scalar_one_or_none()
    if row is not None:
        try:
            _refresh_parent_from_senaite(db, row)
            db.commit()
        except Exception:
            db.rollback()
    return _build_registry_debug_response(db, sample_id)


@app.get("/registry/sample/{sample_id}/details", response_model=RegistrySampleReadResult)
async def get_sample_read_from_registry(
    sample_id: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Sample-details read path: basic-info sourced registry-first (Accu-Mk1
    lims_samples) with per-field SENAITE fallback. Analyses and everything
    else come from the unchanged SENAITE lookup.

    Gated by `get_current_user` (any authenticated user), not `require_admin`
    — it's a read-only projection of data the user already sees via the
    SENAITE lookup it wraps (see spec Access-control). `current_user` is
    resolved before `db` (auth gate before any DB dependency is entered) —
    matches the sibling debug endpoints above.
    """
    from sub_samples.registry_read import registry_row_to_display, OVERLAY_FIELDS

    base = await lookup_senaite_sample(id=sample_id, no_cache=True, db=db, _current_user=current_user)
    payload = base.model_dump()

    row = db.execute(
        select(LimsSample).where(LimsSample.sample_id == sample_id.strip().upper())
    ).scalar_one_or_none()

    field_sources = {f: "senaite" for f in OVERLAY_FIELDS}
    if row is None:
        return RegistrySampleReadResult(**payload, read_source="mk1",
                                        registry_missing=True, field_sources=field_sources)

    overlay = registry_row_to_display(row)
    for field, value in overlay.items():
        # `analytes` is the one OVERLAY_FIELDS entry whose registry shape
        # ({"name", "declared_quantity"}) is NOT the response_model's typed
        # SenaiteAnalyte shape ({"raw_name", "slot_number", ...}). Overlaying
        # it verbatim would raise a Pydantic ValidationError (500) on every
        # sample with registry-populated analytes. Leave SENAITE's typed
        # analytes untouched and keep field_sources["analytes"] == "senaite",
        # which honestly reflects where the shown value came from.
        if field == "analytes":
            continue
        payload[field] = value
        field_sources[field] = "mk1"

    return RegistrySampleReadResult(**payload, read_source="mk1",
                                    registry_missing=False, field_sources=field_sources)


@app.get("/registry/samples", response_model=SenaiteSamplesResponse)
async def list_samples_from_registry(
    review_state: Optional[str] = None,
    limit: int = 50,
    b_start: int = 0,
    search: Optional[str] = None,
    search_field: Optional[str] = None,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Samples-list read sourced from the local lims_samples registry (no SENAITE
    round-trip). Live/SENAITE-only fields (analytes, current review_state) are
    refreshed per-row on the client via progressive backfill."""
    from sub_samples.registry_list import registry_rows_to_list

    stmt = select(LimsSample)
    if review_state:
        states = [s.strip() for s in review_state.split(",") if s.strip()]
        if states:
            stmt = stmt.where(LimsSample.status.in_(states))
    if search:
        s = f"%{search.strip()}%"
        if search_field == "order_number":
            stmt = stmt.where(LimsSample.client_order_number.ilike(s))
        elif search_field == "verification_code":
            # Codes are IS-owned and REPLACED on COA regeneration — resolve ids
            # against the IS DB (parity with /senaite/samples' search) so a
            # regenerated code still finds its sample. The stored column is a
            # stale-prone cache; ILIKE on it is only the IS-down fallback.
            # run_in_threadpool: sync psycopg2 inside an async handler.
            from starlette.concurrency import run_in_threadpool
            from integration_db import search_sample_ids_by_verification_code
            try:
                ids = await run_in_threadpool(
                    search_sample_ids_by_verification_code, search.strip())
                stmt = stmt.where(LimsSample.sample_id.in_(ids))
            except Exception:
                stmt = stmt.where(LimsSample.verification_code.ilike(s))
        else:
            stmt = stmt.where(LimsSample.sample_id.ilike(s))
    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    rows = db.execute(
        stmt.order_by(LimsSample.id.desc()).offset(b_start).limit(limit)
    ).scalars().all()
    items = registry_rows_to_list(rows)
    # Verification codes: overlay the ACTIVE code from the IS DB (one batched
    # query per page). The stored lims_samples copy goes stale when a COA is
    # regenerated (IS-side mutation the registry never sees); on IS-DB failure
    # the stored values stand — the page always renders.
    if items:
        from starlette.concurrency import run_in_threadpool
        from integration_db import fetch_verification_codes_for_samples
        try:
            codes = await run_in_threadpool(
                fetch_verification_codes_for_samples, [it["id"] for it in items])
            for it in items:
                it["verification_code"] = codes.get(it["id"]) or it["verification_code"]
        except Exception:
            pass
    return SenaiteSamplesResponse(items=items, total=total, b_start=b_start)


# ── Peptide requests API (integration-service bridge) ────────────────
# Called server-to-server by integration-service when a WP user submits the
# peptide-request form. Internal service token + idempotency key are required.

@app.post(
    "/peptide-requests",
    response_model=PeptideRequest,
    status_code=status.HTTP_201_CREATED,
)
def create_peptide_request(
    data: PeptideRequestCreate,
    idempotency_key: str = Header(None, alias="Idempotency-Key"),
    _: None = Depends(require_internal_service_token),
):
    if not idempotency_key:
        raise HTTPException(400, "Idempotency-Key header required")
    repo = PeptideRequestRepository()
    cfg = get_peptide_request_config()
    row = repo.create(
        data,
        idempotency_key=idempotency_key,
        clickup_list_id=cfg.clickup_list_id,
    )
    # Best-effort inline ClickUp task creation. On failure the retry job
    # (backend/jobs/clickup_task_retry.py) will pick up the row once it is
    # > 60s old. Never block the 201 response on ClickUp availability.
    if not row.clickup_task_id:
        try:
            client = ClickUpClient(
                api_token=cfg.clickup_api_token,
                list_id=cfg.clickup_list_id,
                accumk1_base_url=os.environ.get("ACCUMK1_BASE_URL", ""),
            )
            task_id = client.create_task_for_request(row)
            repo.update_clickup_task_id(row.id, task_id)
            row = repo.get_by_id(row.id)
        except Exception:
            import logging as _logging
            _logging.getLogger(__name__).exception(
                "inline clickup create failed; retry job will pick up"
            )
    return row


@app.get("/peptide-requests", response_model=PeptideRequestList)
def list_peptide_requests(
    wp_user_id: int,
    status: str | None = None,  # comma-separated
    limit: int = 50,
    offset: int = 0,
    _: None = Depends(require_internal_service_token),
):
    repo = PeptideRequestRepository()
    status_list = status.split(",") if status else None
    items, total = repo.list_by_wp_user(
        wp_user_id, status=status_list, limit=limit, offset=offset
    )
    return PeptideRequestList(total=total, limit=limit, offset=offset, items=items)


@app.get("/peptide-requests/{request_id}", response_model=PeptideRequest)
def get_peptide_request(
    request_id: str,
    _: None = Depends(require_internal_service_token),
):
    repo = PeptideRequestRepository()
    row = repo.get_by_id(UUID(request_id))
    if not row:
        raise HTTPException(404, "not found")
    return row


@app.get(
    "/peptide-requests/{request_id}/history",
    response_model=list[StatusLogEntry],
)
def get_peptide_request_history(
    request_id: str,
    _: None = Depends(require_internal_service_token),
):
    lrepo = StatusLogRepository()
    return lrepo.get_for_request(UUID(request_id))


@app.post("/peptide-requests/{request_id}/retract")
def retract_peptide_request(
    request_id: str,
    data: PeptideRequestRetract,
    _: None = Depends(require_internal_service_token),
):
    """Hard-delete a peptide request that's still in a customer-retractable state.

    Gate: status must be in {"new", "rejected"}. ClickUp comment is
    best-effort (2s timeout, failure logged but not raised). Delete is
    atomic and authoritative. Delete happens before the ClickUp comment
    so a failed delete never leaves an orphaned breadcrumb.
    """
    import logging as _logging
    log = _logging.getLogger(__name__)
    rid = UUID(request_id)
    repo = PeptideRequestRepository()
    row = repo.get_by_id(rid)
    if not row:
        raise HTTPException(
            status_code=404,
            detail={"code": "request_not_found", "message": "Peptide request not found"},
        )
    if row.status not in ("new", "rejected"):
        raise HTTPException(
            status_code=409,
            detail={
                "code": "request_not_retractable",
                "message": "This request can no longer be retracted.",
                "current_status": row.status,
            },
        )

    reason = (data.reason or "").strip()
    if len(reason) > 500:
        reason = reason[:500]

    prior = row.status
    clickup_task_id = row.clickup_task_id
    row_id = row.id
    repo.delete_by_id(rid)

    # Best-effort ClickUp breadcrumb + column move. Delete already
    # succeeded; if either ClickUp call fails we accept the "ghost card"
    # — the spec-blessed failure mode — rather than leaving a half-state
    # where the row is live but already marked retracted on ClickUp.
    # Client is built once, outside both try-blocks, so the second block
    # can reference it even if the first crashes post-build.
    if clickup_task_id:
        try:
            cfg = get_peptide_request_config()
            client = ClickUpClient(
                api_token=cfg.clickup_api_token,
                list_id=cfg.clickup_list_id,
                accumk1_base_url=os.environ.get("ACCUMK1_BASE_URL", ""),
            )
        except Exception:
            log.exception(
                "clickup_retraction_client_init_failed request_id=%s task_id=%s",
                row_id, clickup_task_id,
            )
            client = None
        if client is not None:
            try:
                from datetime import date as _date
                lines = [f"Customer retracted this request on {_date.today().isoformat()}."]
                if reason:
                    lines.append(f"Reason: {reason}")
                client.post_task_comment(clickup_task_id, "\n".join(lines))
                log.info(
                    "clickup_retraction_comment_posted request_id=%s task_id=%s",
                    row_id, clickup_task_id,
                )
            except Exception:
                log.exception(
                    "clickup_retraction_comment_failed request_id=%s task_id=%s",
                    row_id, clickup_task_id,
                )
            try:
                client.set_task_status(clickup_task_id, "retracted")
                log.info(
                    "clickup_retraction_status_moved request_id=%s task_id=%s",
                    row_id, clickup_task_id,
                )
            except Exception:
                log.exception(
                    "clickup_retraction_status_move_failed request_id=%s task_id=%s",
                    row_id, clickup_task_id,
                )
    else:
        log.warning(
            "clickup_retraction_comment_skipped_no_task_id request_id=%s", row_id,
        )

    log.info(
        "peptide_request_retracted request_id=%s prior_status=%s had_reason=%s",
        rid, prior, bool(reason),
    )
    return {"ok": True}


# ── Admin: ClickUp user mapping ──────────────────────────────────────
# Concerns (flagged, not blocking):
#   * Auth: spec called for require_admin_or_service which does not exist.
#     Using require_internal_service_token (same as peptide-request API)
#     means an admin must present the shared service token, not their user
#     bearer. UX is broken for a real admin workflow; consistent with Tasks
#     10/16 pattern. Pre-merge resolution.

@app.get("/admin/clickup-users/unmapped")
def list_unmapped_clickup_users(
    _: None = Depends(require_internal_service_token),
):
    return ClickUpUserMappingRepository().list_unmapped()


@app.post("/admin/clickup-users/{clickup_user_id}/map")
def map_clickup_user(
    clickup_user_id: str,
    accumk1_user_id: int = Body(..., embed=True),
    _: None = Depends(require_internal_service_token),
):
    ClickUpUserMappingRepository().set_mapping(clickup_user_id, accumk1_user_id)
    return {"ok": True}


# ── LIMS UI endpoints (JWT-gated) ────────────────────────────────────
# Parallel to the /api/peptide-requests and /api/admin/clickup-users routes
# above. Those are service-token-gated for integration-service (WP bridge);
# these are JWT-gated for LIMS staff in the React app. No role gating for v1
# — any authenticated user can hit admin routes. A proper role gate
# (lab_manager vs regular) is a follow-up.

@app.get("/lims/peptide-requests", response_model=PeptideRequestList)
def lims_list_peptide_requests(
    wp_user_id: int | None = None,
    status: str | None = None,
    limit: int = 50,
    offset: int = 0,
    _user=Depends(get_current_user),
):
    repo = PeptideRequestRepository()
    status_list = status.split(",") if status else None
    items, total = repo.list_all(
        wp_user_id=wp_user_id, status=status_list, limit=limit, offset=offset
    )
    return PeptideRequestList(total=total, limit=limit, offset=offset, items=items)


@app.get("/lims/peptide-requests/sync/diff")
def lims_peptide_request_sync_diff(_user=Depends(get_current_user)):
    """Compute the 3 discrepancy buckets between ClickUp and Accu-Mk1.

    Declared BEFORE /lims/peptide-requests/{request_id} because FastAPI
    matches routes in declaration order — "sync" would otherwise be
    treated as a UUID path param and 500 on UUID(...) parsing.

    Auth: any logged-in LIMS user. Sync is read-only on this endpoint
    (no side effects until the tech clicks Apply) so admin-gating would
    be unnecessarily restrictive.
    """
    cfg = get_peptide_request_config()
    client = ClickUpClient(
        api_token=cfg.clickup_api_token,
        list_id=cfg.clickup_list_id,
        accumk1_base_url=os.environ.get("ACCUMK1_BASE_URL", ""),
        config=cfg,
    )
    return peptide_request_compute_sync_diff(
        client, PeptideRequestRepository(), cfg
    )


@app.post("/lims/peptide-requests/sync/apply")
def lims_peptide_request_sync_apply(
    body: PeptideRequestSyncApplyRequest,
    _user=Depends(get_current_user),
):
    """Apply tech-selected reconciliation actions.

    Per-item error isolation lives in apply_actions; the route simply
    wires repos + client and returns the resulting counts + errors
    array so the frontend can show a toast with a breakdown.
    """
    cfg = get_peptide_request_config()
    client = ClickUpClient(
        api_token=cfg.clickup_api_token,
        list_id=cfg.clickup_list_id,
        accumk1_base_url=os.environ.get("ACCUMK1_BASE_URL", ""),
        config=cfg,
    )
    return peptide_request_apply_sync_actions(
        body.model_dump(mode="python"),
        client,
        PeptideRequestRepository(),
        StatusLogRepository(),
        ClickUpUserMappingRepository(),
        cfg,
    )


@app.get("/lims/peptide-requests/{request_id}", response_model=PeptideRequest)
def lims_get_peptide_request(
    request_id: str,
    _user=Depends(get_current_user),
):
    repo = PeptideRequestRepository()
    row = repo.get_by_id(UUID(request_id))
    if not row:
        raise HTTPException(404, "not found")
    return row


@app.patch("/lims/peptide-requests/{request_id}")
def lims_update_peptide_request(
    request_id: str,
    body: PeptideRequestUpdate,
    _user=Depends(get_current_user),
):
    """Partial update from the LIMS UI. Currently supports editing sample_id.

    Behaviour:
      * 404 if the row doesn't exist.
      * Always updates the DB column first (source of truth).
      * If the row has a clickup_task_id AND config.clickup_field_sample_id
        is populated, pushes the new value to ClickUp via
        POST /task/{id}/field/{field_id}. Empty string is pushed when
        clearing (sample_id=null).
      * ClickUp failure does NOT roll back the DB. Response returns 200
        with a `warning` field so the UI can surface a retry hint.
        Rationale: DB is the source of truth; a transient ClickUp outage
        shouldn't block a tech from editing a sample id. A future retry
        queue will close the drift window.
    """
    import logging as _logging
    log = _logging.getLogger(__name__)

    repo = PeptideRequestRepository()
    existing = repo.get_by_id(UUID(request_id))
    if not existing:
        raise HTTPException(404, "not found")

    updated = repo.update_sample_id(UUID(request_id), body.sample_id)
    if updated is None:
        # Race: deleted between the get and the update. Treat as 404.
        raise HTTPException(404, "not found")

    warning: Optional[str] = None
    if updated.clickup_task_id:
        cfg = get_peptide_request_config()
        if not cfg.clickup_field_sample_id:
            log.info(
                "PATCH sample_id: CLICKUP_FIELD_SAMPLE_ID unset, skipping ClickUp sync"
            )
        else:
            try:
                client = ClickUpClient(
                    api_token=cfg.clickup_api_token,
                    list_id=cfg.clickup_list_id,
                    accumk1_base_url=os.environ.get("ACCUMK1_BASE_URL", ""),
                    config=cfg,
                )
                client.set_custom_field(
                    updated.clickup_task_id,
                    cfg.clickup_field_sample_id,
                    body.sample_id or "",
                )
            except Exception:
                log.exception(
                    "PATCH sample_id: ClickUp sync failed for task %s",
                    updated.clickup_task_id,
                )
                warning = "ClickUp sync failed; will need manual retry"

    result = updated.model_dump(mode="json")
    if warning:
        result["warning"] = warning
    return result


@app.get(
    "/lims/peptide-requests/{request_id}/history",
    response_model=list[StatusLogEntry],
)
def lims_get_peptide_request_history(
    request_id: str,
    _user=Depends(get_current_user),
):
    lrepo = StatusLogRepository()
    return lrepo.get_for_request(UUID(request_id))


@app.get("/lims/admin/clickup-users/unmapped")
def lims_list_unmapped_clickup_users(
    _user=Depends(require_admin),
):
    return ClickUpUserMappingRepository().list_unmapped()


@app.post("/lims/admin/clickup-users/{clickup_user_id}/map")
def lims_map_clickup_user(
    clickup_user_id: str,
    accumk1_user_id: int = Body(..., embed=True),
    _user=Depends(require_admin),
):
    ClickUpUserMappingRepository().set_mapping(clickup_user_id, accumk1_user_id)
    return {"ok": True}


@app.post("/webhooks/clickup")
async def clickup_webhook(request: Request):
    raw = await request.body()
    sig = request.headers.get("X-Signature")
    cfg = get_peptide_request_config()
    if not verify_signature(raw, sig, cfg.clickup_webhook_secret):
        raise HTTPException(401, "invalid signature")
    payload = json.loads(raw)
    try:
        dispatch_event(
            payload, cfg,
            PeptideRequestRepository(),
            StatusLogRepository(),
            ClickUpUserMappingRepository(),
        )
    except Exception:
        import logging as _logging
        _logging.getLogger(__name__).exception("webhook dispatch failure")
        raise HTTPException(500, "dispatch failed")
    return {"ok": True}


# dropdowns from SENAITE LabContact records.
