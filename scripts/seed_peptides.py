"""
Seed peptides and calibration curves from the lab's HPLC folder structure.

Scans the Peptides folder for calibration Excel files, extracts concentration/area
pairs and retention times, then creates peptides + calibration curves via the API.

Idempotent: skips peptides that already exist (matched by abbreviation).

Usage:
    python scripts/seed_peptides.py [--api-url http://127.0.0.1:8009] [--dry-run]
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl
import requests

# ── Configuration ──────────────────────────────────────────────

PEPTIDES_ROOT = Path(
    r"C:\Users\forre\Valence Analytical\Communication site - Documents"
    r"\Analytical\Lab Reports\Purity and Quantity (HPLC)\Peptides"
)

# Map folder names to (full name, abbreviation)
# Only folders with actual peptide data — skip templates, blends, non-peptides
PEPTIDE_MAP: dict[str, tuple[str, str]] = {
    "AOD 9604":                     ("AOD 9604", "AOD9604"),
    "BPC157":                       ("BPC-157", "BPC157"),
    "Bremelanotide":                ("Bremelanotide (PT-141)", "PT-141"),
    "Cagrilinitide":                ("Cagrilintide", "Cagri"),
    "CJC 1295 (no DAC)":           ("CJC-1295 (no DAC)", "CJC1295"),
    "Delta Sleep Inducing Peptide (DSIP)": ("Delta Sleep Inducing Peptide", "DSIP"),
    "Epithalon":                    ("Epithalon", "Epithalon"),
    "FOX04":                        ("FOXO4-DRI", "FOX04"),
    "GHKCu":                        ("GHK-Cu", "GHK-Cu"),
    "GHRP-6 Acetate":              ("GHRP-6", "GHRP-6"),
    "Hexarelin":                    ("Hexarelin", "Hexarelin"),
    "Ipamorelin":                   ("Ipamorelin", "Ipamorelin"),
    "Kisspeptin":                   ("Kisspeptin-10", "Kisspeptin"),
    "KPV":                          ("Lysine-Proline-Valine", "KPV"),
    "Melanotan II (Acetate)":       ("Melanotan II", "MT-II"),
    "MOTSc":                        ("MOTS-c", "MOTS-c"),
    "NAD500":                       ("NAD+ 500mg", "NAD+"),
    "Oxytocin (acetate)":           ("Oxytocin", "Oxytocin"),
    "Pinealon":                     ("Pinealon", "Pinealon"),
    "Retatrutide":                  ("Retatrutide", "Retatrutide"),
    "Selank":                       ("Selank", "Selank"),
    "Semaglutide":                  ("Semaglutide", "Semaglutide"),
    "Semax":                        ("Semax", "Semax"),
    "Sermorelin":                   ("Sermorelin", "Sermorelin"),
    "TB500":                        ("Thymosin Beta-4 Fragment", "TB-500"),
    "Tesamorelin":                  ("Tesamorelin", "Tesamorelin"),
    "Thymosin Alpha 1":             ("Thymosin Alpha 1", "TA1"),
    "Tirzepatide":                  ("Tirzepatide", "Tirzepatide"),
    "VIP":                          ("Vasoactive Intestinal Peptide", "VIP"),
}

# Sheets to skip when looking for calibration data
SKIP_SHEETS = {"Dissolution method", "Dissolution Method"}
SKIP_SHEET_PREFIXES = ("SOP CalStds", "SOP_CalStds")


# ── Excel parsing ──────────────────────────────────────────────

def _is_number(v) -> bool:
    """Check if a value is a real number (not None, not string, not error)."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        # Handle "NA", "#DIV/0!", etc.
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _to_float(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v))


def try_extract_calibration(ws, max_rows: int = 12) -> dict | None:
    """
    Try to extract calibration data from a worksheet.

    Returns dict with keys: concentrations, areas, rts, format
    or None if no valid data found.
    """
    # Strategy: try multiple column layouts
    # New format (2026+): B=conc, S=area, U=RT  (cols 2, 19, 21)
    # Old format (pre-2026): B=conc, J=area, L=RT  (cols 2, 10, 12)
    # Older format: B=conc, G=area  (cols 2, 7)

    layouts = [
        ("new_S_U", 2, 19, 21),     # B, S, U
        ("old_J_L", 2, 10, 12),     # B, J, L
        ("older_B_G", 2, 7, None),  # B, G, no RT
    ]

    for fmt_name, conc_col, area_col, rt_col in layouts:
        concentrations = []
        areas = []
        rts = []

        for row in range(2, 2 + max_rows):  # Start at row 2 (skip header)
            conc_val = ws.cell(row=row, column=conc_col).value
            area_val = ws.cell(row=row, column=area_col).value

            if not _is_number(conc_val) or not _is_number(area_val):
                continue

            conc = _to_float(conc_val)
            area = _to_float(area_val)

            # Sanity: skip if either is <= 0 or unreasonably large
            if conc <= 0 or area <= 0:
                continue

            concentrations.append(conc)
            areas.append(area)

            if rt_col is not None:
                rt_val = ws.cell(row=row, column=rt_col).value
                if _is_number(rt_val) and _to_float(rt_val) > 0:
                    rts.append(_to_float(rt_val))

        # Need at least 3 valid points for a calibration curve
        if len(concentrations) >= 3:
            # Sanity check: areas should increase with concentration
            # (positive correlation). Skip if areas are flat or inverted.
            max_conc_area = areas[concentrations.index(max(concentrations))]
            min_conc_area = areas[concentrations.index(min(concentrations))]
            if max_conc_area <= min_conc_area * 1.5:
                # Areas don't increase — wrong column, skip this layout
                continue

            return {
                "concentrations": concentrations,
                "areas": areas,
                "rts": rts,
                "format": fmt_name,
                "n_points": len(concentrations),
            }

    return None


def parse_calibration_excel(filepath: Path) -> dict | None:
    """
    Open an Excel file and extract calibration data from the first valid sheet.

    Returns dict with calibration data or None if no valid data found.
    """
    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except Exception as e:
        print(f"    [SKIP] Cannot open {filepath.name}: {e}")
        return None

    for sheet_name in wb.sheetnames:
        # Skip known non-data sheets
        if sheet_name in SKIP_SHEETS:
            continue
        if any(sheet_name.startswith(p) for p in SKIP_SHEET_PREFIXES):
            continue

        ws = wb[sheet_name]
        result = try_extract_calibration(ws)
        if result:
            result["sheet"] = sheet_name
            result["filename"] = filepath.name
            wb.close()
            return result

    wb.close()
    return None


def find_calibration_files(peptide_dir: Path) -> list[Path]:
    """
    Find calibration Excel files for a peptide.
    Looks in Raw Data/**/1290/*.xlsx and any file with "Calibration" in path.
    Also checks for P-0XXX_Std_* files which are the new standard format.
    """
    results = []
    raw_data = peptide_dir / "Raw Data"

    if not raw_data.exists():
        return results

    for xlsx_path in raw_data.rglob("*.xlsx"):
        # Skip temp files
        if xlsx_path.name.startswith("~$"):
            continue

        path_str = str(xlsx_path)

        # Include if in a 1290 folder
        if "1290" in path_str:
            results.append(xlsx_path)
        # Include if path contains "Calibration"
        elif "Calibration" in path_str or "calibration" in path_str:
            results.append(xlsx_path)
        # Include if it's a P-0XXX_Std file (new standard format)
        elif "_Std_" in xlsx_path.name:
            results.append(xlsx_path)

    # Sort by filename (most recent dates tend to sort last)
    results.sort(key=lambda p: p.name)
    return results


# ── API helpers ────────────────────────────────────────────────

def get_existing_peptides(api_url: str) -> dict[str, dict]:
    """Get all existing peptides, keyed by abbreviation."""
    resp = requests.get(f"{api_url}/peptides")
    resp.raise_for_status()
    return {p["abbreviation"]: p for p in resp.json()}


def create_peptide(api_url: str, name: str, abbreviation: str,
                   reference_rt: float | None, diluent_density: float) -> dict:
    """Create a peptide via the API."""
    data = {
        "name": name,
        "abbreviation": abbreviation,
        "reference_rt": reference_rt,
        "rt_tolerance": 0.5,
        "diluent_density": diluent_density,
    }
    resp = requests.post(f"{api_url}/peptides", json=data)
    resp.raise_for_status()
    return resp.json()


def create_calibration(api_url: str, peptide_id: int,
                       concentrations: list[float], areas: list[float],
                       source_filename: str) -> dict:
    """Create a calibration curve via the API."""
    data = {
        "concentrations": concentrations,
        "areas": areas,
        "source_filename": source_filename,
    }
    resp = requests.post(f"{api_url}/peptides/{peptide_id}/calibrations", json=data)
    resp.raise_for_status()
    return resp.json()


# ── Main ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed peptides from lab HPLC folder")
    parser.add_argument("--api-url", default="http://127.0.0.1:8009",
                        help="Backend API URL (default: http://127.0.0.1:8009)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print what would be done without making API calls")
    args = parser.parse_args()

    api_url = args.api_url.rstrip("/")
    dry_run = args.dry_run

    if not PEPTIDES_ROOT.exists():
        print(f"ERROR: Peptides folder not found: {PEPTIDES_ROOT}")
        sys.exit(1)

    # Check API is reachable
    if not dry_run:
        try:
            requests.get(f"{api_url}/health", timeout=5)
        except requests.ConnectionError:
            print(f"ERROR: Cannot reach API at {api_url}")
            print("Start the backend first: cd backend && uvicorn main:app --port 8009")
            sys.exit(1)

    # Get existing peptides
    existing = {} if dry_run else get_existing_peptides(api_url)
    print(f"Existing peptides in DB: {len(existing)} ({', '.join(existing.keys()) or 'none'})\n")

    created = 0
    calibrations_added = 0
    skipped = 0
    no_data = []

    for folder_name, (full_name, abbreviation) in sorted(PEPTIDE_MAP.items()):
        folder_path = PEPTIDES_ROOT / folder_name

        if not folder_path.exists():
            print(f"[MISSING] {folder_name}/ — folder not found")
            continue

        print(f"--- {folder_name} ({abbreviation}) ---")

        # Find calibration Excel files
        cal_files = find_calibration_files(folder_path)
        if not cal_files:
            print(f"  No calibration files found")
            no_data.append(folder_name)
            # Still create the peptide (without calibration)
            if abbreviation in existing:
                print(f"  Peptide already exists (id={existing[abbreviation]['id']})")
                skipped += 1
            elif dry_run:
                print(f"  [DRY RUN] Would create peptide: {full_name} ({abbreviation})")
                created += 1
            else:
                peptide = create_peptide(api_url, full_name, abbreviation, None, 997.1)
                print(f"  Created peptide (id={peptide['id']}) — no calibration data")
                existing[abbreviation] = peptide
                created += 1
            print()
            continue

        print(f"  Found {len(cal_files)} calibration file(s)")

        # Try to parse the most recent calibration file (last in sorted list)
        best_cal = None
        for cal_file in reversed(cal_files):
            result = parse_calibration_excel(cal_file)
            if result:
                best_cal = result
                break

        if not best_cal:
            print(f"  No valid calibration data extracted from any file")
            no_data.append(folder_name)
            # Create peptide without calibration
            if abbreviation in existing:
                print(f"  Peptide already exists (id={existing[abbreviation]['id']})")
                skipped += 1
            elif dry_run:
                print(f"  [DRY RUN] Would create peptide: {full_name} ({abbreviation})")
                created += 1
            else:
                peptide = create_peptide(api_url, full_name, abbreviation, None, 997.1)
                print(f"  Created peptide (id={peptide['id']}) — no valid calibration")
                existing[abbreviation] = peptide
                created += 1
            print()
            continue

        # Extract reference RT (average of all RTs)
        ref_rt = None
        if best_cal["rts"]:
            ref_rt = round(sum(best_cal["rts"]) / len(best_cal["rts"]), 4)

        print(f"  Calibration: {best_cal['filename']} [{best_cal['sheet']}] "
              f"({best_cal['format']}, {best_cal['n_points']} points)")
        print(f"    Concs: {[round(c, 2) for c in best_cal['concentrations']]}")
        print(f"    Areas: {[round(a, 2) for a in best_cal['areas']]}")
        if ref_rt:
            print(f"    Avg RT: {ref_rt} min")

        # Create or skip peptide
        if abbreviation in existing:
            peptide_id = existing[abbreviation]["id"]
            print(f"  Peptide already exists (id={peptide_id})")
            skipped += 1

            # Still add calibration if we have data and they want to update
            # For idempotency, we skip calibration if peptide already has one
            # (user can manually add more via the UI if needed)
            print(f"  Skipping calibration (peptide already seeded)")
            print()
            continue

        if dry_run:
            print(f"  [DRY RUN] Would create peptide: {full_name} ({abbreviation})")
            print(f"  [DRY RUN] Would add calibration: {best_cal['n_points']} points")
            created += 1
            calibrations_added += 1
            print()
            continue

        # Create peptide
        peptide = create_peptide(api_url, full_name, abbreviation, ref_rt, 997.1)
        peptide_id = peptide["id"]
        existing[abbreviation] = peptide
        created += 1
        print(f"  Created peptide (id={peptide_id})")

        # Add calibration curve
        try:
            cal = create_calibration(
                api_url, peptide_id,
                best_cal["concentrations"],
                best_cal["areas"],
                best_cal["filename"],
            )
            calibrations_added += 1
            print(f"  Added calibration: slope={cal.get('slope', '?')}, "
                  f"intercept={cal.get('intercept', '?')}, "
                  f"r_squared={cal.get('r_squared', '?')}")
        except Exception as e:
            print(f"  ERROR adding calibration: {e}")

        print()

    # Summary
    print("=" * 60)
    print(f"SUMMARY")
    print(f"  Peptides created:      {created}")
    print(f"  Calibrations added:    {calibrations_added}")
    print(f"  Skipped (existing):    {skipped}")
    print(f"  No calibration data:   {len(no_data)}")
    if no_data:
        print(f"    {', '.join(no_data)}")
    print(f"  Total in DB:           {len(existing)}")


if __name__ == "__main__":
    main()
